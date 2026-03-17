# -*- coding: utf-8 -*-
"""
驗證 Liteon Forecast 日期計算邏輯 - 全 Plant 完整分析
1. 驗證 processed/6/20260317_195848 (2680 單檔結果)
2. 驗證 processed/6/20260317_200333 (全部 23 Plant 結果)
"""
import os
import sys
sys.stdout.reconfigure(encoding='utf-8')

import pandas as pd
import openpyxl
from datetime import datetime, timedelta, date
from collections import defaultdict

BASE = "d:/github/business_forecasting_lite"

# ── 日期計算邏輯 (與修正後 liteon_forecast_processor.py 一致) ──

WEEKDAY_MAP_FULL = {
    '週一': 0, '禮拜一': 0, '星期一': 0,
    '週二': 1, '禮拜二': 1, '星期二': 1,
    '週三': 2, '禮拜三': 2, '星期三': 2,
    '週四': 3, '禮拜四': 3, '星期四': 3,
    '週五': 4, '禮拜五': 4, '星期五': 4,
    '週六': 5, '禮拜六': 5, '星期六': 5,
    '週日': 6, '禮拜日': 6, '星期日': 6, '禮拜天': 6, '週天': 6,
}

WEEKDAY_MAP_CHAR = {
    '一': 0, '二': 1, '三': 2, '四': 3,
    '五': 4, '六': 5, '日': 6, '天': 6,
}

WEEKDAY_NAMES = ['一', '二', '三', '四', '五', '六', '日']


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    if hasattr(val, 'date') and callable(val.date):
        try:
            return val.date()
        except:
            return None
    if isinstance(val, str):
        val = val.strip()
        if not val or val.lower() in ('nan', 'nat'):
            return None
        for fmt in ['%Y/%m/%d', '%Y-%m-%d', '%m/%d/%Y']:
            try:
                return datetime.strptime(val, fmt).date()
            except:
                continue
    try:
        return pd.to_datetime(val).date()
    except:
        return None


def get_week_end(schedule_date, breakpoint_text):
    if not breakpoint_text or breakpoint_text == 'nan':
        return None
    target_wd = WEEKDAY_MAP_FULL.get(breakpoint_text)
    if target_wd is None:
        return None
    try:
        cur_wd = schedule_date.weekday()
        days_ahead = (target_wd - cur_wd) % 7
        return schedule_date + timedelta(days=int(days_ahead))
    except (TypeError, ValueError):
        return None


def calc_target_from_text(week_end, date_text, on_breakpoint=False):
    """修正邏輯: 排程日在斷點上時不做 -= 7"""
    if date_text.startswith('下下下週') or date_text.startswith('下下下禮拜'):
        weeks_offset = 3; wd_char = date_text[-1]
    elif date_text.startswith('下下週') or date_text.startswith('下下禮拜'):
        weeks_offset = 2; wd_char = date_text[-1]
    elif date_text.startswith('下週') or date_text.startswith('下禮拜'):
        weeks_offset = 1; wd_char = date_text[-1]
    elif date_text.startswith('本週') or date_text.startswith('本禮拜') or date_text.startswith('這週'):
        weeks_offset = 0; wd_char = date_text[-1]
    else:
        return parse_date(date_text)

    target_wd = WEEKDAY_MAP_CHAR.get(wd_char)
    if target_wd is None:
        return None

    bp_wd = week_end.weekday()
    days_diff = (target_wd - bp_wd) % 7
    if days_diff > 0 and not on_breakpoint:
        days_diff -= 7  # 排程日不在斷點上: 斷點是週末，目標在斷點前
    return week_end + timedelta(days=7 * weeks_offset + days_diff)


def calc_target_from_text_OLD(week_end, date_text):
    """舊邏輯 (修正前): 一律做 -= 7"""
    if date_text.startswith('下下下週') or date_text.startswith('下下下禮拜'):
        weeks_offset = 3; wd_char = date_text[-1]
    elif date_text.startswith('下下週') or date_text.startswith('下下禮拜'):
        weeks_offset = 2; wd_char = date_text[-1]
    elif date_text.startswith('下週') or date_text.startswith('下禮拜'):
        weeks_offset = 1; wd_char = date_text[-1]
    elif date_text.startswith('本週') or date_text.startswith('本禮拜') or date_text.startswith('這週'):
        weeks_offset = 0; wd_char = date_text[-1]
    else:
        return parse_date(date_text)

    target_wd = WEEKDAY_MAP_CHAR.get(wd_char)
    if target_wd is None:
        return None

    bp_wd = week_end.weekday()
    days_diff = (target_wd - bp_wd) % 7
    if days_diff > 0:
        days_diff -= 7
    return week_end + timedelta(days=7 * weeks_offset + days_diff)


def calc_erp_target(row, use_new=True):
    """計算 ERP 目標日期"""
    schedule_val = row.get('排程出貨日期')
    schedule_date = parse_date(schedule_val)
    if schedule_date is None:
        return None, "排程出貨日期解析失敗"

    bp_text = str(row.get('排程出貨日期斷點', '')).strip()
    if not bp_text:
        return None, "無斷點"

    week_end = get_week_end(schedule_date, bp_text)
    if week_end is None:
        return None, f"斷點解析失敗: {bp_text}"

    calc_type = str(row.get('日期算法', '')).strip().upper()
    if calc_type == 'ETD':
        date_text = str(row.get('ETD', '')).strip()
    elif calc_type == 'ETA':
        date_text = str(row.get('ETA', '')).strip()
    else:
        date_text = str(row.get('ETD', '')).strip()
        if not date_text:
            date_text = str(row.get('ETA', '')).strip()

    if not date_text or date_text == 'nan':
        return None, "無 ETD/ETA 文字"

    on_breakpoint = (schedule_date == week_end)
    if use_new:
        target = calc_target_from_text(week_end, date_text, on_breakpoint)
    else:
        target = calc_target_from_text_OLD(week_end, date_text)

    if target and target < schedule_date:
        return None, f"目標日 {target} < 排程日 {schedule_date}"

    on_bp_mark = " [ON_BP]" if on_breakpoint else ""
    debug = (f"排程={schedule_date}(週{WEEKDAY_NAMES[schedule_date.weekday()]}), "
             f"斷點={bp_text}, anchor={week_end}(週{WEEKDAY_NAMES[week_end.weekday()]}){on_bp_mark}, "
             f"算法={calc_type}, 文字={date_text}, 目標={target}")
    return target, debug


def read_forecast_dates(filepath):
    """讀取 Forecast 的日期結構"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Daily+Weekly+Monthly']

    plant = str(ws.cell(row=1, column=3).value or '').strip()

    dates = {}
    daily_dates = []
    weekly_dates = []
    monthly_dates = []

    for col in range(11, 70):
        val = ws.cell(row=7, column=col).value
        d = parse_date(val)
        if d:
            dates[col] = d
            if col <= 41:
                daily_dates.append((col, d))
            elif col <= 63:
                weekly_dates.append((col, d))
            else:
                monthly_dates.append((col, d))

    wb.close()
    return {
        'plant': plant,
        'dates': dates,
        'daily': daily_dates,
        'weekly': weekly_dates,
        'monthly': monthly_dates,
        'daily_start': daily_dates[0][1] if daily_dates else None,
        'daily_end': daily_dates[-1][1] if daily_dates else None,
        'weekly_start': weekly_dates[0][1] if weekly_dates else None,
        'weekly_end': weekly_dates[-1][1] if weekly_dates else None,
    }


def find_date_column(dates, target_date):
    """模擬 _find_date_column (含 GAP fallback)"""
    # Daily
    for col, d in dates.items():
        if col > 41:
            continue
        if d == target_date:
            return col, 'Daily'
    # Weekly
    first_weekly_col = None
    first_weekly_date = None
    for col, d in dates.items():
        if col < 42 or col > 63:
            continue
        if first_weekly_col is None or d < first_weekly_date:
            first_weekly_col = col
            first_weekly_date = d
        if d <= target_date <= d + timedelta(days=6):
            return col, 'Weekly'
    # GAP fallback: 日期在第一個 Weekly 之前 → 歸入第一個 Weekly
    if first_weekly_date and target_date < first_weekly_date:
        return first_weekly_col, 'Weekly(GAP)'
    # Monthly
    for col, d in dates.items():
        if col < 64 or col > 69:
            continue
        if d.year == target_date.year and d.month == target_date.month:
            return col, 'Monthly'
    return None, 'NOT_FOUND'


def verify_result_file(result_path, forecast_info):
    """讀取結果檔案中所有非零 Commit cell"""
    wb = openpyxl.load_workbook(result_path, data_only=True)
    ws = wb['Daily+Weekly+Monthly']
    filled = []
    for row in range(8, ws.max_row + 1):
        measure = str(ws.cell(row=row, column=3).value or '').strip()
        if measure != 'Commit':
            continue
        material = str(ws.cell(row=row, column=2).value or '').strip()
        for col in range(11, 70):
            val = ws.cell(row=row, column=col).value
            if val and isinstance(val, (int, float)) and val != 0:
                col_date = forecast_info['dates'].get(col)
                col_type = 'Daily' if col <= 41 else ('Weekly' if col <= 63 else 'Monthly')
                filled.append({
                    'material': material, 'row': row, 'col': col,
                    'value': val, 'date': col_date, 'type': col_type,
                })
    wb.close()
    return filled


# ═══════════════════════════════════════════════════
# Part 1: 驗證 2680 單檔結果
# ═══════════════════════════════════════════════════
print("=" * 80)
print("Part 1: 驗證 2680 單檔 (processed/6/20260317_195848)")
print("=" * 80)

erp_path_2680 = os.path.join(BASE, "processed/6/20260317_195848/integrated_erp.xlsx")
forecast_path_2680 = os.path.join(BASE, "uploads/6/20260317_195848/forecast_data.xlsx")
result_path_2680 = os.path.join(BASE, "processed/6/20260317_195848/forecast_2680_A33.xlsx")

erp_2680 = pd.read_excel(erp_path_2680)
fi_2680 = read_forecast_dates(forecast_path_2680)

print(f"\nPlant: {fi_2680['plant']}")
print(f"Daily: {fi_2680['daily_start']} ~ {fi_2680['daily_end']} ({len(fi_2680['daily'])} cols)")
print(f"Weekly: {fi_2680['weekly_start']} ~ {fi_2680['weekly_end']} ({len(fi_2680['weekly'])} cols)")
print(f"Monthly: {len(fi_2680['monthly'])} cols")
print(f"ERP: {len(erp_2680)} rows")

# 計算新舊邏輯差異
print(f"\n--- 新舊邏輯比較 ---")
diff_count = 0
new_targets = defaultdict(int)
old_targets = defaultdict(int)

for idx, row in erp_2680.iterrows():
    region = str(row.get('客戶需求地區', '')).strip()
    material = str(row.get('客戶料號', '')).strip()
    qty = row.get('淨需求', 0)

    new_target, new_debug = calc_erp_target(row, use_new=True)
    old_target, old_debug = calc_erp_target(row, use_new=False)

    if new_target:
        new_targets[new_target] += 1
    if old_target:
        old_targets[old_target] += 1

    if new_target != old_target:
        diff_count += 1
        if diff_count <= 10:
            print(f"  差異 #{diff_count}: {material} qty={qty}")
            print(f"    舊: {old_target}")
            print(f"    新: {new_target}")
            print(f"    {new_debug}")

print(f"\n  共 {diff_count} 筆差異")

print(f"\n--- 新邏輯目標日期分布 ---")
for d in sorted(new_targets.keys()):
    col, col_type = find_date_column(fi_2680['dates'], d)
    wd = WEEKDAY_NAMES[d.weekday()]
    print(f"  {d} (週{wd}) → {col_type} col={col}: {new_targets[d]} 筆")

print(f"\n--- 舊邏輯目標日期分布 ---")
for d in sorted(old_targets.keys()):
    col, col_type = find_date_column(fi_2680['dates'], d)
    wd = WEEKDAY_NAMES[d.weekday()]
    print(f"  {d} (週{wd}) → {col_type} col={col}: {old_targets[d]} 筆")

# 驗證結果檔案 (用舊邏輯產生的)
if os.path.exists(result_path_2680):
    filled = verify_result_file(result_path_2680, fi_2680)
    print(f"\n--- 結果檔案 (舊邏輯產出) 填入 ---")
    print(f"共 {len(filled)} 個非零 Commit cell")
    date_fills = defaultdict(lambda: {'count': 0, 'total': 0, 'type': ''})
    for f in filled:
        key = f['date']
        date_fills[key]['count'] += 1
        date_fills[key]['total'] += f['value']
        date_fills[key]['type'] = f['type']
    for k in sorted(date_fills.keys()):
        v = date_fills[k]
        wd = WEEKDAY_NAMES[k.weekday()] if k else '?'
        print(f"    {k} (週{wd}) [{v['type']}]: {v['count']} cells, total={v['total']:,.0f}")


# ═══════════════════════════════════════════════════
# Part 2: 全部 Plant 分析 (20260317_200333)
# ═══════════════════════════════════════════════════
print("\n\n" + "=" * 80)
print("Part 2: 全部 Plant 分析 (20260317_200333)")
print("=" * 80)

upload_dir = os.path.join(BASE, "uploads/6/20260317_200333")
processed_dir = os.path.join(BASE, "processed/6/20260317_200333")
erp_path_all = os.path.join(processed_dir, "integrated_erp.xlsx")

erp_all = pd.read_excel(erp_path_all)
print(f"\nERP 總行數: {len(erp_all)}")

# ERP 結構概覽
print(f"\n--- ERP 結構概覽 ---")
print(f"  客戶需求地區 分布:")
for region, cnt in erp_all['客戶需求地區'].value_counts().sort_index().items():
    print(f"    {region}: {cnt} 筆")

print(f"\n  排程出貨日期 分布:")
for d, cnt in erp_all['排程出貨日期'].apply(parse_date).value_counts().sort_index().items():
    if d:
        print(f"    {d} (週{WEEKDAY_NAMES[d.weekday()]}): {cnt} 筆")

print(f"\n  斷點 分布:")
for bp, cnt in erp_all['排程出貨日期斷點'].value_counts().items():
    print(f"    {bp}: {cnt} 筆")

print(f"\n  日期算法 分布:")
for calc, cnt in erp_all['日期算法'].value_counts().items():
    print(f"    {calc}: {cnt} 筆")

print(f"\n  ETD 文字 分布:")
for t, cnt in erp_all['ETD'].value_counts().items():
    print(f"    {t}: {cnt} 筆")

print(f"\n  ETA 文字 分布:")
for t, cnt in erp_all['ETA'].value_counts().items():
    print(f"    {t}: {cnt} 筆")

# 讀各 Forecast 日期結構
print(f"\n--- 各 Forecast 日期結構 ---")
plant_forecasts = {}
for i in range(1, 24):
    fpath = os.path.join(upload_dir, f"forecast_data_{i}.xlsx")
    if not os.path.exists(fpath):
        continue
    fi = read_forecast_dates(fpath)
    plant_forecasts[fi['plant']] = fi
    print(f"  File {i:2d}: Plant={fi['plant']}, "
          f"Daily={fi['daily_start']}~{fi['daily_end']}, "
          f"Weekly starts={fi['weekly_start']}")

# Daily 起始分組
print(f"\n--- Daily 起始日分組 ---")
start_groups = defaultdict(list)
for plant, fi in plant_forecasts.items():
    start_groups[fi['daily_start']].append(plant)
for start, plants in sorted(start_groups.items()):
    end = plant_forecasts[plants[0]]['daily_end']
    ws = plant_forecasts[plants[0]]['weekly_start']
    gap_start = end + timedelta(days=1) if end else None
    gap_days = (ws - gap_start).days if (ws and gap_start) else '?'
    print(f"  起始 {start} → 結束 {end} → GAP {gap_start}~{ws} ({gap_days}天)")
    print(f"    Plants: {', '.join(sorted(plants))}")

# 每個 Plant 的 ERP 日期計算 — 新舊邏輯對比
print(f"\n--- 每個 Plant: 新舊邏輯比較 ---")

all_diff_total = 0
all_summary = {}

for plant in sorted(plant_forecasts.keys()):
    fi = plant_forecasts[plant]
    plant_erp = erp_all[erp_all['客戶需求地區'].astype(str).str.strip() == plant]
    if len(plant_erp) == 0:
        continue

    new_type_counts = defaultdict(int)
    old_type_counts = defaultdict(int)
    new_date_counts = defaultdict(int)
    old_date_counts = defaultdict(int)
    diffs = []
    fill_new = 0
    fill_old = 0
    skip_new = 0
    skip_old = 0

    for idx, row in plant_erp.iterrows():
        material = str(row.get('客戶料號', '')).strip()
        qty = row.get('淨需求', 0)

        new_target, new_debug = calc_erp_target(row, use_new=True)
        old_target, old_debug = calc_erp_target(row, use_new=False)

        if new_target:
            col_new, type_new = find_date_column(fi['dates'], new_target)
            new_type_counts[type_new] += 1
            new_date_counts[new_target] += 1
            fill_new += 1 if col_new else 0
            skip_new += 1 if not col_new else 0
        else:
            skip_new += 1

        if old_target:
            col_old, type_old = find_date_column(fi['dates'], old_target)
            old_type_counts[type_old] += 1
            old_date_counts[old_target] += 1
            fill_old += 1 if col_old else 0
            skip_old += 1 if not col_old else 0
        else:
            skip_old += 1

        if new_target != old_target:
            diffs.append({
                'material': material, 'qty': qty,
                'old': old_target, 'new': new_target,
                'debug': new_debug,
            })

    print(f"\n  Plant {plant}: ERP {len(plant_erp)} 筆")
    print(f"    Daily: {fi['daily_start']} ~ {fi['daily_end']}, Weekly: {fi['weekly_start']}")
    print(f"    新邏輯: 填入={fill_new}, 跳過={skip_new}, 類型={dict(new_type_counts)}")
    print(f"    舊邏輯: 填入={fill_old}, 跳過={skip_old}, 類型={dict(old_type_counts)}")

    if diffs:
        all_diff_total += len(diffs)
        print(f"    ⚠️ 新舊差異: {len(diffs)} 筆")
        for d in diffs[:5]:
            old_wd = f"週{WEEKDAY_NAMES[d['old'].weekday()]}" if d['old'] else 'N/A'
            new_wd = f"週{WEEKDAY_NAMES[d['new'].weekday()]}" if d['new'] else 'N/A'
            print(f"      {d['material']} qty={d['qty']}: "
                  f"舊={d['old']}({old_wd}) → 新={d['new']}({new_wd})")
        if len(diffs) > 5:
            print(f"      ... 還有 {len(diffs)-5} 筆")
    else:
        print(f"    ✅ 新舊邏輯一致")

    # 新邏輯目標日期分布
    if new_date_counts:
        print(f"    新邏輯目標日期:")
        for d in sorted(new_date_counts.keys()):
            col, ctype = find_date_column(fi['dates'], d)
            wd = WEEKDAY_NAMES[d.weekday()]
            in_gap = False
            if fi['daily_end'] and fi['weekly_start']:
                gap_s = fi['daily_end'] + timedelta(days=1)
                if gap_s <= d < fi['weekly_start']:
                    in_gap = True
            gap_mark = " ⚠️GAP" if in_gap else ""
            nf_mark = " ❌NOT_FOUND" if col is None else ""
            print(f"      {d} (週{wd}) → {ctype} col={col} ({new_date_counts[d]}筆){gap_mark}{nf_mark}")

    all_summary[plant] = {
        'erp': len(plant_erp),
        'fill_new': fill_new, 'skip_new': skip_new,
        'fill_old': fill_old, 'skip_old': skip_old,
        'diffs': len(diffs),
    }

# 驗證結果檔案 (舊邏輯產出)
print(f"\n\n--- 驗證結果檔案 (舊邏輯產出) ---")
result_files = sorted([f for f in os.listdir(processed_dir)
                       if f.startswith('forecast_') and f.endswith('.xlsx')])
for rf in result_files:
    parts = rf.replace('forecast_', '').replace('.xlsx', '').split('_')
    plant = parts[0]
    fi = plant_forecasts.get(plant)
    if not fi:
        continue

    rpath = os.path.join(processed_dir, rf)
    filled = verify_result_file(rpath, fi)

    date_fills = defaultdict(lambda: {'count': 0, 'total': 0, 'type': ''})
    for f in filled:
        key = f['date']
        date_fills[key]['count'] += 1
        date_fills[key]['total'] += f['value']
        date_fills[key]['type'] = f['type']

    print(f"\n  {rf}: {len(filled)} cells")
    if filled:
        for k in sorted(date_fills.keys()):
            v = date_fills[k]
            wd = WEEKDAY_NAMES[k.weekday()] if k else '?'
            print(f"    {k} (週{wd}) [{v['type']}]: {v['count']} cells, total={v['total']:,.0f}")


# ═══════════════════════════════════════════════════
# 彙總
# ═══════════════════════════════════════════════════
print("\n\n" + "=" * 80)
print("彙總")
print("=" * 80)

total_diffs = sum(v['diffs'] for v in all_summary.values())
total_erp = sum(v['erp'] for v in all_summary.values())
total_fill_new = sum(v['fill_new'] for v in all_summary.values())
total_fill_old = sum(v['fill_old'] for v in all_summary.values())

print(f"\n全部 Plant ERP: {total_erp} 筆 (ERP 總 {len(erp_all)} 筆)")
print(f"新邏輯填入: {total_fill_new}")
print(f"舊邏輯填入: {total_fill_old}")
print(f"新舊差異筆數: {total_diffs}")

print(f"\n--- 各 Plant 彙總 ---")
print(f"{'Plant':<8} {'ERP':>5} {'新填入':>6} {'舊填入':>6} {'差異':>4}")
print("-" * 35)
for plant in sorted(all_summary.keys()):
    v = all_summary[plant]
    diff_mark = f"⚠️" if v['diffs'] > 0 else "  "
    print(f"{plant:<8} {v['erp']:>5} {v['fill_new']:>6} {v['fill_old']:>6} {v['diffs']:>4} {diff_mark}")
print("-" * 35)
print(f"{'TOTAL':<8} {total_erp:>5} {total_fill_new:>6} {total_fill_old:>6} {total_diffs:>4}")
