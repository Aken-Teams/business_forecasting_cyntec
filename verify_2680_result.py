# -*- coding: utf-8 -*-
"""驗證 test/forecast_2680_A33_20260317_204558.xlsx (新邏輯產出)"""
import sys; sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
import pandas as pd
from datetime import datetime, timedelta, date
from collections import defaultdict

WEEKDAY_MAP_FULL = {
    '週一': 0, '禮拜一': 0, '星期一': 0,
    '週二': 1, '禮拜二': 1, '星期二': 1,
    '週三': 2, '禮拜三': 2, '星期三': 2,
    '週四': 3, '禮拜四': 3, '星期四': 3,
    '週五': 4, '禮拜五': 4, '星期五': 4,
    '週六': 5, '禮拜六': 5, '星期六': 5,
    '週日': 6, '禮拜日': 6, '星期日': 6, '禮拜天': 6, '週天': 6,
}
WEEKDAY_MAP_CHAR = {'一': 0, '二': 1, '三': 2, '四': 3, '五': 4, '六': 5, '日': 6, '天': 6}
WEEKDAY_NAMES = ['一', '二', '三', '四', '五', '六', '日']


def parse_date(val):
    if val is None: return None
    if isinstance(val, date) and not isinstance(val, datetime): return val
    if isinstance(val, datetime): return val.date()
    if hasattr(val, 'date') and callable(val.date):
        try: return val.date()
        except: return None
    if isinstance(val, str):
        val = val.strip()
        if not val or val.lower() in ('nan', 'nat'): return None
        for fmt in ['%Y/%m/%d', '%Y-%m-%d', '%m/%d/%Y']:
            try: return datetime.strptime(val, fmt).date()
            except: continue
    try: return pd.to_datetime(val).date()
    except: return None


def get_week_end(schedule_date, bp_text):
    target_wd = WEEKDAY_MAP_FULL.get(bp_text)
    if target_wd is None: return None
    cur_wd = schedule_date.weekday()
    days_ahead = (target_wd - cur_wd) % 7
    return schedule_date + timedelta(days=int(days_ahead))


def calc_target_NEW(week_end, date_text, on_breakpoint=False):
    if date_text.startswith('下下下週') or date_text.startswith('下下下禮拜'):
        weeks_offset = 3; wd_char = date_text[-1]
    elif date_text.startswith('下下週') or date_text.startswith('下下禮拜'):
        weeks_offset = 2; wd_char = date_text[-1]
    elif date_text.startswith('下週') or date_text.startswith('下禮拜'):
        weeks_offset = 1; wd_char = date_text[-1]
    elif date_text.startswith('本週') or date_text.startswith('本禮拜') or date_text.startswith('這週'):
        weeks_offset = 0; wd_char = date_text[-1]
    else:
        return parse_date(date_text)
    target_wd = WEEKDAY_MAP_CHAR.get(wd_char)
    if target_wd is None: return None
    bp_wd = week_end.weekday()
    days_diff = (target_wd - bp_wd) % 7
    if days_diff > 0 and not on_breakpoint:
        days_diff -= 7
    return week_end + timedelta(days=7 * weeks_offset + days_diff)


def find_col(dates, target_date):
    for col, d in dates.items():
        if col > 41: continue
        if d == target_date: return col, 'Daily'
    first_wk_col, first_wk_date = None, None
    for col, d in dates.items():
        if col < 42 or col > 63: continue
        if first_wk_col is None or d < first_wk_date:
            first_wk_col, first_wk_date = col, d
        if d <= target_date <= d + timedelta(days=6):
            return col, 'Weekly'
    if first_wk_date and target_date < first_wk_date:
        return first_wk_col, 'Weekly(GAP)'
    for col, d in dates.items():
        if col < 64 or col > 69: continue
        if d.year == target_date.year and d.month == target_date.month:
            return col, 'Monthly'
    return None, 'NOT_FOUND'


# ── 讀結果檔 ──
result_path = 'test/forecast_2680_A33_20260317_204558.xlsx'
wb = openpyxl.load_workbook(result_path, data_only=True)
ws = wb['Daily+Weekly+Monthly']

plant = str(ws.cell(row=1, column=3).value or '').strip()
print(f'Plant: {plant}')

dates = {}
for col in range(11, 70):
    val = ws.cell(row=7, column=col).value
    d = parse_date(val)
    if d: dates[col] = d

daily_end = max(d for c, d in dates.items() if c <= 41)
weekly_start = min(d for c, d in dates.items() if 42 <= c <= 63)
print(f'Daily end: {daily_end}, Weekly start: {weekly_start}')

# 讀所有 Commit cells
result_cells = {}
materials_rows = {}
for row in range(8, ws.max_row + 1):
    measure = str(ws.cell(row=row, column=3).value or '').strip()
    if measure != 'Commit':
        continue
    mat = str(ws.cell(row=row, column=2).value or '').strip()
    materials_rows[mat] = row
    for col in range(11, 70):
        val = ws.cell(row=row, column=col).value
        if val and isinstance(val, (int, float)) and val != 0:
            result_cells[(mat, col)] = val
wb.close()

print(f'料號: {len(materials_rows)}, 非零 cells: {len(result_cells)}')

# ── 讀 ERP ──
erp_path = 'processed/6/20260317_195848/integrated_erp.xlsx'
erp = pd.read_excel(erp_path)
erp_plant = erp[erp['客戶需求地區'].astype(str).str.strip() == plant]
print(f'ERP rows for {plant}: {len(erp_plant)}')

# ── 用新邏輯計算預期填入 ──
expected = defaultdict(float)
fill_trace = []

for idx, row in erp_plant.iterrows():
    mat = str(row.get('客戶料號', '')).strip()
    qty = row.get('淨需求', 0)
    try: qty = float(qty)
    except: continue
    if mat not in materials_rows: continue
    if qty <= 0: continue

    schedule = parse_date(row.get('排程出貨日期'))
    if not schedule: continue
    bp = str(row.get('排程出貨日期斷點', '')).strip()
    if not bp or bp == 'nan': continue
    we = get_week_end(schedule, bp)
    if not we: continue

    calc_type = str(row.get('日期算法', '')).strip().upper()
    if calc_type == 'ETA':
        text = str(row.get('ETA', '')).strip()
    elif calc_type == 'ETD':
        text = str(row.get('ETD', '')).strip()
    else:
        text = str(row.get('ETD', '')).strip() or str(row.get('ETA', '')).strip()
    if not text or text == 'nan': continue

    on_bp = (schedule == we)
    target = calc_target_NEW(we, text, on_bp)
    if not target or target < schedule: continue

    col, ctype = find_col(dates, target)
    if not col: continue

    expected[(mat, col)] += qty * 1000
    fill_trace.append({
        'mat': mat, 'qty': qty, 'schedule': schedule,
        'bp': bp, 'on_bp': on_bp, 'text': text, 'calc': calc_type,
        'target': target, 'col': col, 'ctype': ctype,
    })

print(f'新邏輯預期: {len(expected)} cells, {len(fill_trace)} 筆 ERP')

# ── 比對 ──
print(f'\n{"="*60}')
print('預期 vs 結果檔比對')
print(f'{"="*60}')

all_keys = set(expected.keys()) | set(result_cells.keys())
match = mismatch = only_exp = only_res = 0
mismatches = []

for key in sorted(all_keys):
    exp = expected.get(key)
    act = result_cells.get(key)
    mat, col = key
    col_date = dates.get(col)
    if exp and act:
        if abs(exp - act) < 0.01:
            match += 1
        else:
            mismatch += 1
            mismatches.append((mat, col, col_date, exp, act))
    elif exp and not act:
        only_exp += 1
    elif act and not exp:
        only_res += 1

print(f'完全匹配: {match}')
print(f'值不同: {mismatch}')
print(f'預期有/結果無: {only_exp}')
print(f'結果有/預期無: {only_res}')

if mismatch > 0:
    print(f'\n值不同明細:')
    for mat, col, cd, exp, act in mismatches:
        wd = WEEKDAY_NAMES[cd.weekday()] if cd else '?'
        print(f'  {mat} col={col} ({cd} 週{wd}): 預期={exp:,.0f} 實際={act:,.0f} 差={act-exp:+,.0f}')

if only_exp > 0:
    print(f'\n預期有/結果無:')
    for key in sorted(all_keys):
        if key in expected and key not in result_cells:
            mat, col = key
            cd = dates.get(col)
            wd = WEEKDAY_NAMES[cd.weekday()] if cd else '?'
            print(f'  {mat} col={col} ({cd} 週{wd}): 預期={expected[key]:,.0f}')

if only_res > 0:
    print(f'\n結果有/預期無:')
    for key in sorted(all_keys):
        if key not in expected and key in result_cells:
            mat, col = key
            cd = dates.get(col)
            wd = WEEKDAY_NAMES[cd.weekday()] if cd else '?'
            print(f'  {mat} col={col} ({cd} 週{wd}): 實際={result_cells[key]:,.0f}')

# ── 按日期彙總 ──
print(f'\n{"="*60}')
print('結果檔: 按日期彙總')
print(f'{"="*60}')

date_summary = defaultdict(lambda: {'cells': 0, 'total': 0})
for (mat, col), val in result_cells.items():
    cd = dates.get(col)
    ctype = 'Daily' if col <= 41 else ('Weekly' if col <= 63 else 'Monthly')
    key = (cd, ctype, col)
    date_summary[key]['cells'] += 1
    date_summary[key]['total'] += val

for key in sorted(date_summary.keys()):
    cd, ctype, col = key
    info = date_summary[key]
    wd = WEEKDAY_NAMES[cd.weekday()] if cd else '?'
    print(f'  {cd} (週{wd}) [{ctype}] col={col}: {info["cells"]} cells, total={info["total"]:,.0f}')

# ── on_breakpoint trace ──
print(f'\n{"="*60}')
print('on_breakpoint 明細')
print(f'{"="*60}')
bp_traces = [t for t in fill_trace if t['on_bp']]
print(f'{len(bp_traces)} 筆')
for t in bp_traces[:15]:
    twd = WEEKDAY_NAMES[t['target'].weekday()]
    print(f'  {t["mat"]}: 排程={t["schedule"]}({t["bp"]}), {t["calc"]}={t["text"]} '
          f'-> {t["target"]}(週{twd}) [{t["ctype"]}] val={t["qty"]*1000:,.0f}')
if len(bp_traces) > 15:
    print(f'  ... 還有 {len(bp_traces)-15} 筆')

# ── 總計 ──
exp_total = sum(expected.values())
act_total = sum(result_cells.values())
print(f'\n{"="*60}')
print(f'總計: 預期={exp_total:,.0f}, 結果={act_total:,.0f}, 差={act_total-exp_total:+,.0f}')
