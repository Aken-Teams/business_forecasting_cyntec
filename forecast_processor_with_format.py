import openpyxl
from openpyxl import load_workbook
import pandas as pd

def process_forecast_file_with_format():
    """
    處理ForecastDataFile_ALL-0923.xlsx文件，保持原始格式和架構
    1. 當K欄位是"供應數量"時，將L~AW欄位清為0（保持格式）
    2. 當I欄位有"庫存數量"文字時，將下一列I欄位數值清為0（保持格式）
    """
    try:
        # 讀取原始Excel文件
        original_file_path = "forecast/ForecastDataFile_ALL-0923.xlsx"
        output_file_path = "修改後的ForecastDataFile_ALL-0923.xlsx"
        
        print("正在讀取原始Excel文件並保持格式...")
        
        # 使用openpyxl讀取工作簿以保持格式
        wb = load_workbook(original_file_path)
        ws = wb.active
        
        print(f"工作表名稱: {ws.title}")
        print(f"工作表範圍: {ws.max_row} 行 x {ws.max_column} 欄")
        
        # 先讀取數據以進行分析
        df = pd.read_excel(original_file_path)
        print(f"數據分析: {len(df)} 行 x {len(df.columns)} 欄")
        
        # 找到K欄位為"供應數量"的行
        k_column_name = df.columns[10]  # K欄位 (索引10)
        supply_rows = []
        for idx, value in enumerate(df[k_column_name]):
            if value == "供應數量":
                supply_rows.append(idx + 2)  # +2 因為pandas從0開始，Excel從1開始，且第一行可能是標題
        
        print(f"找到 {len(supply_rows)} 行K欄位為'供應數量'的記錄")
        
        # 找到I欄位包含"庫存數量"的行
        i_column_name = df.columns[8]  # I欄位 (索引8)
        inventory_rows = []
        for idx, value in enumerate(df[i_column_name]):
            if str(value).find('庫存數量') != -1:
                inventory_rows.append(idx + 2)  # +2 因為pandas從0開始，Excel從1開始
        
        print(f"找到 {len(inventory_rows)} 行I欄位包含'庫存數量'的記錄")
        
        # 處理供應數量行 - 清空L~AW欄位（列12到列49）
        print("正在處理供應數量行的L~AW欄位...")
        for row_num in supply_rows:
            for col_num in range(12, 50):  # L欄位(12) 到 AW欄位(49)
                cell = ws.cell(row=row_num, column=col_num)
                # 只清空值，保持格式
                if cell.data_type == 'n':  # 數值類型
                    cell.value = 0
                elif cell.data_type == 's':  # 字符串類型
                    cell.value = ""
                else:
                    cell.value = None
        
        print(f"已處理 {len(supply_rows)} 行的L~AW欄位")
        
        # 處理庫存數量行 - 清空下一列的I欄位
        print("正在處理庫存數量行下一列的I欄位...")
        modified_count = 0
        for row_num in inventory_rows:
            next_row = row_num + 1
            if next_row <= ws.max_row:  # 確保下一列存在
                cell = ws.cell(row=next_row, column=9)  # I欄位是第9列
                # 只清空值，保持格式
                if cell.data_type == 'n':  # 數值類型
                    cell.value = 0
                elif cell.data_type == 's':  # 字符串類型
                    cell.value = ""
                else:
                    cell.value = None
                modified_count += 1
        
        print(f"已處理 {modified_count} 行的I欄位")
        
        # 保存文件，保持所有格式
        print("正在保存文件（保持原始格式）...")
        wb.save(output_file_path)
        
        print(f"修改後的文件已保存到: {output_file_path}")
        print("所有原始格式和架構都已保持")
        
        return True
        
    except Exception as e:
        print(f"處理文件時發生錯誤: {e}")
        import traceback
        traceback.print_exc()
        return False

def verify_changes():
    """
    驗證修改結果
    """
    try:
        print("\n正在驗證修改結果...")
        
        # 讀取修改後的文件
        df_modified = pd.read_excel("修改後的ForecastDataFile_ALL-0923.xlsx")
        
        # 檢查供應數量行的L~AW欄位
        k_column_name = df_modified.columns[10]
        supply_mask = df_modified[k_column_name] == "供應數量"
        supply_count = supply_mask.sum()
        
        if supply_count > 0:
            # 檢查L~AW欄位是否為0
            l_to_aw_columns = df_modified.columns[11:49]  # L到AW欄位
            cleared_count = 0
            for col in l_to_aw_columns:
                non_zero_count = df_modified.loc[supply_mask, col].fillna(0).astype(str).str.replace('0', '').str.replace('.0', '').str.strip().astype(bool).sum()
                if non_zero_count == 0:
                    cleared_count += 1
            
            print(f"供應數量行: {supply_count} 行")
            print(f"L~AW欄位完全清空: {cleared_count}/{len(l_to_aw_columns)} 欄位")
        
        # 檢查庫存數量行下一列的I欄位
        i_column_name = df_modified.columns[8]
        inventory_mask = df_modified[i_column_name].astype(str).str.contains('庫存數量', na=False)
        inventory_indices = df_modified[inventory_mask].index.tolist()
        
        cleared_i_count = 0
        for idx in inventory_indices:
            next_idx = idx + 1
            if next_idx < len(df_modified):
                i_value = df_modified.loc[next_idx, i_column_name]
                if pd.isna(i_value) or i_value == 0 or i_value == "":
                    cleared_i_count += 1
        
        print(f"庫存數量行: {len(inventory_indices)} 行")
        print(f"I欄位下一列已清空: {cleared_i_count}/{len(inventory_indices)} 行")
        
    except Exception as e:
        print(f"驗證時發生錯誤: {e}")

if __name__ == "__main__":
    print("開始處理ForecastDataFile_ALL-0923.xlsx文件（保持格式）...")
    
    success = process_forecast_file_with_format()
    
    if success:
        verify_changes()
        print("\n處理完成！文件格式和架構已完全保持。")
    else:
        print("處理失敗")
