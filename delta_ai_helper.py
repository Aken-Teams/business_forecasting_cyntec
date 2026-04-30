"""
Delta Forecast AI 輔助模組
============================

用途:
  當規則式格式偵測 (detect_format) + 指紋比對 (fingerprint) 均失敗時，
  呼叫 DeepSeek API 分析 Excel 結構，識別欄位布局，讓後續 reader/backfill 正常運作。

使用:
  from delta_ai_helper import ai_analyze_file, get_ai_column_hints, clear_ai_cache

  # 取得 AI 分析結果 (含快取)
  result = ai_analyze_file(filepath, file_label='PSW1 XXXX.xlsx')
  if result and result.get('identified'):
      col = result['columns']
      partno_col = col.get('partno')   # 1-based
      marker_col = col.get('marker')   # None for flat format

設計:
  - DeepSeek API (OpenAI-compatible endpoint)，API key 從 .env 讀取
  - 結果 in-memory 快取 (一次 session 內不重複呼叫)
  - Rule-based 優先，AI 只作 fallback
"""

import os
import json
import openpyxl
from datetime import datetime

try:
    from openai import OpenAI
    _HAS_OPENAI = True
except ImportError:
    _HAS_OPENAI = False

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

DEEPSEEK_API_KEY = os.getenv('deepseek_api_key')
DEEPSEEK_BASE_URL = 'https://api.deepseek.com'
DEEPSEEK_MODEL = 'deepseek-chat'

# 行內快取 {abs_path: result_dict or None}
_ai_cache: dict = {}


# ============ 已知格式說明 (給 AI 的 context) ============

_KNOWN_FORMATS_CONTEXT = """Known Delta Electronics buyer forecast file formats:

FORMAT TYPES:
- flat: One row per PARTNO, no marker/type column. Date values are demand quantities.
  Files: EIBG, EISBG, IABG, NBQ1, SVC1+PWC1 DIODE&MOS

- multirow_3: 3 rows per PARTNO in order (Demand, Supply, Balance/Net).
  Typical markers: "Demand"/"Supply"/"Balance", "1.Demand"/"2.Supply"/"3.Net",
                   "A-Demand"/"B-CFM"/"C-Balance", "DEMAND"/"SUPPLY"/"NET AVAIL"
  Files: Ketwadee (PSB5), PSBG, India IAI1&UPI2, FMBG (TPC5), ICTBG-NTL7, ICTBG-PSB9-Siriraht

- multirow_4: 4 rows per PARTNO.
  Typical markers: "GROSS REQTS"/"FIRM ORDERS"/"VENDOR CFM"/"NET AVAIL",
                   "GROSS REQTS"/"FIRM ORDERS"/"VN CFM"/"NET AVAIL",
                   "A-Demand"/"B-Forecast Conf"/"C-Demand"/"D-Net Demand",
                   "Demand"/"Supply"/"Balance"/"Remark"
  Files: Kanyanat (PSB7), Weeraya (PSB7), MWC1+IPC1, DNI-NTL7, ICTBG-PSB9-Kaewarin

COLUMN PATTERNS:
- PARTNO: "PARTNO", "PART NO", "Raw Material(P/N)", "P/N", "料號", "DELTA P/N", "CUSTOMER PART"
- PLANT: "PLANT", "廠區", "廠別" → often ABSENT, must extract from filename
- TYPE/MARKER: "TYPE", "REQUEST ITEM", "類別", unnamed col with demand/supply labels
- VENDOR PART: "VENDOR PARTNO", "VENDOR P/N", "廠商料號", "Vendor Part"
- STOCK: "STOCK", "PLANT STOCK", "Total stock", "庫存"
- ON-WAY: "ON-WAY", "ON THE WAY", "在途"
- DATES: float YYYYMMDD (e.g. 20260420.0), int YYYYMMDD, datetime, "0427", "APR", "PASSDUE"

IMPORTANT: Plant codes look like PSB5, PSB7, PSB9, MWC1, IPC1, NTL7, IAI1, PSW1, CEW1, etc.
If no PLANT column, set plant_from_filename=true.
"""


# ============ 結構抽取 ============

def _extract_file_structure(filepath: str, max_rows: int = 15, max_cols: int = 35) -> dict | None:
    """從 Excel 抽出結構資訊供 AI 分析。"""
    wb = None
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception:
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception:
            return None

    structure = {
        'sheets': wb.sheetnames,
        'sheet_data': {},
    }

    for sheet_name in wb.sheetnames[:4]:
        ws = wb[sheet_name]
        max_row_ws = ws.max_row or 0
        max_col_ws = ws.max_column or 0
        if max_row_ws < 1 or max_col_ws < 1:
            continue

        rows_text = []
        for r in range(1, min(max_rows + 1, max_row_ws + 1)):
            cells = []
            for c in range(1, min(max_cols + 1, max_col_ws + 1)):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                if isinstance(v, datetime):
                    vs = v.strftime('%Y%m%d')
                elif isinstance(v, float):
                    iv = int(v)
                    vs = str(iv) if v == iv else f'{v:.2f}'
                else:
                    vs = str(v).strip()[:60]
                if vs:
                    cells.append(f'C{c}:{vs}')
            if cells:
                rows_text.append('R' + str(r) + ': ' + ' | '.join(cells))

        if rows_text:
            structure['sheet_data'][sheet_name] = rows_text

    if wb:
        try:
            wb.close()
        except Exception:
            pass

    return structure if structure['sheet_data'] else None


# ============ AI 核心分析 ============

def ai_analyze_file(filepath: str, file_label: str | None = None) -> dict | None:
    """
    使用 AI 分析 Excel 欄位布局。快取結果避免重複呼叫。

    Returns:
        dict 或 None (API 失敗時)

        成功時格式:
        {
            "identified": bool,
            "format_type": "flat" | "multirow_3" | "multirow_4" | "unknown",
            "forecast_sheet": str,   # 含預測資料的 sheet 名稱
            "header_row": int,       # 1-based
            "columns": {
                "partno": int | None,
                "plant": int | None,
                "marker": int | None,
                "vendor": int | None,
                "stock": int | None,
                "on_way": int | None,
                "date_start": int | None
            },
            "plant_from_filename": bool,
            "notes": str
        }
    """
    if not _HAS_OPENAI or not DEEPSEEK_API_KEY:
        return None

    cache_key = os.path.abspath(filepath)
    if cache_key in _ai_cache:
        return _ai_cache[cache_key]

    structure = _extract_file_structure(filepath)
    if not structure:
        _ai_cache[cache_key] = None
        return None

    label = file_label or os.path.basename(filepath)

    # 建構 prompt
    data_lines = []
    for sheet_name, rows in structure['sheet_data'].items():
        data_lines.append(f'=== Sheet: {sheet_name} ===')
        data_lines.extend(rows[:15])
        data_lines.append('')

    prompt = f"""{_KNOWN_FORMATS_CONTEXT}

File to analyze: {label}
Available sheets: {structure['sheets']}

Excel data (R=row, C=column, first 15 rows per sheet):
{''.join(line + chr(10) for line in data_lines)}

Task: Identify the column layout of the FORECAST DATA sheet (not purchase orders, not OPO sheets).
Return JSON only, no explanation or markdown.

Required format:
{{
  "identified": true,
  "format_type": "flat|multirow_3|multirow_4|unknown",
  "forecast_sheet": "sheet name",
  "header_row": 1,
  "columns": {{
    "partno": 1,
    "plant": null,
    "marker": 7,
    "vendor": 6,
    "stock": 8,
    "on_way": null,
    "date_start": 10
  }},
  "plant_from_filename": true,
  "notes": "brief observation"
}}

Rules:
1. All column numbers are 1-based
2. Use null if a column does not exist
3. marker=null for flat format (no row type column)
4. date_start = first column with date-like values (YYYYMMDD float/int, month abbreviations, PASSDUE)
5. If PLANT column is absent, set plant_from_filename=true and plant=null
"""

    try:
        client = OpenAI(api_key=DEEPSEEK_API_KEY, base_url=DEEPSEEK_BASE_URL)
        response = client.chat.completions.create(
            model=DEEPSEEK_MODEL,
            messages=[{'role': 'user', 'content': prompt}],
            response_format={'type': 'json_object'},
            temperature=0.1,
            max_tokens=600,
        )
        content = response.choices[0].message.content
        result = json.loads(content)

        _ai_cache[cache_key] = result
        cols = result.get('columns', {})
        print(f"  [AI] {os.path.basename(filepath)}: "
              f"format={result.get('format_type')} "
              f"sheet={result.get('forecast_sheet')} "
              f"partno=C{cols.get('partno')} marker=C{cols.get('marker')} "
              f"date_start=C{cols.get('date_start')}")
        return result

    except Exception as e:
        print(f"  [AI] 分析失敗 [{os.path.basename(filepath)}]: {e}")
        _ai_cache[cache_key] = None
        return None


def get_ai_column_hints(filepath: str, file_label: str | None = None) -> dict:
    """
    取得 AI 偵測的欄位位置 hints (已轉為 scan_headers 相容格式)。

    Returns:
        dict {field: col_1based} 例如 {'partno': 1, 'marker': 7, 'stock': 8}
        若 AI 無法分析則回傳 {}
    """
    result = ai_analyze_file(filepath, file_label)
    if not result or not result.get('identified'):
        return {}
    cols = result.get('columns', {})
    hints = {}
    for field in ('partno', 'plant', 'vendor_part', 'stock', 'on_way'):
        # vendor_part 對應 ai 的 vendor
        ai_key = 'vendor' if field == 'vendor_part' else field
        v = cols.get(ai_key)
        if v is not None:
            hints[field] = int(v)
    return hints


def get_ai_marker_col(filepath: str, file_label: str | None = None) -> int | None:
    """從 AI 結果取 marker 欄位置 (1-based)，flat 格式返回 None。"""
    result = ai_analyze_file(filepath, file_label)
    if not result or not result.get('identified'):
        return None
    m = result.get('columns', {}).get('marker')
    return int(m) if m is not None else None


def get_ai_date_start(filepath: str, file_label: str | None = None) -> int | None:
    """從 AI 結果取日期起始欄 (1-based)。"""
    result = ai_analyze_file(filepath, file_label)
    if not result or not result.get('identified'):
        return None
    ds = result.get('columns', {}).get('date_start')
    return int(ds) if ds is not None else None


def get_ai_forecast_sheet(filepath: str, file_label: str | None = None) -> str | None:
    """從 AI 結果取含預測資料的 sheet 名稱。"""
    result = ai_analyze_file(filepath, file_label)
    if not result or not result.get('identified'):
        return None
    return result.get('forecast_sheet')


def is_ai_available() -> bool:
    """檢查 AI 功能是否可用。"""
    return _HAS_OPENAI and bool(DEEPSEEK_API_KEY)


def clear_ai_cache(filepath: str | None = None) -> None:
    """清除快取 (特定檔案或全部)。"""
    if filepath:
        _ai_cache.pop(os.path.abspath(filepath), None)
    else:
        _ai_cache.clear()
