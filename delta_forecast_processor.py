"""
台達 (Delta) Forecast 合併處理器
================================
將 3 個不同格式的 Buyer Forecast 檔案合併為統一匯總格式。

Buyer 格式:
  Ketwadee (PSB5): MRP sheet, 3 rows/part (Demand/Supply/Net)
  Kanyanat (PSB7): Sheet1, 4 rows/part (A-Demand/B-Supply/D-Net/E-REMARK)
  Weeraya  (PSB7): Sheet1, 5 rows/part (Demand/Firmed/ForecastConf/NetDemand/Remark)

匯總格式: 每料號 3 行 (Demand/Supply/Balance), Balance 用 Excel 公式
"""

import openpyxl
import os
from datetime import datetime, timedelta
from copy import copy
from openpyxl.styles import (Font, Alignment, PatternFill, Border, Side)
from openpyxl.styles.colors import Color


# ---------------------------------------------------------------------------
# 格式常數
# ---------------------------------------------------------------------------

FORMAT_KETWADEE = 'ketwadee'              # MRP sheet, 3 rows/part
FORMAT_KANYANAT = 'kanyanat'              # Sheet1, 4 rows/part (col 24=TYPE)
FORMAT_WEERAYA = 'weeraya'                # Sheet1, 4 rows/part (col 12=TYPE)
FORMAT_INDIA_IAI1 = 'india_iai1'          # PAN JIT, 3 rows/part, 多PLANT
FORMAT_PSW1_CEW1 = 'psw1_cew1'            # Sheet1, 5 rows/part (col 12=Status), 多PLANT
FORMAT_MWC1IPC1 = 'mwc1ipc1'              # Sheet1, 4 rows/part (col 6=REQUEST ITEM), 多PLANT
FORMAT_NBQ1 = 'nbq1'                      # PAN JIT, 1 row/part, 單PLANT檔名
FORMAT_SVC1PWC1_DIODE_MOS = 'svc1pwc1_diode_mos'  # Diode+MOS, 1 row/part, 多PLANT
FORMAT_PSBG = 'psbg'                              # Sheet1, 3 rows/part (col 15=Filter), 單PLANT
FORMAT_EIBG_EISBG = 'eibg_eisbg'                  # Sheet1, col1=ITEM, flat (Demand only), 單PLANT
FORMAT_FMBG = 'fmbg'                              # Sheet1, col12=REQUEST ITEM, 3 rows/part (A-Demand/B-CFM/C-Bal), 多PLANT
FORMAT_IABG = 'iabg'                              # Sheet1, col9=SHIP, flat (Demand only), 多PLANT
FORMAT_ICTBG_NTL7 = 'ictbg_ntl7'                  # Sheet1, col10=REQUEST ITEM, 4 rows/part (GROSS REQTS/...), 多PLANT
FORMAT_ICTBG_PSB9_MRP = 'ictbg_psb9_mrp'          # PSB9_MRP* sheet, col14=Type, 4 rows/part (DEMAND/SUPPLY/NET/Remark), 多PLANT
FORMAT_ICTBG_PSB9_SIRIRAHT = 'ictbg_psb9_siriraht'  # Sheet1, col15=REQUEST ITEM (1Demand/2Supply/3Balance), 3 rows/part, 多PLANT
FORMAT_PRAPAPORN = 'prapaporn'                        # Sheet1, 4 rows/part (A-Demand/B-ForecastConf/D-Net/F-Remark), col12=TYPE, dates@17

FORMAT_LABELS = {
    FORMAT_KETWADEE: 'Ketwadee (PSB5)',
    FORMAT_KANYANAT: 'Kanyanat (PSB7)',
    FORMAT_WEERAYA:  'Weeraya (PSB7)',
    FORMAT_INDIA_IAI1: 'India IAI1/UPI2/DFI1',
    FORMAT_PSW1_CEW1: 'PSW1+CEW1',
    FORMAT_MWC1IPC1:  'MWC1+IPC1',
    FORMAT_NBQ1:      'NBQ1',
    FORMAT_SVC1PWC1_DIODE_MOS: 'SVC1+PWC1 (Diode&MOS)',
    FORMAT_PSBG:     'PSBG (PSB5 PANJIT)',
    FORMAT_EIBG_EISBG: 'EIBG/EISBG (UPW1)',
    FORMAT_FMBG:       'FMBG (TPC5/EMN3)',
    FORMAT_IABG:       'IABG (IMW1)',
    FORMAT_ICTBG_NTL7: 'ICTBG (NTL7)',
    FORMAT_ICTBG_PSB9_MRP:      'ICTBG PSB9 Kaewarin',
    FORMAT_ICTBG_PSB9_SIRIRAHT: 'ICTBG PSB9 Siriraht',
    FORMAT_PRAPAPORN:           'Prapaporn (PSB7)',
}

# 單 PLANT 檔案 (PLANT 從檔名比對)
SINGLE_PLANT_FORMATS = {
    FORMAT_KETWADEE, FORMAT_KANYANAT, FORMAT_WEERAYA, FORMAT_NBQ1, FORMAT_PSBG,
    FORMAT_EIBG_EISBG, FORMAT_PRAPAPORN,
}
# 多 PLANT 檔案 (PLANT 從檔案每列讀)
MULTI_PLANT_FORMATS = {
    FORMAT_INDIA_IAI1, FORMAT_PSW1_CEW1, FORMAT_MWC1IPC1, FORMAT_SVC1PWC1_DIODE_MOS,
    FORMAT_FMBG, FORMAT_IABG, FORMAT_ICTBG_NTL7, FORMAT_ICTBG_PSB9_MRP,
    FORMAT_ICTBG_PSB9_SIRIRAHT,
}


# ---------------------------------------------------------------------------
# 格式自動偵測
# ---------------------------------------------------------------------------

def _cell_str(ws, row, col):
    """安全取 cell 字串值"""
    try:
        v = ws.cell(row=row, column=col).value
        return str(v).strip() if v is not None else ''
    except Exception:
        return ''


def detect_format(filepath):
    """
    自動偵測 Delta Forecast 檔案屬於哪種格式。
    Returns: FORMAT_* 常數之一，無法識別則回傳 None
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheets = wb.sheetnames

        # === 最明確: 兩個 sheet 組合 ===
        if 'Diode' in sheets and 'MOS' in sheets:
            wb.close()
            return FORMAT_SVC1PWC1_DIODE_MOS

        # === PSW1/CEW1 sheet → PSW1+CEW1 variant ===
        if 'PSW1' in sheets or 'CEW1' in sheets:
            wb.close()
            return FORMAT_PSW1_CEW1

        # === MRP sheet: 區分 Ketwadee vs MWC1IPC1 variant ===
        if 'MRP' in sheets:
            ws_mrp = wb['MRP']
            h1_mrp = _cell_str(ws_mrp, 1, 1).upper()
            # MRP sheet 有 REQUEST ITEM → MWC1IPC1 variant (不管 h1 是 PLANT 或 PARTNO)
            if h1_mrp in ('PLANT', 'PARTNO'):
                for c in range(1, 30):
                    if _cell_str(ws_mrp, 1, c).upper() == 'REQUEST ITEM':
                        wb.close()
                        return FORMAT_MWC1IPC1
            # h1=PARTNO 但沒有 REQUEST ITEM → 不是 Ketwadee, 跳過
            if h1_mrp == 'PARTNO':
                pass  # fall through to other checks
            else:
                # Default: Ketwadee (h1=NO 或 BUYER)
                wb.close()
                return FORMAT_KETWADEE

        # === PSB9_MRP* sheet → ICTBG PSB9 Kaewarin ===
        for s in sheets:
            if s.startswith('PSB9_MRP'):
                wb.close()
                return FORMAT_ICTBG_PSB9_MRP

        # === PAN JIT sheet: 分辨 India IAI1 vs NBQ1 ===
        if 'PAN JIT' in sheets:
            ws = wb['PAN JIT']
            h1 = _cell_str(ws, 1, 1)
            h3 = _cell_str(ws, 1, 3)
            h13 = _cell_str(ws, 1, 13)
            # India IAI1: col 3 = PLANT, col 13 = Request
            if h3.upper() == 'PLANT' and h13.lower() == 'request':
                wb.close()
                return FORMAT_INDIA_IAI1
            # NBQ1: col 1 = PARTNO, no PLANT/Request
            if h1.upper() == 'PARTNO':
                wb.close()
                return FORMAT_NBQ1

        # === Sheet1: 多種格式 ===
        if 'Sheet1' in sheets:
            ws = wb['Sheet1']
            h1 = _cell_str(ws, 1, 1)
            h6 = _cell_str(ws, 1, 6)
            h9 = _cell_str(ws, 1, 9)
            h10 = _cell_str(ws, 1, 10)
            h11 = _cell_str(ws, 1, 11)
            h12 = _cell_str(ws, 1, 12)
            h13 = _cell_str(ws, 1, 13)
            h15 = _cell_str(ws, 1, 15)
            h24 = _cell_str(ws, 1, 24)

            # EIBG/EISBG: col 1 = ITEM, col 11 = OTW (flat, Demand only)
            if h1.upper() == 'ITEM' and h11.upper() == 'OTW':
                wb.close()
                return FORMAT_EIBG_EISBG

            # MWC1IPC1: col 1 = PLANT, col 6 = REQUEST ITEM
            if h1.upper() == 'PLANT' and h6.upper() == 'REQUEST ITEM':
                wb.close()
                return FORMAT_MWC1IPC1

            # ICTBG NTL7: col 1 = PLANT, col 10 = REQUEST ITEM (4 rows/part)
            if h1.upper() == 'PLANT' and h10.upper() == 'REQUEST ITEM':
                wb.close()
                return FORMAT_ICTBG_NTL7

            # FMBG: col 1 = PLANT, col 12 = REQUEST ITEM, col 9 = 出貨 (3 rows/part)
            if h1.upper() == 'PLANT' and h12.upper() == 'REQUEST ITEM':
                wb.close()
                return FORMAT_FMBG

            # PSW1+CEW1: col 12 = Status
            if h12 == 'Status':
                wb.close()
                return FORMAT_PSW1_CEW1

            # Weeraya / Prapaporn: col 1 = Plant, col 12 = TYPE
            # 區分方式: Prapaporn marker 帶字母前綴 (A-Demand), Weeraya 無 (Demand)
            if h1.lower() == 'plant' and h12.upper() == 'TYPE':
                first_marker = _cell_str(ws, 2, 12)
                if first_marker and first_marker[0].isalpha() and '-' in first_marker[:3]:
                    wb.close()
                    return FORMAT_PRAPAPORN
                wb.close()
                return FORMAT_WEERAYA

            # ICTBG PSB9 Siriraht: col 15 = REQUEST ITEM (1Demand/2Supply/3Balance)
            if h15.upper() == 'REQUEST ITEM':
                wb.close()
                return FORMAT_ICTBG_PSB9_SIRIRAHT

            # PSBG: col 15 = Filter (values: 1.Demand/2.Supply/3.Net)
            if h15.lower() == 'filter':
                wb.close()
                return FORMAT_PSBG

            # IABG: col 1 = PLANT, col 9 = SHIP, col 13 = PASSDUE (flat)
            if h1.upper() == 'PLANT' and h9.upper() == 'SHIP' and h13.upper() == 'PASSDUE':
                wb.close()
                return FORMAT_IABG

            # Kanyanat: col 24 = TYPE (col 1 = NO, col 3 = Plant)
            if h24.upper() == 'TYPE':
                wb.close()
                return FORMAT_KANYANAT

            # EIBG variant: c1=PLANT, c4=PARTNO, c5=VENDOR PARTNO (multi-row w/ marker)
            h4 = _cell_str(ws, 1, 4).upper()
            h5 = _cell_str(ws, 1, 5).upper()
            if h1.upper() == 'PLANT' and h4 == 'PARTNO' and h5 == 'VENDOR PARTNO':
                wb.close()
                return FORMAT_EIBG_EISBG

        # === 工作表1: 可能是 MWC1IPC1 variant (header 可能在 row 2) ===
        if '工作表1' in sheets:
            ws_tw = wb['工作表1']
            for hr in (1, 2):
                for c in range(1, 25):
                    if _cell_str(ws_tw, hr, c).upper() == 'REQUEST ITEM':
                        wb.close()
                        return FORMAT_MWC1IPC1

        # === Fallback: 掃描所有 sheet 找 Kanyanat-like header (TYPE col + A-Demand) ===
        for sn in sheets:
            ws_fb = wb[sn]
            for hr in range(1, 10):
                for c in range(1, 30):
                    hv = _cell_str(ws_fb, hr, c).upper()
                    if hv == 'TYPE':
                        # 確認下一列有 A-Demand marker
                        marker_val = _cell_str(ws_fb, hr + 1, c)
                        if marker_val == 'A-Demand':
                            wb.close()
                            return FORMAT_KANYANAT
                        break

        wb.close()
        return None
    except Exception:
        return None


def detect_buyer(filepath):
    """
    向後相容: 舊的 detect_buyer() 函式。
    只回傳 'Ketwadee' / 'Kanyanat' / 'Weeraya' / None，其他新格式回傳 None。
    """
    fmt = detect_format(filepath)
    return {
        FORMAT_KETWADEE: 'Ketwadee',
        FORMAT_KANYANAT: 'Kanyanat',
        FORMAT_WEERAYA: 'Weeraya',
    }.get(fmt)


# ---------------------------------------------------------------------------
# 日期欄位工具
# ---------------------------------------------------------------------------

MONTH_NAMES = ('JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN',
               'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC')

# 每個格式: [(sheet_name, date_start_col), ...]
# date_start_col 是 1-based，從該欄開始掃描至 max_col
# 非日期欄會被 _normalize_date_header 自動過濾掉
_FORMAT_SHEETS = {
    FORMAT_KETWADEE:           [('MRP', 16)],
    # Kanyanat: sheet/header row 動態 (Sheet1 或其他 sheet, header 可能不在 row 1)
    FORMAT_KANYANAT:           [],
    FORMAT_WEERAYA:            [('Sheet1', 13)],
    FORMAT_INDIA_IAI1:         [('PAN JIT', 14)],
    FORMAT_PSW1_CEW1:          [('Sheet1', 14), ('PSW1', 10), ('CEW1', 10)],
    # FORMAT_MWC1IPC1: sheet/header 動態 (Sheet1/MRP/工作表1, header row 1 or 2)
    FORMAT_MWC1IPC1:           [],
    FORMAT_NBQ1:               [('PAN JIT', 16)],
    FORMAT_SVC1PWC1_DIODE_MOS: [('Diode', 9), ('MOS', 9)],
    FORMAT_PSBG:               [('Sheet1', 16)],
    FORMAT_EIBG_EISBG:         [('Sheet1', 12)],
    FORMAT_FMBG:               [('Sheet1', 16)],
    FORMAT_IABG:               [('Sheet1', 13)],
    FORMAT_ICTBG_NTL7:         [('Sheet1', 13)],
    # FORMAT_ICTBG_PSB9_MRP: sheet 名稱動態 (PSB9_MRP*)，extract_dates 時特殊處理
    FORMAT_ICTBG_PSB9_MRP:     [],
    FORMAT_ICTBG_PSB9_SIRIRAHT:[('Sheet1', 16)],
    FORMAT_PRAPAPORN:          [('Sheet1', 16)],
}


def _get_ictbg_psb9_mrp_sheet(wb):
    """找到 PSB9_MRP* sheet (名稱可能包含日期，例 'PSB9_MRP 0413')"""
    for s in wb.sheetnames:
        if s.startswith('PSB9_MRP'):
            return s
    return wb.sheetnames[0] if wb.sheetnames else None


def read_date_cols_from_template(template_path):
    """
    從匯總格式模板的 header row 動態讀取日期欄位。(備用，正式流程用 buyer 檔案提取)
    Returns: list of normalized date column names (e.g. ['PASSDUE','20260406',...,'MAR'])
    """
    wb = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    date_cols = []
    for cell in ws[1]:
        if cell.column >= 10 and cell.value is not None:  # J (col 10) onwards
            date_cols.append(_normalize_date_header(cell.value))
    wb.close()
    return [d for d in date_cols if d is not None]


def _normalize_date_header(val):
    """將不同源頭的日期格式統一為匯總格式的 key"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y%m%d')
    # float/int YYYYMMDD (e.g. 20260420.0 → '20260420')
    if isinstance(val, (int, float)):
        ival = int(val)
        s_int = str(ival)
        if s_int.isdigit() and len(s_int) == 8:
            return s_int
    s = str(val).strip()
    if not s:
        return None
    s_upper = s.upper()
    if 'PAST' in s_upper or 'PASSDUE' in s_upper or 'OVER DUE' in s_upper:
        return 'PASSDUE'
    if s.isdigit() and len(s) == 8:
        return s
    # MM/DD/YY 或 MM/DD/YYYY (e.g. '03/30/26' → '20260330')
    if '/' in s and ' ' not in s:
        parts = s.split('/')
        if len(parts) == 3:
            try:
                m, d, y = int(parts[0]), int(parts[1]), int(parts[2])
                if 1 <= m <= 12 and 1 <= d <= 31:
                    if y < 100:
                        y += 2000
                    if 2000 <= y <= 2099:
                        return f'{y:04d}{m:02d}{d:02d}'
            except ValueError:
                pass
    # "2026-JUL" → "JUL" (不限定年份，自動適用任何年度)
    if '-' in s and len(s) >= 5:
        parts = s.split('-')
        if len(parts) == 2 and len(parts[1]) == 3:
            month = parts[1].upper()
            if month in MONTH_NAMES:
                return month
    # 直接就是月份名 (JUL, AUG, ...)
    if s_upper in MONTH_NAMES:
        return s_upper
    return None


def _sort_date_cols(dates, anchor_date=None):
    """
    方案二 (匯總格式模式): 固定 26 欄 = PASSDUE + 16 週 + 9 月

    - PASSDUE: 源檔裡既有的必要欄位 (來自 "PASSDUE"/"PAST DUE" 等標籤)，
               直接對應，不會有週日期折入此欄。
    - W1~W16 (K~Z): **以來源檔最早的週日期 (Monday) 為 W1 起點**，產生 16 個
                    連續週一 (YYYYMMDD)。若沒有任何週日期則 fallback 使用 anchor
                    所在週的週一。
    - M1~M9 (AA~AI): 從 W16 次週起算的 9 個月份標籤 (MMM 縮寫)

    來源檔案的日期欄會依據落點映射到固定欄位：
    - 落在 W1~W16 範圍 → 對應週 Monday (若非 Monday 則取該週週一)
    - 落在 M1~M9 範圍 → 對應月份標籤 (多筆會在 reader 累加)
    - 超出 M9 範圍 (太未來) → 丟棄
    - 早於 W1 (理論上不應發生，因為 W1=最早週) → 丟棄並警告
    - 月份標籤不在 M1~M9 → 丟棄

    Args:
        dates: set/iterable of normalized date keys from source files
        anchor_date: datetime, 備援基準 (預設 = datetime.now()，僅在無任何週日期時使用)

    Returns:
        tuple(date_cols, conversions)
        - date_cols: ['PASSDUE', YYYYMMDD×16, MMM×9] 固定 26 欄
        - conversions: dict {原始 key: 轉換後 key 或 None}
    """
    # 1. 找出來源檔最早的週日期 (Monday) 作為 W1 起點
    weekly_source = []
    for d in dates:
        if isinstance(d, str) and d.isdigit() and len(d) == 8:
            try:
                weekly_source.append(datetime.strptime(d, '%Y%m%d'))
            except ValueError:
                pass

    if weekly_source:
        earliest = min(weekly_source)
        # 對齊到該週的週一 (weekday 0 = Monday)
        first_monday = earliest - timedelta(days=earliest.weekday())
        first_monday = first_monday.replace(hour=0, minute=0, second=0, microsecond=0)
    else:
        # Fallback: 使用 anchor 所在週的週一
        anchor = anchor_date or datetime.now()
        first_monday = anchor - timedelta(days=anchor.weekday())
        first_monday = first_monday.replace(hour=0, minute=0, second=0, microsecond=0)

    # 2. 固定 16 週 Monday (K~Z)
    weekly_mondays = [first_monday + timedelta(weeks=i) for i in range(16)]
    weekly_keys = [m.strftime('%Y%m%d') for m in weekly_mondays]
    last_monday = weekly_mondays[-1]
    w16_sunday = last_monday + timedelta(days=6)

    # 3. 固定 9 個月份 (AA~AI, 從 W16 次週起算)
    w17_monday = last_monday + timedelta(weeks=1)
    monthly_keys = []
    cy, cm = w17_monday.year, w17_monday.month
    for _ in range(9):
        monthly_keys.append(MONTH_NAMES[cm - 1])
        cm += 1
        if cm > 12:
            cm = 1
            cy += 1
    monthly_key_set = set(monthly_keys)

    final_cols = ['PASSDUE'] + weekly_keys + monthly_keys

    # 4. 建立 conversions
    conversions = {}
    rejected = []
    folded_to_month = {}

    for d in dates:
        if d == 'PASSDUE':
            continue
        if d in MONTH_NAMES:
            if d not in monthly_key_set:
                conversions[d] = None
                rejected.append(d)
            continue
        if isinstance(d, str) and d.isdigit() and len(d) == 8:
            try:
                dt = datetime.strptime(d, '%Y%m%d')
            except ValueError:
                conversions[d] = None
                rejected.append(d)
                continue
            if dt < first_monday:
                # 不應發生 (W1 = 最早週)，但若發生則丟棄並警告
                conversions[d] = None
                rejected.append(d)
            elif dt <= w16_sunday:
                days_from_w1 = (dt - first_monday).days
                week_idx = days_from_w1 // 7
                target_key = weekly_keys[week_idx]
                if target_key != d:
                    conversions[d] = target_key
            else:
                # 超出 W16 → 折疊到對應月份
                m_label = MONTH_NAMES[dt.month - 1]
                if m_label in monthly_key_set:
                    conversions[d] = m_label
                    folded_to_month.setdefault(m_label, []).append(d)
                else:
                    conversions[d] = None
                    rejected.append(d)
        else:
            conversions[d] = None
            rejected.append(d)

    print(f"   📅 方案二起始週 (W1): {weekly_keys[0]}, 結束週 (W16): {weekly_keys[-1]}")
    print(f"   📅 月份範圍 (M1~M9): {monthly_keys[0]} ~ {monthly_keys[-1]}")
    if rejected:
        print(f"   ⚠️ 丟棄超出範圍的日期: {rejected}")
    if folded_to_month:
        for m, src_list in folded_to_month.items():
            print(f"   🔄 折疊到 {m}: {src_list}")

    return final_cols, conversions


def extract_dates_from_files(detected_files):
    """
    從每個檔案的 header 動態提取所有日期欄位 (支援全部 15 種格式)。

    Args:
        detected_files: list of (filepath, format_const) tuples

    Returns:
        tuple(date_cols, per_file_dates, warnings)
        - date_cols: list — 排序後的統一日期欄位 (聯集)
        - per_file_dates: dict — {filename: set(date_keys)}
        - warnings: list — 日期不一致的警告訊息
    """
    per_file_dates = {}

    for filepath, fmt in detected_files:
        sheet_specs = _FORMAT_SHEETS.get(fmt, [])
        file_key = os.path.basename(filepath)

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        file_dates = set()

        # ICTBG PSB9 MRP 的 sheet 名稱動態 (PSB9_MRP + 日期)
        if fmt == FORMAT_ICTBG_PSB9_MRP:
            sheet_name = _get_ictbg_psb9_mrp_sheet(wb)
            sheet_specs = [(sheet_name, 15)] if sheet_name else []

        # Kanyanat: sheet/header row 動態
        if fmt == FORMAT_KANYANAT:
            ws_k, hr_k = _find_kanyanat_layout(wb)
            if ws_k:
                cols_k = _scan_header_columns(ws_k, hr_k)
                type_col = cols_k.get('marker', 23)
                scan_start = type_col + 2
                for col_idx, cell in enumerate(ws_k[hr_k], start=1):
                    if col_idx < scan_start:
                        continue
                    v = getattr(cell, 'value', None)
                    if v is None:
                        continue
                    norm = _normalize_date_header(v)
                    if norm is not None:
                        file_dates.add(norm)
                wb.close()
                per_file_dates[file_key] = file_dates
                print(f"  {FORMAT_LABELS.get(fmt, fmt)} [{file_key}]: "
                      f"偵測到 {len(file_dates)} 個日期欄位")
                continue

        # MWC1IPC1: sheet/header 動態 (Sheet1/MRP/工作表1, header row 可能不在 row 1)
        if fmt == FORMAT_MWC1IPC1:
            ws_m, hr_m, cols_m = _find_mwc1ipc1_layout(wb)
            if ws_m and cols_m:
                marker_col = cols_m['marker']
                scan_start = marker_col + 2  # 1-based
                for col_idx, cell in enumerate(ws_m[hr_m], start=1):
                    if col_idx < scan_start:
                        continue
                    v = getattr(cell, 'value', None)
                    if v is None:
                        continue
                    norm = _normalize_date_header(v)
                    if norm is not None:
                        file_dates.add(norm)
                wb.close()
                per_file_dates[file_key] = file_dates
                print(f"  {FORMAT_LABELS.get(fmt, fmt)} [{file_key}]: "
                      f"偵測到 {len(file_dates)} 個日期欄位")
                continue

        for sheet_name, start_col in sheet_specs:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for col_idx, cell in enumerate(ws[1], start=1):
                if col_idx < start_col:
                    continue
                v = getattr(cell, 'value', None)
                if v is None:
                    continue
                norm = _normalize_date_header(v)
                if norm is not None and norm not in file_dates:
                    file_dates.add(norm)
        wb.close()

        per_file_dates[file_key] = file_dates
        print(f"  {FORMAT_LABELS.get(fmt, fmt)} [{file_key}]: 偵測到 {len(file_dates)} 個日期欄位")

    # 聯集 = 所有檔案的日期
    all_dates = set()
    for dates in per_file_dates.values():
        all_dates |= dates

    # 比對差異 (僅顯示警告，不阻擋)
    warnings = []
    keys = list(per_file_dates.keys())
    for i, k1 in enumerate(keys):
        for k2 in keys[i + 1:]:
            only_in_k1 = per_file_dates[k1] - per_file_dates[k2]
            only_in_k2 = per_file_dates[k2] - per_file_dates[k1]
            if only_in_k1:
                msg = f'{k1} 有但 {k2} 沒有的日期: {sorted(only_in_k1)}'
                warnings.append(msg)
            if only_in_k2:
                msg = f'{k2} 有但 {k1} 沒有的日期: {sorted(only_in_k2)}'
                warnings.append(msg)

    if not warnings:
        print(f"  ✅ 所有檔案日期完全一致 ({len(all_dates)} 個日期)")
    else:
        print(f"  ⚠️ 日期不完全一致，共 {len(warnings)} 個差異，已取聯集 ({len(all_dates)} 個日期)")

    date_cols, conversions = _sort_date_cols(all_dates)
    return date_cols, per_file_dates, warnings, conversions


# 舊名稱保留做向後相容 (app.py 可能還在用)
def extract_dates_from_buyer_files(buyer_files):
    """Deprecated: 使用 extract_dates_from_files()"""
    # 將舊 dict 格式轉換為新 list 格式
    old_format_map = {
        'Ketwadee': FORMAT_KETWADEE,
        'Kanyanat': FORMAT_KANYANAT,
        'Weeraya': FORMAT_WEERAYA,
    }
    detected_files = [
        (fp, old_format_map.get(name, FORMAT_KETWADEE))
        for name, fp in buyer_files.items()
    ]
    return extract_dates_from_files(detected_files)


# ---------------------------------------------------------------------------
# Buyer 讀取器
# ---------------------------------------------------------------------------

def _build_date_col_map(ws, start_col, date_cols, conversions=None):
    """
    建立 column → date_key 的映射。

    方案二支援多個 source 欄位映射到同一個 target key (折疊/累加)：
    - 例: 來源 20260406 + 20260407 都映射到同一個週 Monday → reader 需累加兩者
    - 例: 來源 20261015 + 20261101 都映射到 OCT 月份欄 → reader 需累加

    conversions: dict {原始 key: 轉換後 key 或 None (丟棄)}。
    """
    date_col_map = {}
    date_cols_set = set(date_cols)
    for col_idx, cell in enumerate(ws[1], start=1):
        if col_idx < start_col:
            continue
        v = getattr(cell, 'value', None)
        if v is None:
            continue
        norm = _normalize_date_header(v)
        if norm is None:
            continue
        # 套用轉換 (例: 20261015 → OCT, 20260330 → PASSDUE)
        if conversions and norm in conversions:
            norm = conversions[norm]
            if norm is None:
                continue  # 被丟棄
        if norm in date_cols_set:
            date_col_map[col_idx] = norm
    return date_col_map


def _read_row_dates(row, date_col_map):
    """
    從單一資料 row 讀取日期欄位值，支援累加 (多個 source 欄位 → 同一 target key)。

    Args:
        row: tuple of Cell objects (from openpyxl iter_rows, values_only=False)
        date_col_map: dict {col_idx (1-based): target_date_key}

    Returns:
        dict {target_date_key: accumulated_number}
    """
    data = {}
    for col_idx, date_key in date_col_map.items():
        if col_idx - 1 >= len(row):
            continue
        v = row[col_idx - 1].value
        if v is None or v == '':
            continue
        try:
            v_num = float(v)
        except (ValueError, TypeError):
            continue
        data[date_key] = data.get(date_key, 0) + v_num
    return data


def _to_partno(val):
    """將 PARTNO 轉為數字（如果可能），與客戶格式一致"""
    if val is None:
        return ''
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return str(val)


def extract_plant_codes_from_regions(regions):
    """
    從 mapping 表的 region 欄位提取 PLANT 代碼。
    例: ['PSB5 泰國', 'IPC1 東莞'] → ['PSB5', 'IPC1']
    """
    codes = []
    for r in regions:
        if not r:
            continue
        parts = str(r).split()
        if parts:
            codes.append(parts[0])
    return codes


def match_plants_in_filename(filepath, plant_codes):
    """
    從檔名中搜尋 PLANT 代碼 (case-insensitive)。
    回傳檔名中出現的所有 PLANT 代碼，依長度降冪排序 (長的優先)。

    Args:
        filepath: 檔案路徑
        plant_codes: 有效的 PLANT 代碼清單 (如 ['PSB5', 'PSB7', 'IPC1'])

    Returns:
        list of matched codes, e.g. ['PSB5'] or ['IAI1', 'UPI2', 'DFI1']
    """
    if not plant_codes:
        return []
    fn = os.path.splitext(os.path.basename(filepath))[0].upper()
    matched = [code for code in plant_codes if code.upper() in fn]
    matched.sort(key=len, reverse=True)
    return matched


def _scan_header_columns(ws, header_row=1):
    """掃描 header row, 回傳偵測到的欄位位置 (0-indexed dict).

    支援所有 Delta buyer 格式的常見欄位名稱。
    Returns dict with possible keys:
        plant, partno, vendor, stock, otw, marker, buyer_col, item
    """
    cols = {}
    for c, cell in enumerate(ws[header_row]):
        h = str(cell.value or '').strip().upper()
        if not h:
            continue
        # PLANT
        if h == 'PLANT' and 'plant' not in cols:
            cols['plant'] = c
        # PARTNO — 優先 RAW MATERIAL, 再 PARTNO/PART NO/PN/P/N
        elif ('RAW MATERIAL' in h or h in ('PARTNO', 'PART NO', 'PN')) and 'partno' not in cols:
            cols['partno'] = c
        # VENDOR — VENDOR PART / VENDOR PARTNO / MFG / Vendor Part
        elif h in ('VENDOR PART', 'VENDOR PARTNO', 'MFG') and 'vendor' not in cols:
            cols['vendor'] = c
        # STOCK — STOCK / PLANT STOCK / STOCK QTY / PSB9 STOCK / TOTAL STOCK
        elif ('STOCK' in h and 'vendor' not in h.lower()) and 'stock' not in cols:
            cols['stock'] = c
        # OTW — OTW / SHIP IN TRANSIT / ON THE WAY / ON-WAY
        elif (h in ('OTW',) or 'SHIP IN TRANSIT' in h or 'ON THE WAY' in h or 'ON-WAY' in h) and 'otw' not in cols:
            cols['otw'] = c
        # MARKER — TYPE / FILTER / REQUEST / REQUEST ITEM / STATUS / 類別 / Date(PSW1)
        elif h in ('TYPE', 'FILTER', 'REQUEST', 'REQUEST ITEM', 'STATUS', '\u985e\u5225', 'DATE') and 'marker' not in cols:
            cols['marker'] = c
        # BUYER
        elif h == 'BUYER' and 'buyer_col' not in cols:
            cols['buyer_col'] = c
        # ITEM / NO
        elif h in ('NO', 'NO.', 'ITEM') and 'item' not in cols:
            cols['item'] = c
    return cols


def _read_ketwadee(filepath, date_cols, buyer_label=None, plant_code=None, conversions=None):
    """讀取 PSB5 Ketwadee: MRP sheet, 3 rows/part (Demand/Supply/Net).
    動態偵測欄位位置 — 支援有/無 NO 欄、欄位漂移。
    """
    # 某些檔案 read_only 模式 max_row=None, 改用一般模式
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['MRP']

    # --- 動態偵測 header 欄位 ---
    cols = _scan_header_columns(ws, 1)
    partno_idx = cols.get('partno', 1)
    vendor_idx = cols.get('vendor', 2)
    stock_idx = cols.get('stock', 6)
    filter_idx = cols.get('marker', 13)   # Filter = marker
    # 日期從 marker 欄之後掃描
    date_start_col = filter_idx + 2  # 1-based col
    date_col_map = _build_date_col_map(ws, date_start_col, date_cols, conversions)

    results = []
    max_col = ws.max_column or 46
    max_row = ws.max_row or 500
    rows = list(ws.iter_rows(min_row=2, max_row=max_row,
                             min_col=1, max_col=max_col, values_only=False))
    i = 0
    while i < len(rows):
        row = rows[i]
        filter_val = row[filter_idx].value if len(row) > filter_idx else None

        if filter_val == 'Demand':
            part_no = row[partno_idx].value
            vendor_part = row[vendor_idx].value
            stock = row[stock_idx].value or 0

            demand = _read_row_dates(row, date_col_map)
            # 向前掃描找 Supply / Net
            supply = {}
            advance = 1
            for j in range(i + 1, min(i + 5, len(rows))):
                fj = rows[j][filter_idx].value if len(rows[j]) > filter_idx else None
                if fj == 'Supply':
                    supply = _read_row_dates(rows[j], date_col_map)
                elif fj == 'Demand':
                    break
                advance += 1

            results.append({
                'buyer': buyer_label or 'Ketwadee', 'plant': plant_code or 'PSB5',
                'part_no': _to_partno(part_no),
                'vendor_part': str(vendor_part) if vendor_part else '',
                'stock': stock, 'on_way': None,
                'demand': demand, 'supply': supply,
            })
            i += advance
        else:
            i += 1

    wb.close()
    return results


def _find_kanyanat_layout(wb):
    """找到 Kanyanat 的 sheet / header row (支援 Sheet1 或其他 sheet、header 不在 row 1)。

    Returns:
        tuple(ws, header_row) or (None, None)
    """
    for sn in wb.sheetnames:
        ws = wb[sn]
        for hr in range(1, 10):
            cols = _scan_header_columns(ws, hr)
            if 'marker' in cols:
                # 確認下一列有 A-Demand marker
                m_idx = cols['marker']
                try:
                    next_val = ws.cell(hr + 1, m_idx + 1).value
                except Exception:
                    next_val = None
                if next_val == 'A-Demand':
                    return ws, hr
    return None, None


def _read_kanyanat(filepath, date_cols, buyer_label=None, plant_code=None, conversions=None):
    """讀取 PSB7 Kanyanat: 4 rows/part (A-Demand marker).
    動態偵測 sheet、header row、欄位位置。
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws, header_row = _find_kanyanat_layout(wb)
    if ws is None:
        # fallback: Sheet1, row 1
        ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb[wb.sheetnames[0]]
        header_row = 1

    # --- 動態偵測 header 欄位 ---
    cols = _scan_header_columns(ws, header_row)
    partno_idx = cols.get('partno', 4)
    vendor_idx = cols.get('vendor', 5)
    stock_idx = cols.get('stock')
    type_idx = cols.get('marker', 23)  # TYPE column
    date_scan_start = type_idx + 2
    date_col_map = _build_date_col_map(ws, date_scan_start, date_cols, conversions)

    results = []
    max_col = ws.max_column
    rows = list(ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row,
                             min_col=1, max_col=max_col, values_only=False))
    i = 0
    while i < len(rows):
        row = rows[i]
        type_val = row[type_idx].value if len(row) > type_idx else None

        if type_val == 'A-Demand':
            part_no = row[partno_idx].value if len(row) > partno_idx else None
            vendor_part = row[vendor_idx].value if len(row) > vendor_idx else None
            stock = row[stock_idx].value if stock_idx and len(row) > stock_idx else None

            demand = _read_row_dates(row, date_col_map)
            # 向前掃描找 B-Supply
            supply = {}
            advance = 1
            for j in range(i + 1, min(i + 6, len(rows))):
                mj = rows[j][type_idx].value if len(rows[j]) > type_idx else None
                if mj and 'B-' in str(mj):
                    supply = _read_row_dates(rows[j], date_col_map)
                elif mj == 'A-Demand':
                    break
                advance += 1

            results.append({
                'buyer': buyer_label or 'Kanyanat', 'plant': plant_code or 'PSB7',
                'part_no': _to_partno(part_no),
                'vendor_part': str(vendor_part) if vendor_part else '',
                'stock': stock, 'on_way': None,
                'demand': demand, 'supply': supply,
            })
            i += advance
        else:
            i += 1

    wb.close()
    return results


def _read_weeraya(filepath, date_cols, buyer_label=None, plant_code=None, conversions=None):
    """讀取 PSB7 Weeraya: Sheet1, 4~5 rows/part, col 12 = TYPE (Demand marker).
    動態偵測欄位位置 (新版 col2=PN, 舊版 col4=PN):
      Raw Material(P/N) → partno, Vendor Part → vendor, Total stock → stock.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['Sheet1']

    # --- 動態偵測 header 欄位 (0-indexed) ---
    partno_idx = vendor_idx = stock_idx = type_idx = None
    for c, cell in enumerate(ws[1]):
        h = str(cell.value or '').strip().upper()
        if 'RAW MATERIAL' in h or h == 'PART NO':
            partno_idx = c
        elif h == 'VENDOR PART':
            vendor_idx = c
        elif 'TOTAL STOCK' in h:
            stock_idx = c
        elif h == 'TYPE':
            type_idx = c
    if partno_idx is None: partno_idx = 3
    if vendor_idx is None: vendor_idx = 4
    if stock_idx is None: stock_idx = 12
    if type_idx is None: type_idx = 11

    # 日期從 TYPE 欄之後掃描 (跳過 Total stock / PASSDUE 等非日期欄)
    date_scan_start = type_idx + 2   # 1-based col
    date_col_map = _build_date_col_map(ws, date_scan_start, date_cols, conversions)

    results = []
    max_col = ws.max_column
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row,
                             min_col=1, max_col=max_col, values_only=False))
    i = 0
    while i < len(rows):
        row = rows[i]
        type_val = row[type_idx].value if len(row) > type_idx else None

        if type_val == 'Demand':
            part_no = row[partno_idx].value
            vendor_part = row[vendor_idx].value
            stock = row[stock_idx].value or 0

            demand = _read_row_dates(row, date_col_map)

            # 向前掃描找 Forecast Conf / Net Demand (相容 4-row 與 5-row 版本)
            supply = {}
            balance_data = {}
            advance = 1
            for j in range(i + 1, min(i + 6, len(rows))):
                mj = rows[j][type_idx].value if len(rows[j]) > type_idx else None
                if mj == 'Forecast Conf':
                    supply = _read_row_dates(rows[j], date_col_map)
                elif mj == 'Net Demand':
                    balance_data = _read_row_dates(rows[j], date_col_map)
                elif mj == 'Demand':
                    break
                advance += 1

            results.append({
                'buyer': buyer_label or 'Weeraya', 'plant': plant_code or 'PSB7',
                'part_no': _to_partno(part_no),
                'vendor_part': str(vendor_part) if vendor_part else '',
                'stock': stock, 'on_way': None,
                'demand': demand, 'supply': supply,
                'balance_override': balance_data,
            })
            i += advance
        else:
            i += 1

    wb.close()
    return results


def _read_prapaporn(filepath, date_cols, buyer_label=None, plant_code=None, conversions=None):
    """讀取 PSB7 Prapaporn: Sheet1, 4 rows/part (A-Demand/B-Forecast Conf/D-Net Demand/F-Remark)
    動態偵測欄位: Raw Material(P/N) → partno, Vendor Part → vendor,
    Total stock → stock, ON THE WAY → on_way, TYPE = marker.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['Sheet1']

    # --- 動態偵測 header 欄位 (0-indexed) ---
    partno_idx = vendor_idx = stock_idx = type_idx = onway_idx = None
    for c, cell in enumerate(ws[1]):
        h = str(cell.value or '').strip().upper()
        if 'RAW MATERIAL' in h or h == 'PART NO':
            partno_idx = c
        elif h == 'VENDOR PART':
            vendor_idx = c
        elif 'TOTAL STOCK' in h:
            stock_idx = c
        elif h == 'TYPE':
            type_idx = c
        elif 'ON THE WAY' in h or 'ON-WAY' in h or h == 'OTW':
            onway_idx = c
    if partno_idx is None: partno_idx = 3
    if vendor_idx is None: vendor_idx = 4
    if stock_idx is None: stock_idx = 12
    if type_idx is None: type_idx = 11

    date_scan_start = type_idx + 2
    date_col_map = _build_date_col_map(ws, date_scan_start, date_cols, conversions)

    results = []
    max_col = ws.max_column
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row,
                             min_col=1, max_col=max_col, values_only=False))
    i = 0
    while i < len(rows):
        row = rows[i]
        type_val = row[type_idx].value if len(row) > type_idx else None

        if type_val == 'A-Demand':
            part_no = row[partno_idx].value
            vendor_part = row[vendor_idx].value
            stock = row[stock_idx].value or 0
            on_way = row[onway_idx].value if onway_idx and len(row) > onway_idx else None

            demand = _read_row_dates(row, date_col_map)

            # 向前掃描找 B-Forecast Conf / D-Net Demand
            supply = {}
            balance = {}
            advance = 1
            for j in range(i + 1, min(i + 6, len(rows))):
                mj = rows[j][type_idx].value if len(rows[j]) > type_idx else None
                if mj == 'B-Forecast Conf':
                    supply = _read_row_dates(rows[j], date_col_map)
                elif mj == 'D-Net Demand':
                    balance = _read_row_dates(rows[j], date_col_map)
                elif mj == 'A-Demand':
                    break
                advance += 1

            results.append({
                'buyer': buyer_label or 'Prapaporn',
                'plant': plant_code or 'PSB7',
                'part_no': _to_partno(part_no),
                'vendor_part': str(vendor_part) if vendor_part else '',
                'stock': stock, 'on_way': on_way or 0,
                'demand': demand, 'supply': supply,
                'balance_override': balance,
            })
            i += advance
        else:
            i += 1

    wb.close()
    return results


def _read_india_iai1(filepath, date_cols, buyer_label=None, plant_code=None, conversions=None):
    """
    讀取 India IAI1: PAN JIT sheet, 3 rows/part (Demand/Supply/Balance).
    動態偵測欄位位置。多 PLANT 檔案 — 每列讀 PLANT。
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['PAN JIT']

    # --- 動態偵測 header 欄位 ---
    cols = _scan_header_columns(ws, 1)
    plant_idx = cols.get('plant', 2)
    partno_idx = cols.get('partno', 3)
    vendor_idx = cols.get('vendor', 6)
    stock_idx = cols.get('stock', 10)
    marker_idx = cols.get('marker', 12)
    date_scan_start = marker_idx + 2
    date_col_map = _build_date_col_map(ws, date_scan_start, date_cols, conversions)

    results = []
    max_col = ws.max_column
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row,
                             min_col=1, max_col=max_col, values_only=False))
    i = 0
    while i < len(rows):
        row = rows[i]
        marker = row[marker_idx].value if len(row) > marker_idx else None

        if marker == 'Demand':
            row_plant = row[plant_idx].value if len(row) > plant_idx else None
            part_no = row[partno_idx].value if len(row) > partno_idx else None
            vendor_part = row[vendor_idx].value if len(row) > vendor_idx else None
            stock = row[stock_idx].value if len(row) > stock_idx else 0

            demand = _read_row_dates(row, date_col_map)
            # 向前掃描找 Supply / Balance
            supply = {}
            balance = {}
            advance = 1
            for j in range(i + 1, min(i + 5, len(rows))):
                mj = rows[j][marker_idx].value if len(rows[j]) > marker_idx else None
                if mj == 'Supply':
                    supply = _read_row_dates(rows[j], date_col_map)
                elif mj == 'Balance':
                    balance = _read_row_dates(rows[j], date_col_map)
                elif mj == 'Demand':
                    break
                advance += 1

            results.append({
                'buyer': buyer_label or 'India',
                'plant': str(row_plant).strip() if row_plant else (plant_code or ''),
                'part_no': _to_partno(part_no),
                'vendor_part': str(vendor_part) if vendor_part else '',
                'stock': stock or 0, 'on_way': None,
                'demand': demand, 'supply': supply,
                'balance_override': balance,
            })
            i += advance
        else:
            i += 1

    wb.close()
    return results


def _read_psw1_cew1(filepath, date_cols, buyer_label=None, plant_code=None, conversions=None):
    """
    讀取 PSW1+CEW1: 支援 Sheet1 或 PSW1/CEW1 分 sheet。
    動態偵測欄位位置。多 PLANT 檔案 — 每列讀 PLANT。
    支援 A-Demand 和 1.Demand 兩種 marker 風格。
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # 決定要處理哪些 sheet
    target_sheets = []
    if 'Sheet1' in wb.sheetnames:
        target_sheets.append('Sheet1')
    for sn in ('PSW1', 'CEW1'):
        if sn in wb.sheetnames:
            target_sheets.append(sn)
    if not target_sheets:
        target_sheets = [wb.sheetnames[0]]

    results = []
    for sheet_name in target_sheets:
        ws = wb[sheet_name]

        # --- 動態偵測 header 欄位 ---
        cols = _scan_header_columns(ws, 1)
        has_plant_col = 'plant' in cols
        plant_idx = cols.get('plant')  # None if no PLANT column
        partno_idx = cols.get('partno', 5)
        vendor_idx = cols.get('vendor', 7)
        marker_idx = cols.get('marker', 11)  # Status or Date
        stock_idx = cols.get('stock', marker_idx + 1)
        date_scan_start = marker_idx + 2
        if stock_idx and stock_idx > marker_idx:
            date_scan_start = stock_idx + 2
        date_col_map = _build_date_col_map(ws, date_scan_start, date_cols, conversions)

        max_col = ws.max_column
        rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row,
                                 min_col=1, max_col=max_col, values_only=False))
        i = 0
        while i < len(rows):
            row = rows[i]
            marker = row[marker_idx].value if len(row) > marker_idx else None
            m_str = str(marker).strip() if marker else ''

            # 相容 A-Demand 和 1.Demand 兩種風格
            is_demand = m_str in ('A-Demand', '1.Demand')
            if is_demand:
                # 有 PLANT 欄才從 row 讀, 否則用 sheet name
                if has_plant_col and plant_idx is not None:
                    row_plant = row[plant_idx].value if len(row) > plant_idx else None
                else:
                    row_plant = None
                # 沒讀到 PLANT → 用 sheet name 或 plant_code
                if not row_plant and sheet_name in ('PSW1', 'CEW1'):
                    row_plant = sheet_name
                part_no = row[partno_idx].value if len(row) > partno_idx else None
                vendor_part = row[vendor_idx].value if len(row) > vendor_idx else None
                stock = row[stock_idx].value if len(row) > stock_idx else 0

                demand = _read_row_dates(row, date_col_map)
                # 向前掃描找 Supply / Net / Balance
                supply = {}
                balance = {}
                advance = 1
                for j in range(i + 1, min(i + 7, len(rows))):
                    mj = rows[j][marker_idx].value if len(rows[j]) > marker_idx else None
                    mj_str = str(mj).strip() if mj else ''
                    if mj_str in ('B-Supply', '2.Supply'):
                        supply = _read_row_dates(rows[j], date_col_map)
                    elif mj_str in ('C-Net', '3.Net', '3.Balance'):
                        balance = _read_row_dates(rows[j], date_col_map)
                    elif mj_str in ('A-Demand', '1.Demand'):
                        break
                    advance += 1

                results.append({
                    'buyer': buyer_label or 'PSW1+CEW1',
                    'plant': str(row_plant).strip() if row_plant else (plant_code or ''),
                    'part_no': _to_partno(part_no),
                    'vendor_part': str(vendor_part) if vendor_part else '',
                    'stock': stock or 0, 'on_way': None,
                    'demand': demand, 'supply': supply,
                    'balance_override': balance,
                })
                i += advance
            else:
                i += 1

    wb.close()
    return results


def _find_mwc1ipc1_layout(wb):
    """找到 MWC1IPC1 的 sheet / header row / 欄位位置 (支援多 variant)。

    Returns:
        tuple(ws, header_row, cols_dict) or (None, None, None)
        cols_dict keys: plant, partno, vendor, marker, stock (all 0-indexed)
    """
    for sn in ('Sheet1', 'MRP', '工作表1'):
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for hr in (1, 2):
            cols = {}
            for c, cell in enumerate(ws[hr]):
                h = str(cell.value or '').strip().upper()
                if h == 'PLANT':
                    cols['plant'] = c
                elif h == 'PARTNO':
                    cols['partno'] = c
                elif h in ('VENDOR PARTNO', 'VENDOR PART'):
                    cols['vendor'] = c
                elif h == 'REQUEST ITEM':
                    cols['marker'] = c
                elif h == 'PLANT STOCK':
                    cols['stock'] = c
            if 'marker' in cols:
                return ws, hr, cols
    return None, None, None


def _read_mwc1ipc1(filepath, date_cols, buyer_label=None, plant_code=None, conversions=None):
    """
    讀取 MWC1+IPC1 (含 variant): 4 rows/part
    (GROSS REQTS/FIRM ORDERS/VENDOR CFM/NET AVAIL)
    動態偵測 sheet / header row / 欄位位置:
      支援 Sheet1, MRP, 工作表1; header 可在 row 1 或 row 2。
    取 GROSS REQTS→Demand, VENDOR CFM→Supply, NET AVAIL→Balance。
    多 PLANT 檔案 — 每列從 PLANT 欄讀 PLANT。
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws, header_row, cols = _find_mwc1ipc1_layout(wb)
    if ws is None:
        wb.close()
        return []

    plant_idx = cols.get('plant')  # None = no PLANT column (single-plant)
    partno_idx = cols.get('partno', 1)
    vendor_idx = cols.get('vendor', 2)
    marker_idx = cols['marker']
    stock_idx = cols.get('stock', marker_idx + 1)

    # 日期從 marker 欄之後掃描
    date_scan_start = marker_idx + 2   # 1-based col
    date_col_map = _build_date_col_map(ws, date_scan_start, date_cols, conversions)

    results = []
    max_col = ws.max_column
    rows = list(ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row,
                             min_col=1, max_col=max_col, values_only=False))
    i = 0
    while i < len(rows):
        row = rows[i]
        marker = row[marker_idx].value if len(row) > marker_idx else None

        if marker == 'GROSS REQTS':
            row_plant = row[plant_idx].value if plant_idx is not None and len(row) > plant_idx else None
            part_no = row[partno_idx].value if len(row) > partno_idx else None
            vendor_part = row[vendor_idx].value if len(row) > vendor_idx else None
            stock = row[stock_idx].value if len(row) > stock_idx else 0

            demand = _read_row_dates(row, date_col_map)

            # 向前掃描找 VENDOR CFM / VN CFM / NET AVAIL (相容不同 variant)
            supply = {}
            balance = {}
            advance = 1
            for j in range(i + 1, min(i + 6, len(rows))):
                mj = rows[j][marker_idx].value if len(rows[j]) > marker_idx else None
                mj_up = str(mj).strip().upper() if mj else ''
                if ('VENDOR' in mj_up or 'VN' in mj_up) and 'CFM' in mj_up:
                    supply = _read_row_dates(rows[j], date_col_map)
                elif 'NET' in mj_up and 'AVAIL' in mj_up:
                    balance = _read_row_dates(rows[j], date_col_map)
                elif mj_up == 'GROSS REQTS':
                    break
                advance += 1

            results.append({
                'buyer': buyer_label or 'MWC1+IPC1',
                'plant': str(row_plant).strip() if row_plant else (plant_code or ''),
                'part_no': _to_partno(part_no),
                'vendor_part': str(vendor_part) if vendor_part else '',
                'stock': stock or 0, 'on_way': None,
                'demand': demand, 'supply': supply,
                'balance_override': balance,
            })
            i += advance
        else:
            i += 1

    wb.close()
    return results


def _read_nbq1(filepath, date_cols, buyer_label=None, plant_code=None, conversions=None):
    """
    讀取 NBQ1: PAN JIT sheet, 1 row/part (flat, Demand only).
    動態偵測欄位位置。無 PLANT column → 從檔名比對。
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['PAN JIT']

    # --- 動態偵測 header 欄位 ---
    cols = _scan_header_columns(ws, 1)
    partno_idx = cols.get('partno', 0)
    vendor_idx = cols.get('vendor', 2)
    stock_idx = cols.get('stock', 14)
    # 日期從 stock 欄之後開始
    date_scan_start = stock_idx + 2
    date_col_map = _build_date_col_map(ws, date_scan_start, date_cols, conversions)

    results = []
    max_col = ws.max_column
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=1, max_col=max_col, values_only=False):
        part_no = row[partno_idx].value if len(row) > partno_idx else None
        if part_no is None or str(part_no).strip() == '':
            continue
        vendor_part = row[vendor_idx].value if len(row) > vendor_idx else None
        stock = row[stock_idx].value if len(row) > stock_idx else 0

        demand = _read_row_dates(row, date_col_map)

        results.append({
            'buyer': buyer_label or 'NBQ1',
            'plant': plant_code or '',
            'part_no': _to_partno(part_no),
            'vendor_part': str(vendor_part) if vendor_part else '',
            'stock': stock or 0, 'on_way': None,
            'demand': demand, 'supply': {},
        })

    wb.close()
    return results


def _read_svc1pwc1_diode_mos(filepath, date_cols, buyer_label=None, plant_code=None, conversions=None):
    """
    讀取 SVC1+PWC1 DIODE&MOS: Diode + MOS sheets, 1 row/part (flat).
    動態偵測欄位位置。多 PLANT 檔案。
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    results = []

    for sheet_name in ('Diode', 'MOS'):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # --- 動態偵測 header 欄位 ---
        cols = _scan_header_columns(ws, 1)
        plant_idx = cols.get('plant', 0)
        partno_idx = cols.get('partno', 2)
        vendor_idx = cols.get('vendor', 4)
        stock_idx = cols.get('stock', 7)
        date_scan_start = stock_idx + 2
        date_col_map = _build_date_col_map(ws, date_scan_start, date_cols, conversions)
        max_col = ws.max_column

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=1, max_col=max_col, values_only=False):
            row_plant = row[plant_idx].value if len(row) > plant_idx else None
            part_no = row[partno_idx].value if len(row) > partno_idx else None
            if part_no is None or str(part_no).strip() == '':
                continue
            vendor_part = row[vendor_idx].value if len(row) > vendor_idx else None
            stock = row[stock_idx].value if len(row) > stock_idx else 0

            demand = _read_row_dates(row, date_col_map)

            results.append({
                'buyer': buyer_label or 'SVC1+PWC1',
                'plant': str(row_plant).strip() if row_plant else (plant_code or ''),
                'part_no': _to_partno(part_no),
                'vendor_part': str(vendor_part) if vendor_part else '',
                'stock': stock or 0, 'on_way': None,
                'demand': demand, 'supply': {},
            })

    wb.close()
    return results


def _read_psbg(filepath, date_cols, buyer_label=None, plant_code=None, conversions=None):
    """
    讀取 PSBG (PSB5 PANJIT): Sheet1, 3 rows/part (1.Demand/2.Supply/3.Net).
    動態偵測欄位位置。單 PLANT — PLANT 從檔名比對。
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['Sheet1']

    # --- 動態偵測 header 欄位 ---
    cols = _scan_header_columns(ws, 1)
    partno_idx = cols.get('partno', 2)
    vendor_idx = cols.get('vendor', 3)
    stock_idx = cols.get('stock', 7)
    otw_idx = cols.get('otw')
    filter_idx = cols.get('marker', 14)  # Filter = marker
    date_scan_start = filter_idx + 2
    date_col_map = _build_date_col_map(ws, date_scan_start, date_cols, conversions)
    max_col = ws.max_column

    pending = {}
    results = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=1, max_col=max_col, values_only=False):
        part_no = row[partno_idx].value if len(row) > partno_idx else None
        if part_no is None or str(part_no).strip() == '':
            continue
        filter_val = row[filter_idx].value if len(row) > filter_idx else None
        fv = str(filter_val).strip() if filter_val else ''

        pn = _to_partno(part_no)

        if fv == '1.Demand':
            vendor_part = row[vendor_idx].value if len(row) > vendor_idx else None
            stock = row[stock_idx].value if len(row) > stock_idx else 0
            on_way = row[otw_idx].value if otw_idx and len(row) > otw_idx else None
            demand = _read_row_dates(row, date_col_map)
            pending[pn] = {
                'buyer': buyer_label or 'PSBG',
                'plant': plant_code or '',
                'part_no': pn,
                'vendor_part': str(vendor_part) if vendor_part else '',
                'stock': stock or 0,
                'on_way': on_way or 0,
                'demand': demand,
                'supply': {},
            }
        elif fv == '2.Supply' and pn in pending:
            pending[pn]['supply'] = _read_row_dates(row, date_col_map)
        elif fv == '3.Net' and pn in pending:
            results.append(pending.pop(pn))

    results.extend(pending.values())

    wb.close()
    return results


def _read_eibg_eisbg(filepath, date_cols, buyer_label=None, plant_code=None, conversions=None):
    """
    讀取 EIBG/EISBG: Sheet1, 支援 flat (1 row/part) 和 multi-row (Demand/Supply/Balance)。
    動態偵測欄位位置。偵測 marker 欄 → 有則 multi-row, 無則 flat。
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['Sheet1']

    # --- 動態偵測 header 欄位 ---
    cols = _scan_header_columns(ws, 1)
    plant_idx = cols.get('plant')
    partno_idx = cols.get('partno', 2)
    vendor_idx = cols.get('vendor', 3)
    stock_idx = cols.get('stock', 9)
    otw_idx = cols.get('otw')
    marker_idx = cols.get('marker')  # 類別 / None(flat)

    # 日期起始: marker 或 otw 或 stock 之後
    last_meta = stock_idx
    if otw_idx and otw_idx > last_meta:
        last_meta = otw_idx
    if marker_idx and marker_idx > last_meta:
        last_meta = marker_idx
    date_scan_start = last_meta + 2
    date_col_map = _build_date_col_map(ws, date_scan_start, date_cols, conversions)

    results = []
    max_col = ws.max_column

    if marker_idx is not None:
        # === Multi-row mode (Demand/Supply/Balance per part) ===
        rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row,
                                 min_col=1, max_col=max_col, values_only=False))
        i = 0
        while i < len(rows):
            row = rows[i]
            marker = row[marker_idx].value if len(row) > marker_idx else None
            m_str = str(marker).strip() if marker else ''

            if m_str == 'Demand':
                row_plant = row[plant_idx].value if plant_idx is not None and len(row) > plant_idx else None
                part_no = row[partno_idx].value if len(row) > partno_idx else None
                vendor_part = row[vendor_idx].value if len(row) > vendor_idx else None
                stock = row[stock_idx].value if len(row) > stock_idx else 0
                on_way = row[otw_idx].value if otw_idx and len(row) > otw_idx else None

                demand = _read_row_dates(row, date_col_map)
                supply = {}
                balance = {}
                advance = 1
                for j in range(i + 1, min(i + 5, len(rows))):
                    mj = rows[j][marker_idx].value if len(rows[j]) > marker_idx else None
                    mj_str = str(mj).strip() if mj else ''
                    if mj_str == 'Supply':
                        supply = _read_row_dates(rows[j], date_col_map)
                    elif mj_str == 'Balance':
                        balance = _read_row_dates(rows[j], date_col_map)
                    elif mj_str == 'Demand':
                        break
                    advance += 1

                results.append({
                    'buyer': buyer_label or 'EIBG',
                    'plant': str(row_plant).strip() if row_plant else (plant_code or ''),
                    'part_no': _to_partno(part_no),
                    'vendor_part': str(vendor_part) if vendor_part else '',
                    'stock': stock or 0, 'on_way': on_way or 0,
                    'demand': demand, 'supply': supply,
                    'balance_override': balance,
                })
                i += advance
            else:
                i += 1
    else:
        # === Flat mode (1 row/part, Demand only) ===
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=1, max_col=max_col, values_only=False):
            part_no = row[partno_idx].value if len(row) > partno_idx else None
            if part_no is None or str(part_no).strip() == '':
                continue
            vendor_part = row[vendor_idx].value if len(row) > vendor_idx else None
            stock = row[stock_idx].value if len(row) > stock_idx else 0
            on_way = row[otw_idx].value if otw_idx and len(row) > otw_idx else None

            demand = _read_row_dates(row, date_col_map)

            results.append({
                'buyer': buyer_label or 'EIBG',
                'plant': plant_code or '',
                'part_no': _to_partno(part_no),
                'vendor_part': str(vendor_part) if vendor_part else '',
                'stock': stock or 0, 'on_way': on_way or 0,
                'demand': demand, 'supply': {},
            })

    wb.close()
    return results


def _read_fmbg(filepath, date_cols, buyer_label=None, plant_code=None, conversions=None):
    """
    讀取 FMBG: Sheet1, 3 rows/part (A-Demand/B-CFM/C-Bal).
    動態偵測欄位位置。多 PLANT — 每列讀 PLANT。
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['Sheet1']

    # --- 動態偵測 header 欄位 ---
    cols = _scan_header_columns(ws, 1)
    plant_idx = cols.get('plant', 0)
    partno_idx = cols.get('partno', 4)
    vendor_idx = cols.get('vendor', 6)
    marker_idx = cols.get('marker', 11)  # REQUEST ITEM
    stock_idx = cols.get('stock', 14)
    date_scan_start = max(marker_idx, stock_idx) + 2
    date_col_map = _build_date_col_map(ws, date_scan_start, date_cols, conversions)

    results = []
    max_col = ws.max_column
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row,
                             min_col=1, max_col=max_col, values_only=False))
    i = 0
    while i < len(rows):
        row = rows[i]
        marker = row[marker_idx].value if len(row) > marker_idx else None

        if marker == 'A-Demand':
            row_plant = row[plant_idx].value if len(row) > plant_idx else None
            part_no = row[partno_idx].value if len(row) > partno_idx else None
            vendor_part = row[vendor_idx].value if len(row) > vendor_idx else None
            stock = row[stock_idx].value if len(row) > stock_idx else 0

            demand = _read_row_dates(row, date_col_map)
            # 向前掃描找 B-CFM / C-Bal
            supply = {}
            balance = {}
            advance = 1
            for j in range(i + 1, min(i + 5, len(rows))):
                mj = rows[j][marker_idx].value if len(rows[j]) > marker_idx else None
                mj_str = str(mj).strip() if mj else ''
                if mj_str.startswith('B-'):
                    supply = _read_row_dates(rows[j], date_col_map)
                elif mj_str.startswith('C-'):
                    balance = _read_row_dates(rows[j], date_col_map)
                elif mj_str == 'A-Demand':
                    break
                advance += 1

            results.append({
                'buyer': buyer_label or 'FMBG',
                'plant': str(row_plant).strip() if row_plant else (plant_code or ''),
                'part_no': _to_partno(part_no),
                'vendor_part': str(vendor_part) if vendor_part else '',
                'stock': stock or 0, 'on_way': None,
                'demand': demand, 'supply': supply,
                'balance_override': balance,
            })
            i += advance
        else:
            i += 1

    wb.close()
    return results


def _read_iabg(filepath, date_cols, buyer_label=None, plant_code=None, conversions=None):
    """
    讀取 IABG: Sheet1, 1 row/part (flat, Demand only).
    動態偵測欄位位置。多 PLANT — 每列讀 PLANT。
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['Sheet1']

    # --- 動態偵測 header 欄位 ---
    cols = _scan_header_columns(ws, 1)
    plant_idx = cols.get('plant', 0)
    partno_idx = cols.get('partno', 3)
    vendor_idx = cols.get('vendor', 4)
    stock_idx = cols.get('stock', 11)
    date_scan_start = stock_idx + 2
    date_col_map = _build_date_col_map(ws, date_scan_start, date_cols, conversions)

    results = []
    max_col = ws.max_column
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=1, max_col=max_col, values_only=False):
        row_plant = row[plant_idx].value if len(row) > plant_idx else None
        part_no = row[partno_idx].value if len(row) > partno_idx else None
        if part_no is None or str(part_no).strip() == '':
            continue
        vendor_part = row[vendor_idx].value if len(row) > vendor_idx else None
        stock = row[stock_idx].value if len(row) > stock_idx else 0

        demand = _read_row_dates(row, date_col_map)

        results.append({
            'buyer': buyer_label or 'IABG',
            'plant': str(row_plant).strip() if row_plant else (plant_code or ''),
            'part_no': _to_partno(part_no),
            'vendor_part': str(vendor_part) if vendor_part else '',
            'stock': stock or 0, 'on_way': None,
            'demand': demand, 'supply': {},
        })

    wb.close()
    return results


def _read_ictbg_ntl7(filepath, date_cols, buyer_label=None, plant_code=None, conversions=None):
    """
    讀取 ICTBG NTL7: Sheet1, 4 rows/part (GROSS REQTS/FIRM ORDERS/Vendor Cfm/NET AVAIL).
    動態偵測欄位位置。多 PLANT — 每列讀 PLANT。
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['Sheet1']

    # --- 動態偵測 header 欄位 ---
    cols = _scan_header_columns(ws, 1)
    plant_idx = cols.get('plant', 0)
    partno_idx = cols.get('partno', 1)
    vendor_idx = cols.get('vendor', 6)
    marker_idx = cols.get('marker', 9)  # REQUEST ITEM
    stock_idx = cols.get('stock', marker_idx + 1)
    date_scan_start = max(marker_idx, stock_idx) + 2
    date_col_map = _build_date_col_map(ws, date_scan_start, date_cols, conversions)

    results = []
    max_col = ws.max_column
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row,
                             min_col=1, max_col=max_col, values_only=False))
    i = 0
    while i < len(rows):
        row = rows[i]
        marker = row[marker_idx].value if len(row) > marker_idx else None

        if marker == 'GROSS REQTS':
            row_plant = row[plant_idx].value if len(row) > plant_idx else None
            part_no = row[partno_idx].value if len(row) > partno_idx else None
            vendor_part = row[vendor_idx].value if len(row) > vendor_idx else None
            stock = row[stock_idx].value if len(row) > stock_idx else 0

            demand = _read_row_dates(row, date_col_map)
            # 向前掃描找 Vendor Cfm / NET AVAIL
            supply = {}
            balance = {}
            advance = 1
            for j in range(i + 1, min(i + 6, len(rows))):
                mj = rows[j][marker_idx].value if len(rows[j]) > marker_idx else None
                mj_str = str(mj).strip() if mj else ''
                if 'VENDOR' in mj_str.upper() and 'CFM' in mj_str.upper():
                    supply = _read_row_dates(rows[j], date_col_map)
                elif 'NET' in mj_str.upper() and 'AVAIL' in mj_str.upper():
                    balance = _read_row_dates(rows[j], date_col_map)
                elif mj_str == 'GROSS REQTS':
                    break
                advance += 1

            results.append({
                'buyer': buyer_label or 'ICTBG-NTL7',
                'plant': str(row_plant).strip() if row_plant else (plant_code or ''),
                'part_no': _to_partno(part_no),
                'vendor_part': str(vendor_part) if vendor_part else '',
                'stock': stock or 0, 'on_way': None,
                'demand': demand, 'supply': supply,
                'balance_override': balance,
            })
            i += advance
        else:
            i += 1

    wb.close()
    return results


def _read_ictbg_psb9_mrp(filepath, date_cols, buyer_label=None, plant_code=None, conversions=None):
    """
    讀取 ICTBG PSB9 Kaewarin: PSB9_MRP* sheet, 4 rows/part (DEMAND/SUPPLY/NET/Remark).
    動態偵測欄位位置。多 PLANT — 每列讀 PLANT。已有 forward scanning。
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_name = _get_ictbg_psb9_mrp_sheet(wb)
    if not sheet_name or sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]

    # --- 動態偵測 header 欄位 ---
    cols = _scan_header_columns(ws, 1)
    plant_idx = cols.get('plant', 0)
    partno_idx = cols.get('partno', 2)
    vendor_idx = cols.get('vendor', 7)
    stock_idx = cols.get('stock', 10)
    marker_idx = cols.get('marker', 13)  # Type
    date_scan_start = max(marker_idx, stock_idx) + 2
    date_col_map = _build_date_col_map(ws, date_scan_start, date_cols, conversions)

    results = []
    max_col = ws.max_column
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row,
                             min_col=1, max_col=max_col, values_only=False))
    i = 0
    while i < len(rows):
        row = rows[i]
        marker = row[marker_idx].value if len(row) > marker_idx else None
        marker_str = str(marker).strip().upper() if marker else ''

        if marker_str == 'DEMAND':
            row_plant = row[plant_idx].value if len(row) > plant_idx else None
            part_no = row[partno_idx].value if len(row) > partno_idx else None
            vendor_part = row[vendor_idx].value if len(row) > vendor_idx else None
            stock = row[stock_idx].value if len(row) > stock_idx else 0

            demand = _read_row_dates(row, date_col_map)
            supply, balance = {}, {}
            j = i + 1
            next_demand = i + 1
            while j < len(rows) and j < i + 5:
                m = rows[j][marker_idx].value if len(rows[j]) > marker_idx else None
                m_str = str(m).strip().upper() if m else ''
                if m_str == 'DEMAND':
                    next_demand = j
                    break
                if m_str == 'SUPPLY':
                    supply = _read_row_dates(rows[j], date_col_map)
                elif m_str == 'NET':
                    balance = _read_row_dates(rows[j], date_col_map)
                j += 1
                next_demand = j

            results.append({
                'buyer': buyer_label or 'ICTBG-PSB9',
                'plant': str(row_plant).strip() if row_plant else (plant_code or ''),
                'part_no': _to_partno(part_no),
                'vendor_part': str(vendor_part) if vendor_part else '',
                'stock': stock or 0, 'on_way': None,
                'demand': demand, 'supply': supply,
                'balance_override': balance,
            })
            i = next_demand
        else:
            i += 1

    wb.close()
    return results


def _read_ictbg_psb9_siriraht(filepath, date_cols, buyer_label=None, plant_code=None, conversions=None):
    """
    讀取 ICTBG PSB9 Siriraht: Sheet1, 3 rows/part (1Demand/2Supply/3Balance).
    動態偵測欄位位置。多 PLANT — 每列讀 PLANT。
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['Sheet1']

    # --- 動態偵測 header 欄位 ---
    cols = _scan_header_columns(ws, 1)
    plant_idx = cols.get('plant', 0)
    partno_idx = cols.get('partno', 3)
    vendor_idx = cols.get('vendor', 5)
    stock_idx = cols.get('stock', 11)
    marker_idx = cols.get('marker', 14)  # REQUEST ITEM
    date_scan_start = max(marker_idx, stock_idx) + 2
    date_col_map = _build_date_col_map(ws, date_scan_start, date_cols, conversions)

    results = []
    max_col = ws.max_column
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row,
                             min_col=1, max_col=max_col, values_only=False))
    i = 0
    while i < len(rows):
        row = rows[i]
        marker = row[marker_idx].value if len(row) > marker_idx else None
        marker_str = str(marker).strip() if marker else ''

        if marker_str == '1Demand':
            row_plant = row[plant_idx].value if len(row) > plant_idx else None
            part_no = row[partno_idx].value if len(row) > partno_idx else None
            vendor_part = row[vendor_idx].value if len(row) > vendor_idx else None
            stock = row[stock_idx].value if len(row) > stock_idx else 0

            demand = _read_row_dates(row, date_col_map)
            # 向前掃描找 2Supply / 3Balance
            supply = {}
            balance = {}
            advance = 1
            for j in range(i + 1, min(i + 5, len(rows))):
                mj = rows[j][marker_idx].value if len(rows[j]) > marker_idx else None
                mj_str = str(mj).strip() if mj else ''
                if mj_str == '2Supply':
                    supply = _read_row_dates(rows[j], date_col_map)
                elif mj_str == '3Balance':
                    balance = _read_row_dates(rows[j], date_col_map)
                elif mj_str == '1Demand':
                    break
                advance += 1

            results.append({
                'buyer': buyer_label or 'ICTBG-PSB9-S',
                'plant': str(row_plant).strip() if row_plant else (plant_code or ''),
                'part_no': _to_partno(part_no),
                'vendor_part': str(vendor_part) if vendor_part else '',
                'stock': stock or 0, 'on_way': None,
                'demand': demand, 'supply': supply,
                'balance_override': balance,
            })
            i += advance
        else:
            i += 1

    wb.close()
    return results


# ---------------------------------------------------------------------------
# Excel 產出
# ---------------------------------------------------------------------------

def _generate_consolidated_excel(all_source, output_path, reference_path,
                                 date_cols, erp_mapping=None):
    """
    產出匯總格式 Excel, 格式與原始模板完全一致。
    Balance 行使用 Excel 公式。
    C/D 欄位 (ERP 客戶簡稱、ERP 送貨地點) 留空，由第三步驟 forecast 處理時
    透過 customer_mappings 表自動填入，consolidation 階段不寫入。
    """
    # erp_mapping 已棄用，保留參數做相容性

    wb_ref = openpyxl.load_workbook(reference_path)
    ws_ref = wb_ref[wb_ref.sheetnames[0]]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '工作表1'

    # 複製固定表頭 A~I (col 1~9) 從模板
    for col in range(1, 10):
        src = ws_ref.cell(row=1, column=col)
        dst = ws.cell(row=1, column=col, value=src.value)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.border = copy(src.border)

    # 動態寫入日期表頭 J onwards (從 date_cols 產生，不依賴模板)
    # 取得模板中 J 欄的樣式作為日期表頭的參考樣式
    ref_date_cell = ws_ref.cell(row=1, column=10)
    for col_offset, date_key in enumerate(date_cols):
        col_num = 10 + col_offset
        dst = ws.cell(row=1, column=col_num, value=date_key)
        dst.font = copy(ref_date_cell.font)
        dst.fill = copy(ref_date_cell.fill)
        dst.alignment = copy(ref_date_cell.alignment)
        dst.number_format = ref_date_cell.number_format
        dst.border = copy(ref_date_cell.border)

    # 複製固定欄位 A~I 的欄寬
    from openpyxl.utils import get_column_letter
    for col in range(1, 10):
        cl = get_column_letter(col)
        if cl in ws_ref.column_dimensions:
            ws.column_dimensions[cl].width = ws_ref.column_dimensions[cl].width
    # 日期欄位統一寬度 (參考模板 J 欄)
    ref_j_width = ws_ref.column_dimensions.get('J')
    date_col_width = ref_j_width.width if ref_j_width else 11
    for col_offset in range(len(date_cols)):
        cl = get_column_letter(10 + col_offset)
        ws.column_dimensions[cl].width = date_col_width

    # 樣式定義
    font_mingliu_12 = Font(name='新細明體', size=12)
    font_arial_9 = Font(name='Arial', size=9)
    font_arial_10 = Font(name='Arial', size=10)
    nf_data = '#,##0_);[Red]\\(#,##0\\)'
    nf_partno = '000'
    nf_stock = '#,##0'
    align_center = Alignment(vertical='center')
    align_top_left = Alignment(horizontal='left', vertical='top')
    fill_cd = PatternFill(patternType='solid',
                          fgColor=Color(theme=4, tint=0.7999816888943144))
    fill_i = PatternFill(patternType='solid',
                         fgColor=Color(theme=0, tint=0.0))
    fill_supply = PatternFill(patternType='solid',
                              fgColor=Color(theme=9, tint=0.7999816888943144))
    fill_balance = PatternFill(patternType='solid',
                               fgColor=Color(theme=7, tint=0.7999816888943144))
    thin_side = Side(style='thin')
    medium_side = Side(style='medium')
    border_normal = Border(left=thin_side, right=thin_side,
                           top=thin_side, bottom=thin_side)
    border_balance = Border(left=thin_side, right=thin_side,
                            top=thin_side, bottom=medium_side)

    def apply_row_style(row_num, date_type):
        is_balance = (date_type == 'Balance')
        bdr = border_balance if is_balance else border_normal

        for col in [1, 2]:
            c = ws.cell(row=row_num, column=col)
            c.alignment = align_center
            c.border = bdr
        for col in [3, 4]:
            c = ws.cell(row=row_num, column=col)
            c.alignment = align_center
            c.border = bdr
            c.fill = fill_cd
        for col in [5, 6]:
            c = ws.cell(row=row_num, column=col)
            c.alignment = align_top_left
            c.border = bdr
        c = ws.cell(row=row_num, column=7)
        c.font = font_arial_9
        c.alignment = align_top_left
        c.border = bdr
        c = ws.cell(row=row_num, column=8)
        c.font = font_mingliu_12
        c.alignment = align_center
        c.border = bdr
        c = ws.cell(row=row_num, column=9)
        c.alignment = align_top_left
        c.border = bdr
        c.fill = fill_i

        data_fill = fill_supply if date_type == 'Supply' else (
            fill_balance if date_type == 'Balance' else None)
        for col in range(10, 10 + len(date_cols)):
            c = ws.cell(row=row_num, column=col)
            c.alignment = align_center
            c.border = bdr
            if data_fill:
                c.fill = data_fill

    # 寫入資料
    row_num = 2
    for item in all_source:
        buyer = item['buyer']
        plant = item['plant']
        part_no = item['part_no']
        vendor_part = item['vendor_part']
        stock = item['stock']
        # ERP 客戶簡稱、ERP 送貨地點 留空 (第三步驟由 mapping 表填入)
        erp_name, erp_location = '', ''

        # --- Demand ---
        demand_row = row_num
        ws.cell(row=row_num, column=1, value=buyer).font = font_mingliu_12
        ws.cell(row=row_num, column=2, value=plant).font = font_mingliu_12
        ws.cell(row=row_num, column=3, value=erp_name).font = font_mingliu_12
        ws.cell(row=row_num, column=4, value=erp_location).font = font_mingliu_12
        c = ws.cell(row=row_num, column=5, value=part_no)
        c.font = font_arial_9
        c.number_format = nf_partno
        ws.cell(row=row_num, column=6, value=vendor_part).font = font_arial_9
        if stock is not None:
            c = ws.cell(row=row_num, column=7, value=stock)
            c.font = font_arial_9
            c.number_format = nf_stock
        ws.cell(row=row_num, column=9, value='Demand').font = font_arial_10
        for col_offset, date_key in enumerate(date_cols):
            v = item['demand'].get(date_key, 0) or 0
            c = ws.cell(row=row_num, column=10 + col_offset, value=v)
            c.font = font_mingliu_12
            c.number_format = nf_data
        apply_row_style(row_num, 'Demand')
        row_num += 1

        # --- Supply ---
        supply_row = row_num
        ws.cell(row=row_num, column=1, value=buyer).font = font_mingliu_12
        ws.cell(row=row_num, column=2, value=plant).font = font_mingliu_12
        ws.cell(row=row_num, column=3, value=erp_name).font = font_mingliu_12
        ws.cell(row=row_num, column=4, value=erp_location).font = font_mingliu_12
        c = ws.cell(row=row_num, column=5, value=part_no)
        c.font = font_arial_9
        c.number_format = nf_partno
        ws.cell(row=row_num, column=6, value=vendor_part).font = font_arial_9
        ws.cell(row=row_num, column=9, value='Supply').font = font_arial_10
        for col_offset, date_key in enumerate(date_cols):
            v = item['supply'].get(date_key, 0) or 0
            c = ws.cell(row=row_num, column=10 + col_offset,
                        value=v if v != 0 else None)
            c.font = font_mingliu_12
            c.number_format = nf_data
        apply_row_style(row_num, 'Supply')
        row_num += 1

        # --- Balance (公式) ---
        balance_row = row_num
        ws.cell(row=row_num, column=1, value=buyer).font = font_mingliu_12
        ws.cell(row=row_num, column=2, value=plant).font = font_mingliu_12
        ws.cell(row=row_num, column=3, value=erp_name).font = font_mingliu_12
        ws.cell(row=row_num, column=4, value=erp_location).font = font_mingliu_12
        c = ws.cell(row=row_num, column=5, value=part_no)
        c.font = font_arial_9
        c.number_format = nf_partno
        ws.cell(row=row_num, column=6, value=vendor_part).font = font_arial_9
        ws.cell(row=row_num, column=9, value='Balance').font = font_arial_10
        for col_offset, date_key in enumerate(date_cols):
            col_num = 10 + col_offset
            col_letter = openpyxl.utils.get_column_letter(col_num)
            if col_offset == 0:
                formula = f'=G{demand_row}+H{demand_row}-{col_letter}{demand_row}'
            elif col_offset == 1:
                prev_col = openpyxl.utils.get_column_letter(col_num - 1)
                formula = (f'={prev_col}{supply_row}+{prev_col}{balance_row}'
                           f'-{col_letter}{demand_row}')
            else:
                prev_col = openpyxl.utils.get_column_letter(col_num - 1)
                formula = (f'={prev_col}{balance_row}+{prev_col}{supply_row}'
                           f'-{col_letter}{demand_row}')
            c = ws.cell(row=balance_row, column=col_num, value=formula)
            c.font = font_mingliu_12
            c.number_format = nf_data
        apply_row_style(row_num, 'Balance')
        row_num += 1

    # 凍結窗格、自動篩選
    ws.freeze_panes = 'J2'
    last_data_col = 9 + len(date_cols)  # A~I (9) + 日期欄位數
    last_col = openpyxl.utils.get_column_letter(last_data_col)
    ws.auto_filter.ref = f'A1:{last_col}{row_num - 1}'

    # 條件式格式
    for cf in ws_ref.conditional_formatting:
        for rule in cf.rules:
            ws.conditional_formatting.add(str(cf.cells), rule)

    # 複製工作表5 (如果存在)
    if '工作表5' in wb_ref.sheetnames:
        ws5_ref = wb_ref['工作表5']
        ws5 = wb.create_sheet('工作表5')
        for row in ws5_ref.iter_rows(min_row=1, max_row=ws5_ref.max_row,
                                      max_col=ws5_ref.max_column, values_only=False):
            for cell in row:
                dst = ws5.cell(row=cell.row, column=cell.column, value=cell.value)
                dst.font = copy(cell.font)
                dst.fill = copy(cell.fill)
                dst.alignment = copy(cell.alignment)
                dst.number_format = cell.number_format
                dst.border = copy(cell.border)
        for cl, dim in ws5_ref.column_dimensions.items():
            ws5.column_dimensions[cl].width = dim.width

    wb_ref.close()
    wb.save(output_path)
    return len(all_source)  # 料號數


# ---------------------------------------------------------------------------
# 主要入口
# ---------------------------------------------------------------------------

FORMAT_READERS = {
    FORMAT_KETWADEE:           _read_ketwadee,
    FORMAT_KANYANAT:           _read_kanyanat,
    FORMAT_WEERAYA:            _read_weeraya,
    FORMAT_INDIA_IAI1:         _read_india_iai1,
    FORMAT_PSW1_CEW1:          _read_psw1_cew1,
    FORMAT_MWC1IPC1:           _read_mwc1ipc1,
    FORMAT_NBQ1:               _read_nbq1,
    FORMAT_SVC1PWC1_DIODE_MOS: _read_svc1pwc1_diode_mos,
    FORMAT_PSBG:               _read_psbg,
    FORMAT_EIBG_EISBG:         _read_eibg_eisbg,
    FORMAT_FMBG:               _read_fmbg,
    FORMAT_IABG:               _read_iabg,
    FORMAT_ICTBG_NTL7:         _read_ictbg_ntl7,
    FORMAT_ICTBG_PSB9_MRP:     _read_ictbg_psb9_mrp,
    FORMAT_ICTBG_PSB9_SIRIRAHT:_read_ictbg_psb9_siriraht,
    FORMAT_PRAPAPORN:          _read_prapaporn,
}

# 向後相容: 舊 API
BUYER_READERS = {
    'Ketwadee': _read_ketwadee,
    'Kanyanat': _read_kanyanat,
    'Weeraya':  _read_weeraya,
}


def detect_all_formats(forecast_files):
    """
    Pre-detection: 偵測所有檔案的格式，一次回傳所有識別結果。
    Returns:
        tuple(detected, unknown)
        - detected: list of (filepath, format_const) tuples (成功識別)
        - unknown: list of filepaths (無法識別)
    """
    detected, unknown = [], []
    for fp in forecast_files:
        fmt = detect_format(fp)
        if fmt is None:
            unknown.append(fp)
        else:
            detected.append((fp, fmt))
    return detected, unknown


# ---------------------------------------------------------------------------
# 漂移版 fallback (補強 detect_format 認不出時)
# ---------------------------------------------------------------------------

def _extract_dates_from_drift_files(drift_files):
    """用統一 reader 從漂移版檔案提取日期 keys (補強 extract_dates_from_files)。

    Args:
        drift_files: list of (filepath, matched_fmt, score)

    Returns:
        tuple(all_dates, per_file_dates)
    """
    from delta_unified_reader import (
        find_valid_sheets, scan_headers, find_first_date_col, collect_date_cols,
        _read_header_row,
    )
    all_dates = set()
    per_file_dates = {}
    for fp, _matched_fmt, _score in drift_files:
        file_key = os.path.basename(fp)
        try:
            wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        except Exception as e:
            print(f"  ⚠️ 漂移版日期提取失敗 [{file_key}]: {e}")
            per_file_dates[file_key] = set()
            continue
        try:
            sheets = find_valid_sheets(wb)
            file_dates = set()
            for sheet_name, header_row in sheets:
                ws = wb[sheet_name]
                header_values = _read_header_row(ws, header_row)
                _found, headers = scan_headers(header_values)
                date_start = find_first_date_col(headers, header_values)
                date_col_map = collect_date_cols(date_start, header_values)
                file_dates.update(date_col_map.values())
            per_file_dates[file_key] = file_dates
            all_dates.update(file_dates)
            print(f"  漂移版 [{file_key}]: 偵測到 {len(file_dates)} 個日期欄位")
        finally:
            wb.close()
    return all_dates, per_file_dates


def _read_drift_file(fp, date_cols, buyer_label, plant_code, conversions, plant_codes):
    """用統一 reader 讀取漂移版檔案, 輸出格式與 FORMAT_READERS 一致。"""
    from delta_unified_reader import read_buyer_file
    # 統一 reader 用 plant_codes 對檔內無 PLANT 欄的單 PLANT 檔案做檔名 fallback
    use_codes = plant_codes if plant_codes else ([plant_code] if plant_code else None)
    data = read_buyer_file(
        fp, plant_codes=use_codes, date_cols=date_cols, conversions=conversions,
    )
    # 統一 buyer 顯示標籤 (對齊 hardcoded reader 行為)
    if buyer_label:
        for d in data:
            d['buyer'] = buyer_label
    return data


def consolidate(forecast_files, reference_template, output_path,
                erp_mapping=None, plant_codes=None, file_labels=None):
    """
    合併多個 Delta Forecast 檔案為匯總格式 Excel (支援 15 種格式)。

    若 detect_format 認不出, 會嘗試「指紋比對」找已知 15 格式的漂移版本,
    若仍認不出 (真正第 16 格式) 則維持拒絕。

    Args:
        forecast_files: list of file paths (1 個或多個，任何格式組合)
        reference_template: 匯總格式模板路徑 (用於取得表頭格式，不用於日期)
        output_path: 輸出檔案路徑
        erp_mapping: dict {Plant: (ERP客戶簡稱, ERP送貨地點)}, 已棄用
        plant_codes: list of valid PLANT codes, 用於從檔名比對單 PLANT 檔案的 PLANT。
                     建議由 customer_mappings 的 region 欄位提取
                     (例: ['PSB5', 'PSB7', 'IAI1', 'IPC1'])。
        file_labels: dict {filepath: label}，用於自訂 Buyer 欄位顯示名稱。
                     通常由 app.py 傳入原始檔名 (因為暫存檔名被改成 forecast_temp_N)。
                     若未提供則使用 basename without ext。

    Returns:
        dict with keys: success, part_count, format_stats, unknown_files,
                        date_warnings, message
    """
    if not forecast_files:
        return {
            'success': False, 'part_count': 0,
            'message': '未提供任何 Forecast 檔案'
        }

    # 1. 預先偵測所有檔案格式
    detected, unknown = detect_all_formats(forecast_files)

    # 1.5 對 unknown 嘗試指紋比對 (補強: 區分「漂移版」vs「真正第 16 格式」)
    drift_files = []   # list of (fp, matched_fmt, score) — 用統一 reader 處理
    truly_unknown = []  # 真正未知格式 — 維持現有拒絕
    if unknown:
        try:
            from delta_format_fingerprint import match_known_format_fingerprint
            for fp in unknown:
                m, s = match_known_format_fingerprint(fp)
                if m:
                    drift_files.append((fp, m, s))
                    print(f"  🔄 漂移版偵測: {os.path.basename(fp)} → "
                          f"{FORMAT_LABELS.get(m, m)} (相似度 {s})")
                else:
                    truly_unknown.append(fp)
        except Exception as e:
            print(f"  ⚠️ 指紋比對失敗 ({e}), 視所有 unknown 為真正未知")
            truly_unknown = list(unknown)

    if truly_unknown:
        unknown_names = [os.path.basename(fp) for fp in truly_unknown]
        return {
            'success': False, 'part_count': 0,
            'unknown_files': unknown_names,
            'message': '無法識別以下檔案格式 (請確認為 Delta 15 種標準格式): '
                       + ', '.join(unknown_names)
        }

    print(f"Delta 合併: 偵測到 {len(detected)} 個已知格式檔案"
          + (f", {len(drift_files)} 個漂移版檔案" if drift_files else ""))
    for fp, fmt in detected:
        print(f"  [{FORMAT_LABELS.get(fmt, fmt)}] {os.path.basename(fp)}")
    for fp, m, s in drift_files:
        print(f"  [{FORMAT_LABELS.get(m, m)} 漂移版] {os.path.basename(fp)}")

    # 2. 從所有檔案動態提取日期欄位 (取聯集)
    #    conversions: {原始 YYYYMMDD: 月份標籤 或 None}，讀取器要用這個 map
    #    把來源檔的月末日期轉換為月份標籤 (避免週/月欄位重複)。
    date_cols, per_file_dates, date_warnings, conversions = extract_dates_from_files(detected)

    # 2.5 將漂移版的日期合併入聯集後重新 sort
    if drift_files:
        drift_dates, drift_per_file = _extract_dates_from_drift_files(drift_files)
        all_dates = set()
        for s in per_file_dates.values():
            all_dates.update(s)
        all_dates.update(drift_dates)
        per_file_dates.update(drift_per_file)
        date_cols, conversions = _sort_date_cols(all_dates)

    if not date_cols:
        return {
            'success': False, 'part_count': 0,
            'message': '無法從 Forecast 檔案提取日期欄位'
        }

    print(f"  統一日期欄位: {len(date_cols)} 個 ({date_cols[0]} ~ {date_cols[-1]})")
    if conversions:
        non_null_conv = {k: v for k, v in conversions.items() if v is not None}
        rejected_conv = [k for k, v in conversions.items() if v is None]
        if non_null_conv:
            print(f"  月末→月份轉換: {non_null_conv}")
        if rejected_conv:
            print(f"  拒絕日期: {rejected_conv}")

    # 3. 讀取各檔案資料
    # Buyer 欄位暫時使用檔案名稱 (不含副檔名)，等客戶確認代碼→名稱對照後再調整邏輯。
    # PLANT 欄位:
    #   - 單 PLANT 檔案: 從檔名比對 plant_codes (例: 'PSB5 Ketwadee.xlsx' → 'PSB5')
    #   - 多 PLANT 檔案: 每列從工作表欄位讀取 (不使用檔名)
    all_source = []
    format_stats = {}  # {filename: count}

    for fp, fmt in detected:
        reader = FORMAT_READERS.get(fmt)
        if reader is None:
            print(f"  ⚠️ 跳過未註冊 reader 的格式: {fmt} ({os.path.basename(fp)})")
            continue

        # Buyer 顯示名稱 & 檔名比對用途:
        # 優先使用呼叫端傳入的原始檔名 (因為上傳時 app.py 會把檔案改名成 forecast_temp_N)
        label_for_match = file_labels.get(fp) if file_labels else None
        if label_for_match:
            buyer_label = os.path.splitext(label_for_match)[0]
        else:
            buyer_label = os.path.splitext(os.path.basename(fp))[0]
        file_key = os.path.basename(fp)

        # 從檔名比對 PLANT 代碼 (用原始檔名，不用 temp 檔名)
        # 不限 SINGLE_PLANT_FORMATS — 多 PLANT 格式的 reader 也會把 plant_code 當
        # fallback (當檔案沒有 PLANT 欄時使用，如 DNI-NTL7 的 MWC1IPC1 格式)
        plant_code = None
        if plant_codes:
            match_target = label_for_match if label_for_match else fp
            matched = match_plants_in_filename(match_target, plant_codes)
            if matched:
                plant_code = matched[0]
                if len(matched) > 1:
                    print(f"  ⚠️ {file_key}: 檔名中有多個 PLANT {matched}, 取 {plant_code}")

        try:
            data = reader(fp, date_cols,
                          buyer_label=buyer_label, plant_code=plant_code,
                          conversions=conversions)
        except Exception as e:
            return {
                'success': False, 'part_count': 0,
                'message': f'讀取檔案失敗 [{file_key}] ({FORMAT_LABELS.get(fmt, fmt)}): {e}'
            }

        # PLANT 統一驗證: 不在 mapping 表 (plant_codes) 內的一律設為空白
        # 適用所有 reader (單 PLANT / 多 PLANT) — 確保 PLANT 欄完全來自 mapping 表
        if plant_codes:
            valid_upper = {str(p).strip().upper() for p in plant_codes if p}
            invalid_seen = set()
            for item in data:
                p = item.get('plant') or ''
                if p and str(p).strip().upper() not in valid_upper:
                    invalid_seen.add(str(p).strip())
                    item['plant'] = ''
            if invalid_seen:
                print(f"  ⚠️ {file_key}: PLANT 不在 mapping 表，已留空: {sorted(invalid_seen)}")

        format_stats[file_key] = len(data)
        all_source.extend(data)

        if fmt in SINGLE_PLANT_FORMATS:
            # 單 PLANT 顯示: 從 data 看實際使用的 PLANT (可能因 mapping 驗證被清空)
            actual = next((d.get('plant') for d in data if d.get('plant')), None)
            plant_display = actual or '(未比對到)'
            print(f"  {file_key} [{FORMAT_LABELS.get(fmt, fmt)}]: "
                  f"{len(data)} 個料號, PLANT={plant_display}")
        else:
            unique_plants = sorted({d.get('plant', '') for d in data if d.get('plant')})
            print(f"  {file_key} [{FORMAT_LABELS.get(fmt, fmt)}]: "
                  f"{len(data)} 個料號, 多 PLANT={unique_plants}")

    # 3.5 漂移版 — 走統一 reader (補強路徑)
    drift_stats = {}
    for fp, matched_fmt, score in drift_files:
        label_for_match = file_labels.get(fp) if file_labels else None
        if label_for_match:
            buyer_label = os.path.splitext(label_for_match)[0]
        else:
            buyer_label = os.path.splitext(os.path.basename(fp))[0]
        file_key = os.path.basename(fp)

        # 單 PLANT 漂移版檔名比對 PLANT
        plant_code = None
        if matched_fmt in SINGLE_PLANT_FORMATS and plant_codes:
            match_target = label_for_match if label_for_match else fp
            matched = match_plants_in_filename(match_target, plant_codes)
            if matched:
                plant_code = matched[0]

        try:
            data = _read_drift_file(fp, date_cols, buyer_label, plant_code,
                                    conversions, plant_codes)
        except Exception as e:
            return {
                'success': False, 'part_count': 0,
                'message': f'讀取漂移版檔案失敗 [{file_key}] '
                           f'({FORMAT_LABELS.get(matched_fmt, matched_fmt)} 漂移版): {e}'
            }

        # PLANT 統一驗證 (與 detected 路徑一致)
        if plant_codes:
            valid_upper = {str(p).strip().upper() for p in plant_codes if p}
            invalid_seen = set()
            for item in data:
                p = item.get('plant') or ''
                if p and str(p).strip().upper() not in valid_upper:
                    invalid_seen.add(str(p).strip())
                    item['plant'] = ''
            if invalid_seen:
                print(f"  ⚠️ {file_key} (漂移版): PLANT 不在 mapping 表，已留空: {sorted(invalid_seen)}")

        format_stats[file_key] = len(data)
        drift_stats[file_key] = {
            'matched_fmt': matched_fmt,
            'matched_label': FORMAT_LABELS.get(matched_fmt, matched_fmt),
            'score': score,
            'part_count': len(data),
        }
        all_source.extend(data)
        unique_plants = sorted({d.get('plant', '') for d in data if d.get('plant')})
        print(f"  {file_key} [漂移版 → {FORMAT_LABELS.get(matched_fmt, matched_fmt)}]: "
              f"{len(data)} 個料號, PLANT={unique_plants}")

    if not all_source:
        return {
            'success': False, 'part_count': 0,
            'message': '未讀取到任何料號資料 (檔案可能為空或格式異常)'
        }

    # 4. 產出匯總格式 Excel（日期表頭由 date_cols 動態產生）
    part_count = _generate_consolidated_excel(
        all_source, output_path, reference_template, date_cols, erp_mapping
    )

    print(f"  合併完成: {part_count} 個料號 → {os.path.basename(output_path)}")

    total_files = len(detected) + len(drift_files)
    drift_msg = f", 含 {len(drift_files)} 個漂移版" if drift_files else ""
    result = {
        'success': True,
        'part_count': part_count,
        'date_col_count': len(date_cols),
        'format_stats': format_stats,
        # 向後相容: buyer_stats 用舊的 3 Buyer 名稱 filter
        'buyer_stats': {
            name: sum(cnt for fk, cnt in format_stats.items() if name.lower() in fk.lower())
            for name in ('Ketwadee', 'Kanyanat', 'Weeraya')
        },
        'message': f'成功合併 {part_count} 個料號 ({total_files} 個檔案{drift_msg})'
    }
    if date_warnings:
        result['date_warnings'] = date_warnings
    if drift_stats:
        result['drift_stats'] = drift_stats
    return result
