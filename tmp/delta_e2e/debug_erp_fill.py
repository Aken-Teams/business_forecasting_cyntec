"""Debug: 追蹤 ERP 回填對特定料號 204840750234 的行為"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, r'D:\github\business_forecasting_cyntec')

import openpyxl
import pandas as pd
from datetime import datetime

TARGET_PARTNO = '204840750234'
PROC_DIR = r'D:\github\business_forecasting_cyntec\processed\7\20260413_115810'

# =========================================================
# 1. ERP 整合檔中這個料號有幾筆
# =========================================================
print("=" * 70)
print(f"🔍 追蹤料號: {TARGET_PARTNO}")
print("=" * 70)

erp_path = os.path.join(PROC_DIR, 'integrated_erp.xlsx')
erp_df = pd.read_excel(erp_path)
print(f"\n📊 ERP 整合檔: {len(erp_df)} 行, 欄位: {list(erp_df.columns)}")

# 用多種方式找料號 (可能有科學記號)
mask = erp_df['客戶料號'].astype(str).str.strip().str.replace('.0','',regex=False) == TARGET_PARTNO
erp_target = erp_df[mask]
if len(erp_target) == 0:
    # 嘗試數值比對
    try:
        mask2 = erp_df['客戶料號'].apply(lambda x: str(int(float(x))) if pd.notna(x) else '') == TARGET_PARTNO
        erp_target = erp_df[mask2]
    except:
        pass

print(f"\n📌 ERP 中料號 {TARGET_PARTNO} 共 {len(erp_target)} 筆:")
for i, (_, row) in enumerate(erp_target.iterrows()):
    print(f"   [{i}] 客戶簡稱={row.get('客戶簡稱','?')}, "
          f"送貨地點={row.get('送貨地點','?')}, "
          f"客戶需求地區={row.get('客戶需求地區','?')}")
    print(f"        淨需求={row.get('淨需求','?')}, "
          f"排程出貨日期={row.get('排程出貨日期','?')}, "
          f"斷點={row.get('排程出貨日期斷點','?')}, "
          f"ETA={row.get('ETA','?')}")

# =========================================================
# 2. Forecast BEFORE ERP fill (integrated_forecast.xlsx = Step3 結果)
# =========================================================
before_path = os.path.join(PROC_DIR, 'integrated_forecast.xlsx')
wb_before = openpyxl.load_workbook(before_path, data_only=True)
ws_before = wb_before.active

print(f"\n{'='*70}")
print(f"📊 Forecast Step3 結果 (ERP 回填前): integrated_forecast.xlsx")
print(f"{'='*70}")

# Header
headers = {}
for cell in ws_before[1]:
    if cell.value is not None:
        headers[cell.column] = str(cell.value)

print(f"   Col 1~10: {[headers.get(i,'') for i in range(1,11)]}")
print(f"   日期欄: {[(c, headers[c]) for c in sorted(headers) if c >= 10]}")

# 找料號的所有 row
print(f"\n📌 料號 {TARGET_PARTNO} 在 ERP 回填前:")
for r in range(2, ws_before.max_row + 1):
    partno_v = ws_before.cell(row=r, column=5).value
    if partno_v is not None and str(partno_v).strip() == TARGET_PARTNO:
        plant = ws_before.cell(row=r, column=2).value
        customer = ws_before.cell(row=r, column=3).value
        location = ws_before.cell(row=r, column=4).value
        row_type = ws_before.cell(row=r, column=9).value
        stock = ws_before.cell(row=r, column=7).value
        onway = ws_before.cell(row=r, column=8).value

        date_vals = {}
        for col_idx in sorted(headers):
            if col_idx >= 10:
                val = ws_before.cell(row=r, column=col_idx).value
                if val is not None and val != '' and val != 0:
                    date_vals[headers[col_idx]] = val
        print(f"   Row {r}: Type={row_type}, PLANT={plant}, C={customer}, D={location}")
        print(f"           STOCK={stock}, ONWAY={onway}")
        print(f"           日期欄: {date_vals}")
wb_before.close()

# =========================================================
# 3. Forecast AFTER ERP fill (forecast_result.xlsx = Step4 結果)
# =========================================================
after_path = os.path.join(PROC_DIR, 'forecast_result.xlsx')
wb_after = openpyxl.load_workbook(after_path, data_only=True)
ws_after = wb_after.active

print(f"\n{'='*70}")
print(f"📊 Forecast Step4 結果 (ERP 回填後): forecast_result.xlsx")
print(f"{'='*70}")

print(f"\n📌 料號 {TARGET_PARTNO} 在 ERP 回填後:")
for r in range(2, ws_after.max_row + 1):
    partno_v = ws_after.cell(row=r, column=5).value
    if partno_v is not None and str(partno_v).strip() == TARGET_PARTNO:
        plant = ws_after.cell(row=r, column=2).value
        customer = ws_after.cell(row=r, column=3).value
        location = ws_after.cell(row=r, column=4).value
        row_type = ws_after.cell(row=r, column=9).value
        stock = ws_after.cell(row=r, column=7).value
        onway = ws_after.cell(row=r, column=8).value

        date_vals = {}
        for col_idx in sorted(headers):
            if col_idx >= 10:
                val = ws_after.cell(row=r, column=col_idx).value
                if val is not None and val != '' and val != 0:
                    date_vals[headers[col_idx]] = val
        print(f"   Row {r}: Type={row_type}, PLANT={plant}, C={customer}, D={location}")
        print(f"           STOCK={stock}, ONWAY={onway}")
        print(f"           日期欄: {date_vals}")
wb_after.close()

# =========================================================
# 4. 模擬 ETA 目標日期計算
# =========================================================
from delta_forecast_step4 import calculate_eta_target_date, find_fill_col, build_date_col_map

wb_tmp = openpyxl.load_workbook(after_path, data_only=True)
ws_tmp = wb_tmp.active
date_col_map = build_date_col_map(ws_tmp, start_col=10)

print(f"\n{'='*70}")
print(f"🧮 模擬 ERP ETA 計算")
print(f"{'='*70}")

for i, (_, row) in enumerate(erp_target.iterrows()):
    schedule = row.get('排程出貨日期')
    breakpoint_text = row.get('排程出貨日期斷點')
    eta_text = row.get('ETA')
    net = row.get('淨需求')

    target = calculate_eta_target_date(schedule, breakpoint_text, eta_text)
    col = find_fill_col(date_col_map, target) if target else None
    col_header = ws_tmp.cell(row=1, column=col).value if col else None

    print(f"\n   [{i}] 排程出貨日期={schedule}")
    print(f"        斷點={breakpoint_text}, ETA text={eta_text}")
    print(f"        淨需求={net}, ×1000={float(net)*1000 if pd.notna(net) and net else 0}")
    print(f"        → 計算出目標日期={target}")
    print(f"        → 對應欄位 Col {col} (header={col_header})")

# =========================================================
# 5. 比較 Before vs After 差異
# =========================================================
wb_b = openpyxl.load_workbook(before_path, data_only=True)
wb_a = openpyxl.load_workbook(after_path, data_only=True)
ws_b = wb_b.active
ws_a = wb_a.active

print(f"\n{'='*70}")
print(f"📊 Before vs After 差異 (料號 {TARGET_PARTNO})")
print(f"{'='*70}")

for r in range(2, ws_a.max_row + 1):
    partno_v = ws_a.cell(row=r, column=5).value
    if partno_v is not None and str(partno_v).strip() == TARGET_PARTNO:
        row_type = ws_a.cell(row=r, column=9).value
        diffs = {}
        for col_idx in sorted(headers):
            if col_idx >= 10:
                before_val = ws_b.cell(row=r, column=col_idx).value or 0
                after_val = ws_a.cell(row=r, column=col_idx).value or 0
                try:
                    b = float(before_val) if before_val else 0
                    a = float(after_val) if after_val else 0
                except:
                    continue
                if abs(a - b) > 0.01:
                    diffs[headers[col_idx]] = f"{b} → {a} (差={a-b})"
        if diffs:
            print(f"   Row {r} ({row_type}): {diffs}")
        else:
            print(f"   Row {r} ({row_type}): 無差異")

wb_b.close()
wb_a.close()
wb_tmp.close()

print("\n✅ Debug 完成")
