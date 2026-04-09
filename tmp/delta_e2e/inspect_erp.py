import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl

NBSP = '\xa0'
SRC_DIR = r"C:\Users\petty\Desktop\客戶相關資料\01.強茂\台達業務"
ERP_FILE = os.path.join(SRC_DIR, f"0408-上午淨需求{NBSP}(台達).xlsx")

print(f"Reading: {ERP_FILE}")
print(f"Exists: {os.path.exists(ERP_FILE)}")

wb = openpyxl.load_workbook(ERP_FILE, read_only=True, data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} (max_col={ws.max_column}, max_row={ws.max_row}) ===")
    # Print first row as headers
    for col in range(1, min(ws.max_column + 1, 60)):
        v = ws.cell(row=1, column=col).value
        if v is not None:
            print(f"  col {col}: {v!r}")
    # Print 2 data rows
    print(f"  Data rows (2-3):")
    for r in range(2, 4):
        for col in range(1, min(ws.max_column + 1, 10)):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                print(f"    row {r} col {col}: {v!r}")
wb.close()
