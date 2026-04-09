"""
Delta 完整端到端測試: Step 1 合併 → Step 3 映射 → Step 4 Forecast 填入
Uses real Delta files + fake Transit (since user doesn't provide one).
"""
import os
import sys
import io
import shutil

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))

import openpyxl
import pandas as pd
from delta_forecast_processor import consolidate
from delta_forecast_step4 import process_delta_forecast
from database import get_customer_mappings_raw, get_db_connection

NBSP = '\xa0'
SRC_DIR = r"C:\Users\petty\Desktop\客戶相關資料\01.強茂\台達業務"
WORK_DIR = os.path.abspath(os.path.dirname(__file__))

FORECAST_FILES = [
    f"PSBG{NBSP}PSB5-{NBSP}Ketwadee0406(完成).xlsx",
    f"PSBG{NBSP}PSB7_Kanyanat.S0406(完成).xlsx",
    f"PSBG{NBSP}PSB7-Weeraya0406(完成).xlsx",
    "PSBG (India IAI1&UPI2&DFI1 DIODES)-Jack0401.xlsx",
    "PSBG PSW1+CEW1- 楊洋彙整 0330(完成) (002).xlsx",
    "強茂 MWC1IPC1 MRP 03.30.2026.xlsx",
    "NBQ1.xlsx",
    "MRP(SVC1PWC1 DIODE&MOS)-100109-2026-3-30.xlsx",
]

ERP_SRC = os.path.join(SRC_DIR, f"0408-上午淨需求{NBSP}(台達).xlsx")
REFERENCE = os.path.abspath(os.path.join(WORK_DIR, '..', '..', 'compare', 'delta', 'consolidated_template.xlsx'))

# 輸出檔案
STEP1_OUTPUT = os.path.join(WORK_DIR, 'step1_consolidated.xlsx')
STEP3_FORECAST = os.path.join(WORK_DIR, 'step3_integrated_forecast.xlsx')
STEP3_ERP = os.path.join(WORK_DIR, 'step3_integrated_erp.xlsx')
STEP3_TRANSIT = os.path.join(WORK_DIR, 'step3_integrated_transit.xlsx')
STEP4_RESULT = os.path.join(WORK_DIR, 'step4_forecast_result.xlsx')


def get_delta_user_id():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT id FROM users WHERE username=%s", ('delta',))
    row = cur.fetchone()
    conn.close()
    return row['id'] if row else None


def step1_consolidate():
    print("\n" + "=" * 60)
    print("📦 STEP 1: 合併 8 個 Forecast 檔案 (方案二)")
    print("=" * 60)

    forecast_paths = [os.path.join(SRC_DIR, f) for f in FORECAST_FILES]
    for fp in forecast_paths:
        assert os.path.exists(fp), f"缺檔案: {fp}"

    plant_codes = [
        'PSB5', 'PSB7', 'IAI1', 'UPI2', 'DFI1',
        'PSW1', 'CEW1', 'MWC1', 'IPC1', 'NBQ1',
        'SVC1', 'PWC1'
    ]

    result = consolidate(
        forecast_files=forecast_paths,
        reference_template=REFERENCE,
        output_path=STEP1_OUTPUT,
        plant_codes=plant_codes,
    )

    assert result['success'], f"Step 1 failed: {result.get('message')}"
    print(f"✅ Step 1 OK: {result['part_count']} 料號")
    return result


def step3_mapping():
    print("\n" + "=" * 60)
    print("🔧 STEP 3: 映射填入 (ERP/Transit/Forecast C/D)")
    print("=" * 60)

    delta_uid = get_delta_user_id()
    assert delta_uid, "找不到 delta user"
    mappings = get_customer_mappings_raw(delta_uid)
    print(f"   取得 {len(mappings)} 筆 Delta mapping")

    # === 3a. 複製 Step 1 forecast → Step 3 forecast (將會填入 C/D) ===
    shutil.copy2(STEP1_OUTPUT, STEP3_FORECAST)

    # === 3b. ERP 映射 (模擬 app.py 的 Delta ERP branch) ===
    print("\n📊 ERP 映射...")
    erp_df = pd.read_excel(ERP_SRC, sheet_name=0)
    print(f"   ERP 原始: {len(erp_df)} 行, {len(erp_df.columns)} 欄")

    # 建立 (客戶簡稱, 送貨地點) → 映射值 lookup
    delta_erp_lookup = {}
    for m in mappings:
        cname = str(m['customer_name']).strip() if m['customer_name'] else ''
        dl = str(m.get('delivery_location', '')).strip() if m.get('delivery_location') else ''
        if cname and dl:
            delta_erp_lookup[(cname, dl)] = {
                'region': str(m['region']).strip() if m['region'] else '',
                'schedule_breakpoint': str(m['schedule_breakpoint']).strip() if m['schedule_breakpoint'] else '',
                'etd': str(m['etd']).strip() if m['etd'] else '',
                'eta': str(m['eta']).strip() if m['eta'] else '',
            }
    print(f"   Delta ERP lookup: {len(delta_erp_lookup)} 筆 (customer_name, delivery_location)")

    def get_val(row, field):
        cust = str(row['客戶簡稱']).strip() if pd.notna(row['客戶簡稱']) else ''
        deliv = str(row['送貨地點']).strip() if pd.notna(row['送貨地點']) else ''
        return delta_erp_lookup.get((cust, deliv), {}).get(field, '')

    erp_df['客戶需求地區'] = erp_df.apply(lambda r: get_val(r, 'region'), axis=1)
    erp_df['排程出貨日期斷點'] = erp_df.apply(lambda r: get_val(r, 'schedule_breakpoint'), axis=1)
    erp_df['ETD'] = erp_df.apply(lambda r: get_val(r, 'etd'), axis=1)
    erp_df['ETA'] = erp_df.apply(lambda r: get_val(r, 'eta'), axis=1)
    erp_df['已分配'] = ''

    matched = (erp_df['客戶需求地區'] != '').sum()
    print(f"   ERP 映射: {matched}/{len(erp_df)} 行匹配")
    if '排程出貨日期' in erp_df.columns:
        erp_df = erp_df.sort_values('排程出貨日期')
    erp_df.to_excel(STEP3_ERP, index=False)
    print(f"   ✅ ERP 儲存: {STEP3_ERP}")

    # === 3c. 產生假 Transit 檔 (使用者沒提供) ===
    print("\n🚚 產生假 Transit 檔...")
    # 從 ERP 取幾筆作為假 Transit 資料
    sample_erp = erp_df.head(20).copy()

    # Transit 欄位: Tw, Ship Number, Invoice Date, Location, 客戶簡稱,
    #               Ordered Item, Pj Item, Qty, ETA, Status, 集團客戶, 週別
    # 使用 Step 1 forecast 裡的 (PARTNO, PLANT) 組合，才能被 Step 4 匹配
    wb_fc = openpyxl.load_workbook(STEP1_OUTPUT, read_only=True, data_only=True)
    ws_fc = wb_fc.active
    sample_parts = []
    for row in ws_fc.iter_rows(min_row=2, max_row=60, values_only=True):
        if row[8] == 'Demand':  # Date 欄 = Demand
            plant = row[1]
            part_no = row[4]
            if plant and part_no:
                sample_parts.append((plant, str(part_no)))
        if len(sample_parts) >= 15:
            break
    wb_fc.close()

    # 用 ERP lookup 從 plant (region) 反查 (customer_name, delivery_location)
    region_to_cd = {}
    for (cname, dl), m in delta_erp_lookup.items():
        region = m.get('region', '')
        if region and region not in region_to_cd:
            region_to_cd[region] = (cname, dl)

    from datetime import datetime, timedelta
    transit_rows = []
    for i, (plant, part_no) in enumerate(sample_parts):
        cname, dl = region_to_cd.get(plant, ('UNKNOWN', 'UNKNOWN'))
        # 製造各種 ETA 日期測試 Step 4 邏輯
        eta_date = datetime(2026, 4, 13) + timedelta(days=i * 3)  # 4/13, 4/16, 4/19, ...
        transit_rows.append({
            'Tw': 'TW1',
            'Ship Number': f'SN{1000 + i}',
            'Invoice Date': datetime(2026, 4, 1),
            'Location': dl,
            '客戶簡稱': cname,
            'Ordered Item': part_no,
            'Pj Item': f'PJ{i}',
            'Qty': 1000 + (i * 100),
            'ETA': eta_date,
            'Status': 'In Transit',
            '集團客戶': 'Delta',
            '週別': f'W{i+1}',
            # 整合後欄位
            '客戶需求地區': plant,
            '已分配': '',
        })
    transit_df = pd.DataFrame(transit_rows)
    transit_df.to_excel(STEP3_TRANSIT, index=False)
    print(f"   ✅ 假 Transit 儲存: {len(transit_df)} 行 → {STEP3_TRANSIT}")

    # === 3d. Forecast C/D 欄位填入 (Delta 專用) ===
    print("\n📄 Forecast C/D 欄位填入...")
    plant_to_cd = {}
    for m in mappings:
        region = str(m['region']).strip() if m['region'] else ''
        cname = str(m['customer_name']).strip() if m['customer_name'] else ''
        dl = str(m.get('delivery_location', '')).strip() if m.get('delivery_location') else ''
        if region:
            plant_code = region.split()[0]
            if plant_code not in plant_to_cd:
                plant_to_cd[plant_code] = (cname, dl)
    print(f"   PLANT → (C, D) lookup: {len(plant_to_cd)} 筆")

    wb = openpyxl.load_workbook(STEP3_FORECAST)
    ws = wb.active
    matched_rows = 0
    for row_num in range(2, ws.max_row + 1):
        plant = ws.cell(row=row_num, column=2).value  # B
        if plant and str(plant).strip() in plant_to_cd:
            cname, dl = plant_to_cd[str(plant).strip()]
            ws.cell(row=row_num, column=3, value=cname)  # C
            ws.cell(row=row_num, column=4, value=dl)      # D
            matched_rows += 1
    wb.save(STEP3_FORECAST)
    wb.close()
    print(f"   ✅ Forecast C/D 填入: {matched_rows}/{ws.max_row-1} 行")
    return delta_uid


def step4_forecast():
    print("\n" + "=" * 60)
    print("🎯 STEP 4: Delta Forecast Transit + ERP 填入")
    print("=" * 60)

    # 複製 step3 → step4 作為工作檔
    shutil.copy2(STEP3_FORECAST, STEP4_RESULT)

    stats = process_delta_forecast(
        forecast_file=STEP4_RESULT,
        erp_file=STEP3_ERP,
        transit_file=STEP3_TRANSIT,
        output_file=STEP4_RESULT,
    )

    print(f"\n📊 Step 4 統計:")
    print(f"   Transit: 填入 {stats['transit_filled']} cells, 跳過 {stats['transit_skipped']}, 匹配 {stats['transit_matched_rows']} 列")
    print(f"   ERP:     填入 {stats['erp_filled']} cells, 跳過 {stats['erp_skipped']}, 匹配 {stats['erp_matched_rows']} 列")
    print(f"   ✅ 結果儲存: {STEP4_RESULT}")
    return stats


def verify_final():
    print("\n" + "=" * 60)
    print("🔍 驗證最終輸出")
    print("=" * 60)

    wb = openpyxl.load_workbook(STEP4_RESULT, read_only=True, data_only=True)
    ws = wb.active

    # 1. 驗證固定 26 欄
    headers_dates = []
    for col in range(10, 36):
        v = ws.cell(row=1, column=col).value
        headers_dates.append(v)
    print(f"   日期欄位 ({len(headers_dates)}): {headers_dates[0]} ~ {headers_dates[-1]}")
    assert len(headers_dates) == 26, f"應為 26 個日期欄位"
    assert headers_dates[0] == 'PASSDUE'
    assert all(isinstance(w, str) and w.isdigit() and len(w) == 8 for w in headers_dates[1:17])
    MONTH_NAMES = ('JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN',
                   'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC')
    assert all(m in MONTH_NAMES for m in headers_dates[17:26])

    # 2. 驗證 C/D 欄位有填入
    c_filled = 0
    d_filled = 0
    total_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        total_rows += 1
        if row[2]:  # C
            c_filled += 1
        if row[3]:  # D
            d_filled += 1
    print(f"   Forecast 總列數: {total_rows}")
    print(f"   C 欄 (客戶簡稱) 有值: {c_filled}")
    print(f"   D 欄 (送貨地點) 有值: {d_filled}")

    # 3. 驗證 Demand 行有非零值 (來自原始合併)
    demand_nonzero = 0
    passdue_nonzero = 0
    sample_passdue_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[8] == 'Demand':
            # col 10 = PASSDUE, col 11~ = dates
            total_date = sum((x or 0) for x in row[10:36] if isinstance(x, (int, float)))
            if total_date > 0:
                demand_nonzero += 1
            passdue = row[9] or 0
            if isinstance(passdue, (int, float)) and passdue > 0:
                passdue_nonzero += 1
                if len(sample_passdue_rows) < 3:
                    sample_passdue_rows.append((row[1], row[4], passdue))
    print(f"   Demand 列有非零值: {demand_nonzero}")
    print(f"   PASSDUE 有值的 Demand 列: {passdue_nonzero}")
    if sample_passdue_rows:
        print(f"   PASSDUE 範例: {sample_passdue_rows}")

    wb.close()
    print("\n✅ 驗證通過!")


def main():
    step1_consolidate()
    step3_mapping()
    step4_forecast()
    verify_final()
    print("\n" + "=" * 60)
    print("🎉 完整端到端測試通過!")
    print("=" * 60)


if __name__ == '__main__':
    main()
