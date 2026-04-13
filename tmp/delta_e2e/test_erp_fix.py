"""驗證 ERP 修正: 只填 Supply + 已分配標記"""
import sys, io, os, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, r'D:\github\business_forecasting_cyntec')

import openpyxl
import pandas as pd

PROC_DIR = r'D:\github\business_forecasting_cyntec\processed\7\20260413_115810'
TEST_DIR = r'D:\github\business_forecasting_cyntec\tmp\delta_e2e\fix_test'
os.makedirs(TEST_DIR, exist_ok=True)

# 複製 integrated_forecast 和 integrated_erp 到測試目錄
src_forecast = os.path.join(PROC_DIR, 'integrated_forecast.xlsx')
src_erp = os.path.join(PROC_DIR, 'integrated_erp.xlsx')

test_forecast = os.path.join(TEST_DIR, 'forecast_result.xlsx')
test_erp = os.path.join(TEST_DIR, 'integrated_erp.xlsx')

shutil.copy2(src_forecast, test_forecast)
shutil.copy2(src_erp, test_erp)

# 執行 Step 4
from delta_forecast_step4 import process_delta_forecast

print("=" * 70)
print("🔧 執行修正後的 Step 4")
print("=" * 70)

stats = process_delta_forecast(
    forecast_file=test_forecast,
    erp_file=test_erp,
    transit_file=None,
    output_file=test_forecast,
)

print(f"\n📊 統計: {stats}")

# =========================================================
# 驗證 1: 檢查 ERP 已分配
# =========================================================
print(f"\n{'='*70}")
print("🔍 驗證 1: ERP 已分配欄位")
print(f"{'='*70}")

erp_df = pd.read_excel(test_erp)
allocated = erp_df[erp_df['已分配'] == 'Y']
not_allocated = erp_df[erp_df['已分配'] != 'Y']
print(f"   已分配: {len(allocated)} 筆")
print(f"   未分配: {len(not_allocated)} 筆")

# 顯示料號 204840750234 的已分配狀態
TARGET = '204840750234'
mask = erp_df['客戶料號'].astype(str).str.strip().str.replace('.0','',regex=False) == TARGET
if mask.sum() == 0:
    try:
        mask = erp_df['客戶料號'].apply(lambda x: str(int(float(x))) if pd.notna(x) else '') == TARGET
    except:
        pass

target_rows = erp_df[mask]
print(f"\n   料號 {TARGET} 共 {len(target_rows)} 筆:")
for _, row in target_rows.iterrows():
    print(f"     客戶需求地區={row.get('客戶需求地區','?')}, "
          f"淨需求={row.get('淨需求','?')}, "
          f"已分配={row.get('已分配','')}")

# =========================================================
# 驗證 2: Forecast 只有 Supply 被改
# =========================================================
print(f"\n{'='*70}")
print(f"🔍 驗證 2: Before vs After (料號 {TARGET})")
print(f"{'='*70}")

wb_before = openpyxl.load_workbook(src_forecast, data_only=True)
wb_after = openpyxl.load_workbook(test_forecast, data_only=True)
ws_b = wb_before.active
ws_a = wb_after.active

headers = {}
for cell in ws_a[1]:
    if cell.value is not None:
        headers[cell.column] = str(cell.value)

for r in range(2, ws_a.max_row + 1):
    partno_v = ws_a.cell(row=r, column=5).value
    if partno_v is not None and str(partno_v).strip() == TARGET:
        row_type = ws_a.cell(row=r, column=9).value
        plant = ws_a.cell(row=r, column=2).value
        diffs = {}
        for col_idx in sorted(headers):
            if col_idx >= 10:
                before_val = ws_b.cell(row=r, column=col_idx).value or 0
                after_val = ws_a.cell(row=r, column=col_idx).value or 0
                try:
                    b = float(before_val) if before_val else 0
                    a = float(after_val) if after_val else 0
                except:
                    continue
                if abs(a - b) > 0.01:
                    diffs[headers[col_idx]] = f"{b:,.0f} → {a:,.0f} (+{a-b:,.0f})"

        status = "✅ 無變動" if not diffs else f"{'✅' if row_type == 'Supply' else '❌'} 有變動"
        if diffs:
            print(f"   Row {r} [{row_type:8s}] PLANT={plant}: {status}")
            for k, v in diffs.items():
                print(f"      {k}: {v}")
        elif row_type in ('Demand', 'Balance'):
            print(f"   Row {r} [{row_type:8s}] PLANT={plant}: ✅ 正確無變動")

wb_before.close()
wb_after.close()

print(f"\n✅ 驗證完成")
