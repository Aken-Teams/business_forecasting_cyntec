"""
Delta Forecast 回填至原格式檔案
=====================================

目的:
    Step 4 產出 forecast_result.xlsx (匯總格式 Demand/Supply/Balance 3 列/料號)。
    此模組將 Supply 值回填到客戶原本各式 buyer 檔案 (保留樣式/公式/合併儲存格),
    方便直接交回 Delta 各採購單位。

設計:
    1. **完全重用** delta_unified_reader 的欄位定位邏輯 (scan_headers, find_marker_col, ...)
       因 reader 邏輯已能處理 15 格式 + 漂移版, writer 用同邏輯自然一致。
    2. openpyxl load-modify-save 流程, 樣式完整保留。
    3. Flat 格式 (EIBG/IABG/NBQ1/SVC1PWC1) 無 Supply 列 → 自動插入 Supply + Balance 兩列。
    4. 同 (plant, partno) 在 forecast_result 可能多列 (不同 customer/location) → 加總。

使用:
    result = backfill_one_file(original_path, forecast_result_path, output_path,
                               plant_codes=['PSB5', 'PSB7', ...], file_label='PSB5 Ketwadee.xlsx')
    zip_result = backfill_session_to_zip(originals_dir, forecast_result_path, zip_path, plant_codes)
"""
import os
import re
import zipfile
import traceback
from collections import defaultdict
from datetime import datetime

from copy import copy

import openpyxl
from openpyxl.utils import get_column_letter

from delta_unified_reader import (
    find_valid_sheets, scan_headers, find_first_date_col, collect_date_cols,
    find_marker_col, classify_marker, _read_header_row,
)
from delta_forecast_processor import (
    detect_format, FORMAT_LABELS, SINGLE_PLANT_FORMATS,
    match_plants_in_filename, _normalize_date_header, MONTH_NAMES,
)
from delta_format_fingerprint import match_known_format_fingerprint


# ============ forecast_result.xlsx 讀取 ============

# forecast_result.xlsx 欄位結構 (固定):
FR_COL_PLANT = 2       # B
FR_COL_CUSTOMER = 3    # C
FR_COL_LOCATION = 4    # D
FR_COL_PARTNO = 5      # E
FR_COL_ROW_TYPE = 9    # I
FR_DATE_START_COL = 10  # J


def _load_forecast_result_supply(forecast_result_path):
    """讀 forecast_result.xlsx, 回傳 Supply lookup table + 日期欄位 list。

    Returns:
        tuple(lookup, fr_date_keys)
        - lookup: {(plant_upper, partno_str): {date_key: supply_value}}
                  同 (plant, partno) 跨 customer/location 的 Supply 會加總。
        - fr_date_keys: ['PASSDUE', YYYYMMDD, ..., 'AUG', 'SEP', ...]
    """
    wb = openpyxl.load_workbook(forecast_result_path, read_only=True, data_only=True)
    try:
        ws = wb.active

        # 讀 header 取日期欄
        fr_date_keys = []
        fr_date_col_to_key = {}  # {col: key}
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        for c in range(FR_DATE_START_COL, len(header_row) + 1):
            v = header_row[c - 1]
            if v is None:
                continue
            key = _normalize_date_header(v)
            if key is None:
                # 可能是 'PASSDUE' 等特殊值
                s = str(v).strip().upper()
                if 'PASSDUE' in s or 'PAST' in s:
                    key = 'PASSDUE'
                elif s in MONTH_NAMES:
                    key = s
            if key:
                fr_date_keys.append(key)
                fr_date_col_to_key[c] = key

        # 掃 Supply 列建 lookup
        lookup = defaultdict(lambda: defaultdict(float))
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < FR_COL_ROW_TYPE:
                continue
            rt = row[FR_COL_ROW_TYPE - 1]
            if rt is None or str(rt).strip() != 'Supply':
                continue
            plant = row[FR_COL_PLANT - 1]
            partno = row[FR_COL_PARTNO - 1]
            if not plant or not partno:
                continue
            plant_key = str(plant).strip().upper()
            partno_key = str(partno).strip()
            key = (plant_key, partno_key)
            for c, date_key in fr_date_col_to_key.items():
                if c - 1 >= len(row):
                    continue
                v = row[c - 1]
                if v is None or v == '':
                    continue
                try:
                    n = float(v)
                except (TypeError, ValueError):
                    continue
                if n == 0:
                    continue
                lookup[key][date_key] += n

        return dict(lookup), fr_date_keys
    finally:
        wb.close()


# ============ 原檔日期 → forecast_result canonical key 對應 ============

def _build_canonical_map(date_col_map, fr_date_keys):
    """建立「原檔日期欄 → forecast_result 對應 key」的 mapping。

    Args:
        date_col_map: {col: date_str} 來自 collect_date_cols()
        fr_date_keys: forecast_result 的日期 key 清單 (list, order preserved)

    Returns:
        {col: canonical_key}  — 只包含能對應到 fr_date_keys 的欄
    """
    fr_set = set(fr_date_keys)
    # 分離 weekly (YYYYMMDD Monday) 與 monthly (MONTH_NAMES)
    weekly_mondays = []
    for k in fr_date_keys:
        if isinstance(k, str) and k.isdigit() and len(k) == 8:
            try:
                weekly_mondays.append(datetime.strptime(k, '%Y%m%d'))
            except ValueError:
                pass
    weekly_mondays.sort()
    fr_monthly = {k for k in fr_date_keys if k in MONTH_NAMES}
    has_passdue = 'PASSDUE' in fr_set

    col_to_canonical = {}
    for col, date_str in date_col_map.items():
        if date_str == 'PASSDUE':
            if has_passdue:
                col_to_canonical[col] = 'PASSDUE'
            continue
        if date_str in MONTH_NAMES:
            if date_str in fr_monthly:
                col_to_canonical[col] = date_str
            continue
        if isinstance(date_str, str) and date_str.isdigit() and len(date_str) == 8:
            try:
                dt = datetime.strptime(date_str, '%Y%m%d')
            except ValueError:
                continue
            # 1. 嘗試對應到 fr_date_keys 中的週 Monday (允許同週內的日期)
            matched_monday = None
            for mon in weekly_mondays:
                diff = (dt - mon).days
                if 0 <= diff <= 6:
                    matched_monday = mon.strftime('%Y%m%d')
                    break
            if matched_monday and matched_monday in fr_set:
                col_to_canonical[col] = matched_monday
                continue
            # 2. 超出週範圍 → 看是否落在 fr_monthly
            m_label = MONTH_NAMES[dt.month - 1]
            if m_label in fr_monthly:
                col_to_canonical[col] = m_label
            # 3. 否則丟棄
    return col_to_canonical


# ============ 合併儲存格偵測 ============

def _build_merged_cells_lookup(ws):
    """建立 set of (row, col) 代表「此儲存格是合併區但不是左上角」。
    寫入這些儲存格會被 openpyxl 靜默略過, 但我們想明確跳過並計數。
    """
    locked = set()
    for mr in ws.merged_cells.ranges:
        # merged_cells.ranges 是 CellRange 物件, 包含 min_row/max_row/min_col/max_col
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    locked.add((r, c))
    return locked


# ============ Flat 格式回填 (插入 Supply + Balance 列) ============

def _backfill_flat_sheet(ws, header_row, found, date_col_map, col_to_canonical,
                         lookup, filename_plant, result):
    """Flat 格式回填: 每筆 PARTNO 只有 Demand 一列。
    為每筆插入 Supply + Balance 兩列 (Balance 含公式)。

    策略:
      1. 在日期欄前插入一個專用標記欄 → 寫 Demand/Supply/Balance, 不混用原有欄位
      2. 從上往下處理 + offset 追蹤, 因 openpyxl insert_rows 不自動更新公式參照
    """
    partno_col = found['partno']
    plant_col = found.get('plant')
    stock_col = found.get('stock')
    on_way_col = found.get('on_way')
    date_start = min(date_col_map.keys()) if date_col_map else None
    if date_start is None:
        return

    # ---- 插入專用標記欄 (在 stock_col 前, 讓 Demand/Supply/Balance 緊鄰數值區) ----
    if stock_col:
        label_col = stock_col
    else:
        label_col = date_start
    ws.insert_cols(label_col, 1)
    ws.cell(header_row, label_col).value = '類別'

    # 插入欄後, >= label_col 的欄位 index 全部 +1
    col_to_canonical = {c + 1: v for c, v in col_to_canonical.items()}
    sorted_date_cols = sorted(col_to_canonical.keys())
    max_col = ws.max_column or (max(sorted_date_cols) if sorted_date_cols else 20)
    if stock_col and stock_col >= label_col:
        stock_col += 1
    if on_way_col and on_way_col >= label_col:
        on_way_col += 1

    # ---- 收集所有有效 PARTNO 的資料列 (top → bottom) ----
    max_row = ws.max_row or (header_row + 1)
    data_rows = []  # [(original_row_idx, partno, plant)]
    for r in range(header_row + 1, max_row + 1):
        pv = ws.cell(r, partno_col).value
        if pv is None or not str(pv).strip():
            continue
        partno = str(pv).strip()
        plant = filename_plant
        if plant_col:
            plv = ws.cell(r, plant_col).value
            if plv is not None and str(plv).strip():
                plant = str(plv).strip()
        data_rows.append((r, partno, plant))

    if not data_rows:
        return

    offset = 0  # 累計已插入的行數

    for orig_row, partno, plant in data_rows:
        demand_row = orig_row + offset
        # 插入 2 列在 demand_row 下方
        ws.insert_rows(demand_row + 1, 2)
        supply_row = demand_row + 1
        balance_row = demand_row + 2
        offset += 2

        # 複製格式 (從 demand row 到 supply / balance rows)
        for c in range(1, max_col + 1):
            src_cell = ws.cell(demand_row, c)
            for target_r in (supply_row, balance_row):
                dst_cell = ws.cell(target_r, c)
                if src_cell.has_style:
                    dst_cell.font = copy(src_cell.font)
                    dst_cell.fill = copy(src_cell.fill)
                    dst_cell.border = copy(src_cell.border)
                    dst_cell.alignment = copy(src_cell.alignment)
                    dst_cell.number_format = src_cell.number_format

        # 寫標記欄 (Demand / Supply / Balance)
        ws.cell(demand_row, label_col).value = 'Demand'
        ws.cell(supply_row, label_col).value = 'Supply'
        ws.cell(balance_row, label_col).value = 'Balance'

        # 寫 PARTNO (方便辨識)
        ws.cell(supply_row, partno_col).value = partno
        ws.cell(balance_row, partno_col).value = partno

        # Lookup Supply
        plant_key = (plant or '').strip().upper()
        lookup_key = (plant_key, partno)
        supply_dict = lookup.get(lookup_key, {})

        if supply_dict:
            result['n_partno_matched'] += 1

        # 寫 Supply 值 + Balance 公式
        for i, col in enumerate(sorted_date_cols):
            canonical = col_to_canonical[col]
            col_letter = get_column_letter(col)

            # Supply value
            supply_val = supply_dict.get(canonical)
            if supply_val is not None:
                if isinstance(supply_val, float) and supply_val.is_integer():
                    ws.cell(supply_row, col).value = int(supply_val)
                else:
                    ws.cell(supply_row, col).value = supply_val
                result['n_cells_written'] += 1

            # Balance formula
            if i == 0:
                # 首欄: Balance = Stock [+ OnWay] + Supply - Demand
                parts = []
                if stock_col:
                    parts.append(f'{get_column_letter(stock_col)}{demand_row}')
                if on_way_col:
                    parts.append(f'{get_column_letter(on_way_col)}{demand_row}')
                parts.append(f'{col_letter}{supply_row}')
                formula = '=' + '+'.join(parts) + f'-{col_letter}{demand_row}'
            else:
                # 後續: Balance = prev_Balance + Supply - Demand
                prev_col_letter = get_column_letter(sorted_date_cols[i - 1])
                formula = (f'={prev_col_letter}{balance_row}'
                           f'+{col_letter}{supply_row}'
                           f'-{col_letter}{demand_row}')

            ws.cell(balance_row, col).value = formula
            result['n_cells_written'] += 1


# ============ 單檔回填 ============

def backfill_one_file(original_path, forecast_result_path, output_path,
                      plant_codes=None, file_label=None):
    """將 forecast_result.xlsx 的 Supply 回填到一個 buyer 原檔。

    Returns dict:
        success: bool
        format: str or None          (偵測到的格式 FORMAT_* 或 'drift:FORMAT_*')
        format_label: str or None    (人類可讀標籤)
        n_partno_matched: int        (原檔中有對應 forecast 的 partno 列數)
        n_cells_written: int         (實際寫入儲存格數)
        n_skipped_merged: int        (跳過的合併儲存格數)
        n_skipped_formula: int       (跳過的公式儲存格數)
        message: str
        skip_reason: str or None     ('flat_no_supply'/'unknown_format'/'no_valid_sheet'/None)
    """
    result = {
        'success': False, 'format': None, 'format_label': None,
        'n_partno_matched': 0, 'n_cells_written': 0,
        'n_skipped_merged': 0, 'n_skipped_formula': 0,
        'message': '', 'skip_reason': None,
    }
    fname = os.path.basename(original_path)

    # 1. 偵測格式
    fmt = detect_format(original_path)
    is_drift = False
    if fmt is None:
        matched_fmt, score = match_known_format_fingerprint(original_path)
        if matched_fmt:
            fmt = matched_fmt
            is_drift = True
        else:
            result['skip_reason'] = 'unknown_format'
            result['message'] = f'{fname}: 無法識別格式 (非 15 格式亦非漂移版)'
            return result

    result['format'] = ('drift:' + fmt) if is_drift else fmt
    result['format_label'] = FORMAT_LABELS.get(fmt, fmt) + (' (漂移版)' if is_drift else '')

    # 2. 讀 forecast_result supply lookup
    try:
        lookup, fr_date_keys = _load_forecast_result_supply(forecast_result_path)
    except Exception as e:
        result['message'] = f'讀取 forecast_result.xlsx 失敗: {e}'
        return result

    # 3. 決定檔名 PLANT (無 PLANT 欄時的 fallback)
    #    不限 SINGLE_PLANT_FORMATS — 任何沒有 PLANT 欄的檔案都需要 fallback
    filename_plants = []  # 可能有多個 (如 PSW1+CEW1)
    filename_plant = None
    if plant_codes:
        match_target = file_label if file_label else original_path
        matched = match_plants_in_filename(match_target, plant_codes)
        if matched:
            filename_plants = matched
            filename_plant = matched[0]

    # 4. load workbook (保留樣式)
    try:
        wb = openpyxl.load_workbook(original_path)
    except Exception as e:
        result['message'] = f'載入原檔失敗: {e}'
        return result

    try:
        sheet_specs = find_valid_sheets(wb)
        if not sheet_specs:
            result['skip_reason'] = 'no_valid_sheet'
            result['message'] = f'{fname}: 找不到含 PARTNO 與日期欄的 sheet'
            return result

        any_supply_row_found = False

        for sheet_name, header_row in sheet_specs:
            ws = wb[sheet_name]
            header_values = _read_header_row(ws, header_row)
            found, headers = scan_headers(header_values)
            if 'partno' not in found:
                continue
            partno_col = found['partno']
            plant_col = found.get('plant')

            # 決定此 sheet 的 fallback plant:
            #   1. sheet 名本身是 plant code (如 PSW1, CEW1) → 用 sheet 名
            #   2. 否則用檔名比對的 filename_plant
            sheet_plant = filename_plant
            if not plant_col and plant_codes:
                sn_upper = sheet_name.strip().upper()
                for pc in plant_codes:
                    if pc.upper() == sn_upper:
                        sheet_plant = pc
                        break

            date_start = find_first_date_col(headers, header_values)
            date_col_map = collect_date_cols(date_start, header_values)
            if not date_col_map:
                continue

            # 抽前 50 列做 marker 偵測
            rows_sample = []
            for row_values in ws.iter_rows(min_row=header_row + 1, values_only=True):
                rows_sample.append(row_values)
                if len(rows_sample) >= 50:
                    break
            marker_col = find_marker_col(rows_sample)

            # 建 col → canonical key mapping (flat 與 multirow 都需要)
            col_to_canonical = _build_canonical_map(date_col_map, fr_date_keys)
            if not col_to_canonical:
                continue

            if marker_col is None:
                # Flat 格式: 插入 Supply + Balance 列
                _backfill_flat_sheet(ws, header_row, found, date_col_map,
                                    col_to_canonical, lookup, sheet_plant, result)
                any_supply_row_found = True
                continue

            any_supply_row_found = True

            # 合併儲存格 lookup
            merged_locked = _build_merged_cells_lookup(ws)

            # 遍歷 data rows
            last_partno = None
            last_plant = sheet_plant
            for r in range(header_row + 1, ws.max_row + 1):
                # 更新 partno
                pv = ws.cell(r, partno_col).value
                if pv is not None and str(pv).strip():
                    last_partno = str(pv).strip()
                    # 新 partno → 重設 plant 為 sheet plant (避免跨 partno 沿用)
                    last_plant = sheet_plant
                # 更新 plant (多 PLANT 格式)
                if plant_col is not None:
                    plv = ws.cell(r, plant_col).value
                    if plv is not None and str(plv).strip():
                        last_plant = str(plv).strip()

                if not last_partno:
                    continue

                # 是否為 Supply 列
                marker_v = ws.cell(r, marker_col).value
                cat = classify_marker(marker_v)
                if cat != 'supply':
                    continue

                # 找 lookup
                plant_key = (last_plant or '').strip().upper()
                lookup_key = (plant_key, last_partno)
                supply_dict = lookup.get(lookup_key)
                if not supply_dict:
                    continue

                result['n_partno_matched'] += 1

                # 寫入各日期欄
                for col, canonical in col_to_canonical.items():
                    val = supply_dict.get(canonical)
                    if val is None:
                        continue
                    # 跳過合併儲存格
                    if (r, col) in merged_locked:
                        result['n_skipped_merged'] += 1
                        continue
                    cell = ws.cell(r, col)
                    # 跳過公式
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        result['n_skipped_formula'] += 1
                        continue
                    # 寫入 (整數化若可能)
                    if isinstance(val, float) and val.is_integer():
                        cell.value = int(val)
                    else:
                        cell.value = val
                    result['n_cells_written'] += 1

        if not any_supply_row_found:
            result['skip_reason'] = 'flat_no_supply'
            result['message'] = f'{fname}: flat 格式無 Supply 欄, 未回填'
            return result

        # 5. 儲存
        wb.save(output_path)
        result['success'] = True
        result['message'] = (
            f'{fname}: {result["n_partno_matched"]} 個 partno 對應, '
            f'寫入 {result["n_cells_written"]} 個儲存格'
            + (f', 跳過 {result["n_skipped_merged"]} 合併儲存格'
               if result['n_skipped_merged'] else '')
            + (f', 跳過 {result["n_skipped_formula"]} 公式儲存格'
               if result['n_skipped_formula'] else '')
        )
        return result
    except Exception as e:
        tb = traceback.format_exc()
        result['message'] = f'{fname}: 回填過程例外: {e}\n{tb}'
        return result
    finally:
        wb.close()


# ============ Session 層級: 打包 ZIP ============

def backfill_session_to_zip(originals_dir, forecast_result_path, zip_path,
                            plant_codes=None):
    """讀 originals_dir 下所有 .xlsx + forecast_result_path → 產出 zip_path。

    Returns dict:
        success: bool
        zip_path: str
        manifest: list of per-file result dict
        n_files_total: int
        n_files_success: int
        n_files_skipped: int
        message: str
    """
    if not os.path.isdir(originals_dir):
        return {'success': False, 'message': f'找不到 originals 資料夾: {originals_dir}',
                'manifest': [], 'n_files_total': 0, 'n_files_success': 0, 'n_files_skipped': 0}
    if not os.path.exists(forecast_result_path):
        return {'success': False, 'message': f'找不到 forecast_result.xlsx: {forecast_result_path}',
                'manifest': [], 'n_files_total': 0, 'n_files_success': 0, 'n_files_skipped': 0}

    source_files = sorted([f for f in os.listdir(originals_dir)
                           if f.lower().endswith('.xlsx') and not f.startswith('~')])
    if not source_files:
        return {'success': False, 'message': 'originals 資料夾內無 .xlsx 檔案',
                'manifest': [], 'n_files_total': 0, 'n_files_success': 0, 'n_files_skipped': 0}

    # 暫存資料夾用於產出 per-file backfilled
    import tempfile
    tmp_dir = tempfile.mkdtemp(prefix='delta_backfill_')

    manifest = []
    try:
        for fname in source_files:
            src = os.path.join(originals_dir, fname)
            out_name = _make_backfilled_name(fname)
            out_path = os.path.join(tmp_dir, out_name)
            res = backfill_one_file(
                src, forecast_result_path, out_path,
                plant_codes=plant_codes, file_label=fname,
            )
            res['original_name'] = fname
            # 成功回填 → 用新檔名; 跳過 → 用原檔名 (直接放原檔到 ZIP)
            res['output_name'] = out_name if res['success'] else fname
            manifest.append(res)

        # 打包 ZIP: 所有原檔都進 ZIP,無論回填成功或跳過
        n_success = sum(1 for r in manifest if r['success'])
        n_skipped = sum(1 for r in manifest if r['skip_reason'])
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for r in manifest:
                if r['success']:
                    # 回填成功 → 用暫存檔
                    src_path = os.path.join(tmp_dir, r['output_name'])
                else:
                    # 跳過或失敗 → 直接放原檔 (保留客戶送來的格式)
                    src_path = os.path.join(originals_dir, r['original_name'])
                if os.path.exists(src_path):
                    zf.write(src_path, arcname=r['output_name'])
            # 寫入 README
            readme = _build_readme(manifest, forecast_result_path)
            zf.writestr('README.txt', readme)

        return {
            'success': True,
            'zip_path': zip_path,
            'manifest': manifest,
            'n_files_total': len(source_files),
            'n_files_success': n_success,
            'n_files_skipped': n_skipped,
            'message': f'已產出 ZIP: {n_success} 成功 / {n_skipped} 跳過 / {len(source_files)} 總數',
        }
    finally:
        import shutil
        shutil.rmtree(tmp_dir, ignore_errors=True)


def _make_backfilled_name(original_name):
    """原名 'Ketwadee0406.xlsx' → 'Ketwadee0406_backfilled.xlsx'"""
    base, ext = os.path.splitext(original_name)
    if not ext:
        ext = '.xlsx'
    return f'{base}_backfilled{ext}'


def _build_readme(manifest, forecast_result_path):
    """產出 README.txt 摘要報告"""
    lines = []
    lines.append('Delta Forecast 回填報告')
    lines.append('=' * 60)
    lines.append(f'產出時間: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    lines.append(f'來源: {os.path.basename(forecast_result_path)} (匯總結果)')
    lines.append(f'處理檔案: {len(manifest)} 個')
    lines.append('')

    ok_items = [r for r in manifest if r['success']]
    skip_items = [r for r in manifest if r['skip_reason']]
    fail_items = [r for r in manifest if not r['success'] and not r['skip_reason']]

    if ok_items:
        lines.append('成功回填:')
        lines.append('-' * 60)
        for r in ok_items:
            lines.append(
                f'  [OK] [{r["format_label"]}] {r["original_name"]}: '
                f'{r["n_partno_matched"]} 個 partno, 寫入 {r["n_cells_written"]} 個儲存格'
            )
            if r['n_skipped_merged']:
                lines.append(f'       (跳過 {r["n_skipped_merged"]} 合併儲存格)')
            if r['n_skipped_formula']:
                lines.append(f'       (跳過 {r["n_skipped_formula"]} 公式儲存格)')
        lines.append('')

    if skip_items:
        lines.append('跳過 (未回填, 原檔原封不動放入 ZIP):')
        lines.append('-' * 60)
        for r in skip_items:
            reason_zh = {
                'flat_no_supply': 'flat 格式無 Supply 欄, 原檔結構不支援回填',
                'unknown_format': '無法識別格式 (非 15 標準格式亦非漂移版)',
                'no_valid_sheet': '找不到含 PARTNO 與日期欄的 sheet',
            }.get(r['skip_reason'], r['skip_reason'])
            lines.append(f'  [SKIP] {r["original_name"]}: {reason_zh}')
        lines.append('')

    if fail_items:
        lines.append('失敗:')
        lines.append('-' * 60)
        for r in fail_items:
            lines.append(f'  [FAIL] {r["original_name"]}: {r["message"]}')
        lines.append('')

    lines.append('=' * 60)
    lines.append('說明:')
    lines.append('  - Supply 值已回填至原檔對應儲存格 (其他欄位保留原值)')
    lines.append('  - Balance 公式若存在會自動重算 (開啟檔案時)')
    lines.append('  - Flat 格式 (EIBG/IABG/NBQ1/SVC1PWC1 等) 原檔無 Supply 欄 → 自動插入 Supply + Balance 列 (Balance 含公式)')
    lines.append('  - 合併儲存格與公式儲存格會被跳過 (避免破壞原檔結構)')
    return '\n'.join(lines)
