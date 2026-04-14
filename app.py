from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash, make_response, session
from functools import wraps
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import datetime, timedelta
import os
import json
import warnings
import hashlib
import time
from dotenv import load_dotenv
warnings.filterwarnings('ignore')

# 載入環境變數
load_dotenv()

# 導入資料庫模組
from database import (
    get_db_connection, init_database, create_default_users,
    verify_user, log_activity, log_upload, log_process, get_user_by_id,
    get_customer_mappings, get_customer_mappings_raw, save_customer_mappings, save_customer_mappings_list, has_customer_mappings,
    # IT/Admin 管理介面函數
    get_upload_records, get_process_records, get_activity_logs_filtered,
    get_all_customer_mappings, get_users_with_company, get_all_users,
    # 用戶管理函數
    create_user, update_user, delete_user, update_activity_logs_enum,
    # 管理者客戶映射 CRUD 函數
    admin_create_customer_mapping, admin_update_customer_mapping,
    admin_delete_customer_mapping, admin_get_customer_mapping_by_id,
    # 規則管理 CRUD 函數
    get_all_processing_rules, get_processing_rules_by_category,
    get_processing_rule_by_id, update_processing_rule,
    create_processing_rule, delete_processing_rule, toggle_processing_rule_status,
    get_processing_rules_by_user
)

def normalize_date_for_mapping(date_value):
    """
    為 mapping 階段標準化日期格式
    處理各種日期格式：文字格式、pandas Timestamp、datetime 對象等
    """
    try:
        # 如果是空值或 NaN
        if pd.isna(date_value) or date_value is None:
            return None
        
        # 如果已經是 datetime 對象，轉換為標準字串格式
        if isinstance(date_value, (datetime, pd.Timestamp)):
            return date_value.strftime("%Y/%m/%d")
        
        # 如果是字串，嘗試解析並標準化
        if isinstance(date_value, str):
            date_str = str(date_value).strip()
            if not date_str or date_str.lower() in ['nan', 'none', '']:
                return None
            
            # 嘗試多種日期格式
            date_formats = [
                "%Y/%m/%d",      # 2025/10/01
                "%Y-%m-%d",      # 2025-10-01
                "%m/%d/%Y",      # 10/01/2025
                "%d/%m/%Y",      # 01/10/2025
                "%Y%m%d",        # 20251001
            ]
            
            for fmt in date_formats:
                try:
                    parsed_date = datetime.strptime(date_str, fmt)
                    return parsed_date.strftime("%Y/%m/%d")
                except ValueError:
                    continue
            
            # 如果所有格式都失敗，使用 pandas 自動解析
            try:
                parsed_date = pd.to_datetime(date_str)
                return parsed_date.strftime("%Y/%m/%d")
            except:
                print(f"    ⚠️ 無法解析日期格式: {date_str}")
                return None
        
        # 其他類型，嘗試用 pandas 轉換
        try:
            parsed_date = pd.to_datetime(date_value)
            return parsed_date.strftime("%Y/%m/%d")
        except:
            print(f"    ⚠️ 無法處理的日期類型: {type(date_value)} - {date_value}")
            return None
            
    except Exception as e:
        print(f"    ❌ 日期標準化失敗: {e}")
        return None

app = Flask(__name__)
app.secret_key = os.getenv('FLASK_SECRET_KEY', 'default_secret_key')
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)  # Session 有效期 8 小時

# ========================================
# 登入驗證裝飾器
# ========================================

def login_required(f):
    """登入驗證裝飾器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            # 如果是 API 請求，返回 JSON
            if request.path.startswith('/api/') or request.path.startswith('/upload') or request.path.startswith('/process') or request.path.startswith('/download'):
                return jsonify({'success': False, 'message': '請先登入', 'redirect': '/login'}), 401
            # 否則重定向到登入頁面
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    """管理者權限裝飾器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'success': False, 'message': '請先登入', 'redirect': '/login'}), 401
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            if request.path.startswith('/api/'):
                return jsonify({'success': False, 'message': '權限不足，僅限管理者'}), 403
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function


def it_or_admin_required(f):
    """IT 或管理者權限裝飾器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'success': False, 'message': '請先登入', 'redirect': '/login'}), 401
            return redirect(url_for('login'))
        if session.get('role') not in ['admin', 'it']:
            if request.path.startswith('/api/'):
                return jsonify({'success': False, 'message': '權限不足，僅限 IT 或管理者'}), 403
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def get_current_user():
    """獲取當前登入用戶"""
    if 'user_id' in session:
        return {
            'id': session.get('user_id'),
            'username': session.get('username'),
            'display_name': session.get('display_name'),
            'role': session.get('role'),
            'company': session.get('company')
        }
    return None

def get_client_ip():
    """獲取客戶端 IP"""
    if request.headers.get('X-Forwarded-For'):
        return request.headers.get('X-Forwarded-For').split(',')[0].strip()
    return request.remote_addr

# ========================================
# 文件格式驗證功能
# ========================================

# 範本文件路徑
COMPARE_FOLDER = 'compare'

def get_template_columns(template_type, username=None):
    """
    獲取範本文件的欄位結構
    template_type: 'erp', 'transit', 'forecast'
    username: 用戶名稱，用於指定客戶專屬模板目錄

    範本路徑優先順序：
    1. compare/{username}/{template_type}.xlsx（客戶專屬模板）
    2. compare/{template_type}.xlsx（通用模板，向後兼容）
    """
    try:
        # 優先使用客戶專屬模板
        if username:
            template_file = os.path.join(COMPARE_FOLDER, username, f'{template_type}.xlsx')
            if os.path.exists(template_file):
                print(f"📋 使用客戶專屬模板: {template_file}")
            else:
                # 客戶專屬模板不存在，回退到通用模板
                template_file = os.path.join(COMPARE_FOLDER, f'{template_type}.xlsx')
                print(f"⚠️ 客戶專屬模板不存在，使用通用模板: {template_file}")
        else:
            template_file = os.path.join(COMPARE_FOLDER, f'{template_type}.xlsx')

        if not os.path.exists(template_file):
            # 提供更清楚的錯誤訊息
            if username:
                expected_path = f'compare\\{username}\\{template_type}.xlsx'
            else:
                expected_path = f'compare\\{template_type}.xlsx'
            return None, f'範本文件不存在: {expected_path}'

        if template_type == 'forecast':
            # Forecast 使用 header=None 讀取原始結構
            # 光寶：只讀取 Daily+Weekly+Monthly sheet
            sheet_name = 'Daily+Weekly+Monthly' if username and username.lower() == 'liteon' else 0
            df = pd.read_excel(template_file, nrows=20, header=None, sheet_name=sheet_name)
            return df, None
        else:
            # ERP 和 Transit 讀取欄位名稱
            df = pd.read_excel(template_file, nrows=1)
            return list(df.columns), None
    except Exception as e:
        return None, f'讀取範本文件失敗: {str(e)}'

def find_column_by_name(df, patterns, required=True):
    """
    根據欄位名稱模式動態查找 DataFrame 欄位

    參數:
        df: pandas DataFrame
        patterns: str 或 list[str]，欄位名稱模式（多個模式為 AND 邏輯）
        required: bool，是否為必填欄位

    返回:
        tuple: (欄位名稱, 錯誤訊息)
    """
    if isinstance(patterns, str):
        patterns = [patterns]

    patterns_clean = [str(p).strip().replace('\n', '') for p in patterns]

    for col in df.columns:
        col_clean = str(col).strip().replace('\n', '')
        if all(pattern in col_clean for pattern in patterns_clean):
            return col, None

    if required:
        if len(patterns) == 1:
            return None, f"找不到欄位：「{patterns[0]}」"
        else:
            return None, f"找不到包含「{'」和「'.join(patterns)}」的欄位"
    return None, None

def validate_erp_format(uploaded_file_path, username=None):
    """
    驗證 ERP 文件格式
    只檢查必要欄位名稱是否存在（不卡控欄位數量和順序）
    username: 用戶名稱，用於指定客戶專屬模板目錄
    """
    try:
        # 獲取範本欄位（根據用戶名稱取得對應模板）
        template_columns, error = get_template_columns('erp', username)
        if error:
            return False, error, []

        # 讀取上傳的文件
        uploaded_df = pd.read_excel(uploaded_file_path, nrows=1)
        uploaded_columns = list(uploaded_df.columns)

        # 標準化欄位名稱（去除空白和換行符）
        template_columns_clean = [str(col).strip().replace('\n', '') for col in template_columns]
        uploaded_columns_clean = [str(col).strip().replace('\n', '') for col in uploaded_columns]

        # 檢查必要欄位是否存在（不卡控順序和數量）
        missing_columns = []
        for template_col in template_columns_clean:
            if template_col not in uploaded_columns_clean:
                missing_columns.append(template_col)

        if missing_columns:
            error_details = []
            for col in missing_columns[:5]:  # 最多顯示5個錯誤
                error_details.append(f"缺少欄位：「{col}」")

            if len(missing_columns) > 5:
                error_details.append(f"...還有 {len(missing_columns)-5} 個欄位缺少")

            return False, '缺少必要欄位', error_details

        return True, 'ERP 文件格式驗證通過', []

    except Exception as e:
        return False, f'驗證過程發生錯誤: {str(e)}', []

def validate_transit_format(uploaded_file_path, username=None):
    """
    驗證在途文件格式
    只檢查必要欄位名稱是否存在（不卡控欄位數量和順序）
    username: 用戶名稱，用於指定客戶專屬模板目錄
    """
    try:
        # 獲取範本欄位（根據用戶名稱取得對應模板）
        template_columns, error = get_template_columns('transit', username)
        if error:
            return False, error, []

        # 讀取上傳的文件
        uploaded_df = pd.read_excel(uploaded_file_path, nrows=1)
        uploaded_columns = list(uploaded_df.columns)

        # 標準化欄位名稱（去除空白和換行符）
        template_columns_clean = [str(col).strip().replace('\n', '') for col in template_columns]
        uploaded_columns_clean = [str(col).strip().replace('\n', '') for col in uploaded_columns]

        # 檢查必要欄位是否存在（不卡控順序和數量）
        missing_columns = []
        for template_col in template_columns_clean:
            if template_col not in uploaded_columns_clean:
                missing_columns.append(template_col)

        if missing_columns:
            error_details = []
            for col in missing_columns[:5]:  # 最多顯示5個錯誤
                error_details.append(f"缺少欄位：「{col}」")

            if len(missing_columns) > 5:
                error_details.append(f"...還有 {len(missing_columns)-5} 個欄位缺少")

            return False, '缺少必要欄位', error_details

        return True, '在途文件格式驗證通過', []

    except Exception as e:
        return False, f'驗證過程發生錯誤: {str(e)}', []

def validate_forecast_format(uploaded_file_path, username=None):
    """
    驗證 Forecast 文件格式
    只檢查欄位數量是否與範本一致（不比對具體數據內容，因為資料會變動）
    光寶：只驗證 Daily+Weekly+Monthly sheet 是否存在且有資料（欄位數會隨日期變動）
    username: 用戶名稱，用於指定客戶專屬模板目錄
    """
    try:
        # 光寶：驗證目標 sheet 是否存在，並比對固定標題欄位
        if username and username.lower() == 'liteon':
            target_sheet = 'Daily+Weekly+Monthly'
            try:
                uploaded_df = pd.read_excel(uploaded_file_path, nrows=3, header=None, sheet_name=target_sheet)
            except ValueError:
                return False, f'找不到工作表「{target_sheet}」', []

            # 比對灰色區塊的固定標題（位置固定不變）
            expected_labels = {
                (0, 1): 'Plant:',
                (0, 3): 'Buyer Code:',
                (0, 5): 'Released Time:',
                (1, 1): 'ERP Vendor Code:',
                (1, 3): 'Report Date:',
                (1, 5): 'VDS Num:',
                (2, 1): 'Vendor Name:',
                (2, 5): 'Version:',
            }
            missing_labels = []
            for (row, col), expected in expected_labels.items():
                if row < len(uploaded_df) and col < len(uploaded_df.columns):
                    actual = str(uploaded_df.iloc[row, col]).strip() if pd.notna(uploaded_df.iloc[row, col]) else ''
                    if actual != expected:
                        missing_labels.append(f'({row+1},{col+1}) 預期「{expected}」，實際「{actual}」')
                else:
                    missing_labels.append(f'({row+1},{col+1}) 缺少「{expected}」')

            if missing_labels:
                return False, '標題欄位不符', missing_labels[:5]

            return True, 'Forecast 文件格式驗證通過', []

        # 其他客戶：比對欄位數量
        # 獲取範本結構（根據用戶名稱取得對應模板）
        template_df, error = get_template_columns('forecast', username)
        if error:
            return False, error, []

        # 讀取上傳的文件（不使用 header）
        uploaded_df = pd.read_excel(uploaded_file_path, nrows=20, header=None)

        # 只檢查欄位數量是否一致（不比對資料內容，因為資料會變動）
        if len(uploaded_df.columns) != len(template_df.columns):
            return False, f'欄位數量不符：預期 {len(template_df.columns)} 個欄位，實際 {len(uploaded_df.columns)} 個欄位', []

        # 檢查列數是否足夠（至少要與範本列數一致）
        min_rows = len(template_df)
        if len(uploaded_df) < min_rows:
            return False, f'資料列數不足：至少需要 {min_rows} 列，實際只有 {len(uploaded_df)} 列', []

        return True, 'Forecast 文件格式驗證通過', []

    except Exception as e:
        return False, f'驗證過程發生錯誤: {str(e)}', []


def check_transit_requirements_from_forecast(forecast_file_paths, user_id, username):
    """
    檢查 Forecast 文件中的 Plant 欄位（F欄，索引5）是否需要在途文件
    方案 C：只提醒不卡控

    Pegatron 專用邏輯：
    - 讀取 Forecast 的 F 欄位（Plant）
    - 與 customer_mappings 的 region 欄位比對
    - 根據 requires_transit 欄位決定是否需要提醒

    參數:
        forecast_file_paths: 單個檔案路徑(str)或多個檔案路徑列表(list)
        user_id: 用戶 ID
        username: 用戶名稱

    返回: {
        'has_transit_requirement': bool,  # 是否有需要在途的項目
        'transit_required_regions': [],   # 需要在途的 region 列表
        'transit_not_required_regions': [], # 不需要在途的 region 列表
        'unmapped_regions': [],           # 沒有映射的 region 列表
        'message': str                    # 提醒訊息
    }
    """
    result = {
        'has_transit_requirement': False,
        'transit_required_regions': [],
        'transit_not_required_regions': [],
        'unmapped_regions': [],
        'message': ''
    }

    # 只針對 Pegatron 用戶進行檢查
    if username.lower() != 'pegatron':
        return result

    # 統一處理：將單個路徑轉為列表
    if isinstance(forecast_file_paths, str):
        file_paths = [forecast_file_paths]
    else:
        file_paths = forecast_file_paths

    try:
        # 收集所有檔案中的 Plant 值
        all_plant_values = set()

        for forecast_file_path in file_paths:
            # 讀取 Forecast 文件
            forecast_ext = os.path.splitext(forecast_file_path)[1].lower()
            if forecast_ext == '.xls':
                forecast_df = pd.read_excel(forecast_file_path, header=None, engine='xlrd')
            else:
                forecast_df = pd.read_excel(forecast_file_path, header=None)

            # 獲取 F 欄位（索引 5）的所有唯一值（跳過標題行）
            # Pegatron Forecast 的 F 欄位是 Plant（如：3A32）
            if len(forecast_df.columns) <= 5:
                print(f"⚠️ Forecast 文件 {os.path.basename(forecast_file_path)} 欄位數不足，跳過")
                continue

            # 從第 3 行開始讀取（跳過標題行，通常前 2 行是標題）
            plant_values = forecast_df.iloc[2:, 5].dropna().unique()
            plant_values = [str(v).strip() for v in plant_values if str(v).strip()]

            if plant_values:
                all_plant_values.update(plant_values)
                print(f"📋 {os.path.basename(forecast_file_path)} 中的 Plant 值: {plant_values}")

        plant_values = list(all_plant_values)

        if not plant_values:
            print(f"⚠️ 所有 Forecast 文件 F 欄位都沒有有效數據")
            return result

        print(f"📋 所有 Forecast 文件中的 Plant 值（合併後）: {plant_values}")

        # 獲取用戶的 mapping 資料
        from database import get_customer_mapping_list
        mappings = get_customer_mapping_list(user_id)

        if not mappings:
            # 沒有映射資料，所有 Plant 都視為未映射
            result['unmapped_regions'] = list(plant_values)
            result['message'] = f'提醒：發現 {len(plant_values)} 個廠區尚未配置映射資料'
            return result

        # 建立 region -> requires_transit 的對照表
        region_transit_map = {}
        for mapping in mappings:
            region = mapping.get('region', '')
            requires_transit = mapping.get('requires_transit', True)
            # 如果是 0 或 False 則不需要在途
            if requires_transit == 0 or requires_transit is False:
                requires_transit = False
            else:
                requires_transit = True
            region_transit_map[region] = requires_transit

        # 檢查每個 Plant 值
        for plant in plant_values:
            if plant in region_transit_map:
                if region_transit_map[plant]:
                    result['transit_required_regions'].append(plant)
                else:
                    result['transit_not_required_regions'].append(plant)
            else:
                result['unmapped_regions'].append(plant)

        # 判斷是否需要在途
        result['has_transit_requirement'] = len(result['transit_required_regions']) > 0

        # 組合提醒訊息
        messages = []
        if result['transit_required_regions']:
            regions_str = '、'.join(result['transit_required_regions'])
            messages.append(f"廠區 [{regions_str}] 需要上傳在途文件")
        if result['transit_not_required_regions']:
            regions_str = '、'.join(result['transit_not_required_regions'])
            messages.append(f"廠區 [{regions_str}] 不需要上傳在途文件")
        if result['unmapped_regions']:
            regions_str = '、'.join(result['unmapped_regions'])
            messages.append(f"廠區 [{regions_str}] 尚未配置映射")

        result['message'] = '；'.join(messages)

        print(f"🔍 在途檢查結果: 需要={result['transit_required_regions']}, 不需要={result['transit_not_required_regions']}, 未映射={result['unmapped_regions']}")

        return result

    except Exception as e:
        print(f"❌ 檢查在途需求時發生錯誤: {e}")
        return result

# ========================================

# 生成版本號用於防止快取
def get_version():
    """生成基於時間戳的版本號"""
    return str(int(time.time()))

# 添加版本控制到模板上下文
@app.context_processor
def inject_version():
    return dict(version=get_version())

# 全局變量存儲上傳的文件
UPLOAD_FOLDER = 'uploads'
PROCESSED_FOLDER = 'processed'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
FILE_RETENTION_DAYS = 30  # 檔案保留天數

# 確保文件夾存在
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)
os.makedirs('templates', exist_ok=True)
os.makedirs('static/css', exist_ok=True)
os.makedirs('static/js', exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_file_extension(filename):
    """
    獲取檔案的副檔名（小寫）
    返回: '.xls' 或 '.xlsx'
    """
    if '.' in filename:
        ext = '.' + filename.rsplit('.', 1)[1].lower()
        if ext in ['.xls', '.xlsx']:
            return ext
    return '.xlsx'  # 預設

def find_file_with_extensions(folder, base_name):
    """
    查找檔案，支持 .xls 和 .xlsx 兩種副檔名
    例如: find_file_with_extensions('/path', 'forecast_data')
    會嘗試尋找 forecast_data.xlsx 和 forecast_data.xls

    返回: 找到的檔案完整路徑，或 None
    """
    for ext in ['.xlsx', '.xls']:
        filepath = os.path.join(folder, base_name + ext)
        if os.path.exists(filepath):
            return filepath
    return None

def extract_plant_mrp_from_forecast(forecast_file):
    """
    從 Forecast 檔案中提取 Plant 和 MRP ID
    F 欄 (col index 5): Plant (例如 3A32)
    G 欄 (col index 6): MRP ID (例如 A00Y)

    返回: (plant, mrp_id) 或 (None, None) 如果無法提取
    """
    try:
        # 根據檔案格式選擇引擎
        ext = os.path.splitext(forecast_file)[1].lower()
        if ext == '.xls':
            df = pd.read_excel(forecast_file, header=None, engine='xlrd', nrows=20)
        else:
            df = pd.read_excel(forecast_file, header=None, nrows=20)

        # 尋找 WEEK# 行來定位資料
        for row_idx in range(min(15, len(df))):
            m_val = df.iloc[row_idx, 12] if row_idx < len(df) and 12 < len(df.columns) and pd.notna(df.iloc[row_idx, 12]) else ''
            if m_val == 'WEEK#':
                f_val = str(df.iloc[row_idx, 5]).strip() if pd.notna(df.iloc[row_idx, 5]) else ''
                g_val = str(df.iloc[row_idx, 6]).strip() if pd.notna(df.iloc[row_idx, 6]) else ''
                if f_val and g_val:
                    print(f"  📋 提取 Plant={f_val}, MRP_ID={g_val} from {os.path.basename(forecast_file)}")
                    return f_val, g_val

        return None, None
    except Exception as e:
        print(f"  ⚠️ 無法從 {forecast_file} 提取 Plant/MRP ID: {e}")
        return None, None

def is_xls_format(file_path):
    """
    檢查檔案是否為舊版 .xls 格式
    返回: True 表示是 .xls 格式，False 表示是 .xlsx 格式
    """
    import zipfile
    try:
        with zipfile.ZipFile(file_path, 'r') as z:
            return False  # 是有效的 xlsx（zip 格式）
    except zipfile.BadZipFile:
        return True  # 不是 zip 格式，是舊版 .xls

def cleanup_xls_file(file_path, output_path, username):
    """
    清理 .xls 格式的檔案（使用 LibreOffice 跨平台方案）
    完整保留格式、公式，只修改指定儲存格的值為 0
    支援 Windows 和 Linux 環境
    """
    from libreoffice_utils import cleanup_xls_file_libreoffice
    return cleanup_xls_file_libreoffice(file_path, output_path, username)

def get_or_create_session_folder(user_id, folder_type='uploads', upload_session_id=None):
    """
    獲取或建立用戶的 session 資料夾
    資料夾結構: {folder_type}/{user_id}/{YYYYMMDD_HHMMSS}/

    優先使用前端傳來的 upload_session_id，確保同一批上傳的檔案都在同一個資料夾。
    這樣不再依賴 Flask session cookie，避免 cookie 大小限制或跨請求丟失的問題。

    參數:
        user_id: 用戶 ID
        folder_type: 'uploads' 或 'processed'
        upload_session_id: 前端傳來的上傳 session ID（優先使用）
    """
    # 優先使用前端傳來的 upload_session_id
    if upload_session_id:
        session_timestamp = upload_session_id
        # 同步更新到 Flask session（供後續處理步驟使用）
        session['current_session_timestamp'] = session_timestamp
        session.modified = True
        print(f"📁 使用前端傳來的 session ID: {session_timestamp}")
    else:
        # 如果沒有前端傳來的 ID，嘗試從 Flask session 獲取
        session_key = 'current_session_timestamp'
        session_timestamp = session.get(session_key)

        if not session_timestamp:
            # 建立新的 session 時間戳（後備方案）
            session_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            session[session_key] = session_timestamp
            session.modified = True
            print(f"📁 建立新的 session 時間戳（後備）: {session_timestamp}")
        else:
            print(f"📁 使用 Flask session 時間戳: {session_timestamp}")

    print(f"🔍 get_or_create_session_folder: user_id={user_id}, folder_type={folder_type}, session_id={session_timestamp}")

    # 建立資料夾路徑
    if folder_type == 'uploads':
        folder_path = os.path.join(UPLOAD_FOLDER, str(user_id), session_timestamp)
    else:
        folder_path = os.path.join(PROCESSED_FOLDER, str(user_id), session_timestamp)

    # 確保資料夾存在
    os.makedirs(folder_path, exist_ok=True)

    print(f"   folder_path: {folder_path}")

    return folder_path, session_timestamp

def reset_session_folder():
    """
    重置 session 資料夾時間戳
    當用戶進入上傳頁面或重新開始流程時調用此函數
    確保每次新的工作流程使用新的資料夾
    """
    old_timestamp = session.get('current_session_timestamp')
    keys_to_clear = [
        'current_session_timestamp',
        # 舊格式（完整路徑）- 保留向下相容
        'current_erp_file',
        'current_forecast_file',
        'current_forecast_files',
        'current_transit_file',
        'current_processed_folder',
        # 新格式（上傳標記）
        'uploaded_erp',
        'uploaded_forecast',
        'uploaded_transit',
        'forecast_merge_mode',
        'forecast_file_count'
    ]

    cleared = []
    for key in keys_to_clear:
        if key in session:
            session.pop(key)
            cleared.append(key)

    if cleared:
        session.modified = True
        print(f"🔄 已重置 session (舊 timestamp: {old_timestamp})")
        print(f"   清除的 keys: {cleared}")
        print(f"   session size after reset: {len(str(dict(session)))} bytes")

def get_session_folder_path(user_id, folder_type='uploads'):
    """
    獲取當前 session 的資料夾路徑（不建立新的）
    """
    session_key = 'current_session_timestamp'
    session_timestamp = session.get(session_key)

    if not session_timestamp:
        return None, None

    if folder_type == 'uploads':
        folder_path = os.path.join(UPLOAD_FOLDER, str(user_id), session_timestamp)
    else:
        folder_path = os.path.join(PROCESSED_FOLDER, str(user_id), session_timestamp)

    return folder_path, session_timestamp

def cleanup_old_folders():
    """清理超過保留期限的資料夾"""
    import shutil
    try:
        cutoff_date = datetime.now() - timedelta(days=FILE_RETENTION_DAYS)
        cleaned_count = 0

        for base_folder in [UPLOAD_FOLDER, PROCESSED_FOLDER]:
            if not os.path.exists(base_folder):
                continue

            # 遍歷用戶資料夾
            for user_folder in os.listdir(base_folder):
                user_folder_path = os.path.join(base_folder, user_folder)
                if not os.path.isdir(user_folder_path):
                    # 處理舊格式的檔案（直接在 uploads/processed 下的檔案）
                    if os.path.isfile(user_folder_path):
                        file_mtime = datetime.fromtimestamp(os.path.getmtime(user_folder_path))
                        if file_mtime < cutoff_date:
                            os.remove(user_folder_path)
                            cleaned_count += 1
                            print(f"🗑️ 已清理過期檔案: {user_folder}")
                    continue

                # 遍歷 session 時間戳資料夾
                for session_folder in os.listdir(user_folder_path):
                    session_folder_path = os.path.join(user_folder_path, session_folder)
                    if not os.path.isdir(session_folder_path):
                        continue

                    # 嘗試解析資料夾名稱中的時間戳
                    try:
                        folder_date = datetime.strptime(session_folder, '%Y%m%d_%H%M%S')
                        if folder_date < cutoff_date:
                            shutil.rmtree(session_folder_path)
                            cleaned_count += 1
                            print(f"🗑️ 已清理過期資料夾: {base_folder}/{user_folder}/{session_folder}")
                    except ValueError:
                        # 無法解析時間戳，使用資料夾修改時間
                        folder_mtime = datetime.fromtimestamp(os.path.getmtime(session_folder_path))
                        if folder_mtime < cutoff_date:
                            shutil.rmtree(session_folder_path)
                            cleaned_count += 1
                            print(f"🗑️ 已清理過期資料夾: {base_folder}/{user_folder}/{session_folder}")

                # 如果用戶資料夾為空，也刪除
                if os.path.exists(user_folder_path) and not os.listdir(user_folder_path):
                    os.rmdir(user_folder_path)
                    print(f"🗑️ 已清理空用戶資料夾: {base_folder}/{user_folder}")

        if cleaned_count > 0:
            print(f"✅ 共清理 {cleaned_count} 個過期資料夾/檔案")
        return cleaned_count
    except Exception as e:
        print(f"❌ 清理過期資料夾失敗: {e}")
        return 0

def get_user_file_path(file_type):
    """
    從 session 動態計算當前用戶的檔案路徑
    不再直接存儲完整路徑，而是根據 session timestamp 和檔案類型來計算
    這樣可以避免 session cookie 超過 4KB 限制

    注意：由於 Flask session cookie 可能遺失，此函數會直接檢查檔案是否存在，
    而不再依賴 session 中的 uploaded_* 標記。
    """
    user = get_current_user()
    if not user:
        print(f"⚠️ get_user_file_path({file_type}): 用戶未登入")
        return None

    session_timestamp = session.get('current_session_timestamp')
    if not session_timestamp:
        print(f"⚠️ get_user_file_path({file_type}): session_timestamp 不存在")
        return None

    print(f"🔍 get_user_file_path({file_type}): session_timestamp={session_timestamp}, user_id={user['id']}")

    # 不再依賴 uploaded_* 標記（因為 session cookie 可能遺失）
    # 直接根據 session_timestamp 計算路徑並檢查檔案是否存在

    # 根據檔案類型決定可能的基礎檔名（不含副檔名）
    possible_base_names = []
    if file_type == 'erp':
        possible_base_names = ['erp_data']
    elif file_type == 'forecast':
        # 檢查是否為合併模式
        if session.get('forecast_merge_mode', True):
            possible_base_names = ['forecast_data', 'forecast_data_1']
        else:
            # 非合併模式，返回第一個檔案
            possible_base_names = ['forecast_data_1', 'forecast_data']
    elif file_type == 'transit':
        possible_base_names = ['transit_data']
    else:
        return None

    # 嘗試每個可能的基礎檔名（支援 .xlsx 和 .xls）
    upload_folder = os.path.join(UPLOAD_FOLDER, str(user['id']), session_timestamp)
    for base_name in possible_base_names:
        filepath = find_file_with_extensions(upload_folder, base_name)
        if filepath:
            print(f"✅ get_user_file_path({file_type}): 找到檔案 {filepath}")
            return filepath

    print(f"❌ get_user_file_path({file_type}): 在 {upload_folder} 中找不到檔案")
    return None

def set_user_file_path(file_type, filepath):
    """
    設置檔案上傳標記到 session（不存儲完整路徑）
    只存儲上傳標記，路徑由 get_user_file_path 動態計算
    """
    # 設置上傳標記
    uploaded_key = f'uploaded_{file_type}'
    session[uploaded_key] = True
    session.modified = True
    print(f"📝 set_user_file_path: {uploaded_key} = True (filepath: {filepath})")
    print(f"   session keys after set: {list(session.keys())}")
    print(f"   session size estimate: {len(str(dict(session)))} bytes")

def get_forecast_files():
    """
    獲取 Forecast 檔案列表（支援多檔案模式）
    根據 session timestamp 動態計算檔案路徑

    注意：由於 Flask session cookie 可能遺失，此函數會直接掃描資料夾，
    而不再依賴 session 中的 uploaded_forecast 標記。
    """
    user = get_current_user()
    if not user:
        return []

    session_timestamp = session.get('current_session_timestamp')
    if not session_timestamp:
        return []

    upload_folder = os.path.join(UPLOAD_FOLDER, str(user['id']), session_timestamp)
    print(f"🔍 get_forecast_files: 掃描資料夾 {upload_folder}")

    # 不再依賴 uploaded_forecast 標記，直接掃描資料夾
    if not os.path.exists(upload_folder):
        print(f"❌ get_forecast_files: 資料夾不存在")
        return []

    # 先檢查合併模式的單一檔案（支援 .xlsx 和 .xls）
    single_file = find_file_with_extensions(upload_folder, 'forecast_data')
    if single_file:
        print(f"✅ get_forecast_files: 找到合併檔案 {single_file}")
        return [single_file]

    # 再檢查多檔案模式（forecast_data_1.xlsx/.xls, forecast_data_2.xlsx/.xls, ...）
    files = []
    for i in range(1, 100):  # 最多支援 99 個檔案
        filepath = find_file_with_extensions(upload_folder, f'forecast_data_{i}')
        if filepath:
            files.append(filepath)
        else:
            break  # 遇到不存在的編號就停止

    if files:
        print(f"✅ get_forecast_files: 找到 {len(files)} 個多檔案")
        return files

    print(f"❌ get_forecast_files: 在資料夾中找不到 Forecast 檔案")
    return []

def get_user_processed_folder():
    """獲取當前用戶的 processed 資料夾路徑"""
    user = get_current_user()
    if not user:
        return PROCESSED_FOLDER

    folder_path, _ = get_session_folder_path(user['id'], 'processed')
    if folder_path and os.path.exists(folder_path):
        return folder_path

    # 如果還沒有 session 資料夾，建立一個
    folder_path, _ = get_or_create_session_folder(user['id'], 'processed')
    return folder_path

# ========================================
# 登入/登出路由
# ========================================

@app.route('/login')
def login():
    """登入頁面"""
    # 如果已登入，跳轉到首頁
    if 'user_id' in session:
        return redirect(url_for('index'))
    return render_template('login.html')

@app.route('/api/login', methods=['POST'])
def api_login():
    """登入 API"""
    try:
        data = request.get_json()
        username = data.get('username', '').strip()
        password = data.get('password', '')

        if not username or not password:
            return jsonify({'success': False, 'message': '請輸入帳號和密碼'})

        # 驗證用戶
        user = verify_user(username, password)

        if user:
            # 設置 Session
            session.permanent = True
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['display_name'] = user['display_name']
            session['role'] = user['role']
            session['company'] = user['company']

            # 記錄登入活動
            log_activity(
                user_id=user['id'],
                username=user['username'],
                action_type='login',
                action_detail=f"用戶 {user['display_name']} 登入成功",
                ip_address=get_client_ip(),
                user_agent=request.headers.get('User-Agent')
            )

            print(f"✅ 用戶登入成功: {user['username']} ({user['display_name']})")

            # 根據角色決定跳轉頁面
            redirect_url = '/'
            if user['role'] == 'admin':
                redirect_url = '/admin'
            elif user['role'] == 'it':
                redirect_url = '/it'

            return jsonify({
                'success': True,
                'message': '登入成功',
                'redirect_url': redirect_url,
                'user': {
                    'username': user['username'],
                    'display_name': user['display_name'],
                    'role': user['role'],
                    'company': user['company']
                }
            })
        else:
            # 記錄登入失敗
            log_activity(
                user_id=None,
                username=username,
                action_type='login_failed',
                action_detail=f"登入失敗：帳號 {username}",
                ip_address=get_client_ip(),
                user_agent=request.headers.get('User-Agent')
            )

            print(f"❌ 登入失敗: {username}")
            return jsonify({'success': False, 'message': '帳號或密碼錯誤'})

    except Exception as e:
        print(f"❌ 登入處理錯誤: {e}")
        return jsonify({'success': False, 'message': '系統錯誤，請稍後再試'})

@app.route('/api/logout', methods=['POST'])
@login_required
def api_logout():
    """登出 API"""
    try:
        user = get_current_user()
        if user:
            # 記錄登出活動
            log_activity(
                user_id=user['id'],
                username=user['username'],
                action_type='logout',
                action_detail=f"用戶 {user['display_name']} 登出",
                ip_address=get_client_ip(),
                user_agent=request.headers.get('User-Agent')
            )
            print(f"✅ 用戶登出: {user['username']}")

        # 清除 Session
        session.clear()
        return jsonify({'success': True, 'message': '已登出'})

    except Exception as e:
        print(f"❌ 登出處理錯誤: {e}")
        session.clear()
        return jsonify({'success': True, 'message': '已登出'})

@app.route('/logout')
def logout():
    """登出頁面"""
    user = get_current_user()
    if user:
        log_activity(
            user_id=user['id'],
            username=user['username'],
            action_type='logout',
            action_detail=f"用戶 {user['display_name']} 登出",
            ip_address=get_client_ip(),
            user_agent=request.headers.get('User-Agent')
        )
    session.clear()
    return redirect(url_for('login'))

@app.route('/api/user')
@login_required
def api_get_user():
    """獲取當前用戶資訊"""
    user = get_current_user()
    if user:
        return jsonify({'success': True, 'user': user})
    return jsonify({'success': False, 'message': '未登入'})

@app.route('/api/reset_session', methods=['POST'])
@login_required
def api_reset_session():
    """重置當前的 session 資料夾時間戳（當用戶重新開始流程時調用）"""
    try:
        reset_session_folder()
        return jsonify({'success': True, 'message': 'Session 已重置'})
    except Exception as e:
        print(f"❌ 重置 session 失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/delete_upload', methods=['POST'])
@login_required
def api_delete_upload():
    """刪除已上傳的檔案（當用戶想重新上傳時調用）"""
    user = get_current_user()
    try:
        data = request.get_json()
        file_type = data.get('type')  # 'erp', 'forecast', 'transit'
        upload_session_id = data.get('upload_session_id')
        filename = data.get('filename')  # 可選：指定單一檔案名稱

        if not file_type:
            return jsonify({'success': False, 'message': '未指定檔案類型'})

        if not upload_session_id:
            return jsonify({'success': False, 'message': '未指定 session ID'})

        # 建立資料夾路徑
        upload_folder = os.path.join(UPLOAD_FOLDER, str(user['id']), upload_session_id)

        if not os.path.exists(upload_folder):
            return jsonify({'success': True, 'message': '資料夾不存在，無需刪除'})

        # 根據類型決定要刪除的檔案
        deleted_files = []
        remaining_files = []  # 剩餘的檔案（用於 forecast 單檔刪除）

        if file_type == 'erp':
            # 刪除 erp_data.xlsx 或 erp_data.xls
            for ext in ['.xlsx', '.xls']:
                filepath = os.path.join(upload_folder, f'erp_data{ext}')
                if os.path.exists(filepath):
                    os.remove(filepath)
                    deleted_files.append(f'erp_data{ext}')
                    print(f"🗑️ 已刪除 ERP 檔案: {filepath}")

        elif file_type == 'forecast':
            transit_check = None  # 用於儲存重新檢查的在途需求

            if filename:
                # 單檔刪除模式：只刪除指定的檔案
                filepath = os.path.join(upload_folder, filename)
                if os.path.exists(filepath):
                    os.remove(filepath)
                    deleted_files.append(filename)
                    print(f"🗑️ 已刪除單一 Forecast 檔案: {filepath}")

                # 計算剩餘的 forecast 檔案
                for f in os.listdir(upload_folder):
                    if f.startswith('forecast_data'):
                        remaining_files.append(f)

                # 如果還有剩餘檔案，重新檢查在途需求
                if remaining_files:
                    remaining_paths = [os.path.join(upload_folder, f) for f in remaining_files]
                    transit_check = check_transit_requirements_from_forecast(remaining_paths, user['id'], user['username'])
            else:
                # 全部刪除模式：刪除所有 forecast_data 開頭的檔案
                for f in os.listdir(upload_folder):
                    if f.startswith('forecast_data'):
                        filepath = os.path.join(upload_folder, f)
                        os.remove(filepath)
                        deleted_files.append(f)
                        print(f"🗑️ 已刪除 Forecast 檔案: {filepath}")

        elif file_type == 'transit':
            # 刪除 transit_data.xlsx 或 transit_data.xls
            for ext in ['.xlsx', '.xls']:
                filepath = os.path.join(upload_folder, f'transit_data{ext}')
                if os.path.exists(filepath):
                    os.remove(filepath)
                    deleted_files.append(f'transit_data{ext}')
                    print(f"🗑️ 已刪除在途檔案: {filepath}")

        else:
            return jsonify({'success': False, 'message': f'不支援的檔案類型: {file_type}'})

        response_data = {
            'success': True,
            'message': f'已刪除 {len(deleted_files)} 個檔案',
            'deleted_files': deleted_files,
            'remaining_files': remaining_files,
            'remaining_count': len(remaining_files)
        }

        # 如果有重新檢查在途需求，加入回應
        if file_type == 'forecast' and 'transit_check' in dir() and transit_check:
            response_data['transit_check'] = transit_check

        return jsonify(response_data)

    except Exception as e:
        print(f"❌ 刪除上傳檔案失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/migrate_mapping', methods=['POST'])
@login_required
def api_migrate_mapping():
    """
    將現有的 mapping_data.json 遷移到資料庫
    僅限管理員或 IT 人員使用
    """
    user = get_current_user()

    # 檢查權限（僅管理員和 IT 可執行）
    if user['role'] not in ['admin', 'it']:
        return jsonify({'success': False, 'message': '權限不足，僅管理員和 IT 人員可執行此操作'})

    try:
        mapping_file = os.path.join('mapping', 'mapping_data.json')

        if not os.path.exists(mapping_file):
            return jsonify({'success': False, 'message': 'mapping_data.json 檔案不存在'})

        # 讀取 JSON 檔案
        with open(mapping_file, 'r', encoding='utf-8') as f:
            mapping_data = json.load(f)

        # 儲存到當前用戶的資料庫
        if save_customer_mappings(user['id'], mapping_data):
            print(f"✅ 已將 mapping_data.json 遷移至資料庫 (user: {user['username']})")

            # 記錄活動
            log_activity(user['id'], user['username'], 'mapping_start',
                       f"遷移 mapping_data.json 至資料庫", get_client_ip(), request.headers.get('User-Agent'))

            return jsonify({
                'success': True,
                'message': f'成功將 mapping 資料遷移至資料庫（用戶：{user["display_name"]}）',
                'customers_count': len(set(
                    list(mapping_data.get('regions', {}).keys()) +
                    list(mapping_data.get('schedule_breakpoints', {}).keys()) +
                    list(mapping_data.get('etd', {}).keys()) +
                    list(mapping_data.get('eta', {}).keys())
                ))
            })
        else:
            return jsonify({'success': False, 'message': '儲存至資料庫失敗'})

    except Exception as e:
        print(f"❌ 遷移 mapping 失敗: {e}")
        return jsonify({'success': False, 'message': f'遷移失敗: {str(e)}'})

# ========================================
# 主要頁面路由
# ========================================

@app.route('/')
@login_required
def index():
    user = get_current_user()
    response = make_response(render_template('index.html', user=user))
    # 設置不緩存HTML頁面
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/upload_erp', methods=['POST'])
@login_required
def upload_erp():
    user = get_current_user()
    original_filename = ''
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '沒有選擇文件'})

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '沒有選擇文件'})

        original_filename = file.filename

        # 獲取前端傳來的 upload_session_id
        upload_session_id = request.form.get('upload_session_id')
        print(f"📥 ERP 上傳 - 前端傳來的 session_id: {upload_session_id}")

        if file and allowed_file(file.filename):
            # ========== 檢查測試模式 ==========
            test_mode = request.form.get('test_mode') == 'true'
            customer_id = request.form.get('customer_id')
            template_username = user['username']
            # IT 測試模式：檔案放在 IT 人員的資料夾，但使用客戶的模板驗證

            if test_mode and customer_id and user['role'] in ['admin', 'it']:
                # 測試模式：使用被測試客戶的模板進行驗證，但檔案仍放在 IT 人員資料夾
                test_customer = get_user_by_id(int(customer_id))
                if test_customer:
                    template_username = test_customer['username']
                    print(f"[IT測試模式] ERP: 使用客戶 {template_username} 的模板驗證，檔案放在 IT 人員 (ID: {user['id']}) 資料夾")

            # 使用資料夾管理結構：uploads/{user_id}/{session_timestamp}/erp_data.xlsx/.xls
            # IT 測試模式下也使用 IT 人員的 user_id
            upload_folder, session_timestamp = get_or_create_session_folder(user['id'], 'uploads', upload_session_id)
            original_ext = get_file_extension(original_filename)  # 保留原始副檔名
            filename = 'erp_data' + original_ext
            filepath = os.path.join(upload_folder, filename)

            # 清理超過保留期限的資料夾
            cleanup_old_folders()

            # 保存文件
            file.save(filepath)
            print(f"ERP文件已保存到: {filepath} (session: {session_timestamp})")

            # 檢查文件是否真的存在
            if not os.path.exists(filepath):
                return jsonify({'success': False, 'message': '文件保存失敗'})

            # ========== 格式驗證 ==========

            print(f"開始驗證 ERP 文件格式...（用戶: {template_username}）")
            is_valid, message, details = validate_erp_format(filepath, template_username)

            if not is_valid:
                # 驗證失敗，刪除已上傳的文件
                os.remove(filepath)
                print(f"❌ ERP 文件格式驗證失敗: {message}")
                if details:
                    for detail in details:
                        print(f"   - {detail}")

                # 記錄上傳失敗
                log_upload(user['id'], 'erp', original_filename, 0, 0, 0, 'validation_failed', message)
                log_activity(user['id'], user['username'], 'upload_erp_failed',
                           f"ERP 文件上傳失敗：{message}", get_client_ip(), request.headers.get('User-Agent'))

                return jsonify({
                    'success': False,
                    'message': f'ERP 文件格式驗證失敗：{message}',
                    'details': details,
                    'validation_error': True
                })

            print(f"✅ ERP 文件格式驗證通過")
            # ========== 格式驗證結束 ==========

            # 讀取文件並返回基本信息
            try:
                df = pd.read_excel(filepath)
                file_size = os.path.getsize(filepath)
                print(f"ERP文件讀取成功: {len(df)} 行, {len(df.columns)} 欄位")

                # 將檔案路徑存入 session，供後續處理使用
                set_user_file_path('erp', filepath)

                # 記錄上傳成功
                log_upload(user['id'], 'erp', original_filename, file_size, len(df), len(df.columns), 'success')
                log_activity(user['id'], user['username'], 'upload_erp',
                           f"ERP 文件上傳成功：{original_filename} -> {filename}", get_client_ip(), request.headers.get('User-Agent'))

                return jsonify({
                    'success': True,
                    'message': 'ERP文件上傳成功（格式驗證通過）',
                    'rows': len(df),
                    'columns': list(df.columns),
                    'file_size': file_size,
                    'saved_filename': filename
                })
            except Exception as e:
                print(f"ERP文件讀取失敗: {str(e)}")
                log_upload(user['id'], 'erp', original_filename, 0, 0, 0, 'failed', str(e))
                return jsonify({'success': False, 'message': f'文件讀取失敗: {str(e)}'})

        return jsonify({'success': False, 'message': '不支持的文件格式'})

    except Exception as e:
        print(f"ERP上傳處理錯誤: {str(e)}")
        if user:
            log_upload(user['id'], 'erp', original_filename, 0, 0, 0, 'failed', str(e))
        return jsonify({'success': False, 'message': f'上傳處理失敗: {str(e)}'})

@app.route('/upload_forecast', methods=['POST'])
@login_required
def upload_forecast():
    """
    上傳 Forecast 文件（支援多檔案上傳）
    - 多檔案：使用 'files' 欄位
    - 單檔案：使用 'file' 欄位（向後兼容）
    - 多個 Forecast 檔案會合併成一個 forecast_data.xlsx
    """
    user = get_current_user()
    original_filenames = []
    try:
        # 檢查是否有上傳檔案（支援多檔案 'files' 或單檔案 'file'）
        files_list = []
        if 'files' in request.files:
            files_list = request.files.getlist('files')
        elif 'file' in request.files:
            files_list = [request.files['file']]

        if not files_list or all(f.filename == '' for f in files_list):
            return jsonify({'success': False, 'message': '沒有選擇文件'})

        # 過濾掉空檔名
        files_list = [f for f in files_list if f.filename != '']
        if not files_list:
            return jsonify({'success': False, 'message': '沒有選擇文件'})

        # 檢查所有文件格式
        for file in files_list:
            if not allowed_file(file.filename):
                return jsonify({'success': False, 'message': f'不支持的文件格式: {file.filename}'})
            original_filenames.append(file.filename)

        # 獲取前端傳來的 upload_session_id
        upload_session_id = request.form.get('upload_session_id')
        print(f"📥 Forecast 上傳 - 前端傳來的 session_id: {upload_session_id}")

        # 取得測試模式參數
        test_mode = request.form.get('test_mode') == 'true'
        customer_id = request.form.get('customer_id')
        template_username = user['username']
        # IT 測試模式：檔案放在 IT 人員的資料夾，但使用客戶的模板驗證

        if test_mode and customer_id and user['role'] in ['admin', 'it']:
            test_customer = get_user_by_id(int(customer_id))
            if test_customer:
                template_username = test_customer['username']
                print(f"[IT測試模式] Forecast: 使用客戶 {template_username} 的模板驗證，檔案放在 IT 人員 (ID: {user['id']}) 資料夾")

        # 使用資料夾管理結構（IT 測試模式下也使用 IT 人員的 user_id）
        upload_folder, session_timestamp = get_or_create_session_folder(user['id'], 'uploads', upload_session_id)

        # Delta 台達: 支援 8 種格式，1+ 檔案任何組合。統一走多檔合併路徑，
        # 不論實際上傳幾個檔案。
        is_delta_upload_early = (template_username == 'delta')

        # ========== 處理多檔案上傳 ==========
        if len(files_list) == 1 and not is_delta_upload_early:
            # 單檔案上傳：保留原始副檔名
            file = files_list[0]
            original_filename = file.filename
            original_ext = get_file_extension(original_filename)  # 取得原始副檔名 (.xls 或 .xlsx)
            filename = 'forecast_data' + original_ext
            filepath = os.path.join(upload_folder, filename)

            file.save(filepath)
            print(f"Forecast文件已保存到: {filepath} (session: {session_timestamp})")

            if not os.path.exists(filepath):
                return jsonify({'success': False, 'message': '文件保存失敗'})

            # 格式驗證
            print(f"開始驗證 Forecast 文件格式...（用戶: {template_username}）")
            is_valid, message, details = validate_forecast_format(filepath, template_username)

            if not is_valid:
                os.remove(filepath)
                print(f"❌ Forecast 文件格式驗證失敗: {message}")
                log_upload(user['id'], 'forecast', original_filename, 0, 0, 0, 'validation_failed', message)
                log_activity(user['id'], user['username'], 'upload_forecast_failed',
                           f"Forecast 文件上傳失敗：{message}", get_client_ip(), request.headers.get('User-Agent'))
                return jsonify({
                    'success': False,
                    'message': f'Forecast 文件格式驗證失敗：{message}',
                    'details': details,
                    'validation_error': True
                })

            print(f"✅ Forecast 文件格式驗證通過")

            # 讀取並返回
            df = pd.read_excel(filepath)
            file_size = os.path.getsize(filepath)
            set_user_file_path('forecast', filepath)

            log_upload(user['id'], 'forecast', original_filename, file_size, len(df), len(df.columns), 'success')
            log_activity(user['id'], user['username'], 'upload_forecast',
                       f"Forecast 文件上傳成功：{original_filename}", get_client_ip(), request.headers.get('User-Agent'))

            # 檢查在途需求（方案 C：只提醒不卡控）
            # IT 測試模式下，使用客戶的 ID 來查詢映射資料
            mapping_user_id = int(customer_id) if test_mode and customer_id else user['id']
            transit_check = check_transit_requirements_from_forecast(filepath, mapping_user_id, template_username)

            response_data = {
                'success': True,
                'message': 'Forecast文件上傳成功（格式驗證通過）',
                'rows': len(df),
                'columns': list(df.columns),
                'file_size': file_size,
                'file_count': 1,
                'saved_filename': filename
            }

            # 如果有在途相關提醒，加入回應
            if transit_check['message']:
                response_data['transit_check'] = transit_check

            return jsonify(response_data)

        else:
            # 多檔案上傳模式
            # 取得合併選項（預設為合併）
            merge_files = request.form.get('merge_files', 'true') == 'true'
            print(f"=== 多檔案上傳模式：收到 {len(files_list)} 個 Forecast 文件，合併模式: {merge_files} ===")

            all_dataframes = []
            files_info = []
            total_size = 0
            validation_errors = []

            # 先儲存所有檔案到暫存位置（保留原始副檔名）
            temp_files = []
            first_file_ext = None  # 記錄第一個檔案的副檔名，用於合併後的檔案
            for idx, file in enumerate(files_list):
                original_ext = get_file_extension(file.filename)
                if idx == 0:
                    first_file_ext = original_ext  # 記錄第一個檔案的副檔名
                temp_filename = f'forecast_temp_{idx}{original_ext}'
                temp_filepath = os.path.join(upload_folder, temp_filename)
                file.save(temp_filepath)
                temp_files.append((file.filename, temp_filepath, original_ext))
                print(f"  暫存文件 {idx + 1}: {file.filename} -> {temp_filepath}")

            # ========== Delta 台達：8 種格式自動偵測 (1+ 檔案任何組合) ==========
            is_delta_upload = (template_username == 'delta')

            if is_delta_upload:
                from delta_forecast_processor import (
                    detect_format, FORMAT_LABELS, consolidate as delta_consolidate,
                )

                # .xls → .xlsx 自動轉換 (openpyxl 不支援 .xls)
                from libreoffice_utils import convert_xls_to_xlsx
                converted_temp_files = []
                for original_name, temp_path, ext in temp_files:
                    if ext.lower() == '.xls':
                        print(f"  🔄 轉換 .xls → .xlsx: {original_name}")
                        xlsx_path = convert_xls_to_xlsx(temp_path, os.path.dirname(temp_path))
                        if xlsx_path and os.path.exists(xlsx_path):
                            os.remove(temp_path)
                            converted_temp_files.append((original_name, xlsx_path, '.xlsx'))
                        else:
                            converted_temp_files.append((original_name, temp_path, ext))
                            print(f"  ⚠️ 轉換失敗，使用原始 .xls: {original_name}")
                    else:
                        converted_temp_files.append((original_name, temp_path, ext))
                temp_files = converted_temp_files

                # 偵測每個檔案的格式 (8 種之一)
                detected_formats = []  # list of (original_name, temp_path, fmt)
                for original_name, temp_path, _ in temp_files:
                    fmt = detect_format(temp_path)
                    if fmt is None:
                        validation_errors.append({
                            'filename': original_name,
                            'message': '無法辨識 Delta Forecast 格式',
                            'details': [
                                '支援格式 (共 8 種): Ketwadee (PSB5) / Kanyanat (PSB7) / '
                                'Weeraya (PSB7) / India IAI1+UPI2 / PSW1+CEW1 / '
                                'MWC1+IPC1 / NBQ1 / SVC1+PWC1 (Diode&MOS)'
                            ]
                        })
                    else:
                        detected_formats.append((original_name, temp_path, fmt))
                        file_size = os.path.getsize(temp_path)
                        total_size += file_size
                        files_info.append({
                            'name': original_name,
                            'rows': 0,
                            'columns': 0,
                            'size': file_size,
                            'format': FORMAT_LABELS.get(fmt, fmt),
                        })
                        print(f"  ✅ {original_name}: {FORMAT_LABELS.get(fmt, fmt)}")

                # 至少要有 1 個有效格式
                if not validation_errors and not detected_formats:
                    validation_errors.append({
                        'filename': '(整體)',
                        'message': '未找到任何可識別的 Delta Forecast 檔案',
                        'details': [],
                    })

            else:
                # 非 Delta：使用原有驗證邏輯
                for original_name, temp_path, _ in temp_files:
                    print(f"驗證 Forecast 文件: {original_name}（用戶: {template_username}）")
                    is_valid, message, details = validate_forecast_format(temp_path, template_username)

                    if not is_valid:
                        validation_errors.append({
                            'filename': original_name,
                            'message': message,
                            'details': details
                        })
                        continue

                    try:
                        df = pd.read_excel(temp_path)
                        file_size = os.path.getsize(temp_path)
                        total_size += file_size

                        all_dataframes.append(df)
                        files_info.append({
                            'name': original_name,
                            'rows': len(df),
                            'columns': len(df.columns),
                            'size': file_size
                        })
                        print(f"  ✅ {original_name}: {len(df)} 行, {len(df.columns)} 欄位")

                    except Exception as e:
                        validation_errors.append({
                            'filename': original_name,
                            'message': f'讀取失敗: {str(e)}',
                            'details': []
                        })

            # 如果有驗證錯誤
            if validation_errors:
                # 清理暫存檔案
                for _, temp_path, _ in temp_files:
                    if os.path.exists(temp_path):
                        os.remove(temp_path)

                error_details = []
                for err in validation_errors:
                    error_details.append(f"{err['filename']}: {err['message']}")
                    if err['details']:
                        error_details.extend([f"  - {d}" for d in err['details']])

                log_activity(user['id'], user['username'], 'upload_forecast_failed',
                           f"Forecast 多檔案上傳失敗：{len(validation_errors)} 個文件驗證失敗", get_client_ip(), request.headers.get('User-Agent'))

                return jsonify({
                    'success': False,
                    'message': f'{len(validation_errors)} 個文件格式驗證失敗',
                    'details': error_details,
                    'validation_error': True
                })

            # 檢查是否有有效檔案
            if not is_delta_upload and not all_dataframes:
                return jsonify({'success': False, 'message': '沒有有效的 Forecast 文件'})

            # ========== Delta 台達：8 格式自動合併 ==========
            if is_delta_upload:
                import time
                print(f"=== Delta 自動合併模式：合併 {len(temp_files)} 個 Forecast 檔案 ===")
                merge_start = time.time()

                try:
                    # 從 customer_mappings 讀取 PLANT 代碼 (region 欄位)，
                    # 用於單 PLANT 檔案的檔名比對 (如 Ketwadee/Kanyanat/Weeraya/NBQ1)
                    from database import get_customer_mappings_raw
                    mapping_user_id = int(customer_id) if test_mode and customer_id else user['id']
                    raw_mappings = get_customer_mappings_raw(mapping_user_id) or []
                    plant_codes = []
                    for m in raw_mappings:
                        region = m.get('region') if isinstance(m, dict) else None
                        if region:
                            # 容許兩種格式: 純代碼 'PSB5' 或 舊格式 'PSB5 泰國' (取第一個 token)
                            first = str(region).split()[0] if str(region).split() else ''
                            if first and first not in plant_codes:
                                plant_codes.append(first)
                    print(f"  從 mapping 載入 {len(plant_codes)} 個 PLANT 代碼: {plant_codes}")

                    file_paths = [temp_path for _, temp_path, _ in temp_files]
                    # 原始檔名對應: 暫存檔 forecast_temp_N.xlsx → 使用者上傳的原檔名
                    # 用於 Buyer 欄位顯示 (避免出現 'forecast_temp_0')
                    file_labels = {
                        temp_path: original_name
                        for original_name, temp_path, _ in temp_files
                    }
                    reference_template = os.path.join('compare', 'delta', 'consolidated_template.xlsx')
                    final_filename = 'forecast_data.xlsx'
                    final_filepath = os.path.join(upload_folder, final_filename)

                    result = delta_consolidate(
                        file_paths, reference_template, final_filepath,
                        plant_codes=plant_codes,
                        file_labels=file_labels,
                    )

                    if not result['success']:
                        for _, temp_path, _ in temp_files:
                            if os.path.exists(temp_path):
                                os.remove(temp_path)
                        return jsonify({
                            'success': False,
                            'message': f'Delta 合併失敗: {result["message"]}'
                        })

                    # 保留客戶原始檔案（合併前），存到 originals 子資料夾
                    originals_folder = os.path.join(upload_folder, 'originals')
                    os.makedirs(originals_folder, exist_ok=True)
                    for original_name, temp_path, _ in temp_files:
                        if os.path.exists(temp_path) and temp_path != final_filepath:
                            import shutil
                            dest = os.path.join(originals_folder, original_name)
                            shutil.copy2(temp_path, dest)

                    # 清理暫存檔案
                    for _, temp_path, _ in temp_files:
                        if os.path.exists(temp_path) and temp_path != final_filepath:
                            os.remove(temp_path)

                    merged_size = os.path.getsize(final_filepath)
                    duration = time.time() - merge_start
                    print(f"=== Delta 合併完成：{result['part_count']} 個料號，耗時 {duration:.2f} 秒 ===")

                    # 更新 files_info 的 rows/columns
                    date_col_count = result.get('date_col_count', 0)
                    for fi in files_info:
                        fi['rows'] = result['part_count']
                        fi['columns'] = date_col_count

                    # 儲存 session
                    set_user_file_path('forecast', final_filepath)
                    session['forecast_merge_mode'] = False
                    session.modified = True

                    # 記錄日誌
                    filenames_str = ', '.join(original_filenames)
                    log_upload(user['id'], 'forecast', filenames_str, merged_size, result['part_count'], 0, 'success')
                    log_activity(user['id'], user['username'], 'upload_forecast',
                               f"Delta Forecast 合併上傳成功：{len(files_list)} 個檔案合併為匯總格式", get_client_ip(), request.headers.get('User-Agent'))

                    # 檢查在途需求
                    transit_check = check_transit_requirements_from_forecast(final_filepath, mapping_user_id, template_username)

                    format_stats = result.get('format_stats', {})
                    format_info = ', '.join([f'{os.path.basename(k)}: {v} 筆'
                                             for k, v in format_stats.items()])
                    # 讀取合併後的欄位名稱供前端顯示
                    import openpyxl as _opx
                    _wb_info = _opx.load_workbook(final_filepath, read_only=True)
                    _ws_info = _wb_info.active
                    _columns_list = [c.value for c in _ws_info[1] if c.value is not None]
                    _wb_info.close()

                    response_data = {
                        'success': True,
                        'message': f'Delta {len(files_list)} 個 Forecast 檔案合併成功（{result["part_count"]} 個料號）',
                        'file_count': len(files_list),
                        'rows': result['part_count'],
                        'columns': _columns_list,
                        'total_rows': result['part_count'],
                        'total_size': merged_size,
                        'files': files_info,
                        'merge_mode': False,
                        'saved_filename': final_filename,
                        'delta_consolidation': True,
                        'format_stats': format_info,
                        'buyer_stats': format_info,  # 向後相容 UI
                    }
                    if result.get('date_warnings'):
                        response_data['date_warnings'] = result['date_warnings']

                    if transit_check['message']:
                        response_data['transit_check'] = transit_check

                    return jsonify(response_data)

                except Exception as e:
                    print(f"Delta 合併處理錯誤: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    for _, temp_path, _ in temp_files:
                        if os.path.exists(temp_path):
                            os.remove(temp_path)
                    return jsonify({
                        'success': False,
                        'message': f'Delta 合併處理失敗: {str(e)}'
                    })

            # ========== 根據合併選項處理 ==========
            # Liteon: 不用 LibreOffice 合併（會丟失 Plant/Buyer 且合併全部 sheets）
            # Liteon 的智慧合併在 run_forecast 時由 merge_liteon_forecast_files() 處理
            is_liteon_upload = (template_username == 'liteon')
            if merge_files and not is_liteon_upload and not is_delta_upload:
                # 非 Liteon 合併模式：使用 LibreOffice 跨平台方案，保留格式
                import shutil
                import time
                from libreoffice_utils import merge_excel_files_libreoffice

                print(f"開始合併 {len(temp_files)} 個 Forecast 檔案（使用 LibreOffice 保留格式）...")
                merge_start = time.time()

                # 準備檔案路徑列表
                file_paths = [temp_path for _, temp_path, _ in temp_files]
                final_filename = 'forecast_data' + first_file_ext
                final_filepath = os.path.join(upload_folder, final_filename)

                # 使用 LibreOffice 合併檔案
                if len(temp_files) > 1:
                    try:
                        success = merge_excel_files_libreoffice(file_paths, final_filepath, skip_header=True)
                        if not success:
                            # 清理暫存檔案
                            for _, temp_path, _ in temp_files:
                                if os.path.exists(temp_path):
                                    os.remove(temp_path)
                            return jsonify({
                                'success': False,
                                'message': '合併 Forecast 檔案失敗',
                                'details': '請確認 LibreOffice 已正確安裝'
                            })
                    except Exception as merge_err:
                        # 清理暫存檔案
                        for _, temp_path, _ in temp_files:
                            if os.path.exists(temp_path):
                                os.remove(temp_path)
                        return jsonify({
                            'success': False,
                            'message': f'合併 Forecast 檔案失敗: {str(merge_err)}',
                            'details': '請確認 LibreOffice 已正確安裝'
                        })
                else:
                    # 只有一個檔案，直接複製
                    shutil.copy2(file_paths[0], final_filepath)

                # 計算總行數
                total_rows = sum(df.shape[0] for df in all_dataframes)
                print(f"=== 多檔案合併完成：{final_filepath}，總行數：{total_rows}，總耗時 {time.time() - merge_start:.2f} 秒 ===")

                # 清理暫存檔案
                for _, temp_path, _ in temp_files:
                    if os.path.exists(temp_path) and temp_path != final_filepath:
                        os.remove(temp_path)

                # 取得合併後的檔案大小
                merged_size = os.path.getsize(final_filepath)

                # 儲存上傳標記到 session（不存儲完整路徑，避免 cookie 超過 4KB）
                set_user_file_path('forecast', final_filepath)
                session['forecast_merge_mode'] = True
                session.modified = True

                # 記錄日誌
                filenames_str = ', '.join(original_filenames)
                # 計算總欄數（從第一個 dataframe 取得）
                total_columns = len(all_dataframes[0].columns) if all_dataframes else 0
                log_upload(user['id'], 'forecast', filenames_str, merged_size, total_rows, total_columns, 'success')
                log_activity(user['id'], user['username'], 'upload_forecast',
                           f"Forecast 多檔案上傳成功：{len(files_list)} 個文件已合併", get_client_ip(), request.headers.get('User-Agent'))

                # 檢查在途需求（方案 C：只提醒不卡控）
                # IT 測試模式下，使用客戶的 ID 來查詢映射資料
                mapping_user_id = int(customer_id) if test_mode and customer_id else user['id']
                transit_check = check_transit_requirements_from_forecast(final_filepath, mapping_user_id, template_username)

                response_data = {
                    'success': True,
                    'message': f'{len(files_list)} 個 Forecast 文件上傳並合併成功',
                    'file_count': len(files_list),
                    'total_rows': total_rows,
                    'total_size': merged_size,
                    'files': files_info,
                    'merge_mode': True,
                    'saved_filename': final_filename
                }

                # 如果有在途相關提醒，加入回應
                if transit_check['message']:
                    response_data['transit_check'] = transit_check

                return jsonify(response_data)

            else:
                # 不合併模式（或 Liteon 合併模式）：將暫存檔案重新命名為正式檔案（保留原始副檔名）
                # Liteon 勾合併時走此路徑，保留分檔，合併在 run_forecast 時處理
                saved_files = []
                total_rows = 0

                for idx, (original_name, temp_path, file_ext) in enumerate(temp_files):
                    # 產生正式檔名：forecast_data_1.xls/.xlsx, forecast_data_2.xls/.xlsx, ...（保留原始副檔名）
                    final_filename = f'forecast_data_{idx + 1}{file_ext}'
                    final_filepath = os.path.join(upload_folder, final_filename)

                    # 移動暫存檔案到正式位置
                    import shutil
                    shutil.move(temp_path, final_filepath)

                    saved_files.append({
                        'original_name': original_name,
                        'saved_name': final_filename,
                        'path': final_filepath,
                        'rows': files_info[idx]['rows']
                    })
                    total_rows += files_info[idx]['rows']

                    print(f"  儲存文件 {idx + 1}: {original_name} -> {final_filename}")

                print(f"=== 多檔案分開儲存完成：{len(saved_files)} 個文件 ===")

                # 儲存上傳標記到 session（只存儲檔案數量，不存儲完整路徑列表）
                set_user_file_path('forecast', saved_files[0]['path'] if saved_files else None)
                session['forecast_merge_mode'] = (merge_files and is_liteon_upload)  # Liteon 合併模式時為 True
                session['forecast_file_count'] = len(saved_files)
                session.modified = True

                # 記錄日誌
                filenames_str = ', '.join(original_filenames)
                log_upload(user['id'], 'forecast', filenames_str, total_size, total_rows, files_info[0]['columns'] if files_info else 0, 'success')
                log_activity(user['id'], user['username'], 'upload_forecast',
                           f"Forecast 多檔案上傳成功：{len(files_list)} 個文件（不合併）", get_client_ip(), request.headers.get('User-Agent'))

                # 檢查在途需求（方案 C：只提醒不卡控）- 檢查所有檔案
                # IT 測試模式下，使用客戶的 ID 來查詢映射資料
                all_file_paths = [f['path'] for f in saved_files]
                mapping_user_id = int(customer_id) if test_mode and customer_id else user['id']
                transit_check = check_transit_requirements_from_forecast(all_file_paths, mapping_user_id, template_username) if saved_files else {'message': ''}

                # 將 saved_name 加入 files_info 中，方便前端單檔刪除
                for idx, saved_file in enumerate(saved_files):
                    if idx < len(files_info):
                        files_info[idx]['saved_name'] = saved_file['saved_name']

                response_data = {
                    'success': True,
                    'message': f'{len(files_list)} 個 Forecast 文件上傳成功（不合併）',
                    'file_count': len(files_list),
                    'total_rows': total_rows,
                    'total_size': total_size,
                    'files': files_info,
                    'merge_mode': False,
                    'saved_files': [f['saved_name'] for f in saved_files]
                }

                # 如果有在途相關提醒，加入回應
                if transit_check['message']:
                    response_data['transit_check'] = transit_check

                return jsonify(response_data)

    except Exception as e:
        print(f"Forecast上傳處理錯誤: {str(e)}")
        if user:
            log_upload(user['id'], 'forecast', original_filename, 0, 0, 0, 'failed', str(e))
        return jsonify({'success': False, 'message': f'上傳處理失敗: {str(e)}'})


def merge_liteon_forecast_files(cleaned_files, output_path):
    """
    合併多個 cleaned Liteon forecast 為一個檔案，前置 Plant + Buyer Code 欄。

    原始每檔結構: C1=Plant, E1=Buyer, Row 7=headers, Row 8+=data
    合併後結構: Row 1 = [Plant, Buyer Code, 原 Row 7 headers], Row 2+ = data

    注意：不同檔案的日期欄位可能起始日不同，需用日期 remapping 對齊。
    保留原始格式（字型、填色、數字格式、對齊、邊框）和公式。
    """
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.formula.translate import Translator
    from copy import copy
    from datetime import datetime as _dt

    SHEET_NAME = 'Daily+Weekly+Monthly'
    DATE_COL_START = 11  # Column K (1-based)

    merged_wb = Workbook()
    merged_ws = merged_wb.active
    merged_ws.title = SHEET_NAME

    header_written = False
    current_row = 2  # Row 1 = header, Row 2+ = data
    master_date_cols = {}  # date_value -> merged_col (1-based, already +2)
    plant_daily_end_dates = {}  # plant_code -> last daily date (for each file's daily range)

    def _to_date(val):
        """Convert cell value to comparable date key"""
        if val is None:
            return None
        if isinstance(val, _dt):
            return val.date()
        if hasattr(val, 'date') and callable(val.date):
            try:
                return val.date()
            except:
                return None
        if isinstance(val, str):
            val = val.strip()
            for fmt in ['%Y/%m/%d', '%Y-%m-%d', '%m/%d/%Y']:
                try:
                    return _dt.strptime(val, fmt).date()
                except ValueError:
                    continue
        return None

    def _copy_cell(src_cell, tgt_cell, value_override=None):
        """複製 cell 的值（或公式）和樣式"""
        tgt_cell.value = value_override if value_override is not None else src_cell.value
        if src_cell.has_style:
            tgt_cell.font = copy(src_cell.font)
            tgt_cell.fill = copy(src_cell.fill)
            tgt_cell.number_format = src_cell.number_format
            tgt_cell.border = copy(src_cell.border)
            tgt_cell.alignment = copy(src_cell.alignment)
            tgt_cell.protection = copy(src_cell.protection)

    for file_idx, filepath in enumerate(cleaned_files):
        try:
            wb = openpyxl.load_workbook(filepath)
            if SHEET_NAME in wb.sheetnames:
                ws = wb[SHEET_NAME]
            else:
                ws = wb.active

            # Read Plant (C1) and Buyer Code (E1)
            plant_code = str(ws.cell(row=1, column=3).value or '').strip()
            buyer_code = str(ws.cell(row=1, column=5).value or '').strip()

            # Write header from first file's Row 7 (保留格式)
            if not header_written:
                # Plant / Buyer Code header (用第一個資料 cell 的樣式)
                header_style_cell = ws.cell(row=7, column=1)
                plant_hdr = merged_ws.cell(row=1, column=1, value='Plant')
                buyer_hdr = merged_ws.cell(row=1, column=2, value='Buyer Code')
                if header_style_cell.has_style:
                    for hdr_cell in [plant_hdr, buyer_hdr]:
                        hdr_cell.font = copy(header_style_cell.font)
                        hdr_cell.fill = copy(header_style_cell.fill)
                        hdr_cell.number_format = header_style_cell.number_format
                        hdr_cell.border = copy(header_style_cell.border)
                        hdr_cell.alignment = copy(header_style_cell.alignment)

                for col in range(1, ws.max_column + 1):
                    src_cell = ws.cell(row=7, column=col)
                    merged_col = col + 2
                    tgt_cell = merged_ws.cell(row=1, column=merged_col)
                    _copy_cell(src_cell, tgt_cell)
                    # 記錄日期欄位映射
                    if col >= DATE_COL_START:
                        d = _to_date(src_cell.value)
                        if d:
                            master_date_cols[d] = merged_col

                # 複製欄寬 (原始欄 +2 偏移)
                merged_ws.column_dimensions['A'].width = 8   # Plant
                merged_ws.column_dimensions['B'].width = 12  # Buyer Code
                for col in range(1, ws.max_column + 1):
                    src_letter = get_column_letter(col)
                    tgt_letter = get_column_letter(col + 2)
                    if src_letter in ws.column_dimensions:
                        merged_ws.column_dimensions[tgt_letter].width = ws.column_dimensions[src_letter].width

                # 複製列高 (Row 7 -> Row 1)
                if ws.row_dimensions[7].height:
                    merged_ws.row_dimensions[1].height = ws.row_dimensions[7].height

                header_written = True
                max_merged_col = ws.max_column + 2

            # 建立此檔案的 source_col -> merged_col 映射
            col_map = {}
            for col in range(1, DATE_COL_START):
                col_map[col] = col + 2
            for col in range(DATE_COL_START, ws.max_column + 1):
                val = ws.cell(row=7, column=col).value
                d = _to_date(val)
                if d and d in master_date_cols:
                    col_map[col] = master_date_cols[d]

            # 記錄此 Plant 的 Daily 結束日期（原始檔案的 daily 範圍）
            # Daily columns: col 11 (K) ~ 41 (AO) in original file
            DAILY_END_COL_ORIG = 41
            last_daily_date = None
            for col in range(DATE_COL_START, min(DAILY_END_COL_ORIG + 1, ws.max_column + 1)):
                d = _to_date(ws.cell(row=7, column=col).value)
                if d:
                    if last_daily_date is None or d > last_daily_date:
                        last_daily_date = d
            if plant_code and last_daily_date:
                plant_daily_end_dates[plant_code] = last_daily_date
                print(f"[Merge]   Plant {plant_code} daily range ends: {last_daily_date}")

            # Copy data rows (Row 8+) with Plant and Buyer prepended (保留格式和公式)
            for row in range(8, ws.max_row + 1):
                # Skip completely empty rows
                has_data = False
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=row, column=col).value is not None:
                        has_data = True
                        break
                if not has_data:
                    continue

                # Plant / Buyer Code (用該列第一個 cell 的樣式)
                row_style_cell = ws.cell(row=row, column=1)
                plant_cell = merged_ws.cell(row=current_row, column=1, value=plant_code)
                buyer_cell = merged_ws.cell(row=current_row, column=2, value=buyer_code)
                if row_style_cell.has_style:
                    for cell in [plant_cell, buyer_cell]:
                        cell.font = copy(row_style_cell.font)
                        cell.fill = copy(row_style_cell.fill)
                        cell.border = copy(row_style_cell.border)
                        cell.alignment = copy(row_style_cell.alignment)

                # 複製資料欄位（值 + 格式 + 公式，公式用 Translator 調整行列參照）
                for col in range(1, ws.max_column + 1):
                    target_col = col_map.get(col)
                    if target_col:
                        src_cell = ws.cell(row=row, column=col)
                        tgt_cell = merged_ws.cell(row=current_row, column=target_col)
                        val = src_cell.value
                        if isinstance(val, str) and val.startswith('='):
                            # 公式：用 Translator 自動調整 row/col 參照
                            try:
                                orig_coord = f"{get_column_letter(col)}{row}"
                                tgt_coord = f"{get_column_letter(target_col)}{current_row}"
                                adjusted = Translator(val, orig_coord).translate_formula(tgt_coord)
                                _copy_cell(src_cell, tgt_cell, value_override=adjusted)
                            except Exception:
                                _copy_cell(src_cell, tgt_cell)  # fallback: 原樣複製
                        else:
                            _copy_cell(src_cell, tgt_cell)

                # 複製列高
                if ws.row_dimensions[row].height:
                    merged_ws.row_dimensions[current_row].height = ws.row_dimensions[row].height

                current_row += 1

            wb.close()
            remapped = sum(1 for c in range(DATE_COL_START, ws.max_column + 1) if c in col_map)
            print(f"[Merge] 檔案 {file_idx + 1}: Plant={plant_code}, Buyer={buyer_code}, "
                  f"資料列={current_row - 2}, 日期對齊={remapped}欄")

        except Exception as e:
            print(f"[Merge] 檔案 {file_idx + 1} 合併失敗: {e}")
            import traceback
            traceback.print_exc()
            continue

    merged_wb.save(output_path)
    print(f"[Merge] 合併完成: {current_row - 2} 列 → {os.path.basename(output_path)}")
    print(f"[Merge] Plant Daily 結束日期: {plant_daily_end_dates}")
    return current_row - 2, plant_daily_end_dates  # return (total_rows, plant_daily_end_dates)


@app.route('/merge_forecast_files', methods=['POST'])
@login_required
def merge_forecast_files():
    """
    合併多個 Forecast 檔案
    當用戶在確認對話框勾選「合併」但上傳時沒有合併時，呼叫此 API
    """
    user = get_current_user()
    try:
        data = request.get_json() or {}
        upload_session_id = data.get('upload_session_id')

        if not upload_session_id:
            return jsonify({'success': False, 'message': '缺少 upload_session_id'})

        # 取得上傳資料夾
        upload_folder = os.path.join(UPLOAD_FOLDER, str(user['id']), upload_session_id)
        if not os.path.exists(upload_folder):
            return jsonify({'success': False, 'message': '找不到上傳資料夾'})

        print(f"=== 開始合併 Forecast 檔案 (session: {upload_session_id}) ===")

        # 檢查是否有多個分開的檔案
        multi_files = []
        first_file_ext = None
        for i in range(1, 100):
            filepath = find_file_with_extensions(upload_folder, f'forecast_data_{i}')
            if filepath:
                multi_files.append(filepath)
                if first_file_ext is None:
                    first_file_ext = os.path.splitext(filepath)[1].lower()
            else:
                break

        if len(multi_files) < 2:
            # 已經是單一檔案或沒有檔案
            return jsonify({
                'success': True,
                'message': '不需要合併（已是單一檔案）',
                'merged': False
            })

        print(f"  找到 {len(multi_files)} 個待合併檔案")

        # 判斷是否為 Liteon（檢查 test_mode 或 user 本身）
        test_mode = data.get('test_mode', False)
        customer_id = data.get('customer_id')
        if test_mode and customer_id and user['role'] in ['admin', 'it']:
            target_user = get_user_by_id(int(customer_id))
            is_liteon_merge = (target_user and target_user['username'] == 'liteon')
        else:
            is_liteon_merge = (user['username'] == 'liteon')

        if is_liteon_merge:
            # Liteon: 不做 LibreOffice 合併（會丟失 Plant/Buyer 且合併全部 sheets）
            # 只設定 session flag，保留分檔，run_forecast 時由 merge_liteon_forecast_files() 處理
            print(f"  Liteon 模式：跳過 LibreOffice 合併，保留 {len(multi_files)} 個分檔")

            session['forecast_merge_mode'] = True
            session.modified = True

            log_activity(user['id'], user['username'], 'merge_forecast',
                       f"Liteon 合併模式標記：{len(multi_files)} 個檔案（保留分檔）", get_client_ip(), request.headers.get('User-Agent'))

            return jsonify({
                'success': True,
                'message': f'{len(multi_files)} 個檔案已標記合併（Liteon 模式）',
                'merged': True,
                'merge_time': 0
            })

        # 非 Liteon：使用 LibreOffice 進行合併（跨平台方案）
        import shutil
        import time
        from libreoffice_utils import merge_excel_files_libreoffice

        merge_start = time.time()

        # 設定輸出檔案路徑
        final_filename = 'forecast_data' + first_file_ext
        final_filepath = os.path.join(upload_folder, final_filename)

        # 使用 LibreOffice 合併檔案
        try:
            success = merge_excel_files_libreoffice(multi_files, final_filepath, skip_header=True)
            if not success:
                return jsonify({
                    'success': False,
                    'message': '合併檔案失敗',
                    'details': '請確認 LibreOffice 已正確安裝'
                })
        except Exception as merge_err:
            return jsonify({
                'success': False,
                'message': f'合併檔案失敗: {str(merge_err)}',
                'details': '請確認 LibreOffice 已正確安裝'
            })

        # 刪除原來的分開檔案
        for filepath in multi_files:
            if os.path.exists(filepath):
                os.remove(filepath)
                print(f"  已刪除: {os.path.basename(filepath)}")

        merge_time = time.time() - merge_start
        print(f"=== 合併完成，耗時 {merge_time:.2f} 秒 ===")

        # 更新 session
        session['forecast_merge_mode'] = True
        session.modified = True

        log_activity(user['id'], user['username'], 'merge_forecast',
                   f"Forecast 檔案合併成功：{len(multi_files)} 個檔案", get_client_ip(), request.headers.get('User-Agent'))

        return jsonify({
            'success': True,
            'message': f'{len(multi_files)} 個檔案已合併',
            'merged': True,
            'merged_filename': final_filename,
            'merge_time': round(merge_time, 2)
        })

    except Exception as e:
        print(f"合併 Forecast 檔案失敗: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'合併失敗: {str(e)}'})


@app.route('/upload_transit', methods=['POST'])
@login_required
def upload_transit():
    user = get_current_user()
    original_filename = ''
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '沒有選擇文件'})

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '沒有選擇文件'})

        original_filename = file.filename

        # 獲取前端傳來的 upload_session_id
        upload_session_id = request.form.get('upload_session_id')
        print(f"📥 Transit 上傳 - 前端傳來的 session_id: {upload_session_id}")

        if file and allowed_file(file.filename):
            # ========== 檢查測試模式 ==========
            test_mode = request.form.get('test_mode') == 'true'
            customer_id = request.form.get('customer_id')
            template_username = user['username']
            # IT 測試模式：檔案放在 IT 人員的資料夾，但使用客戶的模板驗證

            if test_mode and customer_id and user['role'] in ['admin', 'it']:
                # 測試模式：使用被測試客戶的模板進行驗證，但檔案仍放在 IT 人員資料夾
                test_customer = get_user_by_id(int(customer_id))
                if test_customer:
                    template_username = test_customer['username']
                    print(f"[IT測試模式] Transit: 使用客戶 {template_username} 的模板驗證，檔案放在 IT 人員 (ID: {user['id']}) 資料夾")

            # 使用資料夾管理結構：uploads/{user_id}/{session_timestamp}/transit_data.xlsx/.xls
            # IT 測試模式下也使用 IT 人員的 user_id
            upload_folder, session_timestamp = get_or_create_session_folder(user['id'], 'uploads', upload_session_id)
            original_ext = get_file_extension(original_filename)  # 保留原始副檔名
            filename = 'transit_data' + original_ext
            filepath = os.path.join(upload_folder, filename)

            # 保存文件
            file.save(filepath)
            print(f"在途文件已保存到: {filepath} (session: {session_timestamp})")

            # 檢查文件是否真的存在
            if not os.path.exists(filepath):
                return jsonify({'success': False, 'message': '文件保存失敗'})

            # ========== 格式驗證 ==========

            print(f"開始驗證在途文件格式...（用戶: {template_username}）")
            is_valid, message, details = validate_transit_format(filepath, template_username)

            if not is_valid:
                # 驗證失敗，刪除已上傳的文件
                os.remove(filepath)
                print(f"❌ 在途文件格式驗證失敗: {message}")
                if details:
                    for detail in details:
                        print(f"   - {detail}")

                # 記錄上傳失敗
                log_upload(user['id'], 'transit', original_filename, 0, 0, 0, 'validation_failed', message)
                log_activity(user['id'], user['username'], 'upload_transit_failed',
                           f"在途文件上傳失敗：{message}", get_client_ip(), request.headers.get('User-Agent'))

                return jsonify({
                    'success': False,
                    'message': f'在途文件格式驗證失敗：{message}',
                    'details': details,
                    'validation_error': True
                })

            print(f"✅ 在途文件格式驗證通過")
            # ========== 格式驗證結束 ==========

            # 讀取文件並返回基本信息
            try:
                df = pd.read_excel(filepath)
                file_size = os.path.getsize(filepath)
                print(f"在途文件讀取成功: {len(df)} 行, {len(df.columns)} 欄位")

                # 顯示欄位資訊
                print(f"✅ 在途文件欄位資訊:")
                for i, col in enumerate(df.columns):
                    print(f"   索引{i}: {col}")

                # 將檔案路徑存入 session，供後續處理使用
                set_user_file_path('transit', filepath)
                print(f"✅ Transit 檔案路徑已設置到 session: {filepath}")
                print(f"   Session keys: {list(session.keys())}")
                print(f"   current_transit_file: {session.get('current_transit_file')}")

                # 記錄上傳成功
                log_upload(user['id'], 'transit', original_filename, file_size, len(df), len(df.columns), 'success')
                log_activity(user['id'], user['username'], 'upload_transit',
                           f"在途文件上傳成功：{original_filename} -> {filename}", get_client_ip(), request.headers.get('User-Agent'))

                return jsonify({
                    'success': True,
                    'message': '在途文件上傳成功（格式驗證通過）',
                    'rows': len(df),
                    'columns': list(df.columns),
                    'file_size': file_size,
                    'saved_filename': filename
                })
            except Exception as e:
                print(f"在途文件讀取失敗: {str(e)}")
                log_upload(user['id'], 'transit', original_filename, 0, 0, 0, 'failed', str(e))
                return jsonify({'success': False, 'message': f'文件讀取失敗: {str(e)}'})

        return jsonify({'success': False, 'message': '不支持的文件格式'})

    except Exception as e:
        print(f"在途文件上傳處理錯誤: {str(e)}")
        if user:
            log_upload(user['id'], 'transit', original_filename, 0, 0, 0, 'failed', str(e))
        return jsonify({'success': False, 'message': f'上傳處理失敗: {str(e)}'})

@app.route('/mapping')
@login_required
def mapping():
    user = get_current_user()
    response = make_response(render_template('mapping.html', user=user))
    # 設置不緩存HTML頁面
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/get_mapping_data')
@login_required
def get_mapping_data():
    user = get_current_user()
    try:
        # 檢查是否為 IT 測試模式
        test_mode = request.args.get('test_mode') == 'true'
        customer_id = request.args.get('customer_id')

        # 決定使用哪個用戶的 mapping 資料
        mapping_user_id = user['id']
        if test_mode and customer_id and user['role'] in ['admin', 'it']:
            mapping_user_id = int(customer_id)
            print(f"[IT測試模式] 載入客戶 ID {customer_id} 的 mapping 資料")

        # 1. 首先嘗試從資料庫讀取用戶的 mapping 資料（使用 raw 格式以支援多記錄）
        if has_customer_mappings(mapping_user_id):
            print(f"從資料庫讀取用戶 ID {mapping_user_id} 的 mapping 資料...")
            raw_mappings = get_customer_mappings_raw(mapping_user_id)

            if raw_mappings:
                # 返回列表格式，每筆記錄為一行
                mapping_list = []
                for row in raw_mappings:
                    # 處理 requires_transit，確保正確轉換布林值
                    requires_transit = row.get('requires_transit', True)
                    if requires_transit is None or requires_transit == 1:
                        requires_transit = True
                    elif requires_transit == 0:
                        requires_transit = False

                    item = {
                        'customer_name': row['customer_name'] or '',
                        'region': row['region'] or '',
                        'schedule_breakpoint': row['schedule_breakpoint'] or '',
                        'etd': row['etd'] or '',
                        'eta': row['eta'] or '',
                        'requires_transit': requires_transit
                    }
                    # 光寶專用欄位（其他客戶這些欄位為 NULL，不影響）
                    if row.get('order_type'):
                        item['order_type'] = row['order_type']
                    if row.get('warehouse'):
                        item['warehouse'] = row['warehouse']
                    if row.get('date_calc_type'):
                        item['date_calc_type'] = row['date_calc_type']
                    if row.get('delivery_location'):
                        item['delivery_location'] = row['delivery_location']
                    mapping_list.append(item)

                return jsonify({
                    'success': True,
                    'mapping_list': mapping_list,
                    'customer_column': '客戶簡稱',
                    'source': 'database',
                    'format': 'list'
                })

        # 2. 如果資料庫沒有資料，嘗試從 mapping 表 Excel 讀取（向後相容）
        mapping_file = os.path.join('mapping', 'mapping表.xlsx')
        if os.path.exists(mapping_file):
            print("資料庫無資料，從 mapping表.xlsx 讀取...")
            mapping_df = pd.read_excel(mapping_file)

            # 找到客戶簡稱欄位（通常是A欄位）
            customer_col = mapping_df.columns[0]  # A欄位

            # 找到其他欄位
            region_col = None
            schedule_col = None
            etd_col = None
            eta_col = None

            for col in mapping_df.columns:
                if '地區' in str(col) or 'region' in str(col).lower():
                    region_col = col
                elif '排程' in str(col) or '斷點' in str(col):
                    schedule_col = col
                elif 'ETD' in str(col):
                    etd_col = col
                elif 'ETA' in str(col):
                    eta_col = col

            # 構建現有的映射數據
            existing_mapping = {
                'regions': {},
                'schedule_breakpoints': {},
                'etd': {},
                'eta': {}
            }

            for idx, row in mapping_df.iterrows():
                customer = str(row[customer_col])
                if region_col and pd.notna(row[region_col]):
                    existing_mapping['regions'][customer] = str(row[region_col])
                if schedule_col and pd.notna(row[schedule_col]):
                    existing_mapping['schedule_breakpoints'][customer] = str(row[schedule_col])
                if etd_col and pd.notna(row[etd_col]):
                    existing_mapping['etd'][customer] = str(row[etd_col])
                if eta_col and pd.notna(row[eta_col]):
                    existing_mapping['eta'][customer] = str(row[eta_col])

            return jsonify({
                'success': True,
                'customers': list(existing_mapping['regions'].keys()),
                'customer_column': customer_col,
                'existing_mapping': existing_mapping,
                'source': 'mapping_table'
            })

        # 3. 如果都沒有，則從ERP文件獲取客戶列表
        erp_file = get_user_file_path('erp')
        if not erp_file or not os.path.exists(erp_file):
            return jsonify({'success': False, 'message': '請先上傳ERP文件或配置映射表'})

        print("從ERP文件獲取客戶數據...")
        df = pd.read_excel(erp_file)

        # 找到客戶簡稱欄位
        customer_col = None
        for col in df.columns:
            if '客戶' in str(col) and '簡稱' in str(col):
                customer_col = col
                break

        if customer_col is None:
            return jsonify({'success': False, 'message': '找不到客戶簡稱欄位'})

        # 獲取唯一的客戶簡稱
        unique_customers = df[customer_col].dropna().unique().tolist()

        return jsonify({
            'success': True,
            'customers': unique_customers,
            'customer_column': customer_col,
            'existing_mapping': {},
            'source': 'erp_file'
        })
    except Exception as e:
        print(f"獲取映射數據失敗: {str(e)}")
        return jsonify({'success': False, 'message': f'獲取映射數據失敗: {str(e)}'})

@app.route('/save_mapping', methods=['POST'])
@login_required
def save_mapping():
    user = get_current_user()
    try:
        mapping_data = request.json

        # 統計客戶數量
        all_customers = set()
        all_customers.update(mapping_data.get('regions', {}).keys())
        all_customers.update(mapping_data.get('schedule_breakpoints', {}).keys())
        all_customers.update(mapping_data.get('etd', {}).keys())
        all_customers.update(mapping_data.get('eta', {}).keys())
        customer_count = len(all_customers)

        # 1. 儲存到資料庫（主要儲存方式，按用戶區分）
        if save_customer_mappings(user['id'], mapping_data):
            print(f"✅ 已儲存 mapping 資料到資料庫 (user: {user['username']})")

            # 記錄 LOG：mapping 配置保存成功
            log_activity(
                user_id=user['id'],
                username=user['username'],
                action_type='mapping_config_save',
                action_detail=f"用戶 {user['display_name']} 保存映射配置，共 {customer_count} 個客戶",
                ip_address=get_client_ip(),
                user_agent=request.headers.get('User-Agent')
            )
        else:
            # 記錄 LOG：mapping 配置保存失敗
            log_activity(
                user_id=user['id'],
                username=user['username'],
                action_type='mapping_config_failed',
                action_detail=f"用戶 {user['display_name']} 保存映射配置失敗：儲存至資料庫失敗",
                ip_address=get_client_ip(),
                user_agent=request.headers.get('User-Agent')
            )
            return jsonify({'success': False, 'message': '儲存映射資料到資料庫失敗'})

        # 2. 同時更新 Excel 文件（向後相容，保持舊格式）
        mapping_excel_file = os.path.join('mapping', 'mapping表.xlsx')
        if os.path.exists(mapping_excel_file):
            # 讀取現有Excel文件
            mapping_df = pd.read_excel(mapping_excel_file)

            # 更新數據
            for idx, row in mapping_df.iterrows():
                customer = str(row.iloc[0])  # 客戶簡稱欄位

                # 更新各個欄位
                if customer in mapping_data.get('regions', {}):
                    mapping_df.iloc[idx, 1] = mapping_data['regions'][customer]  # 客戶需求地區
                if customer in mapping_data.get('schedule_breakpoints', {}):
                    mapping_df.iloc[idx, 3] = mapping_data['schedule_breakpoints'][customer]  # 排程出貨日期斷點
                if customer in mapping_data.get('etd', {}):
                    mapping_df.iloc[idx, 4] = mapping_data['etd'][customer]  # ETD
                if customer in mapping_data.get('eta', {}):
                    mapping_df.iloc[idx, 5] = mapping_data['eta'][customer]  # ETA

            # 保存更新後的Excel文件
            mapping_df.to_excel(mapping_excel_file, index=False)
            print(f"✅ 已同步更新 Excel 文件: {mapping_excel_file}")

        return jsonify({'success': True, 'message': '映射表保存成功（已儲存至資料庫）'})
    except Exception as e:
        print(f"❌ 保存映射表失敗: {str(e)}")
        # 記錄 LOG：mapping 配置保存異常
        log_activity(
            user_id=user['id'],
            username=user['username'],
            action_type='mapping_config_failed',
            action_detail=f"用戶 {user['display_name']} 保存映射配置異常：{str(e)}",
            ip_address=get_client_ip(),
            user_agent=request.headers.get('User-Agent')
        )
        return jsonify({'success': False, 'message': f'保存映射表失敗: {str(e)}'})


@app.route('/save_mapping_list', methods=['POST'])
@login_required
def save_mapping_list():
    """
    新版保存映射 API - 支援列表格式
    每筆記錄以 (customer_name, region) 為唯一 key
    """
    user = get_current_user()
    try:
        data = request.json
        mapping_list = data.get('mapping_list', [])

        # 檢查是否為 IT 測試模式
        test_mode = data.get('test_mode', False)
        customer_id = data.get('customer_id')

        # 決定要保存到哪個用戶的映射資料
        mapping_user_id = user['id']
        if test_mode and customer_id and user['role'] in ['admin', 'it']:
            mapping_user_id = int(customer_id)

        if not mapping_list:
            return jsonify({'success': False, 'message': '沒有資料需要保存'})

        # 直接儲存列表格式到資料庫
        if save_customer_mappings_list(mapping_user_id, mapping_list):
            print(f"✅ 已儲存 {len(mapping_list)} 筆 mapping 資料到資料庫 (user: {user['username']})")

            # 記錄 LOG
            log_activity(
                user_id=user['id'],
                username=user['username'],
                action_type='mapping_config_save',
                action_detail=f"用戶 {user['display_name']} 保存映射配置，共 {len(mapping_list)} 筆記錄",
                ip_address=get_client_ip(),
                user_agent=request.headers.get('User-Agent')
            )

            return jsonify({'success': True, 'message': f'映射表保存成功，共 {len(mapping_list)} 筆記錄'})
        else:
            log_activity(
                user_id=user['id'],
                username=user['username'],
                action_type='mapping_config_failed',
                action_detail=f"用戶 {user['display_name']} 保存映射配置失敗：儲存至資料庫失敗",
                ip_address=get_client_ip(),
                user_agent=request.headers.get('User-Agent')
            )
            return jsonify({'success': False, 'message': '儲存映射資料到資料庫失敗'})

    except Exception as e:
        print(f"❌ 保存映射表失敗: {str(e)}")
        log_activity(
            user_id=user['id'],
            username=user['username'],
            action_type='mapping_config_failed',
            action_detail=f"用戶 {user['display_name']} 保存映射配置異常：{str(e)}",
            ip_address=get_client_ip(),
            user_agent=request.headers.get('User-Agent')
        )
        return jsonify({'success': False, 'message': f'保存映射表失敗: {str(e)}'})


@app.route('/process_forecast_cleanup', methods=['POST'])
@login_required
def process_forecast_cleanup():
    user = get_current_user()
    start_time = time.time()
    try:
        # 檢查是否為測試模式（支援無 body 的請求）
        data = {}
        try:
            data = request.get_json(silent=True) or {}
        except:
            pass
        test_mode = data.get('test_mode', False)
        customer_id = data.get('customer_id')
        # IT 測試模式：檔案放在 IT 人員的資料夾（與上傳時一致）

        if test_mode and customer_id and user['role'] in ['admin', 'it']:
            print(f"[IT測試模式] Cleanup: 檔案放在 IT 人員 (ID: {user['id']}) 資料夾")

        # 獲取前端傳來的 upload_session_id
        upload_session_id = data.get('upload_session_id')
        print(f"📥 Cleanup - 前端傳來的 session_id: {upload_session_id}")

        # 如果有前端傳來的 session_id，先同步到 Flask session
        if upload_session_id:
            session['current_session_timestamp'] = upload_session_id
            session.modified = True
            print(f"📁 已同步 session_id 到 Flask session: {upload_session_id}")

        # 記錄開始處理（使用 IT 人員的身份記錄 log）
        log_activity(user['id'], user['username'], 'cleanup_start',
                   f"開始 Forecast 數據清理{' (IT測試模式, 客戶ID: ' + str(customer_id) + ')' if test_mode else ''}", get_client_ip(), request.headers.get('User-Agent'))

        # 使用資料夾管理結構（IT 測試模式下也使用 IT 人員的 user_id）
        processed_folder, session_timestamp = get_or_create_session_folder(user['id'], 'processed', upload_session_id)

        # 存儲 processed 資料夾路徑到 session
        session['current_processed_folder'] = processed_folder

        # 根據實際上傳的檔案來判斷是否為多檔案模式（不依賴 session 標記）
        # 掃描上傳資料夾，檢查是否有多個 forecast_data_*.xlsx/.xls 檔案
        # IT 測試模式下也使用 IT 人員的資料夾
        upload_folder = os.path.join(UPLOAD_FOLDER, str(user['id']), session_timestamp)
        print(f"🔍 Cleanup - 掃描上傳資料夾: {upload_folder}")

        # 檢查是否有編號的多檔案 (forecast_data_1.xlsx/.xls, forecast_data_2.xlsx/.xls, ...)
        multi_files = []
        for i in range(1, 100):
            filepath = find_file_with_extensions(upload_folder, f'forecast_data_{i}')
            if filepath:
                multi_files.append(filepath)
            else:
                break

        # 檢查是否有單一合併檔案（支援 .xlsx 和 .xls）
        single_file = find_file_with_extensions(upload_folder, 'forecast_data')
        has_single_file = single_file is not None

        print(f"📁 多檔案數量: {len(multi_files)}, 單一檔案存在: {has_single_file}")

        # 判斷模式：如果有多個編號檔案，就是多檔案分開模式
        is_multi_file_mode = len(multi_files) > 1

        if is_multi_file_mode:
            forecast_files_list = multi_files
            print(f"=== 多檔案清理模式：{len(forecast_files_list)} 個檔案 ===")
        elif len(multi_files) == 1:
            # 只有一個編號檔案，當作單檔案處理
            forecast_file = multi_files[0]
            print(f"=== 單檔案清理模式（編號檔案）：{forecast_file} ===")
        elif has_single_file:
            forecast_file = single_file
            print(f"=== 單檔案清理模式（合併檔案）：{forecast_file} ===")
        else:
            log_process(user['id'], 'cleanup', 'failed', '請先上傳Forecast文件')
            return jsonify({'success': False, 'message': '請先上傳Forecast文件'})

        # 根據用戶決定清理邏輯
        # IT 測試模式下，使用客戶的用戶名來決定清理邏輯
        username = user['username']
        if test_mode and customer_id and user['role'] in ['admin', 'it']:
            test_customer = get_user_by_id(int(customer_id))
            if test_customer:
                username = test_customer['username']
                print(f"[IT測試模式] Cleanup: 使用客戶 {username} 的清理邏輯")
        print(f"📋 用戶: {username}，使用對應的清理邏輯")

        # 多檔案分開模式
        if is_multi_file_mode:
            print(f"=== 多檔案清理模式：{len(forecast_files_list)} 個檔案 ===")

            total_cleaned_count = 0
            cleaned_files_info = []

            # 分離 .xls 和 .xlsx 檔案
            xls_files = []
            xlsx_files = []
            for idx, file_path in enumerate(forecast_files_list):
                if file_path and os.path.exists(file_path):
                    if is_xls_format(file_path):
                        xls_files.append((idx, file_path))
                    else:
                        xlsx_files.append((idx, file_path))
                else:
                    original_name = os.path.basename(file_path) if file_path else f'file_{idx}'
                    print(f"  ⚠️ 檔案不存在: {file_path}")
                    cleaned_files_info.append({
                        'name': original_name,
                        'cleaned_cells': 0,
                        'status': 'error',
                        'message': '檔案不存在'
                    })

            # ========== 批次處理所有 .xls 檔案（使用 LibreOffice 跨平台方案）==========
            if xls_files:
                print(f"  📁 批次處理 {len(xls_files)} 個 .xls 檔案（使用 LibreOffice）...")
                from libreoffice_utils import cleanup_xls_file_libreoffice

                for idx, file_path in xls_files:
                    original_name = os.path.basename(file_path)
                    print(f"  清理檔案 {idx + 1}/{len(forecast_files_list)}: {original_name}")

                    try:
                        # 設定輸出檔案路徑（輸出為 .xlsx 格式）
                        cleaned_filename = f'cleaned_forecast_{idx + 1}.xlsx'
                        cleaned_file = os.path.join(processed_folder, cleaned_filename)

                        # 使用 LibreOffice 清理（輸出為 xlsx）
                        cleaned_count = cleanup_xls_file_libreoffice(file_path, cleaned_file, username)

                        total_cleaned_count += cleaned_count
                        cleaned_files_info.append({
                            'name': original_name,
                            'cleaned_cells': cleaned_count,
                            'status': 'success',
                            'cleaned_path': cleaned_file
                        })
                        print(f"    ✅ 清理了 {cleaned_count} 個單元格")

                    except Exception as file_error:
                        print(f"    ❌ 清理失敗: {str(file_error)}")
                        cleaned_files_info.append({
                            'name': original_name,
                            'cleaned_cells': 0,
                            'status': 'error',
                            'message': str(file_error)
                        })

            # ========== 處理 .xlsx 檔案 ==========
            for idx, file_path in xlsx_files:
                original_name = os.path.basename(file_path)
                print(f"  清理檔案 {idx + 1}/{len(forecast_files_list)}: {original_name}")

                try:
                    wb = load_workbook(file_path)
                    cleaned_count = 0

                    if username == 'delta':
                        # 台達：I欄="Supply" 的整列，從 J 欄到最後一欄清零
                        ws = wb.active
                        for row_idx in range(1, ws.max_row + 1):
                            i_cell = ws.cell(row=row_idx, column=9)  # I欄 = column 9
                            if i_cell.value and str(i_cell.value).strip() == 'Supply':
                                for col_idx in range(10, ws.max_column + 1):  # J=10 到最後
                                    cell = ws.cell(row=row_idx, column=col_idx)
                                    if cell.value is not None and cell.value != 0:
                                        cell.value = 0
                                        cleaned_count += 1

                    elif username == 'liteon':
                        # 光寶：指定讀取 Daily+Weekly+Monthly sheet
                        # 清理條件：C欄(column 3) = "Commit" 時，清零 J~BY 欄(column 10~77)
                        if 'Daily+Weekly+Monthly' in wb.sheetnames:
                            ws = wb['Daily+Weekly+Monthly']
                        else:
                            ws = wb.active
                        for row_idx in range(1, ws.max_row + 1):
                            c_cell = ws.cell(row=row_idx, column=3)
                            if c_cell.value and str(c_cell.value).strip() == "Commit":
                                for col_idx in range(10, min(78, ws.max_column + 1)):  # J=10, BY=77
                                    cell = ws.cell(row=row_idx, column=col_idx)
                                    if cell.value is not None and cell.value != 0:
                                        cell.value = 0
                                        cleaned_count += 1
                    elif username == 'pegatron':
                        # 和碩：M欄(column 13) = "ETA QTY" 時，清零 N~DN 欄(column 14~118)
                        ws = wb.active
                        for row_idx in range(1, ws.max_row + 1):
                            m_cell = ws.cell(row=row_idx, column=13)
                            if m_cell.value and str(m_cell.value).strip() == "ETA QTY":
                                for col_idx in range(14, min(119, ws.max_column + 1)):
                                    cell = ws.cell(row=row_idx, column=col_idx)
                                    if cell.value is not None and cell.value != 0:
                                        cell.value = 0
                                        cleaned_count += 1
                    else:
                        # 廣達：K欄(column 11) = "供應數量" 時，清零 L~AW 欄(column 12~49)
                        ws = wb.active
                        for row_idx in range(1, ws.max_row + 1):
                            k_cell = ws.cell(row=row_idx, column=11)
                            if k_cell.value and str(k_cell.value) == "供應數量":
                                for col_idx in range(12, min(50, ws.max_column + 1)):
                                    cell = ws.cell(row=row_idx, column=col_idx)
                                    if cell.value != 0:
                                        cell.value = 0
                                        cleaned_count += 1

                            i_cell = ws.cell(row=row_idx, column=9)
                            if i_cell.value and "庫存數量" in str(i_cell.value):
                                next_row_i_cell = ws.cell(row=row_idx + 1, column=9)
                                if next_row_i_cell.value != 0:
                                    next_row_i_cell.value = 0
                                    cleaned_count += 1

                    cleaned_filename = f'cleaned_forecast_{idx + 1}.xlsx'
                    cleaned_file = os.path.join(processed_folder, cleaned_filename)
                    wb.save(cleaned_file)

                    total_cleaned_count += cleaned_count
                    cleaned_files_info.append({
                        'name': original_name,
                        'cleaned_cells': cleaned_count,
                        'status': 'success',
                        'cleaned_path': cleaned_file
                    })
                    print(f"    ✅ 清理了 {cleaned_count} 個單元格")

                except Exception as file_error:
                    print(f"    ❌ 清理失敗: {str(file_error)}")
                    cleaned_files_info.append({
                        'name': original_name,
                        'cleaned_cells': 0,
                        'status': 'error',
                        'message': str(file_error)
                    })

            # 統計成功清理的檔案數量
            cleaned_paths = [f for f in cleaned_files_info if f['status'] == 'success']

            duration = time.time() - start_time
            print(f"=== 多檔案清理完成：共清理 {total_cleaned_count} 個單元格 ===")

            log_process(user['id'], 'cleanup', 'success', f'清理了 {len(cleaned_paths)} 個檔案，共 {total_cleaned_count} 個單元格', duration)
            log_activity(user['id'], user['username'], 'cleanup_success',
                       f"Forecast 多檔案數據清理成功，{len(cleaned_paths)} 個檔案，共 {total_cleaned_count} 個單元格", get_client_ip(), request.headers.get('User-Agent'))

            return jsonify({
                'success': True,
                'message': f'Forecast數據清理完成，清理了 {len(cleaned_paths)} 個檔案，共 {total_cleaned_count} 個單元格',
                'multi_file': True,
                'file_count': len(cleaned_paths),
                'files': cleaned_files_info,
                'total_cleaned_cells': total_cleaned_count
            })

        # 單檔案模式（合併模式或原本就只上傳一個檔案）
        else:
            if not forecast_file or not os.path.exists(forecast_file):
                log_process(user['id'], 'cleanup', 'failed', '請先上傳Forecast文件')
                return jsonify({'success': False, 'message': '請先上傳Forecast文件'})

            # 檢查檔案格式
            is_xls = is_xls_format(forecast_file)

            print("開始清理Forecast數據，保持原始格式...")

            if is_xls:
                # ========== .xls 格式：使用 LibreOffice 跨平台方案保留格式 ==========
                # 注意：輸出會是 .xlsx 格式（避免轉換回 .xls 失敗）
                print(f"ℹ️ 檔案為 .xls 格式，使用 LibreOffice 處理以保留格式和公式")
                cleaned_file = os.path.join(processed_folder, 'cleaned_forecast.xlsx')
                cleaned_count = cleanup_xls_file(forecast_file, cleaned_file, username)
            else:
                # ========== .xlsx 格式：使用 openpyxl 保持格式 ==========
                wb = load_workbook(forecast_file)
                ws = wb.active

                # 清理數據
                cleaned_count = 0

                if username == 'delta':
                    # ========== Delta 專屬清理邏輯 ==========
                    # I欄="Supply" 的整列，從 J 欄到最後一欄清零
                    for row_idx in range(1, ws.max_row + 1):
                        i_cell = ws.cell(row=row_idx, column=9)  # I欄 = column 9
                        if i_cell.value and str(i_cell.value).strip() == 'Supply':
                            for col_idx in range(10, ws.max_column + 1):  # J=10 到最後
                                cell = ws.cell(row=row_idx, column=col_idx)
                                if cell.value is not None and cell.value != 0:
                                    cell.value = 0
                                    cleaned_count += 1

                elif username == 'liteon':
                    # ========== liteon 專屬清理邏輯 ==========
                    # 指定讀取 Daily+Weekly+Monthly sheet
                    # 清理條件：C欄(column 3) = "Commit" 時，清零 J~BY 欄(column 10~77)
                    if 'Daily+Weekly+Monthly' in wb.sheetnames:
                        ws = wb['Daily+Weekly+Monthly']
                    for row_idx in range(1, ws.max_row + 1):
                        c_cell = ws.cell(row=row_idx, column=3)
                        if c_cell.value and str(c_cell.value).strip() == "Commit":
                            for col_idx in range(10, min(78, ws.max_column + 1)):  # J=10, BY=77
                                cell = ws.cell(row=row_idx, column=col_idx)
                                if cell.value is not None and cell.value != 0:
                                    cell.value = 0
                                    cleaned_count += 1
                else:
                    for row_idx in range(1, ws.max_row + 1):
                        # ========== pegatron 專屬清理邏輯 ==========
                        if username == 'pegatron':
                            # 檢查M欄位（第13列）是否為 "ETA QTY"
                            m_cell = ws.cell(row=row_idx, column=13)
                            if m_cell.value and str(m_cell.value).strip() == "ETA QTY":
                                # 清空N~DN欄位（第14列到第118列）設為 0
                                for col_idx in range(14, min(119, ws.max_column + 1)):
                                    cell = ws.cell(row=row_idx, column=col_idx)
                                    if cell.value is not None and cell.value != 0:
                                        cell.value = 0
                                        cleaned_count += 1
                        # ========== quanta 原有清理邏輯 ==========
                        else:
                            # 檢查K欄位（第11列）是否為"供應數量"
                            k_cell = ws.cell(row=row_idx, column=11)
                            if k_cell.value and str(k_cell.value) == "供應數量":
                                # 清空L~AW欄位（第12列到第49列）
                                for col_idx in range(12, min(50, ws.max_column + 1)):
                                    cell = ws.cell(row=row_idx, column=col_idx)
                                    if cell.value != 0:
                                        cell.value = 0
                                        cleaned_count += 1

                            # 檢查I欄位（第9列）是否包含"庫存數量"
                            i_cell = ws.cell(row=row_idx, column=9)
                            if i_cell.value and "庫存數量" in str(i_cell.value):
                                next_row_i_cell = ws.cell(row=row_idx + 1, column=9)
                                if next_row_i_cell.value != 0:
                                    next_row_i_cell.value = 0
                                    cleaned_count += 1

                # 保存清理後的文件
                cleaned_file = os.path.join(processed_folder, 'cleaned_forecast.xlsx')
                wb.save(cleaned_file)

            duration = time.time() - start_time
            print(f"Forecast數據清理完成，清理了 {cleaned_count} 個單元格")

            # 記錄處理成功
            log_process(user['id'], 'cleanup', 'success', f'清理了 {cleaned_count} 個單元格', duration)
            log_activity(user['id'], user['username'], 'cleanup_success',
                       f"Forecast 數據清理成功，清理了 {cleaned_count} 個單元格", get_client_ip(), request.headers.get('User-Agent'))

            return jsonify({
                'success': True,
                'message': f'Forecast數據清理完成，清理了 {cleaned_count} 個單元格',
                'multi_file': False,
                'file': 'cleaned_forecast.xlsx',
                'cleaned_cells': cleaned_count
            })

    except Exception as e:
        duration = time.time() - start_time
        print(f"Forecast數據清理失敗: {str(e)}")
        log_process(user['id'], 'cleanup', 'failed', str(e), duration)
        log_activity(user['id'], user['username'], 'cleanup_failed',
                   f"Forecast 數據清理失敗：{str(e)}", get_client_ip(), request.headers.get('User-Agent'))
        return jsonify({'success': False, 'message': f'數據清理失敗: {str(e)}'})

@app.route('/process_erp_mapping', methods=['POST'])
@login_required
def process_erp_mapping():
    user = get_current_user()
    start_time = time.time()
    try:
        # 檢查是否為測試模式（silent=True 避免沒有 JSON body 時報錯）
        data = request.get_json(silent=True) or {}
        test_mode = data.get('test_mode', False)
        test_customer_id = data.get('customer_id')

        # 獲取在途文件是否必填的參數（預設為必填）
        transit_required = data.get('transit_required', True)
        print(f"📋 Transit Required: {transit_required}")
        print(f"🔍 /process_erp_mapping: test_mode={test_mode}, test_customer_id={test_customer_id}")

        # 獲取前端傳來的 upload_session_id
        upload_session_id = data.get('upload_session_id')
        print(f"📥 Mapping - 前端傳來的 session_id: {upload_session_id}")

        # 如果有前端傳來的 session_id，先同步到 Flask session
        if upload_session_id:
            session['current_session_timestamp'] = upload_session_id
            session.modified = True
            print(f"📁 已同步 session_id 到 Flask session: {upload_session_id}")

        # 決定使用哪個用戶的 mapping 資料
        # IT 測試模式：使用客戶的 mapping 資料，但檔案放在 IT 人員資料夾
        mapping_user_id = user['id']
        if test_mode and test_customer_id and (user['role'] in ['admin', 'it']):
            mapping_user_id = test_customer_id
            print(f"[IT測試模式] Mapping: 使用客戶 ID {test_customer_id} 的 mapping 資料，檔案放在 IT 人員 (ID: {user['id']}) 資料夾")

        # 記錄開始處理（使用 IT 人員的身份記錄 log）
        log_activity(user['id'], user['username'], 'mapping_start',
                   f"開始 ERP 和在途數據整合{' (IT測試模式, 客戶ID: ' + str(test_customer_id) + ')' if test_mode else ''}", get_client_ip(), request.headers.get('User-Agent'))

        # 計算檔案路徑（IT 測試模式下使用 IT 人員的資料夾）
        session_timestamp = session.get('current_session_timestamp')
        # IT 測試模式：檔案在 IT 人員資料夾
        upload_folder = os.path.join(UPLOAD_FOLDER, str(user['id']), session_timestamp)
        erp_file = find_file_with_extensions(upload_folder, 'erp_data')
        transit_file = find_file_with_extensions(upload_folder, 'transit_data')

        mapping_excel_file = os.path.join('mapping', 'mapping表.xlsx')

        # 除錯日誌
        print(f"[映射整合] 檢查檔案路徑:")
        print(f"  - ERP 檔案路徑: {erp_file}")
        print(f"  - Transit 檔案路徑: {transit_file}")
        print(f"  - ERP 檔案存在: {os.path.exists(erp_file) if erp_file else False}")
        print(f"  - Transit 檔案存在: {os.path.exists(transit_file) if transit_file else False}")
        print(f"  - Session keys: {list(session.keys())}")

        if not erp_file or not os.path.exists(erp_file):
            log_process(user['id'], 'mapping', 'failed', '請先上傳ERP文件')
            return jsonify({'success': False, 'message': '請先上傳ERP文件'})

        # 在途文件檢查（根據 transit_required 決定）
        transit_skipped = False
        if not transit_file or not os.path.exists(transit_file):
            if transit_required:
                log_process(user['id'], 'mapping', 'failed', '請先上傳在途文件')
                return jsonify({'success': False, 'message': '請先上傳在途文件'})
            else:
                # 在途文件為選填且未上傳，標記為跳過
                transit_skipped = True
                print("⏭️ 在途文件為選填且未上傳，跳過在途整合")

        # 檢查 mapping 資料來源（優先資料庫）
        mapping_data = None

        # 1. 優先從資料庫讀取（使用 mapping_user_id）
        if has_customer_mappings(mapping_user_id):
            print(f"從資料庫讀取用戶 ID {mapping_user_id} 的 mapping 資料...")
            mapping_data = get_customer_mappings(mapping_user_id)

        # 2. 如果資料庫沒有，嘗試從 JSON 檔案讀取（向後相容）
        if not mapping_data:
            mapping_file = os.path.join('mapping', 'mapping_data.json')
            if os.path.exists(mapping_file):
                print("從 mapping_data.json 讀取映射資料...")
                with open(mapping_file, 'r', encoding='utf-8') as f:
                    mapping_data = json.load(f)

        # 3. 檢查是否有 mapping 資料
        if not mapping_data and not os.path.exists(mapping_excel_file):
            log_process(user['id'], 'mapping', 'failed', '請先配置映射表')
            return jsonify({'success': False, 'message': '請先配置映射表'})

        # === 1. 處理 ERP 數據整合 ===
        print("開始整合 ERP 數據...")
        erp_df = pd.read_excel(erp_file)

        # 標準化排程出貨日期欄位（處理文字/日期格式不一致問題）
        if '排程出貨日期' in erp_df.columns:
            print("🔧 標準化排程出貨日期欄位格式...")

            # 顯示處理前的數據類型統計
            date_col = erp_df['排程出貨日期']
            print(f"   處理前數據類型統計:")
            print(f"   - 總記錄數: {len(date_col)}")
            print(f"   - 非空記錄數: {date_col.notna().sum()}")
            print(f"   - 字串類型: {sum(isinstance(x, str) for x in date_col if pd.notna(x))}")
            print(f"   - 日期類型: {sum(isinstance(x, (datetime, pd.Timestamp)) for x in date_col if pd.notna(x))}")

            # 標準化日期格式
            erp_df['排程出貨日期'] = erp_df['排程出貨日期'].apply(normalize_date_for_mapping)

            # 顯示處理後的結果
            processed_col = erp_df['排程出貨日期']
            print(f"   處理後結果:")
            print(f"   - 成功標準化: {processed_col.notna().sum()}")
            print(f"   - 標準化失敗: {processed_col.isna().sum()}")
            if processed_col.notna().sum() > 0:
                sample_dates = processed_col.dropna().head(3).tolist()
                print(f"   - 範例日期: {sample_dates}")

            print("✅ 排程出貨日期欄位已標準化")

        # 找到客戶簡稱欄位
        customer_col = None
        for col in erp_df.columns:
            if '客戶' in str(col) and '簡稱' in str(col):
                customer_col = col
                break

        if customer_col is None:
            return jsonify({'success': False, 'message': 'ERP文件找不到客戶簡稱欄位'})

        # 判斷客戶類型（使用不同的映射邏輯）
        # IT 測試模式下，根據目標客戶 ID 判斷
        if test_mode and test_customer_id:
            from database import get_user_by_id
            target_user = get_user_by_id(test_customer_id)
            target_username = target_user.get('username', '').lower() if target_user else ''
            is_pegatron = target_username == 'pegatron'
            is_liteon = target_username == 'liteon'
            is_delta = target_username == 'delta'
            print(f"IT: test_customer_id={test_customer_id}, is_pegatron={is_pegatron}, is_liteon={is_liteon}, is_delta={is_delta}")
        else:
            target_username = user['username'].lower()
            is_pegatron = target_username == 'pegatron'
            is_liteon = target_username == 'liteon'
            is_delta = target_username == 'delta'
            print(f"user={user['username']}, is_pegatron={is_pegatron}, is_liteon={is_liteon}, is_delta={is_delta}")

        if is_pegatron:
            # === Pegatron 專用映射邏輯 ===
            # ERP 匹配規則: D欄(客戶簡稱) + M欄(Line 客戶採購單號)前4字 + AG欄(送貨地點)
            print("🔧 使用 Pegatron 專用映射邏輯...")

            # 取得原始 mapping 記錄
            mapping_records = get_customer_mappings_raw(mapping_user_id)
            print(f"   取得 {len(mapping_records)} 筆 mapping 記錄")

            # 建立兩種 lookup:
            # 1. 完整匹配: (customer_name, region, delivery_location) -> mapping values
            # 2. 簡化匹配: (customer_name, region) -> mapping values (當 delivery_location 為空時)
            pegatron_mapping_lookup = {}  # 完整 3 欄位匹配
            pegatron_mapping_lookup_simple = {}  # 簡化 2 欄位匹配（不需要送貨地點）
            for m in mapping_records:
                customer_name = str(m['customer_name']).strip() if m['customer_name'] else ''
                region = str(m['region']).strip() if m['region'] else ''
                delivery_location = str(m['delivery_location']).strip() if m['delivery_location'] else ''

                mapping_values = {
                    'region': region,
                    'schedule_breakpoint': str(m['schedule_breakpoint']).strip() if m['schedule_breakpoint'] else '',
                    'etd': str(m['etd']).strip() if m['etd'] else '',
                    'eta': str(m['eta']).strip() if m['eta'] else ''
                }

                if delivery_location:
                    # 有送貨地點的用完整 3 欄位匹配
                    key = (customer_name, region, delivery_location)
                    pegatron_mapping_lookup[key] = mapping_values
                else:
                    # 沒有送貨地點的用簡化 2 欄位匹配
                    key = (customer_name, region)
                    pegatron_mapping_lookup_simple[key] = mapping_values

            print(f"   建立 {len(pegatron_mapping_lookup)} 筆完整匹配 + {len(pegatron_mapping_lookup_simple)} 筆簡化匹配")

            # 動態查找必要欄位
            line_po_col, err = find_column_by_name(erp_df, 'Line 客戶採購單號')
            if err:
                return jsonify({'success': False, 'message': f'Pegatron ERP {err}'})

            delivery_col, err = find_column_by_name(erp_df, '送貨地點')
            if err:
                return jsonify({'success': False, 'message': f'Pegatron ERP {err}'})

            # 應用 Pegatron 映射
            def get_pegatron_mapping(row, field):
                customer = str(row[customer_col]).strip() if pd.notna(row[customer_col]) else ''
                line_po = str(row[line_po_col]).strip() if pd.notna(row[line_po_col]) else ''
                delivery = str(row[delivery_col]).strip() if pd.notna(row[delivery_col]) else ''

                # 取 Line 客戶採購單號的前 4 字作為 region key
                region_key = line_po[:4] if len(line_po) >= 4 else line_po

                # 先嘗試完整 3 欄位匹配
                key_full = (customer, region_key, delivery)
                mapping = pegatron_mapping_lookup.get(key_full)

                # 如果完整匹配失敗，嘗試簡化 2 欄位匹配（不需要送貨地點）
                if not mapping:
                    key_simple = (customer, region_key)
                    mapping = pegatron_mapping_lookup_simple.get(key_simple, {})

                return mapping.get(field, '') if mapping else ''

            erp_df['客戶需求地區'] = erp_df.apply(lambda row: get_pegatron_mapping(row, 'region'), axis=1)
            erp_df['排程出貨日期斷點'] = erp_df.apply(lambda row: get_pegatron_mapping(row, 'schedule_breakpoint'), axis=1)
            erp_df['ETD'] = erp_df.apply(lambda row: get_pegatron_mapping(row, 'etd'), axis=1)
            erp_df['ETA'] = erp_df.apply(lambda row: get_pegatron_mapping(row, 'eta'), axis=1)

            # 統計匹配結果
            matched_count = (erp_df['客戶需求地區'] != '').sum()
            print(f"   ✅ Pegatron ERP 映射完成: {matched_count}/{len(erp_df)} 行匹配成功")
        elif is_liteon:
            # === Liteon 專用映射邏輯 ===
            # 11一般訂單: 送貨地點(AG) + 客戶簡稱(D) + 訂單型態(AM) = "11"
            # 32HUB補貨單: 倉庫(AL) + 客戶簡稱(D) + 訂單型態(AM) = "32"
            print("使用 Liteon 專用映射邏輯...")

            mapping_records = get_customer_mappings_raw(mapping_user_id)
            print(f"   取得 {len(mapping_records)} 筆 mapping 記錄")

            # 建立兩種 lookup (key 全部轉小寫/去空白 做模糊比對)
            # type 11: (customer_name, delivery_location, "11") -> mapping values
            # type 32: (customer_name, warehouse, "32") -> mapping values
            liteon_lookup_11 = {}
            liteon_lookup_32 = {}
            for m in mapping_records:
                cname = str(m['customer_name']).strip() if m['customer_name'] else ''
                order_type = str(m.get('order_type', '')).strip()
                delivery_loc = str(m.get('delivery_location', '')).strip() if m.get('delivery_location') else ''
                warehouse = str(m.get('warehouse', '')).strip() if m.get('warehouse') else ''

                mapping_values = {
                    'region': str(m['region']).strip() if m['region'] else '',
                    'schedule_breakpoint': str(m['schedule_breakpoint']).strip() if m['schedule_breakpoint'] else '',
                    'etd': str(m['etd']).strip() if m['etd'] else '',
                    'eta': str(m['eta']).strip() if m['eta'] else '',
                    'date_calc_type': str(m.get('date_calc_type', '')).strip() if m.get('date_calc_type') else ''
                }

                if order_type == '11' and delivery_loc:
                    key = (cname, delivery_loc)
                    liteon_lookup_11[key] = mapping_values
                elif order_type == '32' and warehouse:
                    key = (cname, warehouse)
                    liteon_lookup_32[key] = mapping_values

            print(f"   建立 {len(liteon_lookup_11)} 筆 type-11 lookup + {len(liteon_lookup_32)} 筆 type-32 lookup")

            # 動態查找 ERP 欄位
            delivery_col, err = find_column_by_name(erp_df, '送貨地點')
            if err:
                return jsonify({'success': False, 'message': f'Liteon ERP {err}'})

            order_type_col, err = find_column_by_name(erp_df, '訂單型態')
            if err:
                return jsonify({'success': False, 'message': f'Liteon ERP {err}'})

            # 倉庫欄位 (可能不存在，嘗試查找)
            warehouse_col, _ = find_column_by_name(erp_df, '倉庫', required=False)

            print(f"   送貨地點欄位: {delivery_col}")
            print(f"   訂單型態欄位: {order_type_col}")
            print(f"   倉庫欄位: {warehouse_col}")

            # 應用 Liteon 映射
            def get_liteon_mapping(row, field):
                customer = str(row[customer_col]).strip() if pd.notna(row[customer_col]) else ''
                order_type_val = str(row[order_type_col]).strip() if pd.notna(row[order_type_col]) else ''

                # 提取訂單型態前綴: "11一般訂單" -> "11", "32HUB補貨單" -> "32"
                ot_prefix = order_type_val[:2] if len(order_type_val) >= 2 else order_type_val

                if ot_prefix == '11':
                    delivery = str(row[delivery_col]).strip() if pd.notna(row[delivery_col]) else ''
                    key = (customer, delivery)
                    mapping = liteon_lookup_11.get(key, {})
                elif ot_prefix == '32':
                    wh = str(row[warehouse_col]).strip() if warehouse_col and pd.notna(row[warehouse_col]) else ''
                    key = (customer, wh)
                    mapping = liteon_lookup_32.get(key, {})
                else:
                    mapping = {}

                return mapping.get(field, '') if mapping else ''

            erp_df['客戶需求地區'] = erp_df.apply(lambda row: get_liteon_mapping(row, 'region'), axis=1)
            erp_df['排程出貨日期斷點'] = erp_df.apply(lambda row: get_liteon_mapping(row, 'schedule_breakpoint'), axis=1)
            erp_df['ETD'] = erp_df.apply(lambda row: get_liteon_mapping(row, 'etd'), axis=1)
            erp_df['ETA'] = erp_df.apply(lambda row: get_liteon_mapping(row, 'eta'), axis=1)
            erp_df['日期算法'] = erp_df.apply(lambda row: get_liteon_mapping(row, 'date_calc_type'), axis=1)

            matched_count = (erp_df['客戶需求地區'] != '').sum()
            print(f"   Liteon ERP 映射完成: {matched_count}/{len(erp_df)} 行匹配成功")
        elif is_delta:
            # === Delta 專用 ERP 映射邏輯 ===
            # 匹配規則: D欄(客戶簡稱) + AG欄(送貨地點) → mapping 表
            # 帶入: 客戶需求地區、排程出貨日期斷點、ETD、ETA
            print("🔧 使用 Delta 專用 ERP 映射邏輯...")

            mapping_records = get_customer_mappings_raw(mapping_user_id)
            print(f"   取得 {len(mapping_records)} 筆 mapping 記錄")

            # 建立 lookup: (customer_name, delivery_location) -> mapping values
            delta_erp_lookup = {}
            for m in mapping_records:
                cname = str(m['customer_name']).strip() if m['customer_name'] else ''
                dl = str(m.get('delivery_location', '')).strip() if m.get('delivery_location') else ''
                mapping_values = {
                    'region': str(m['region']).strip() if m['region'] else '',
                    'schedule_breakpoint': str(m['schedule_breakpoint']).strip() if m['schedule_breakpoint'] else '',
                    'etd': str(m['etd']).strip() if m['etd'] else '',
                    'eta': str(m['eta']).strip() if m['eta'] else '',
                }
                if cname and dl:
                    delta_erp_lookup[(cname, dl)] = mapping_values

            print(f"   建立 {len(delta_erp_lookup)} 筆 (客戶簡稱, 送貨地點) lookup")

            # 動態查找 ERP 送貨地點欄位
            delivery_col, err = find_column_by_name(erp_df, '送貨地點')
            if err:
                return jsonify({'success': False, 'message': f'Delta ERP {err}'})

            print(f"   客戶簡稱欄位: {customer_col}")
            print(f"   送貨地點欄位: {delivery_col}")

            # 應用 Delta 映射
            def get_delta_erp_mapping(row, field):
                cust = str(row[customer_col]).strip() if pd.notna(row[customer_col]) else ''
                deliv = str(row[delivery_col]).strip() if pd.notna(row[delivery_col]) else ''
                mapping = delta_erp_lookup.get((cust, deliv), {})
                return mapping.get(field, '')

            erp_df['客戶需求地區'] = erp_df.apply(lambda row: get_delta_erp_mapping(row, 'region'), axis=1)
            erp_df['排程出貨日期斷點'] = erp_df.apply(lambda row: get_delta_erp_mapping(row, 'schedule_breakpoint'), axis=1)
            erp_df['ETD'] = erp_df.apply(lambda row: get_delta_erp_mapping(row, 'etd'), axis=1)
            erp_df['ETA'] = erp_df.apply(lambda row: get_delta_erp_mapping(row, 'eta'), axis=1)

            matched_count = (erp_df['客戶需求地區'] != '').sum()
            print(f"   ✅ Delta ERP 映射完成: {matched_count}/{len(erp_df)} 行匹配成功")
        else:
            # === 原有映射邏輯（Quanta 等其他客戶）===
            # 應用映射到 ERP（使用客戶簡稱單欄位匹配）
            erp_df['客戶需求地區'] = erp_df[customer_col].map(mapping_data.get('regions', {}))
            erp_df['排程出貨日期斷點'] = erp_df[customer_col].map(mapping_data.get('schedule_breakpoints', {}))
            erp_df['ETD'] = erp_df[customer_col].map(mapping_data.get('etd', {}))
            erp_df['ETA'] = erp_df[customer_col].map(mapping_data.get('eta', {}))

        # 按排程出貨日期排序
        if '排程出貨日期' in erp_df.columns:
            erp_df = erp_df.sort_values('排程出貨日期')

        # 新增「已分配」欄位（用於 1 對 1 分配邏輯追蹤）
        erp_df['已分配'] = ''
        print(f"📋 ERP 新增「已分配」欄位，目前欄位數: {len(erp_df.columns)}，欄位列表: {list(erp_df.columns)}")

        # 使用資料夾管理結構：processed/{user_id}/{session_timestamp}/integrated_erp.xlsx
        processed_folder, session_timestamp = get_or_create_session_folder(user['id'], 'processed')

        # 保存整合後的 ERP 文件
        integrated_erp_file = os.path.join(processed_folder, 'integrated_erp.xlsx')
        erp_df.to_excel(integrated_erp_file, index=False)
        print(f"✅ ERP數據整合完成: {len(erp_df)} 行 (session: {session_timestamp})")
        
        # === 2. 處理 Transit 數據整合 ===
        transit_rows = 0
        if transit_skipped:
            print("⏭️ 跳過在途數據整合（選填項目）")
        elif is_pegatron:
            # === Pegatron 專用 Transit 映射邏輯 ===
            print("開始整合在途數據...")
            transit_df = pd.read_excel(transit_file)
            # Transit 匹配規則: L欄(Line 客戶採購單號) + E欄(Ordered Item) 匹配 ERP 的 M欄 + N欄
            print("🔧 使用 Pegatron 專用 Transit 映射邏輯...")

            # 動態查找 Transit 必要欄位
            transit_ordered_item_col, err = find_column_by_name(transit_df, 'Ordered Item')
            if err:
                return jsonify({'success': False, 'message': f'Pegatron Transit {err}'})

            transit_line_po_col, err = find_column_by_name(transit_df, 'Line 客戶採購單號')
            if err:
                return jsonify({'success': False, 'message': f'Pegatron Transit {err}'})

            print(f"   Transit Ordered Item 欄位: {transit_ordered_item_col}")
            print(f"   Transit Line 客戶採購單號欄位: {transit_line_po_col}")

            # 動態查找 ERP 欄位（用於 Transit 匹配）
            erp_line_po_col, err = find_column_by_name(erp_df, 'Line 客戶採購單號')
            if err:
                return jsonify({'success': False, 'message': f'ERP {err}'})

            erp_pn_col, err = find_column_by_name(erp_df, '客戶料號')
            if err:
                return jsonify({'success': False, 'message': f'ERP {err}'})

            print(f"   ERP Line 客戶採購單號欄位: {erp_line_po_col}")
            print(f"   ERP 客戶料號欄位: {erp_pn_col}")

            # 建立 ERP lookup: (Line 客戶採購單號, 客戶料號) -> mapping values
            erp_lookup = {}
            for idx, row in erp_df.iterrows():
                line_po = str(row[erp_line_po_col]).strip() if pd.notna(row[erp_line_po_col]) else ''
                pn = str(row[erp_pn_col]).strip() if pd.notna(row[erp_pn_col]) else ''

                if line_po and pn:
                    key = (line_po, pn)
                    if key not in erp_lookup:  # 保留第一筆匹配
                        erp_lookup[key] = {
                            'region': str(row.get('客戶需求地區', '')).strip() if pd.notna(row.get('客戶需求地區', '')) else '',
                            'schedule_breakpoint': str(row.get('排程出貨日期斷點', '')).strip() if pd.notna(row.get('排程出貨日期斷點', '')) else '',
                            'etd': str(row.get('ETD', '')).strip() if pd.notna(row.get('ETD', '')) else '',
                            'eta': str(row.get('ETA', '')).strip() if pd.notna(row.get('ETA', '')) else ''
                        }

            print(f"   建立 ERP lookup: {len(erp_lookup)} 筆")

            # 應用 Transit 映射
            def get_pegatron_transit_mapping(row, field):
                line_po = str(row[transit_line_po_col]).strip() if pd.notna(row[transit_line_po_col]) else ''
                ordered_item = str(row[transit_ordered_item_col]).strip() if pd.notna(row[transit_ordered_item_col]) else ''

                key = (line_po, ordered_item)
                mapping = erp_lookup.get(key, {})
                return mapping.get(field, '')

            transit_df['客戶需求地區'] = transit_df.apply(lambda row: get_pegatron_transit_mapping(row, 'region'), axis=1)
            transit_df['排程出貨日期斷點'] = transit_df.apply(lambda row: get_pegatron_transit_mapping(row, 'schedule_breakpoint'), axis=1)
            transit_df['ETD'] = transit_df.apply(lambda row: get_pegatron_transit_mapping(row, 'etd'), axis=1)
            transit_df['ETA_mapping'] = transit_df.apply(lambda row: get_pegatron_transit_mapping(row, 'eta'), axis=1)

            # 統計匹配結果
            matched_count = (transit_df['客戶需求地區'] != '').sum()
            print(f"   ✅ Pegatron Transit 映射完成: {matched_count}/{len(transit_df)} 行匹配成功")

            # 新增「已分配」欄位（用於 1 對 1 分配邏輯追蹤）
            transit_df['已分配'] = ''
            print(f"📋 Transit 新增「已分配」欄位，目前欄位數: {len(transit_df.columns)}")

            # 保存整合後的 Transit 文件（使用同一個 session 資料夾）
            integrated_transit_file = os.path.join(processed_folder, 'integrated_transit.xlsx')
            transit_df.to_excel(integrated_transit_file, index=False)
            print(f"✅ 在途數據整合完成: {len(transit_df)} 行 (session: {session_timestamp})")
            transit_rows = len(transit_df)
        elif is_liteon:
            # === Liteon Transit 映射邏輯 ===
            # Transit 檔案已包含: K=訂單型態, L=送貨地點, M=倉庫
            # K=11 → 用 L(送貨地點) 查 mapping → 客戶需求地區
            # K=32 → 用 M(倉庫) 查 mapping → 客戶需求地區
            print("使用 Liteon Transit 映射邏輯...")
            transit_df = pd.read_excel(transit_file)

            # 從 mapping 表建立兩個 lookup: 送貨地點->region, 倉庫->region
            dl_to_region = {}
            wh_to_region = {}
            liteon_transit_mappings = get_customer_mappings_raw(mapping_user_id)
            for m in liteon_transit_mappings:
                ot = str(m.get('order_type', '')).strip()
                dl = str(m.get('delivery_location', '')).strip() if m.get('delivery_location') else ''
                wh = str(m.get('warehouse', '')).strip() if m.get('warehouse') else ''
                region = str(m['region']).strip() if m['region'] else ''
                if ot == '11' and dl:
                    dl_to_region[dl] = region
                elif ot == '32' and wh:
                    wh_to_region[wh] = region

            print(f"   送貨地點 lookup: {len(dl_to_region)} 筆, 倉庫 lookup: {len(wh_to_region)} 筆")

            # 動態查找 Transit 欄位（用欄位名稱，不再用 hardcoded index）
            transit_ot_col, err = find_column_by_name(transit_df, '訂單型態')
            if err:
                return jsonify({'success': False, 'message': f'Liteon Transit {err}'})

            transit_dl_col, err = find_column_by_name(transit_df, '送貨地點')
            if err:
                return jsonify({'success': False, 'message': f'Liteon Transit {err}'})

            transit_wh_col, _ = find_column_by_name(transit_df, '倉庫', required=False)

            print(f"   Transit 訂單型態欄位: {transit_ot_col}")
            print(f"   Transit 送貨地點欄位: {transit_dl_col}")
            print(f"   Transit 倉庫欄位: {transit_wh_col}")

            def get_liteon_transit_region(row):
                ot_val = str(row[transit_ot_col]).strip() if pd.notna(row[transit_ot_col]) else ''
                ot_prefix = ot_val[:2] if len(ot_val) >= 2 else ot_val

                if ot_prefix == '11':
                    dl_val = str(row[transit_dl_col]).strip() if pd.notna(row[transit_dl_col]) else ''
                    return dl_to_region.get(dl_val, '')
                elif ot_prefix == '32':
                    wh_val = str(row[transit_wh_col]).strip() if transit_wh_col and pd.notna(row[transit_wh_col]) else ''
                    return wh_to_region.get(wh_val, '')
                else:
                    return ''

            transit_df['客戶需求地區'] = transit_df.apply(get_liteon_transit_region, axis=1)

            matched_count = (transit_df['客戶需求地區'] != '').sum()
            print(f"   Liteon Transit 映射完成: {matched_count}/{len(transit_df)} 行匹配成功")

            transit_df['已分配'] = ''
            integrated_transit_file = os.path.join(processed_folder, 'integrated_transit.xlsx')
            transit_df.to_excel(integrated_transit_file, index=False)
            print(f"在途數據整合完成: {len(transit_df)} 行 (session: {session_timestamp})")
            transit_rows = len(transit_df)
        elif is_delta:
            # === Delta Transit 映射邏輯 ===
            # 匹配規則: D欄(送貨地點) + K欄(客戶簡稱) → mapping 表 → 客戶需求地區
            print("🔧 使用 Delta Transit 映射邏輯...")
            transit_df = pd.read_excel(transit_file)

            # 建立 lookup: (customer_name, delivery_location) -> region
            delta_transit_lookup = {}
            delta_mappings = get_customer_mappings_raw(mapping_user_id)
            for m in delta_mappings:
                cname = str(m['customer_name']).strip() if m['customer_name'] else ''
                dl = str(m.get('delivery_location', '')).strip() if m.get('delivery_location') else ''
                region = str(m['region']).strip() if m['region'] else ''
                if cname and dl:
                    delta_transit_lookup[(cname, dl)] = region

            print(f"   建立 {len(delta_transit_lookup)} 筆 (客戶簡稱, 送貨地點) lookup")

            # 動態查找 Transit 欄位
            transit_customer_col, err = find_column_by_name(transit_df, ['客戶', '簡稱'])
            if err:
                return jsonify({'success': False, 'message': f'Delta Transit {err}'})

            # Delta Transit 的送貨地點欄位使用英文 'Location' (D 欄)
            transit_dl_col, err = find_column_by_name(transit_df, 'Location')
            if err:
                # 若找不到 Location，再嘗試中文欄位名
                transit_dl_col, err = find_column_by_name(transit_df, '送貨地點')
                if err:
                    return jsonify({'success': False, 'message': f'Delta Transit 找不到送貨地點欄位 (Location)'})

            print(f"   客戶簡稱欄位: {transit_customer_col}")
            print(f"   送貨地點欄位: {transit_dl_col}")

            def get_delta_transit_region(row):
                cust = str(row[transit_customer_col]).strip() if pd.notna(row[transit_customer_col]) else ''
                deliv = str(row[transit_dl_col]).strip() if pd.notna(row[transit_dl_col]) else ''
                return delta_transit_lookup.get((cust, deliv), '')

            transit_df['客戶需求地區'] = transit_df.apply(get_delta_transit_region, axis=1)

            matched_count = (transit_df['客戶需求地區'] != '').sum()
            print(f"   ✅ Delta Transit 映射完成: {matched_count}/{len(transit_df)} 行匹配成功")

            transit_df['已分配'] = ''
            integrated_transit_file = os.path.join(processed_folder, 'integrated_transit.xlsx')
            transit_df.to_excel(integrated_transit_file, index=False)
            print(f"✅ 在途數據整合完成: {len(transit_df)} 行 (session: {session_timestamp})")
            transit_rows = len(transit_df)
        else:
            # === 原有 Transit 映射邏輯（Quanta 等其他客戶）===
            print("開始整合在途數據...")
            transit_df = pd.read_excel(transit_file)
            # 建立 mapping 字典（從 mapping_data 轉換格式）
            mapping_dict = {}
            if mapping_data:
                # 從資料庫格式的 mapping_data 建立字典
                all_customers = set()
                all_customers.update(mapping_data.get('regions', {}).keys())
                all_customers.update(mapping_data.get('schedule_breakpoints', {}).keys())
                all_customers.update(mapping_data.get('etd', {}).keys())
                all_customers.update(mapping_data.get('eta', {}).keys())

                for customer in all_customers:
                    mapping_dict[customer] = {
                        'region': mapping_data.get('regions', {}).get(customer, ''),
                        'schedule_breakpoint': mapping_data.get('schedule_breakpoints', {}).get(customer, ''),
                        'etd': mapping_data.get('etd', {}).get(customer, ''),
                        'eta': mapping_data.get('eta', {}).get(customer, '')
                    }
                print(f"從資料庫/JSON 建立 mapping 字典，共 {len(mapping_dict)} 個客戶")
            elif os.path.exists(mapping_excel_file):
                # 向後相容：從 Excel 讀取 mapping
                mapping_excel_df = pd.read_excel(mapping_excel_file)

                # 動態查找 Mapping 表欄位
                mapping_customer_col, err = find_column_by_name(mapping_excel_df, ['客戶', '簡稱'])
                if err:
                    return jsonify({'success': False, 'message': f'Mapping 表 {err}'})

                mapping_region_col, _ = find_column_by_name(mapping_excel_df, '地區', required=False)
                mapping_schedule_col, _ = find_column_by_name(mapping_excel_df, '斷點', required=False)
                mapping_etd_col, _ = find_column_by_name(mapping_excel_df, 'ETD', required=False)
                mapping_eta_col, _ = find_column_by_name(mapping_excel_df, 'ETA', required=False)

                for idx, row in mapping_excel_df.iterrows():
                    customer = str(row[mapping_customer_col])
                    mapping_dict[customer] = {
                        'region': str(row[mapping_region_col]) if mapping_region_col and pd.notna(row[mapping_region_col]) else '',
                        'schedule_breakpoint': str(row[mapping_schedule_col]) if mapping_schedule_col and pd.notna(row[mapping_schedule_col]) else '',
                        'etd': str(row[mapping_etd_col]) if mapping_etd_col and pd.notna(row[mapping_etd_col]) else '',
                        'eta': str(row[mapping_eta_col]) if mapping_eta_col and pd.notna(row[mapping_eta_col]) else ''
                    }
                print(f"從 Excel 建立 mapping 字典，共 {len(mapping_dict)} 個客戶")

            # 動態查找 Transit 客戶欄位
            transit_customer_col, err = find_column_by_name(transit_df, ['客戶', '簡稱'])
            if err:
                return jsonify({'success': False, 'message': f'在途文件 {err}'})
            print(f"在途文件客戶簡稱欄位: {transit_customer_col}")

            # 新版在途文件結構（12欄位）：
            # 索引0-11: Tw, Ship Number, Invoice Date, Location, 客戶簡稱, Ordered Item, Pj Item, Qty, ETA, Stauts, 集團客戶, 週別
            # 整合後會新增4個欄位到末尾，變成索引12-15：客戶需求地區, 排程出貨日期斷點, ETD, ETA

            # 應用映射到 Transit（新增到文件末尾）
            transit_df['客戶需求地區'] = transit_df[transit_customer_col].apply(
                lambda x: mapping_dict.get(str(x), {}).get('region', '') if pd.notna(x) else ''
            )
            transit_df['排程出貨日期斷點'] = transit_df[transit_customer_col].apply(
                lambda x: mapping_dict.get(str(x), {}).get('schedule_breakpoint', '') if pd.notna(x) else ''
            )
            transit_df['ETD'] = transit_df[transit_customer_col].apply(
                lambda x: mapping_dict.get(str(x), {}).get('etd', '') if pd.notna(x) else ''
            )
            transit_df['ETA_mapping'] = transit_df[transit_customer_col].apply(
                lambda x: mapping_dict.get(str(x), {}).get('eta', '') if pd.notna(x) else ''
            )

            # 顯示整合後的欄位結構
            print(f"✅ 在途數據整合完成，欄位結構:")
            for i, col in enumerate(transit_df.columns):
                print(f"   索引{i}: {col}")
            print(f"   整合後總欄位數: {len(transit_df.columns)}")

            # 新增「已分配」欄位（用於 1 對 1 分配邏輯追蹤）
            transit_df['已分配'] = ''
            print(f"📋 Transit 新增「已分配」欄位，目前欄位數: {len(transit_df.columns)}，最後 3 欄: {list(transit_df.columns[-3:])}")

            # 注意：整合後的結構（總共17個欄位，索引0-16）
            # 索引8: ETA (原始文件中的ETA)
            # 索引12: 客戶需求地區 (整合後新增)
            # 索引13: 排程出貨日期斷點 (整合後新增)
            # 索引14: ETD (整合後新增)
            # 索引15: ETA_mapping (整合後新增，來自mapping表)
            # 索引16: 已分配 (整合後新增，用於1對1分配追蹤)

            # 保存整合後的 Transit 文件（使用同一個 session 資料夾）
            integrated_transit_file = os.path.join(processed_folder, 'integrated_transit.xlsx')
            transit_df.to_excel(integrated_transit_file, index=False)
            print(f"✅ 在途數據整合完成: {len(transit_df)} 行 (session: {session_timestamp})")
            transit_rows = len(transit_df)

        # === Delta Forecast C/D 欄位填入 (Delta 專用) ===
        # 依 PLANT(B) 查 mapping 表的 region，帶入 C(客戶簡稱) 和 D(送貨地點)
        forecast_cd_msg = ''
        if is_delta:
            # 優先使用 cleaned_forecast (Step 2 已清零 Supply)，避免殘留舊值
            cleaned_forecast_file = os.path.join(processed_folder, 'cleaned_forecast.xlsx')
            if os.path.exists(cleaned_forecast_file):
                forecast_file = cleaned_forecast_file
            else:
                forecast_file = find_file_with_extensions(upload_folder, 'forecast_data')
            if forecast_file and os.path.exists(forecast_file):
                print("🔧 開始填入 Delta Forecast C/D 欄位...")

                # 建立 PLANT 代碼 -> (客戶簡稱, 送貨地點) lookup
                plant_to_cd = {}
                for m in get_customer_mappings_raw(mapping_user_id):
                    region = str(m['region']).strip() if m['region'] else ''
                    cname = str(m['customer_name']).strip() if m['customer_name'] else ''
                    dl = str(m.get('delivery_location', '')).strip() if m.get('delivery_location') else ''
                    if region:
                        # region 可能是 "PSB5" 或 "PSB5 泰國"，取第一段作為 PLANT 代碼
                        parts = region.split()
                        plant_code = parts[0] if parts else region
                        if plant_code not in plant_to_cd:
                            plant_to_cd[plant_code] = (cname, dl)

                print(f"   建立 {len(plant_to_cd)} 筆 PLANT → (C, D) lookup")

                # 用 openpyxl 修改 forecast_data.xlsx (保留樣式/公式)
                import openpyxl
                wb_fc = openpyxl.load_workbook(forecast_file)
                ws_fc = wb_fc.active

                matched_rows = 0
                total_rows = 0
                for r in range(2, ws_fc.max_row + 1):
                    plant_val = ws_fc.cell(row=r, column=2).value  # B = PLANT
                    if plant_val is None or plant_val == '':
                        continue
                    total_rows += 1
                    plant_str = str(plant_val).strip()
                    if plant_str in plant_to_cd:
                        cname, dl = plant_to_cd[plant_str]
                        ws_fc.cell(row=r, column=3, value=cname)  # C = 客戶簡稱
                        ws_fc.cell(row=r, column=4, value=dl)     # D = 送貨地點
                        matched_rows += 1

                wb_fc.save(forecast_file)

                # 另存一份到 processed 資料夾作為紀錄
                import shutil
                integrated_forecast_file = os.path.join(processed_folder, 'integrated_forecast.xlsx')
                shutil.copy2(forecast_file, integrated_forecast_file)

                forecast_cd_msg = f', Forecast C/D 填入: {matched_rows}/{total_rows} 行'
                print(f"   ✅ Delta Forecast C/D 填入完成: {matched_rows}/{total_rows} 行匹配")
            else:
                print("⚠️ 未找到 Forecast 檔案，跳過 C/D 填入")

        # 存儲 processed 資料夾路徑到 session
        session['current_processed_folder'] = processed_folder

        duration = time.time() - start_time
        # 記錄處理成功
        transit_log_msg = f'ERP: {len(erp_df)} 行' + (', Transit: 已跳過' if transit_skipped else f', Transit: {transit_rows} 行') + forecast_cd_msg
        log_process(user['id'], 'mapping', 'success', transit_log_msg, duration)
        log_activity(user['id'], user['username'], 'mapping_success',
                   f"ERP 數據整合成功" + ("（在途已跳過）" if transit_skipped else " 和在途數據整合成功"), get_client_ip(), request.headers.get('User-Agent'))

        return jsonify({
            'success': True,
            'message': 'ERP 數據整合完成' + ('（在途已跳過）' if transit_skipped else ' 和在途數據整合完成'),
            'erp_file': 'integrated_erp.xlsx',
            'transit_file': '' if transit_skipped else 'integrated_transit.xlsx',
            'erp_rows': len(erp_df),
            'transit_rows': transit_rows,
            'transit_skipped': transit_skipped
        })
    except Exception as e:
        duration = time.time() - start_time
        print(f"數據整合失敗: {str(e)}")
        import traceback
        traceback.print_exc()
        log_process(user['id'], 'mapping', 'failed', str(e), duration)
        log_activity(user['id'], user['username'], 'mapping_failed',
                   f"ERP 和在途數據整合失敗：{str(e)}", get_client_ip(), request.headers.get('User-Agent'))
        return jsonify({'success': False, 'message': f'數據整合失敗: {str(e)}'})

@app.route('/run_forecast', methods=['POST'])
@login_required
def run_forecast():
    user = get_current_user()
    start_time = time.time()
    try:
        # 檢查是否為測試模式（silent=True 避免沒有 JSON body 時報錯）
        data = request.get_json(silent=True) or {}
        test_mode = data.get('test_mode', False)
        test_customer_id = data.get('customer_id')

        # 獲取在途文件是否必填的參數（預設為必填）
        transit_required = data.get('transit_required', True)
        print(f"📋 Forecast - Transit Required: {transit_required}")

        # 獲取前端傳來的 upload_session_id
        upload_session_id = data.get('upload_session_id')
        print(f"📥 Forecast - 前端傳來的 session_id: {upload_session_id}")

        # 如果有前端傳來的 session_id，先同步到 Flask session
        if upload_session_id:
            session['current_session_timestamp'] = upload_session_id
            session.modified = True
            print(f"📁 已同步 session_id 到 Flask session: {upload_session_id}")

        # IT 測試模式：決定使用哪個處理器（根據被測試客戶的 ID）
        # 但檔案仍放在 IT 人員的資料夾
        processor_user_id = user['id']  # 用於決定使用哪個處理器（Pegatron 或通用）
        if test_mode and test_customer_id and (user['role'] in ['admin', 'it']):
            processor_user_id = int(test_customer_id)
            print(f"[IT測試模式] Forecast: 使用客戶 ID {test_customer_id} 的處理器，檔案放在 IT 人員 (ID: {user['id']}) 資料夾")

        # 記錄開始處理（使用 IT 人員的身份記錄 log）
        log_activity(user['id'], user['username'], 'forecast_start',
                   f"開始 FORECAST 處理{' (IT測試模式, 客戶ID: ' + str(test_customer_id) + ')' if test_mode else ''}", get_client_ip(), request.headers.get('User-Agent'))

        # 計算 processed 資料夾路徑（IT 測試模式下使用 IT 人員的資料夾）
        processed_folder = session.get('current_processed_folder')
        if not processed_folder:
            # 如果有前端傳來的 session_id，直接計算路徑（使用 IT 人員的 user_id）
            if upload_session_id:
                processed_folder = os.path.join(PROCESSED_FOLDER, str(user['id']), upload_session_id)
            else:
                # 嘗試使用 session 資料夾路徑
                processed_folder, _ = get_session_folder_path(user['id'], 'processed')

        if not processed_folder or not os.path.exists(processed_folder):
            log_process(user['id'], 'forecast', 'failed', '請先完成數據清理和整合')
            return jsonify({'success': False, 'message': '請先完成數據清理和整合'})

        # 檢查必要文件是否存在
        integrated_erp = os.path.join(processed_folder, 'integrated_erp.xlsx')
        integrated_transit = os.path.join(processed_folder, 'integrated_transit.xlsx')

        if not os.path.exists(integrated_erp):
            log_process(user['id'], 'forecast', 'failed', '請先完成ERP數據整合')
            return jsonify({'success': False, 'message': '請先完成ERP數據整合'})

        # 檢查是否有 Transit 文件
        has_transit = os.path.exists(integrated_transit)
        # 如果在途為選填且沒有在途文件，標記為跳過
        transit_skipped = not transit_required and not has_transit
        if transit_skipped:
            print("⏭️ 在途數據跳過（選填項目且未上傳）")

        # 根據 processed 資料夾中的檔案來判斷是否為多檔案模式（不依賴 session 標記）
        # 掃描 cleaned_forecast_*.xlsx 或 cleaned_forecast_*.xls 檔案
        print(f"🔍 Forecast - 掃描 processed 資料夾: {processed_folder}")

        multi_cleaned_files = []
        for i in range(1, 100):
            filepath = find_file_with_extensions(processed_folder, f'cleaned_forecast_{i}')
            if filepath:
                multi_cleaned_files.append(filepath)
            else:
                break

        single_cleaned_file = find_file_with_extensions(processed_folder, 'cleaned_forecast')
        has_single_cleaned = single_cleaned_file is not None

        print(f"📁 多檔案數量: {len(multi_cleaned_files)}, 單一清理檔案存在: {has_single_cleaned}")

        # 判斷模式：如果有多個 cleaned_forecast_*.xlsx，就是多檔案模式
        is_multi_file_mode = len(multi_cleaned_files) > 1

        # 執行FORECAST處理
        from ultra_fast_forecast_processor import UltraFastForecastProcessor

        # Pegatron (user_id=5) 使用專用處理器
        # Liteon (user_id=6) 使用專用處理器
        # Delta: 使用 username 判斷（不依賴 hardcoded ID）
        # 在 IT 測試模式下，使用被測試客戶的 ID 來判斷
        is_pegatron = processor_user_id == 5
        is_liteon = processor_user_id == 6
        processor_user = get_user_by_id(processor_user_id)
        is_delta = processor_user and processor_user['username'] == 'delta'

        if is_delta:
            # ===== Delta 專用處理：映射過的 forecast + ERP/Transit 填入 =====
            import shutil
            from delta_forecast_step4 import process_delta_forecast

            # Delta 的來源 forecast：優先使用映射階段產生的 integrated_forecast.xlsx
            integrated_forecast = os.path.join(processed_folder, 'integrated_forecast.xlsx')
            if os.path.exists(integrated_forecast):
                forecast_source = integrated_forecast
            elif has_single_cleaned:
                forecast_source = single_cleaned_file
            elif multi_cleaned_files:
                forecast_source = multi_cleaned_files[0]
            else:
                log_process(user['id'], 'forecast', 'failed', 'Delta: 找不到 forecast 來源檔案')
                return jsonify({'success': False, 'message': 'Delta: 找不到 forecast 來源檔案'})

            output_filename = 'forecast_result.xlsx'
            result_file = os.path.join(processed_folder, output_filename)

            # 先複製一份作為工作檔，避免修改到 integrated_forecast
            shutil.copy2(forecast_source, result_file)

            try:
                stats = process_delta_forecast(
                    forecast_file=result_file,
                    erp_file=integrated_erp,
                    transit_file=integrated_transit if has_transit else None,
                    output_file=result_file,
                )
            except Exception as e:
                import traceback
                traceback.print_exc()
                log_process(user['id'], 'forecast', 'failed', f'Delta Forecast 處理失敗: {str(e)}')
                return jsonify({'success': False, 'message': f'Delta Forecast 處理失敗: {str(e)}'})

            file_size = os.path.getsize(result_file)
            duration = time.time() - start_time

            print(f"=== Delta Forecast 處理完成：耗時 {duration:.2f} 秒 ===")

            log_activity(user['id'], user['username'], 'run_forecast',
                       f"Delta Forecast 處理完成 (ERP 填入 {stats['erp_filled']}, Transit 填入 {stats['transit_filled']})",
                       get_client_ip(), request.headers.get('User-Agent'))

            return jsonify({
                'success': True,
                'message': f'Delta Forecast 處理完成',
                'stats': {
                    'erp_filled': stats['erp_filled'],
                    'erp_skipped': stats['erp_skipped'],
                    'transit_filled': stats['transit_filled'],
                    'transit_skipped': stats['transit_skipped'],
                    'erp_source': f"ERP 匹配 {stats['erp_matched_rows']} 列",
                    'transit_source': f"Transit 匹配 {stats['transit_matched_rows']} 列" if has_transit else '無 Transit'
                },
                'duration': round(duration, 2),
                'file_size': file_size,
                'output_filename': output_filename
            })

        elif is_liteon:
            # ===== Liteon 專用處理：使用 LiteonForecastProcessor =====
            from liteon_forecast_processor import LiteonForecastProcessor

            # 單檔案 fallback：如果沒有 cleaned_forecast_*.xlsx 但有 cleaned_forecast.xlsx
            if not multi_cleaned_files and has_single_cleaned:
                multi_cleaned_files = [single_cleaned_file]

            # 檢查是否為合併模式
            merge_mode = session.get('forecast_merge_mode', False)
            print(f"=== Liteon 模式：{len(multi_cleaned_files)} 個檔案, 合併模式: {merge_mode} ===")

            total_erp_filled = 0
            total_transit_filled = 0
            total_erp_skipped = 0
            total_transit_skipped = 0
            processed_files = []
            failed_files = []

            if merge_mode and len(multi_cleaned_files) > 1:
                # ===== 合併模式：先合併再處理 =====
                print(f"\n=== Liteon 合併模式：合併 {len(multi_cleaned_files)} 個檔案 ===")
                merged_file = os.path.join(processed_folder, 'merged_forecast.xlsx')

                try:
                    total_rows, plant_daily_end_dates = merge_liteon_forecast_files(multi_cleaned_files, merged_file)
                    print(f"合併完成: {total_rows} 列資料")

                    output_filename = 'forecast_merged.xlsx'
                    processor = LiteonForecastProcessor(
                        forecast_file=merged_file,
                        erp_file=integrated_erp,
                        transit_file=integrated_transit if has_transit else None,
                        output_folder=processed_folder,
                        output_filename=output_filename,
                        merged_mode=True,
                        plant_daily_end_dates=plant_daily_end_dates
                    )

                    success = processor.process_all_blocks()

                    if success:
                        result_file = os.path.join(processed_folder, output_filename)
                        if os.path.exists(result_file):
                            file_size = os.path.getsize(result_file)
                            processed_files.append({
                                'input': 'merged_forecast.xlsx',
                                'output': output_filename,
                                'erp_filled': processor.total_filled,
                                'transit_filled': processor.total_transit_filled,
                                'file_size': file_size
                            })
                            total_erp_filled = processor.total_filled
                            total_erp_skipped = processor.total_skipped
                            total_transit_filled = processor.total_transit_filled
                            total_transit_skipped = processor.total_transit_skipped
                            print(f"  ✅ 合併處理成功: ERP填入 {processor.total_filled}, Transit填入 {processor.total_transit_filled}")
                    else:
                        failed_files.append({'input': 'merged_forecast.xlsx', 'error': '合併處理失敗'})

                except Exception as e:
                    print(f"  ❌ 合併處理失敗: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    failed_files.append({'input': 'merged_forecast.xlsx', 'error': str(e)})

            else:
                # ===== 逐檔模式：原有邏輯 =====
                for idx, forecast_file in enumerate(multi_cleaned_files, 1):
                    file_basename = os.path.basename(forecast_file)
                    import re
                    match = re.search(r'cleaned_forecast_(\d+)\.xlsx?', file_basename)
                    file_num = match.group(1) if match else str(idx)

                    # Liteon: 從 Forecast C1 (Plant) + E1 (Buyer Code) 作為檔名
                    try:
                        _tmp_wb = openpyxl.load_workbook(forecast_file, read_only=True)
                        _tmp_ws = _tmp_wb['Daily+Weekly+Monthly']
                        plant_code = str(_tmp_ws.cell(row=1, column=3).value or '').strip()
                        buyer_code = str(_tmp_ws.cell(row=1, column=5).value or '').strip()
                        _tmp_wb.close()
                        if plant_code and buyer_code:
                            output_filename = f'forecast_{plant_code}_{buyer_code}.xlsx'
                        elif plant_code:
                            output_filename = f'forecast_{plant_code}.xlsx'
                        else:
                            output_filename = f'forecast_result_{file_num}.xlsx'
                    except:
                        output_filename = f'forecast_result_{file_num}.xlsx'

                    print(f"\n--- 處理檔案 {idx}/{len(multi_cleaned_files)}: {file_basename} ---")

                    try:
                        processor = LiteonForecastProcessor(
                            forecast_file=forecast_file,
                            erp_file=integrated_erp,
                            transit_file=integrated_transit if has_transit else None,
                            output_folder=processed_folder,
                            output_filename=output_filename
                        )

                        success = processor.process_all_blocks()

                        if success:
                            result_file = os.path.join(processed_folder, processor.output_filename)
                            if os.path.exists(result_file):
                                file_size = os.path.getsize(result_file)
                                processed_files.append({
                                    'input': file_basename,
                                    'output': processor.output_filename,
                                    'erp_filled': processor.total_filled,
                                    'transit_filled': processor.total_transit_filled,
                                    'file_size': file_size
                                })
                                total_erp_filled += processor.total_filled
                                total_erp_skipped += processor.total_skipped
                                total_transit_filled += processor.total_transit_filled
                                total_transit_skipped += processor.total_transit_skipped
                                print(f"  ✅ 成功: ERP填入 {processor.total_filled}, Transit填入 {processor.total_transit_filled}")
                            else:
                                print(f"  ❌ 結果文件未找到: {result_file}")
                                failed_files.append({'input': file_basename, 'error': '結果文件未生成'})
                        else:
                            failed_files.append({'input': file_basename, 'error': '處理失敗'})

                    except Exception as e:
                        print(f"  ❌ 處理失敗: {str(e)}")
                        import traceback
                        traceback.print_exc()
                        failed_files.append({'input': file_basename, 'error': str(e)})

            duration = time.time() - start_time

            if processed_files:
                mode_label = '合併處理' if (merge_mode and len(multi_cleaned_files) > 1) else '多檔案處理'
                log_process(user['id'], 'forecast', 'success',
                          f'Liteon {mode_label}: {len(processed_files)} 成功, ERP填入: {total_erp_filled}, Transit填入: {total_transit_filled}', duration)
                log_activity(user['id'], user['username'], 'forecast_success',
                           f"Liteon FORECAST {mode_label}成功: {len(processed_files)} 個檔案", get_client_ip(), request.headers.get('User-Agent'))

                return jsonify({
                    'success': True,
                    'message': f'FORECAST處理完成：{len(processed_files)} 個檔案',
                    'multi_file': True,
                    'merged_mode': merge_mode and len(multi_cleaned_files) > 1,
                    'files': processed_files,
                    'failed_files': failed_files,
                    'file_count': len(multi_cleaned_files),
                    'success_count': len(processed_files),
                    'total_erp_filled': total_erp_filled,
                    'total_erp_skipped': total_erp_skipped,
                    'total_transit_filled': total_transit_filled,
                    'total_transit_skipped': total_transit_skipped,
                    'transit_file_skipped': transit_skipped
                })
            else:
                log_process(user['id'], 'forecast', 'failed', '所有檔案處理失敗', duration)
                return jsonify({'success': False, 'message': '所有檔案處理失敗'})

        elif is_pegatron:
            # ===== Pegatron 專用處理：使用 PegatronForecastProcessor =====
            from pegatron_forecast_processor import PegatronForecastProcessor

            # 判斷是否為多檔案模式
            if is_multi_file_mode:
                # ===== Pegatron 多檔案模式 =====
                print(f"=== Pegatron 多檔案模式：{len(multi_cleaned_files)} 個檔案 ===")

                total_erp_filled = 0
                total_transit_filled = 0
                processed_files = []
                failed_files = []

                for idx, forecast_file in enumerate(multi_cleaned_files, 1):
                    # 從檔名提取編號
                    file_basename = os.path.basename(forecast_file)
                    import re
                    match = re.search(r'cleaned_forecast_(\d+)\.xlsx?', file_basename)
                    file_num = match.group(1) if match else str(idx)

                    # 從 Forecast 檔案提取 Plant 和 MRP ID 來生成檔名
                    plant, mrp_id = extract_plant_mrp_from_forecast(forecast_file)
                    if plant and mrp_id:
                        output_filename = f'forecast_{plant}_{mrp_id}.xlsx'
                    else:
                        output_filename = f'forecast_result_{file_num}.xlsx'

                    print(f"\n--- 處理檔案 {idx}/{len(multi_cleaned_files)}: {file_basename} ---")

                    try:
                        processor = PegatronForecastProcessor(
                            forecast_file=forecast_file,
                            erp_file=integrated_erp,
                            transit_file=integrated_transit if has_transit else None,
                            output_folder=processed_folder,
                            output_filename=output_filename
                        )

                        success = processor.process_all_blocks()

                        if success:
                            # 使用 processor.output_filename 確保檔名正確
                            result_file = os.path.join(processed_folder, processor.output_filename)
                            if os.path.exists(result_file):
                                file_size = os.path.getsize(result_file)
                                processed_files.append({
                                    'input': file_basename,
                                    'output': processor.output_filename,
                                    'erp_filled': processor.total_filled,
                                    'transit_filled': processor.total_transit_filled,
                                    'file_size': file_size
                                })
                                total_erp_filled += processor.total_filled
                                total_transit_filled += processor.total_transit_filled
                                print(f"  ✅ 成功: ERP填入 {processor.total_filled}, Transit填入 {processor.total_transit_filled}")
                            else:
                                print(f"  ❌ 結果文件未找到: {result_file}")
                                failed_files.append({'input': file_basename, 'error': '結果文件未生成'})
                        else:
                            failed_files.append({'input': file_basename, 'error': '處理失敗'})

                    except Exception as e:
                        print(f"  ❌ 處理失敗: {str(e)}")
                        failed_files.append({'input': file_basename, 'error': str(e)})

                duration = time.time() - start_time

                if processed_files:
                    log_process(user['id'], 'forecast', 'success',
                              f'多檔案處理: {len(processed_files)} 成功, ERP填入: {total_erp_filled}, Transit填入: {total_transit_filled}', duration)
                    log_activity(user['id'], user['username'], 'forecast_success',
                               f"Pegatron FORECAST 多檔案處理成功: {len(processed_files)} 個檔案", get_client_ip(), request.headers.get('User-Agent'))

                    return jsonify({
                        'success': True,
                        'message': f'FORECAST處理完成：{len(processed_files)} 個檔案',
                        'multi_file': True,
                        'files': processed_files,
                        'failed_files': failed_files,
                        'file_count': len(multi_cleaned_files),
                        'success_count': len(processed_files),
                        'total_erp_filled': total_erp_filled,
                        'total_erp_skipped': 0,
                        'total_transit_filled': total_transit_filled,
                        'total_transit_skipped': 0,
                        'transit_file_skipped': transit_skipped
                    })
                else:
                    log_process(user['id'], 'forecast', 'failed', '所有檔案處理失敗', duration)
                    return jsonify({'success': False, 'message': '所有檔案處理失敗'})

            else:
                # ===== Pegatron 單檔案模式 =====
                is_merged_forecast = False
                cleaned_forecast = find_file_with_extensions(processed_folder, 'cleaned_forecast')
                if cleaned_forecast:
                    is_merged_forecast = True  # 這是合併的檔案
                else:
                    cleaned_forecast = find_file_with_extensions(processed_folder, 'cleaned_forecast_1')

                if not cleaned_forecast:
                    log_process(user['id'], 'forecast', 'failed', '請先完成Forecast數據清理')
                    return jsonify({'success': False, 'message': '請先完成Forecast數據清理'})

                # 決定輸出檔名
                if is_merged_forecast:
                    # 合併檔案用 forecast_ALL
                    output_filename = 'forecast_ALL.xlsx'
                else:
                    # 分開檔案用 Plant_MRP ID
                    plant, mrp_id = extract_plant_mrp_from_forecast(cleaned_forecast)
                    if plant and mrp_id:
                        output_filename = f'forecast_{plant}_{mrp_id}.xlsx'
                    else:
                        output_filename = 'forecast_result.xlsx'

                print("開始 Pegatron FORECAST 處理...")
                print(f"清理後的Forecast文件: {cleaned_forecast}")
                print(f"整合後的ERP文件: {integrated_erp}")
                if has_transit:
                    print(f"整合後的Transit文件: {integrated_transit}")

                processor = PegatronForecastProcessor(
                    forecast_file=cleaned_forecast,
                    erp_file=integrated_erp,
                    transit_file=integrated_transit if has_transit else None,
                    output_folder=processed_folder,
                    output_filename=output_filename,
                    is_merged=is_merged_forecast
                )

                success = processor.process_all_blocks()

                if success:
                    # 注意：PegatronForecastProcessor 會強制輸出 .xlsx
                    result_file = os.path.join(processed_folder, processor.output_filename)
                    if os.path.exists(result_file):
                        file_size = os.path.getsize(result_file)
                        duration = time.time() - start_time
                        print(f"Pegatron FORECAST處理完成，結果文件: {result_file} (大小: {file_size} bytes)")

                        log_process(user['id'], 'forecast', 'success',
                                  f'ERP填入: {processor.total_filled}, Transit填入: {processor.total_transit_filled}', duration)
                        log_activity(user['id'], user['username'], 'forecast_success',
                                   f"Pegatron FORECAST 處理成功", get_client_ip(), request.headers.get('User-Agent'))

                        result_data = {
                            'success': True,
                            'message': 'FORECAST處理完成',
                            'file': processor.output_filename,
                            'erp_filled': processor.total_filled,
                            'erp_skipped': processor.total_skipped,
                            'transit_filled': processor.total_transit_filled,
                            'transit_skipped': processor.total_transit_skipped,
                            'transit_file_skipped': transit_skipped,
                            'file_size': file_size
                        }
                        return jsonify(result_data)
                    else:
                        duration = time.time() - start_time
                        print(f"錯誤：結果文件未找到: {result_file}")
                        log_process(user['id'], 'forecast', 'failed', '結果文件未生成', duration)
                        return jsonify({'success': False, 'message': 'FORECAST處理完成但結果文件未生成'})
                else:
                    duration = time.time() - start_time
                    log_process(user['id'], 'forecast', 'failed', '處理失敗', duration)
                    return jsonify({'success': False, 'message': 'FORECAST處理失敗'})

        elif not is_multi_file_mode:
            # ===== 單檔案模式：處理單一 cleaned_forecast.xlsx/.xls 或 cleaned_forecast_1.xlsx/.xls =====
            # 先檢查合併檔案，再檢查編號檔案（支援 .xlsx 和 .xls）
            is_merged_forecast = False
            cleaned_forecast = find_file_with_extensions(processed_folder, 'cleaned_forecast')
            if cleaned_forecast:
                is_merged_forecast = True  # 這是合併的檔案
            else:
                # 嘗試找 cleaned_forecast_1.xlsx/.xls（只有一個檔案的分開模式）
                cleaned_forecast = find_file_with_extensions(processed_folder, 'cleaned_forecast_1')

            if not cleaned_forecast:
                log_process(user['id'], 'forecast', 'failed', '請先完成Forecast數據清理')
                return jsonify({'success': False, 'message': '請先完成Forecast數據清理'})

            # 決定輸出檔名（廣達等非 Pegatron 客戶）
            if is_merged_forecast:
                output_filename = 'forecast_ALL.xlsx'
            else:
                # 分開模式維持原本檔名格式
                output_filename = 'forecast_result.xlsx'

            print("開始FORECAST處理（單檔案模式）...")
            print(f"清理後的Forecast文件: {cleaned_forecast}")
            print(f"輸出檔名: {output_filename}")
            print(f"整合後的ERP文件: {integrated_erp}")
            if has_transit:
                print(f"整合後的Transit文件: {integrated_transit}")
            else:
                print("⚠️ 未找到Transit文件，將跳過Transit數據處理")

            processor = UltraFastForecastProcessor(
                forecast_file=cleaned_forecast,
                erp_file=integrated_erp,
                transit_file=integrated_transit if has_transit else None,
                output_folder=processed_folder,
                output_filename=output_filename
            )

            success = processor.process_all_blocks()

            if success:
                result_file = os.path.join(processed_folder, output_filename)
                if os.path.exists(result_file):
                    file_size = os.path.getsize(result_file)
                    duration = time.time() - start_time
                    print(f"FORECAST處理完成，結果文件: {result_file} (大小: {file_size} bytes)")

                    log_process(user['id'], 'forecast', 'success',
                              f'ERP填入: {processor.total_filled}, Transit填入: {processor.total_transit_filled if has_transit else 0}', duration)
                    log_activity(user['id'], user['username'], 'forecast_success',
                               f"FORECAST 處理成功", get_client_ip(), request.headers.get('User-Agent'))

                    result_data = {
                        'success': True,
                        'message': 'FORECAST處理完成',
                        'file': output_filename,
                        'erp_filled': processor.total_filled,
                        'erp_skipped': processor.total_skipped,
                        'file_size': file_size,
                        'transit_file_skipped': transit_skipped
                    }

                    if has_transit:
                        result_data['transit_filled'] = processor.total_transit_filled
                        result_data['transit_skipped'] = processor.total_transit_skipped

                    return jsonify(result_data)
                else:
                    duration = time.time() - start_time
                    print("錯誤：結果文件未生成")
                    log_process(user['id'], 'forecast', 'failed', '結果文件未生成', duration)
                    return jsonify({'success': False, 'message': 'FORECAST處理完成但結果文件未生成'})
            else:
                duration = time.time() - start_time
                print("FORECAST處理失敗")
                log_process(user['id'], 'forecast', 'failed', '處理失敗', duration)
                log_activity(user['id'], user['username'], 'forecast_failed',
                           "FORECAST 處理失敗", get_client_ip(), request.headers.get('User-Agent'))
                return jsonify({'success': False, 'message': 'FORECAST處理失敗'})

        else:
            # ===== 多檔案模式：分別處理每個 cleaned_forecast_N.xlsx/.xls =====
            import glob

            # 尋找所有 cleaned_forecast_*.xlsx 和 cleaned_forecast_*.xls 檔案
            forecast_files = []
            for ext in ['.xlsx', '.xls']:
                forecast_pattern = os.path.join(processed_folder, f'cleaned_forecast_*{ext}')
                forecast_files.extend(glob.glob(forecast_pattern))
            forecast_files = sorted(set(forecast_files))  # 去重並排序

            if not forecast_files:
                log_process(user['id'], 'forecast', 'failed', '請先完成Forecast數據清理（未找到清理後的Forecast檔案）')
                return jsonify({'success': False, 'message': '請先完成Forecast數據清理（未找到清理後的Forecast檔案）'})

            print(f"開始FORECAST處理（多檔案模式）... 共 {len(forecast_files)} 個檔案")
            print(f"整合後的ERP文件: {integrated_erp}")
            if has_transit:
                print(f"整合後的Transit文件: {integrated_transit}")
            else:
                print("⚠️ 未找到Transit文件，將跳過Transit數據處理")

            # 統計資料
            total_erp_filled = 0
            total_erp_skipped = 0
            total_transit_filled = 0
            total_transit_skipped = 0
            processed_files = []
            failed_files = []

            for idx, forecast_file in enumerate(forecast_files, 1):
                # 從檔名提取編號，例如 cleaned_forecast_1.xlsx 或 cleaned_forecast_1.xls -> 1
                file_basename = os.path.basename(forecast_file)
                # 提取數字部分（支援 .xlsx 和 .xls）
                import re
                match = re.search(r'cleaned_forecast_(\d+)\.xlsx?', file_basename)
                if match:
                    file_num = match.group(1)
                else:
                    file_num = str(idx)

                # 輸出檔案使用 .xlsx 格式（廣達等非 Pegatron 客戶維持原本檔名格式）
                output_filename = f'forecast_result_{file_num}.xlsx'

                print(f"\n處理檔案 {idx}/{len(forecast_files)}: {file_basename}")

                processor = UltraFastForecastProcessor(
                    forecast_file=forecast_file,
                    erp_file=integrated_erp,
                    transit_file=integrated_transit if has_transit else None,
                    output_folder=processed_folder,
                    output_filename=output_filename  # 指定輸出檔名
                )

                success = processor.process_all_blocks()

                if success:
                    result_file = os.path.join(processed_folder, output_filename)
                    if os.path.exists(result_file):
                        file_size = os.path.getsize(result_file)
                        print(f"✓ 檔案 {file_basename} 處理完成 (大小: {file_size} bytes)")

                        # 累加統計
                        total_erp_filled += processor.total_filled
                        total_erp_skipped += processor.total_skipped
                        if has_transit:
                            total_transit_filled += processor.total_transit_filled
                            total_transit_skipped += processor.total_transit_skipped

                        processed_files.append({
                            'input': file_basename,
                            'output': output_filename,
                            'erp_filled': processor.total_filled,
                            'erp_skipped': processor.total_skipped,
                            'transit_filled': processor.total_transit_filled if has_transit else 0,
                            'transit_skipped': processor.total_transit_skipped if has_transit else 0,
                            'file_size': file_size
                        })
                    else:
                        print(f"✗ 檔案 {file_basename} 結果未生成")
                        failed_files.append(file_basename)
                else:
                    print(f"✗ 檔案 {file_basename} 處理失敗")
                    failed_files.append(file_basename)

            duration = time.time() - start_time

            if processed_files:
                # 至少有一個檔案處理成功
                success_count = len(processed_files)
                fail_count = len(failed_files)

                log_message = f'多檔案模式: {success_count}/{len(forecast_files)} 成功, ERP填入: {total_erp_filled}, Transit填入: {total_transit_filled}'
                log_process(user['id'], 'forecast', 'success', log_message, duration)
                log_activity(user['id'], user['username'], 'forecast_success',
                           f"FORECAST 多檔案處理成功 ({success_count}/{len(forecast_files)})", get_client_ip(), request.headers.get('User-Agent'))

                result_data = {
                    'success': True,
                    'message': f'FORECAST處理完成（{success_count}/{len(forecast_files)} 個檔案成功）',
                    'multi_file': True,
                    'files': processed_files,
                    'failed_files': failed_files,
                    'total_erp_filled': total_erp_filled,
                    'total_erp_skipped': total_erp_skipped,
                    'file_count': len(forecast_files),
                    'success_count': success_count,
                    'fail_count': fail_count,
                    'transit_file_skipped': transit_skipped
                }

                if has_transit:
                    result_data['total_transit_filled'] = total_transit_filled
                    result_data['total_transit_skipped'] = total_transit_skipped

                return jsonify(result_data)
            else:
                # 所有檔案都處理失敗
                log_process(user['id'], 'forecast', 'failed', f'多檔案模式: 全部 {len(forecast_files)} 個檔案處理失敗', duration)
                log_activity(user['id'], user['username'], 'forecast_failed',
                           "FORECAST 多檔案處理全部失敗", get_client_ip(), request.headers.get('User-Agent'))
                return jsonify({'success': False, 'message': 'FORECAST處理失敗：所有檔案都處理失敗'})

    except Exception as e:
        duration = time.time() - start_time
        print(f"FORECAST處理異常: {str(e)}")
        import traceback
        traceback.print_exc()
        log_process(user['id'], 'forecast', 'failed', str(e), duration)
        log_activity(user['id'], user['username'], 'forecast_failed',
                   f"FORECAST 處理失敗：{str(e)}", get_client_ip(), request.headers.get('User-Agent'))
        return jsonify({'success': False, 'message': f'FORECAST處理失敗: {str(e)}'})

@app.route('/check_files')
@login_required
def check_files():
    try:
        # 從 session 獲取當前用戶的檔案路徑
        erp_file = get_user_file_path('erp')
        forecast_file = get_user_file_path('forecast')
        transit_file = get_user_file_path('transit')

        erp_exists = erp_file and os.path.exists(erp_file)
        forecast_exists = forecast_file and os.path.exists(forecast_file)
        transit_exists = transit_file and os.path.exists(transit_file)

        return jsonify({
            'success': True,
            'erp_exists': erp_exists,
            'forecast_exists': forecast_exists,
            'transit_exists': transit_exists,
            'erp_size': os.path.getsize(erp_file) if erp_exists else 0,
            'forecast_size': os.path.getsize(forecast_file) if forecast_exists else 0,
            'transit_size': os.path.getsize(transit_file) if transit_exists else 0
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/download/<filename>')
@login_required
def download_file(filename):
    user = get_current_user()

    # 從 session 獲取當前的 processed 資料夾
    processed_folder = session.get('current_processed_folder')
    if not processed_folder:
        processed_folder, _ = get_session_folder_path(user['id'], 'processed')

    # 嘗試從用戶的 session 資料夾下載
    if processed_folder:
        file_path = os.path.join(processed_folder, filename)
        if os.path.exists(file_path):
            # 記錄下載活動
            file_size = os.path.getsize(file_path)
            log_activity(user['id'], user['username'], 'download',
                       f"下載文件：{filename} ({file_size} bytes)", get_client_ip(), request.headers.get('User-Agent'))

            # 為下載檔案加上時間戳
            session_timestamp = session.get('current_session_timestamp', '')
            if session_timestamp:
                # 從檔名分離副檔名
                name, ext = os.path.splitext(filename)
                download_filename = f"{name}_{session_timestamp}{ext}"
            else:
                download_filename = filename

            return send_file(file_path, as_attachment=True, download_name=download_filename)

    # 向後相容：嘗試從舊的 processed 資料夾下載
    file_path = os.path.join(PROCESSED_FOLDER, filename)
    if os.path.exists(file_path):
        file_size = os.path.getsize(file_path)
        log_activity(user['id'], user['username'], 'download',
                   f"下載文件：{filename} ({file_size} bytes)", get_client_ip(), request.headers.get('User-Agent'))
        return send_file(file_path, as_attachment=True)

    return jsonify({'success': False, 'message': '文件不存在'})

# 靜態資源路由，添加版本控制
@app.route('/static/<path:filename>')
def static_files(filename):
    """提供靜態文件並設置適當的緩存策略"""
    file_path = os.path.join('static', filename)
    if os.path.exists(file_path):
        response = make_response(send_file(file_path))

        # 根據文件類型設置不同的緩存策略
        if filename.endswith(('.css', '.js')):
            # CSS和JS文件設置較短的緩存時間，但允許緩存
            response.headers['Cache-Control'] = 'public, max-age=3600'  # 1小時
        elif filename.endswith(('.png', '.jpg', '.jpeg', '.gif', '.ico')):
            # 圖片文件可以緩存更長時間
            response.headers['Cache-Control'] = 'public, max-age=86400'  # 24小時
        else:
            # 其他文件不緩存
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'

        return response
    else:
        return jsonify({'error': 'File not found'}), 404


# ========================================
# IT/Admin 管理介面路由
# ========================================

@app.route('/admin')
@admin_required
def admin_dashboard():
    """管理者首頁"""
    user = get_current_user()
    response = make_response(render_template('admin.html', user=user))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@app.route('/it')
@it_or_admin_required
def it_dashboard():
    """IT 人員首頁"""
    user = get_current_user()
    response = make_response(render_template('it_dashboard.html', user=user))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@app.route('/logs')
@it_or_admin_required
def logs_page():
    """LOG 查看頁面"""
    user = get_current_user()
    response = make_response(render_template('logs.html', user=user))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@app.route('/users_manage')
@admin_required
def users_manage_page():
    """用戶管理頁面（僅管理者）"""
    user = get_current_user()
    response = make_response(render_template('users_manage.html', user=user))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@app.route('/mappings_view')
@admin_required
def mappings_view_page():
    """客戶映射檢視頁面（僅管理者）"""
    user = get_current_user()
    response = make_response(render_template('mappings_view.html', user=user))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@app.route('/test_function')
@it_or_admin_required
def test_function_page():
    """測試功能頁面"""
    user = get_current_user()
    response = make_response(render_template('test_function.html', user=user))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


# ========================================
# IT/Admin 管理介面 API
# ========================================

@app.route('/api/logs/activity')
@it_or_admin_required
def api_get_activity_logs():
    """取得活動日誌"""
    try:
        user_id = request.args.get('user_id', type=int)
        action_type = request.args.get('action_type')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)

        offset = (page - 1) * per_page
        records, total = get_activity_logs_filtered(
            user_id=user_id, action_type=action_type,
            start_date=start_date, end_date=end_date,
            limit=per_page, offset=offset
        )

        # 處理日期格式（JSON 序列化）
        for record in records:
            if record.get('created_at'):
                record['created_at'] = record['created_at'].strftime('%Y-%m-%d %H:%M:%S')

        return jsonify({
            'success': True,
            'records': records,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': (total + per_page - 1) // per_page if total > 0 else 0
        })
    except Exception as e:
        print(f"❌ 取得活動日誌失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/logs/upload')
@it_or_admin_required
def api_get_upload_records():
    """取得上傳記錄"""
    try:
        user_id = request.args.get('user_id', type=int)
        file_type = request.args.get('file_type')
        status = request.args.get('status')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)

        offset = (page - 1) * per_page
        records, total = get_upload_records(
            user_id=user_id, file_type=file_type, status=status,
            start_date=start_date, end_date=end_date,
            limit=per_page, offset=offset
        )

        # 處理日期格式
        for record in records:
            if record.get('created_at'):
                record['created_at'] = record['created_at'].strftime('%Y-%m-%d %H:%M:%S')

        return jsonify({
            'success': True,
            'records': records,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': (total + per_page - 1) // per_page if total > 0 else 0
        })
    except Exception as e:
        print(f"❌ 取得上傳記錄失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/logs/process')
@it_or_admin_required
def api_get_process_records():
    """取得處理記錄"""
    try:
        user_id = request.args.get('user_id', type=int)
        process_type = request.args.get('process_type')
        status = request.args.get('status')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)

        offset = (page - 1) * per_page
        records, total = get_process_records(
            user_id=user_id, process_type=process_type, status=status,
            start_date=start_date, end_date=end_date,
            limit=per_page, offset=offset
        )

        # 處理日期格式
        for record in records:
            if record.get('created_at'):
                record['created_at'] = record['created_at'].strftime('%Y-%m-%d %H:%M:%S')

        return jsonify({
            'success': True,
            'records': records,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': (total + per_page - 1) // per_page if total > 0 else 0
        })
    except Exception as e:
        print(f"❌ 取得處理記錄失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/admin/users')
@admin_required
def api_get_all_users():
    """取得所有用戶列表（僅管理者）"""
    try:
        users = get_all_users()

        # 處理日期格式
        for user in users:
            if user.get('created_at'):
                user['created_at'] = user['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            if user.get('last_login'):
                user['last_login'] = user['last_login'].strftime('%Y-%m-%d %H:%M:%S')

        return jsonify({'success': True, 'users': users})
    except Exception as e:
        print(f"❌ 取得用戶列表失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/admin/users', methods=['POST'])
@admin_required
def api_create_user():
    """建立新用戶（僅管理者）"""
    try:
        data = request.get_json()

        # 驗證必填欄位
        required_fields = ['username', 'password', 'display_name']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'message': f'缺少必填欄位: {field}'})

        # 驗證角色
        role = data.get('role', 'user')
        if role not in ['admin', 'it', 'user']:
            return jsonify({'success': False, 'message': '無效的角色'})

        # 建立用戶
        success, message, user_id = create_user(
            username=data['username'],
            password=data['password'],
            display_name=data['display_name'],
            role=role,
            company=data.get('company'),
            is_active=data.get('is_active', True)
        )

        if success:
            # 建立用戶專屬的 compare 資料夾
            user_compare_folder = os.path.join(app.root_path, 'compare', data['username'])
            try:
                os.makedirs(user_compare_folder, exist_ok=True)
                print(f"✅ 已建立用戶範本資料夾: {user_compare_folder}")
            except Exception as folder_error:
                print(f"⚠️ 建立用戶範本資料夾失敗: {folder_error}")

            # 記錄操作日誌
            admin_user = session.get('user', {})
            log_activity(
                user_id=admin_user.get('id'),
                username=admin_user.get('username'),
                action_type='user_create',
                action_detail=f"建立用戶: {data['username']} ({data['display_name']}), 角色: {role}",
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent')
            )
            return jsonify({'success': True, 'message': message, 'user_id': user_id})
        else:
            return jsonify({'success': False, 'message': message})

    except Exception as e:
        print(f"❌ 建立用戶失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/admin/users/<int:user_id>', methods=['PUT'])
@admin_required
def api_update_user(user_id):
    """更新用戶資料（僅管理者）"""
    try:
        data = request.get_json()

        # 取得原用戶資料（用於日誌記錄）
        original_user = get_user_by_id(user_id)
        if not original_user:
            return jsonify({'success': False, 'message': '用戶不存在'})

        # 驗證角色（如果有提供）
        if 'role' in data and data['role'] not in ['admin', 'it', 'user']:
            return jsonify({'success': False, 'message': '無效的角色'})

        # 建立更新參數
        update_params = {}
        if 'username' in data:
            update_params['username'] = data['username']
        if 'display_name' in data:
            update_params['display_name'] = data['display_name']
        if 'password' in data and data['password']:
            update_params['password'] = data['password']
        if 'role' in data:
            update_params['role'] = data['role']
        if 'company' in data:
            update_params['company'] = data['company']
        if 'is_active' in data:
            update_params['is_active'] = data['is_active']

        if not update_params:
            return jsonify({'success': False, 'message': '沒有要更新的資料'})

        # 更新用戶
        success, message = update_user(user_id, **update_params)

        if success:
            # 記錄操作日誌
            admin_user = session.get('user', {})

            # 判斷是切換狀態還是一般更新
            if len(update_params) == 1 and 'is_active' in update_params:
                action_type = 'user_toggle_status'
                status_text = '啟用' if update_params['is_active'] else '停用'
                action_detail = f"切換用戶狀態: {original_user['username']} -> {status_text}"
            else:
                action_type = 'user_update'
                changes = []
                if 'username' in update_params:
                    changes.append(f"用戶名: {original_user['username']} -> {update_params['username']}")
                if 'display_name' in update_params:
                    changes.append(f"顯示名稱: {update_params['display_name']}")
                if 'role' in update_params:
                    changes.append(f"角色: {update_params['role']}")
                if 'company' in update_params:
                    changes.append(f"公司: {update_params['company']}")
                if 'password' in update_params:
                    changes.append("密碼已更新")
                if 'is_active' in update_params:
                    changes.append(f"狀態: {'啟用' if update_params['is_active'] else '停用'}")
                action_detail = f"更新用戶 {original_user['username']}: " + ", ".join(changes)

            log_activity(
                user_id=admin_user.get('id'),
                username=admin_user.get('username'),
                action_type=action_type,
                action_detail=action_detail,
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent')
            )
            return jsonify({'success': True, 'message': message})
        else:
            return jsonify({'success': False, 'message': message})

    except Exception as e:
        print(f"❌ 更新用戶失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@admin_required
def api_delete_user(user_id):
    """刪除用戶（僅管理者）"""
    try:
        # 取得原用戶資料（用於日誌記錄）
        original_user = get_user_by_id(user_id)
        if not original_user:
            return jsonify({'success': False, 'message': '用戶不存在'})

        # 防止刪除自己
        current_user = session.get('user', {})
        if current_user.get('id') == user_id:
            return jsonify({'success': False, 'message': '無法刪除自己的帳號'})

        # 刪除用戶
        success, message, deleted_username = delete_user(user_id)

        if success:
            # 記錄操作日誌
            admin_user = session.get('user', {})
            log_activity(
                user_id=admin_user.get('id'),
                username=admin_user.get('username'),
                action_type='user_delete',
                action_detail=f"刪除用戶: {deleted_username} ({original_user['display_name']})",
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent')
            )
            return jsonify({'success': True, 'message': message})
        else:
            return jsonify({'success': False, 'message': message})

    except Exception as e:
        print(f"❌ 刪除用戶失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/admin/mappings')
@admin_required
def api_get_all_mappings():
    """取得所有用戶的客戶映射（僅管理者）"""
    try:
        mappings = get_all_customer_mappings()

        # 處理日期格式
        for mapping in mappings:
            if mapping.get('created_at'):
                mapping['created_at'] = mapping['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            if mapping.get('updated_at'):
                mapping['updated_at'] = mapping['updated_at'].strftime('%Y-%m-%d %H:%M:%S')

        return jsonify({'success': True, 'mappings': mappings})
    except Exception as e:
        print(f"❌ 取得客戶映射失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/admin/mappings', methods=['POST'])
@admin_required
def api_create_mapping():
    """新增客戶映射（僅管理者）"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        customer_name = data.get('customer_name')
        delivery_location = data.get('delivery_location')
        region = data.get('region')
        schedule_breakpoint = data.get('schedule_breakpoint')
        etd = data.get('etd')
        eta = data.get('eta')
        requires_transit = data.get('requires_transit', True)  # 預設為 True

        if not user_id or not customer_name:
            return jsonify({'success': False, 'message': '用戶ID和客戶簡稱為必填'})

        if not region:
            return jsonify({'success': False, 'message': '客戶廠區為必填'})

        success, message, mapping_id = admin_create_customer_mapping(
            user_id, customer_name, delivery_location, region, schedule_breakpoint, etd, eta, requires_transit
        )

        if success:
            # 記錄活動
            admin_user = get_current_user()
            log_activity(admin_user['id'], 'mapping_create',
                        f"新增客戶映射: {customer_name} (user_id={user_id})",
                        request.remote_addr)

        return jsonify({'success': success, 'message': message, 'mapping_id': mapping_id})
    except Exception as e:
        print(f"❌ 新增客戶映射失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/admin/mappings/<int:mapping_id>', methods=['PUT'])
@admin_required
def api_update_mapping(mapping_id):
    """更新客戶映射（僅管理者）"""
    try:
        data = request.get_json()

        # 只傳遞有值的欄位
        update_data = {}
        for field in ['customer_name', 'delivery_location', 'region', 'schedule_breakpoint', 'etd', 'eta', 'requires_transit']:
            if field in data:
                update_data[field] = data[field]

        if not update_data:
            return jsonify({'success': False, 'message': '沒有提供要更新的欄位'})

        success, message = admin_update_customer_mapping(mapping_id, **update_data)

        if success:
            # 記錄活動
            admin_user = get_current_user()
            log_activity(admin_user['id'], 'mapping_update',
                        f"更新客戶映射 ID={mapping_id}",
                        request.remote_addr)

        return jsonify({'success': success, 'message': message})
    except Exception as e:
        print(f"❌ 更新客戶映射失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/admin/mappings/<int:mapping_id>', methods=['DELETE'])
@admin_required
def api_delete_mapping(mapping_id):
    """刪除客戶映射（僅管理者）"""
    try:
        # 先取得映射資訊用於記錄
        mapping = admin_get_customer_mapping_by_id(mapping_id)

        success, message = admin_delete_customer_mapping(mapping_id)

        if success and mapping:
            # 記錄活動
            admin_user = get_current_user()
            log_activity(admin_user['id'], 'mapping_delete',
                        f"刪除客戶映射: {mapping['customer_name']} (company={mapping.get('company', 'N/A')})",
                        request.remote_addr)

        return jsonify({'success': success, 'message': message})
    except Exception as e:
        print(f"❌ 刪除客戶映射失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/admin/mappings/<int:mapping_id>')
@admin_required
def api_get_mapping(mapping_id):
    """取得單一客戶映射（僅管理者）"""
    try:
        mapping = admin_get_customer_mapping_by_id(mapping_id)
        if not mapping:
            return jsonify({'success': False, 'message': '映射不存在'})

        # 處理日期格式
        if mapping.get('created_at'):
            mapping['created_at'] = mapping['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        if mapping.get('updated_at'):
            mapping['updated_at'] = mapping['updated_at'].strftime('%Y-%m-%d %H:%M:%S')

        return jsonify({'success': True, 'mapping': mapping})
    except Exception as e:
        print(f"❌ 取得客戶映射失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


# ==================== 規則管理 API ====================

@app.route('/rules')
@admin_required
def rules_view():
    """規則管理頁面"""
    user = get_current_user()
    return render_template('rules.html', user=user)


@app.route('/api/admin/rules')
@admin_required
def api_get_all_rules():
    """取得所有處理規則"""
    try:
        category = request.args.get('category')
        user_id = request.args.get('user_id', type=int)

        if user_id:
            if category:
                rules = get_processing_rules_by_category(category, user_id)
            else:
                rules = get_processing_rules_by_user(user_id)
        else:
            if category:
                rules = get_processing_rules_by_category(category)
            else:
                rules = get_all_processing_rules()

        # 處理日期格式
        for rule in rules:
            if rule.get('created_at'):
                rule['created_at'] = rule['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            if rule.get('updated_at'):
                rule['updated_at'] = rule['updated_at'].strftime('%Y-%m-%d %H:%M:%S')

        return jsonify({'success': True, 'rules': rules})
    except Exception as e:
        print(f"❌ 取得處理規則失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/admin/rules/<int:rule_id>')
@admin_required
def api_get_rule(rule_id):
    """取得單一處理規則"""
    try:
        rule = get_processing_rule_by_id(rule_id)
        if not rule:
            return jsonify({'success': False, 'message': '規則不存在'})

        if rule.get('created_at'):
            rule['created_at'] = rule['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        if rule.get('updated_at'):
            rule['updated_at'] = rule['updated_at'].strftime('%Y-%m-%d %H:%M:%S')

        return jsonify({'success': True, 'rule': rule})
    except Exception as e:
        print(f"❌ 取得處理規則失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/admin/rules', methods=['POST'])
@admin_required
def api_create_rule():
    """新增處理規則"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        rule_name = data.get('rule_name', '').strip()
        rule_category = data.get('rule_category', '').strip()
        rule_description = data.get('rule_description', '').strip()
        rule_config = data.get('rule_config')
        display_order = data.get('display_order', 0)

        if not user_id:
            return jsonify({'success': False, 'message': '用戶 ID 為必填'})

        if not rule_name:
            return jsonify({'success': False, 'message': '規則名稱為必填'})

        if rule_category not in ['erp', 'transit', 'forecast', 'mapping', 'cleanup']:
            return jsonify({'success': False, 'message': '無效的規則類別'})

        success, message, rule_id = create_processing_rule(
            user_id, rule_name, rule_category, rule_description, rule_config, display_order
        )

        if success:
            return jsonify({'success': True, 'message': message, 'rule_id': rule_id})
        return jsonify({'success': False, 'message': message})
    except Exception as e:
        print(f"❌ 新增處理規則失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/admin/rules/<int:rule_id>', methods=['PUT'])
@admin_required
def api_update_rule(rule_id):
    """更新處理規則"""
    try:
        data = request.get_json()
        update_data = {}

        if 'rule_name' in data:
            update_data['rule_name'] = data['rule_name'].strip()
        if 'rule_description' in data:
            update_data['rule_description'] = data['rule_description'].strip()
        if 'rule_config' in data:
            update_data['rule_config'] = data['rule_config']
        if 'is_active' in data:
            update_data['is_active'] = data['is_active']
        if 'display_order' in data:
            update_data['display_order'] = data['display_order']

        success, message = update_processing_rule(rule_id, **update_data)

        return jsonify({'success': success, 'message': message})
    except Exception as e:
        print(f"❌ 更新處理規則失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/admin/rules/<int:rule_id>', methods=['DELETE'])
@admin_required
def api_delete_rule(rule_id):
    """刪除處理規則"""
    try:
        success, message = delete_processing_rule(rule_id)
        return jsonify({'success': success, 'message': message})
    except Exception as e:
        print(f"❌ 刪除處理規則失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/admin/rules/<int:rule_id>/toggle', methods=['POST'])
@admin_required
def api_toggle_rule_status(rule_id):
    """切換規則啟用狀態"""
    try:
        success, message, new_status = toggle_processing_rule_status(rule_id)
        return jsonify({'success': success, 'message': message, 'is_active': new_status})
    except Exception as e:
        print(f"❌ 切換規則狀態失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/test/customers')
@it_or_admin_required
def api_get_test_customers():
    """取得可測試的客戶列表（依據 LOGIN_ALLOWED_CUSTOMERS 過濾）"""
    try:
        customers = get_users_with_company()

        # 從環境變數讀取允許的客戶清單（逗號分隔，空白=全部顯示）
        allowed = os.environ.get('LOGIN_ALLOWED_CUSTOMERS', '').strip()
        allowed_list = [name.strip().lower() for name in allowed.split(',') if name.strip()] if allowed else []

        if allowed_list:
            customers = [c for c in customers if c['username'].lower() in allowed_list]

        return jsonify({'success': True, 'customers': customers})
    except Exception as e:
        print(f"❌ 取得客戶列表失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/test/run', methods=['POST'])
@it_or_admin_required
def api_run_test():
    """執行測試流程"""
    import shutil

    user = get_current_user()
    data = request.get_json()
    customer_id = data.get('customer_id')

    if not customer_id:
        return jsonify({'success': False, 'message': '請選擇客戶'})

    try:
        # 取得客戶資訊
        customer = get_user_by_id(customer_id)
        if not customer:
            return jsonify({'success': False, 'message': '客戶不存在'})

        company = customer['company']

        # 檢查測試檔案是否存在
        test_folder = os.path.join('test_data', company)
        if not os.path.exists(test_folder):
            return jsonify({'success': False, 'message': f'測試資料夾不存在: test_data/{company}'})

        # 檢查必要檔案是否存在（支援 .xlsx 和 .xls）
        required_base_names = ['erp_data', 'forecast_data', 'transit_data']
        missing_files = []
        found_files = {}  # 記錄找到的實際檔案名稱
        for base_name in required_base_names:
            found = find_file_with_extensions(test_folder, base_name)
            if found:
                found_files[base_name] = os.path.basename(found)
            else:
                missing_files.append(f'{base_name}.xlsx 或 {base_name}.xls')

        if missing_files:
            return jsonify({
                'success': False,
                'message': f'缺少測試檔案: {", ".join(missing_files)}'
            })

        # 建立測試 session
        test_session_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S') + '_test'
        upload_folder = os.path.join(UPLOAD_FOLDER, str(customer_id), test_session_timestamp)
        processed_folder = os.path.join(PROCESSED_FOLDER, str(customer_id), test_session_timestamp)

        os.makedirs(upload_folder, exist_ok=True)
        os.makedirs(processed_folder, exist_ok=True)

        # 複製測試檔案（保留原始副檔名）
        for base_name, actual_filename in found_files.items():
            shutil.copy(
                os.path.join(test_folder, actual_filename),
                os.path.join(upload_folder, actual_filename)
            )

        # 取得客戶的 username 用於模板驗證
        customer_username = customer['username']

        # ========== 格式驗證（使用客戶專屬模板，支援 .xlsx 和 .xls）==========
        validation_errors = []

        # 驗證 ERP 格式
        erp_test_file = find_file_with_extensions(upload_folder, 'erp_data')
        if erp_test_file:
            is_valid, message, details = validate_erp_format(erp_test_file, customer_username)
            if not is_valid:
                validation_errors.append({
                    'file': 'ERP',
                    'message': message,
                    'details': details
                })

        # 驗證 Forecast 格式
        forecast_test_file = find_file_with_extensions(upload_folder, 'forecast_data')
        if forecast_test_file:
            is_valid, message, details = validate_forecast_format(forecast_test_file, customer_username)
            if not is_valid:
                validation_errors.append({
                    'file': 'Forecast',
                    'message': message,
                    'details': details
                })

        # 驗證在途格式
        transit_test_file = find_file_with_extensions(upload_folder, 'transit_data')
        if transit_test_file:
            is_valid, message, details = validate_transit_format(transit_test_file, customer_username)
            if not is_valid:
                validation_errors.append({
                    'file': '在途',
                    'message': message,
                    'details': details
                })

        # 如果有驗證錯誤，返回錯誤訊息
        if validation_errors:
            # 清理已複製的測試檔案
            shutil.rmtree(upload_folder, ignore_errors=True)
            shutil.rmtree(processed_folder, ignore_errors=True)

            error_messages = []
            for err in validation_errors:
                error_messages.append(f"{err['file']}: {err['message']}")
                if err['details']:
                    for detail in err['details'][:3]:
                        error_messages.append(f"  - {detail}")

            return jsonify({
                'success': False,
                'message': '測試檔案格式驗證失敗',
                'validation_errors': validation_errors,
                'details': error_messages
            })

        print(f"✅ 測試檔案格式驗證通過（使用模板: compare/{customer_username}/）")
        # ========== 格式驗證結束 ==========

        # 記錄測試開始
        log_activity(
            user_id=user['id'],
            username=user['username'],
            action_type='forecast_start',
            action_detail=f"執行測試：{company} (customer_id: {customer_id})",
            ip_address=get_client_ip(),
            user_agent=request.headers.get('User-Agent')
        )

        # 執行完整流程
        # 1. 數據清理（支援 .xlsx 和 .xls）
        from openpyxl import load_workbook as openpyxl_load_workbook
        forecast_file = find_file_with_extensions(upload_folder, 'forecast_data')
        if not forecast_file:
            return jsonify({'success': False, 'message': '找不到 Forecast 檔案'})

        # 根據檔案類型選擇處理方式
        is_xls = is_xls_format(forecast_file)
        if is_xls:
            # .xls 格式使用 LibreOffice 跨平台方案（輸出為 xlsx）
            cleaned_file = os.path.join(processed_folder, 'cleaned_forecast.xlsx')
            cleaned_count = cleanup_xls_file(forecast_file, cleaned_file, customer_username)
        else:
            # .xlsx 格式使用 openpyxl
            wb = openpyxl_load_workbook(forecast_file)
            ws = wb.active

            cleaned_count = 0
            for row_idx in range(1, ws.max_row + 1):
                k_cell = ws.cell(row=row_idx, column=11)
                if k_cell.value and str(k_cell.value) == "供應數量":
                    for col_idx in range(12, min(50, ws.max_column + 1)):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value != 0:
                            cell.value = 0
                            cleaned_count += 1

                i_cell = ws.cell(row=row_idx, column=9)
                if i_cell.value and "庫存數量" in str(i_cell.value):
                    next_row_i_cell = ws.cell(row=row_idx + 1, column=9)
                    if next_row_i_cell.value != 0:
                        next_row_i_cell.value = 0
                        cleaned_count += 1

            cleaned_file = os.path.join(processed_folder, 'cleaned_forecast.xlsx')
            wb.save(cleaned_file)

        # 2. ERP 和 Transit 整合（支援 .xlsx 和 .xls）
        erp_file = find_file_with_extensions(upload_folder, 'erp_data')
        transit_file = find_file_with_extensions(upload_folder, 'transit_data')

        erp_df = pd.read_excel(erp_file)
        transit_df = pd.read_excel(transit_file)

        # 取得客戶的 mapping 資料
        mapping_data = get_customer_mappings(customer_id)
        if not mapping_data:
            mapping_data = {'regions': {}, 'schedule_breakpoints': {}, 'etd': {}, 'eta': {}}

        # 標準化日期格式
        if '排程出貨日期' in erp_df.columns:
            erp_df['排程出貨日期'] = erp_df['排程出貨日期'].apply(normalize_date_for_mapping)

        # 找到客戶簡稱欄位
        customer_col = None
        for col in erp_df.columns:
            if '客戶' in str(col) and '簡稱' in str(col):
                customer_col = col
                break

        if customer_col:
            erp_df['客戶需求地區'] = erp_df[customer_col].map(mapping_data.get('regions', {}))
            erp_df['排程出貨日期斷點'] = erp_df[customer_col].map(mapping_data.get('schedule_breakpoints', {}))
            erp_df['ETD'] = erp_df[customer_col].map(mapping_data.get('etd', {}))
            erp_df['ETA'] = erp_df[customer_col].map(mapping_data.get('eta', {}))

        if '排程出貨日期' in erp_df.columns:
            erp_df = erp_df.sort_values('排程出貨日期')

        integrated_erp_file = os.path.join(processed_folder, 'integrated_erp.xlsx')
        erp_df.to_excel(integrated_erp_file, index=False)

        # Transit 整合
        mapping_dict = {}
        all_customers = set()
        all_customers.update(mapping_data.get('regions', {}).keys())
        for cust in all_customers:
            mapping_dict[cust] = {
                'region': mapping_data.get('regions', {}).get(cust, ''),
                'schedule_breakpoint': mapping_data.get('schedule_breakpoints', {}).get(cust, ''),
                'etd': mapping_data.get('etd', {}).get(cust, ''),
                'eta': mapping_data.get('eta', {}).get(cust, '')
            }

        # 判斷是否為 Pegatron（檢查是否有 Line 客戶採購單號 欄位）
        is_pegatron_transit = False
        transit_line_po_col, _ = find_column_by_name(transit_df, 'Line 客戶採購單號', required=False)
        transit_ordered_item_col, _ = find_column_by_name(transit_df, 'Ordered Item', required=False)

        if transit_line_po_col and transit_ordered_item_col:
            # Pegatron Transit 映射邏輯：用 Line 客戶採購單號 + Ordered Item 匹配 ERP
            is_pegatron_transit = True
            print(f"🔧 IT測試: 使用 Pegatron Transit 映射邏輯")

            # 動態查找 ERP 欄位
            erp_line_po_col, _ = find_column_by_name(erp_df, 'Line 客戶採購單號', required=False)
            erp_pn_col, _ = find_column_by_name(erp_df, '客戶料號', required=False)

            if erp_line_po_col and erp_pn_col:
                # 建立 ERP lookup
                erp_lookup = {}
                for idx, row in erp_df.iterrows():
                    line_po = str(row[erp_line_po_col]).strip() if pd.notna(row[erp_line_po_col]) else ''
                    pn = str(row[erp_pn_col]).strip() if pd.notna(row[erp_pn_col]) else ''
                    if line_po and pn:
                        key = (line_po, pn)
                        if key not in erp_lookup:
                            erp_lookup[key] = {
                                'region': str(row.get('客戶需求地區', '')).strip() if pd.notna(row.get('客戶需求地區', '')) else '',
                                'schedule_breakpoint': str(row.get('排程出貨日期斷點', '')).strip() if pd.notna(row.get('排程出貨日期斷點', '')) else '',
                                'etd': str(row.get('ETD', '')).strip() if pd.notna(row.get('ETD', '')) else '',
                                'eta': str(row.get('ETA', '')).strip() if pd.notna(row.get('ETA', '')) else ''
                            }

                # 應用 Transit 映射
                def get_pegatron_transit_mapping(row, field):
                    line_po = str(row[transit_line_po_col]).strip() if pd.notna(row[transit_line_po_col]) else ''
                    ordered_item = str(row[transit_ordered_item_col]).strip() if pd.notna(row[transit_ordered_item_col]) else ''
                    key = (line_po, ordered_item)
                    mapping = erp_lookup.get(key, {})
                    return mapping.get(field, '')

                transit_df['客戶需求地區'] = transit_df.apply(lambda row: get_pegatron_transit_mapping(row, 'region'), axis=1)
                transit_df['排程出貨日期斷點'] = transit_df.apply(lambda row: get_pegatron_transit_mapping(row, 'schedule_breakpoint'), axis=1)
                transit_df['ETD'] = transit_df.apply(lambda row: get_pegatron_transit_mapping(row, 'etd'), axis=1)
                transit_df['ETA_mapping'] = transit_df.apply(lambda row: get_pegatron_transit_mapping(row, 'eta'), axis=1)

        if not is_pegatron_transit:
            # Quanta 等其他客戶：用客戶簡稱映射
            transit_customer_col, _ = find_column_by_name(transit_df, ['客戶', '簡稱'], required=False)
            if transit_customer_col:
                print(f"🔧 IT測試: 使用 Quanta Transit 映射邏輯")
                transit_df['客戶需求地區'] = transit_df[transit_customer_col].apply(
                    lambda x: mapping_dict.get(str(x), {}).get('region', '') if pd.notna(x) else ''
                )
                transit_df['排程出貨日期斷點'] = transit_df[transit_customer_col].apply(
                    lambda x: mapping_dict.get(str(x), {}).get('schedule_breakpoint', '') if pd.notna(x) else ''
                )
                transit_df['ETD'] = transit_df[transit_customer_col].apply(
                    lambda x: mapping_dict.get(str(x), {}).get('etd', '') if pd.notna(x) else ''
                )
                transit_df['ETA_mapping'] = transit_df[transit_customer_col].apply(
                    lambda x: mapping_dict.get(str(x), {}).get('eta', '') if pd.notna(x) else ''
                )
            else:
                print(f"⚠️ IT測試: Transit 找不到可用的映射欄位")

        integrated_transit_file = os.path.join(processed_folder, 'integrated_transit.xlsx')
        transit_df.to_excel(integrated_transit_file, index=False)

        # 3. 執行 FORECAST 處理
        from ultra_fast_forecast_processor import UltraFastForecastProcessor

        processor = UltraFastForecastProcessor(
            forecast_file=cleaned_file,
            erp_file=integrated_erp_file,
            transit_file=integrated_transit_file,
            output_folder=processed_folder
        )

        success = processor.process_all_blocks()

        if success:
            result_file = os.path.join(processed_folder, 'forecast_result.xlsx')
            if os.path.exists(result_file):
                # 記錄測試成功
                log_activity(
                    user_id=user['id'],
                    username=user['username'],
                    action_type='forecast_success',
                    action_detail=f"測試完成：{company}，ERP填入: {processor.total_filled}，Transit填入: {processor.total_transit_filled}",
                    ip_address=get_client_ip(),
                    user_agent=request.headers.get('User-Agent')
                )

                return jsonify({
                    'success': True,
                    'message': f'測試完成：{company}',
                    'result': {
                        'company': company,
                        'cleaned_cells': cleaned_count,
                        'erp_filled': processor.total_filled,
                        'erp_skipped': processor.total_skipped,
                        'transit_filled': processor.total_transit_filled,
                        'transit_skipped': processor.total_transit_skipped,
                        'result_folder': processed_folder,
                        'download_path': f'/api/files/download/{customer_id}/{test_session_timestamp}/forecast_result.xlsx'
                    }
                })

        return jsonify({'success': False, 'message': '測試執行失敗'})

    except Exception as e:
        print(f"❌ 測試執行失敗: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'測試執行失敗: {str(e)}'})


@app.route('/api/files/browse')
@it_or_admin_required
def api_browse_files():
    """瀏覽 uploads 或 processed 資料夾"""
    try:
        folder_type = request.args.get('type', 'processed')
        user_id = request.args.get('user_id')

        base_folder = UPLOAD_FOLDER if folder_type == 'uploads' else PROCESSED_FOLDER

        # 取得用戶 ID 到名稱的映射
        user_mapping = {}
        all_users = get_all_users()
        for u in all_users:
            user_mapping[str(u['id'])] = f"{u['display_name']} ({u['company']})"

        if user_id:
            base_folder = os.path.join(base_folder, str(user_id))

        if not os.path.exists(base_folder):
            return jsonify({'success': True, 'files': [], 'folders': [], 'user_mapping': user_mapping})

        files = []
        folders = []

        # 列出第一層目錄
        for item in os.listdir(base_folder):
            item_path = os.path.join(base_folder, item)
            if os.path.isdir(item_path):
                # 計算資料夾大小和檔案數
                file_count = 0
                total_size = 0
                for root, dirs, filenames in os.walk(item_path):
                    for f in filenames:
                        file_count += 1
                        total_size += os.path.getsize(os.path.join(root, f))

                folders.append({
                    'name': item,
                    'path': item if not user_id else f"{user_id}/{item}",
                    'file_count': file_count,
                    'total_size': total_size,
                    'modified': datetime.fromtimestamp(os.path.getmtime(item_path)).strftime('%Y-%m-%d %H:%M:%S')
                })
            else:
                files.append({
                    'name': item,
                    'path': item if not user_id else f"{user_id}/{item}",
                    'size': os.path.getsize(item_path),
                    'modified': datetime.fromtimestamp(os.path.getmtime(item_path)).strftime('%Y-%m-%d %H:%M:%S')
                })

        return jsonify({'success': True, 'files': files, 'folders': folders, 'user_mapping': user_mapping})
    except Exception as e:
        print(f"❌ 瀏覽檔案失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/files/browse/<path:folder_path>')
@it_or_admin_required
def api_browse_folder(folder_path):
    """瀏覽指定資料夾的內容"""
    try:
        folder_type = request.args.get('type', 'processed')
        base_folder = UPLOAD_FOLDER if folder_type == 'uploads' else PROCESSED_FOLDER

        # 取得用戶 ID 到名稱的映射
        user_mapping = {}
        all_users = get_all_users()
        for u in all_users:
            user_mapping[str(u['id'])] = f"{u['display_name']} ({u['company']})"

        # 安全性檢查
        if '..' in folder_path:
            return jsonify({'success': False, 'message': '無效的路徑'}), 400

        # 處理 Windows 路徑分隔符
        folder_path = folder_path.replace('\\', '/')
        target_folder = os.path.join(base_folder, folder_path)

        print(f"[檔案瀏覽] 目標路徑: {target_folder}")

        if not os.path.exists(target_folder) or not os.path.isdir(target_folder):
            print(f"[檔案瀏覽] 路徑不存在或不是目錄")
            return jsonify({'success': True, 'files': [], 'folders': [], 'user_mapping': user_mapping})

        files = []
        folders = []

        for item in os.listdir(target_folder):
            item_path = os.path.join(target_folder, item)
            # 使用正斜線統一路徑
            rel_path = f"{folder_path}/{item}"

            if os.path.isdir(item_path):
                # 計算子資料夾檔案數
                file_count = len([f for f in os.listdir(item_path) if os.path.isfile(os.path.join(item_path, f))])
                folders.append({
                    'name': item,
                    'path': rel_path,
                    'file_count': file_count,
                    'modified': datetime.fromtimestamp(os.path.getmtime(item_path)).strftime('%Y-%m-%d %H:%M:%S')
                })
            else:
                files.append({
                    'name': item,
                    'path': rel_path,
                    'size': os.path.getsize(item_path),
                    'modified': datetime.fromtimestamp(os.path.getmtime(item_path)).strftime('%Y-%m-%d %H:%M:%S')
                })

        print(f"[檔案瀏覽] 找到 {len(folders)} 個資料夾, {len(files)} 個檔案")
        return jsonify({'success': True, 'files': files, 'folders': folders, 'current_path': folder_path, 'user_mapping': user_mapping})
    except Exception as e:
        print(f"❌ 瀏覽資料夾失敗: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/files/download/<path:filepath>')
@it_or_admin_required
def api_download_admin_file(filepath):
    """下載指定檔案（IT/Admin 專用）"""
    user = get_current_user()

    # 安全性檢查
    if '..' in filepath or filepath.startswith('/'):
        return jsonify({'success': False, 'message': '無效的檔案路徑'}), 400

    folder_type = request.args.get('type', 'processed')
    base_folder = UPLOAD_FOLDER if folder_type == 'uploads' else PROCESSED_FOLDER

    full_path = os.path.join(base_folder, filepath)

    if os.path.exists(full_path) and os.path.isfile(full_path):
        # 記錄下載活動
        file_size = os.path.getsize(full_path)
        log_activity(
            user_id=user['id'],
            username=user['username'],
            action_type='download',
            action_detail=f"[管理] 下載文件：{filepath} ({file_size} bytes)",
            ip_address=get_client_ip(),
            user_agent=request.headers.get('User-Agent')
        )
        return send_file(full_path, as_attachment=True)

    return jsonify({'success': False, 'message': '檔案不存在'}), 404


@app.route('/api/users/list')
@it_or_admin_required
def api_get_users_list():
    """取得用戶列表（用於篩選下拉選單）"""
    try:
        users = get_all_users()
        user_list = [{'id': u['id'], 'display_name': u['display_name'], 'company': u['company']} for u in users]
        return jsonify({'success': True, 'users': user_list})
    except Exception as e:
        print(f"❌ 取得用戶列表失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/customers/login-options')
def api_get_login_customers():
    """取得登入頁面的客戶選項列表（公開 API，不需登入）"""
    try:
        users = get_users_with_company()

        # 從環境變數讀取允許的客戶清單（逗號分隔，空白=全部顯示）
        allowed = os.environ.get('LOGIN_ALLOWED_CUSTOMERS', '').strip()
        allowed_list = [name.strip().lower() for name in allowed.split(',') if name.strip()] if allowed else []

        customer_list = [
            {
                'username': u['username'],
                'display_name': u['display_name'],
                'company': u['company'] or u['display_name']
            }
            for u in users
            if not allowed_list or u['username'].lower() in allowed_list
        ]
        return jsonify({'success': True, 'customers': customer_list})
    except Exception as e:
        print(f"❌ 取得登入客戶列表失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})


if __name__ == '__main__':
    # 初始化資料庫
    print("正在初始化資料庫...")
    if init_database():
        print("✅ 資料庫初始化成功")
        # 更新 activity_logs ENUM（新增用戶管理操作類型）
        print("正在更新 activity_logs ENUM...")
        update_activity_logs_enum()
        # 建立預設帳號
        print("正在建立預設帳號...")
        if create_default_users():
            print("✅ 預設帳號建立成功")
        else:
            print("⚠️ 預設帳號建立失敗或已存在")
    else:
        print("❌ 資料庫初始化失敗")

    # 啟動時清理超過保留期限的資料夾
    print(f"正在清理超過 {FILE_RETENTION_DAYS} 天的資料夾...")
    cleanup_old_folders()

    app.run(debug=True, host='0.0.0.0', port=12058, use_reloader=False)
