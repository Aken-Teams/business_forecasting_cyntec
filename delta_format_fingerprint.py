"""
台達 Buyer Forecast 格式指紋比對
====================================

目的:
    detect_format() 認不出來時 (例如客戶把 PARTNO 改名、插欄、改 sheet 名)，
    用「結構指紋」比對檔案是否是已知 15 格式之一的漂移變形版本。
    若是 → 由統一 reader 處理；若否 (真正第 16 種格式) → 維持現有拒絕。

設計:
    每個格式預先建立 6 維指紋:
      - sheet_signature: sheet 名稱簽名 (Diode+MOS / MRP / PSB9_MRP* / PAN JIT / Sheet1 / OTHER)
      - partno_col: PARTNO 欄位約略位置
      - date_start_col: 日期欄起始位置
      - date_count: 日期欄數量
      - marker_col: marker 欄位置 (None 表 flat 格式)
      - layout: 資料排列 (flat / multirow_3 / multirow_4 / multirow_5)

    新檔案抽相同 6 維 → 與 15 指紋比對 → 取最高分 → 若 >= threshold 視為漂移版。
"""
import os
import openpyxl

from delta_unified_reader import (
    find_valid_sheets, scan_headers, find_first_date_col, collect_date_cols,
    find_marker_col, classify_marker, _read_header_row, _get_cell_value,
)


# ============ 預先 compute 的 15 格式指紋 ============
# 由 build_fingerprints.py 從 15 個原檔產生 (見模組底部)
FINGERPRINTS = {}  # 由 _load_fingerprints() 填入


# ============ Sheet 簽名規則 (與 detect_format 對齊) ============
def canonicalize_sheets(sheet_names):
    """把 sheet 名稱集合 → 簽名字串 (與 detect_format 偵測順序一致)"""
    s = set(sheet_names)
    if 'Diode' in s and 'MOS' in s:
        return 'DIODE_MOS'
    if 'MRP' in s:
        return 'MRP'
    for name in s:
        if str(name).startswith('PSB9_MRP'):
            return 'PSB9_MRP_DYNAMIC'
    if 'PAN JIT' in s:
        return 'PAN_JIT'
    if 'Sheet1' in s:
        return 'SHEET1'
    return 'OTHER'


# ============ 指紋抽取 ============
def extract_fingerprint(filepath):
    """從檔案抽出結構指紋 (6 維)。

    回傳 dict 或 None (檔案無合格 sheet/欄位)。
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception:
        return None

    try:
        sheet_signature = canonicalize_sheets(wb.sheetnames)
        sheets = find_valid_sheets(wb)
        if not sheets:
            # 某些檔案 read_only 模式 max_row=None → 改用一般模式重試
            wb.close()
            try:
                wb = openpyxl.load_workbook(filepath, data_only=True)
                sheets = find_valid_sheets(wb)
            except Exception:
                return None
        if not sheets:
            return None

        sheet_name, header_row = sheets[0]
        ws = wb[sheet_name]
        header_values = _read_header_row(ws, header_row)
        found, headers = scan_headers(header_values)
        if 'partno' not in found:
            return None
        date_start = find_first_date_col(headers, header_values)
        date_col_map = collect_date_cols(date_start, header_values)
        if not date_col_map:
            return None

        # 抽前 50 列做 marker 偵測 + layout 判斷
        rows_sample = []
        for row_values in ws.iter_rows(min_row=header_row + 1, values_only=True):
            rows_sample.append(row_values)
            if len(rows_sample) >= 50:
                break
        marker_col = find_marker_col(rows_sample)

        # 推算 layout: 同一 partno 對應幾列 marker
        partno_col = found['partno']
        if marker_col is None:
            layout = 'flat'
        else:
            groups = {}
            last_partno = None
            for row_values in rows_sample:
                p = _get_cell_value(row_values, partno_col)
                if p and str(p).strip():
                    last_partno = str(p).strip()
                if last_partno:
                    m = classify_marker(_get_cell_value(row_values, marker_col))
                    if m and m != 'skip':
                        groups[last_partno] = groups.get(last_partno, 0) + 1
            if groups:
                avg = sum(groups.values()) / len(groups)
                if avg < 1.5:
                    layout = 'flat'
                elif avg < 3.5:
                    layout = 'multirow_3'
                elif avg < 4.5:
                    layout = 'multirow_4'
                else:
                    layout = 'multirow_5'
            else:
                layout = 'flat'

        return {
            'sheet_signature': sheet_signature,
            'partno_col': found.get('partno'),
            'date_start_col': date_start,
            'date_count': len(date_col_map),
            'marker_col': marker_col,
            'layout': layout,
        }
    finally:
        wb.close()


# ============ 比對演算法 ============
def _score(new_fp, ref_fp):
    """計算 new_fp 與 ref_fp 的相似度分數 (連續分數, 自我比對 = 100)。

    每個維度都用「越近越高分」的連續函數, 避免階梯式造成 tie。
    """
    score = 0

    # 1. sheet_signature: 完全一致 +30 (主要鑑別點)
    if new_fp['sheet_signature'] == ref_fp['sheet_signature']:
        score += 30

    # 2. partno_col: 連續分數, max 25 (diff=0 → 25, 每差 1 欄扣 5, 最多扣到 0)
    diff = abs((new_fp.get('partno_col') or 0) - (ref_fp.get('partno_col') or 0))
    score += max(0, 25 - diff * 5)

    # 3. date_start_col: 連續分數, max 20 (每差 1 欄扣 4)
    diff = abs((new_fp.get('date_start_col') or 0) - (ref_fp.get('date_start_col') or 0))
    score += max(0, 20 - diff * 4)

    # 4. date_count: ratio 越接近 1 越高分, max 15
    nc = new_fp.get('date_count', 0) or 0
    rc = ref_fp.get('date_count', 0) or 0
    if rc > 0 and nc > 0:
        ratio = nc / rc if nc < rc else rc / nc  # 永遠取 <= 1
        score += int(ratio * 15)

    # 5. marker_col: 兩邊皆 None → +10; 兩邊非 None → 連續分數 max 10
    nm = new_fp.get('marker_col')
    rm = ref_fp.get('marker_col')
    if nm is None and rm is None:
        score += 10
    elif nm is not None and rm is not None:
        diff = abs(nm - rm)
        score += max(0, 10 - diff * 3)

    # 6. layout: 一致 +10
    if new_fp.get('layout') == ref_fp.get('layout'):
        score += 10

    return score


def match_known_format_fingerprint(filepath, threshold=70):
    """嘗試把檔案匹配到已知 15 格式之一 (寬鬆版 detect_format)。

    Args:
        filepath: 檔案路徑
        threshold: 最低分數 (預設 70 分)

    Returns:
        tuple(fmt_or_None, score)
        - fmt: 匹配到的格式常數 (與 delta_forecast_processor.FORMAT_* 一致)
        - score: 0-100
    """
    new_fp = extract_fingerprint(filepath)
    if new_fp is None:
        return None, 0

    if not FINGERPRINTS:
        return None, 0

    best_fmt, best_score = None, 0
    for fmt, ref_fp in FINGERPRINTS.items():
        s = _score(new_fp, ref_fp)
        if s > best_score:
            best_fmt, best_score = fmt, s

    if best_score >= threshold:
        return best_fmt, best_score

    # ── AI fallback: 指紋分數不足 → 用 AI 確認 layout 類型後重新比對 ──
    try:
        from delta_ai_helper import ai_analyze_file
        ai = ai_analyze_file(filepath)
        if ai and ai.get('identified'):
            ai_layout = ai.get('format_type', 'unknown')
            # 在與 AI layout 一致的格式中找最高分, 門檻降低 25 分
            constrained_best_fmt, constrained_best_score = None, 0
            for fmt, ref_fp in FINGERPRINTS.items():
                if ref_fp.get('layout') != ai_layout:
                    continue
                s = _score(new_fp, ref_fp)
                if s > constrained_best_score:
                    constrained_best_fmt, constrained_best_score = fmt, s
            ai_threshold = max(45, threshold - 25)
            if constrained_best_score >= ai_threshold:
                print(f"  [AI+FP] {os.path.basename(filepath)}: "
                      f"指紋={constrained_best_score} AI={ai_layout} → {constrained_best_fmt}")
                return constrained_best_fmt, constrained_best_score
    except Exception:
        pass

    return None, best_score


# ============ 預先 compute 的指紋表 (15 格式 × 6 維) ============
# 來源: 對 D:/tmp 下 15 個原檔 + delta_forecast_processor.FORMAT_* 對照, 由 build_fingerprints.py 產生.
# 若新增格式 → 更新此表.
def _load_fingerprints():
    """填入 15 個格式的預設指紋。"""
    # 由 build_fingerprints.py 從 D:/tmp 下 15 個原檔產生 (14 個 fmt, EIBG_EISBG 有 2 樣本但同 fmt)
    return {
        'eibg_eisbg':          {'sheet_signature': 'SHEET1',           'partno_col': 3,  'date_start_col': 12, 'date_count': 27, 'marker_col': None, 'layout': 'flat'},
        'fmbg':                {'sheet_signature': 'SHEET1',           'partno_col': 5,  'date_start_col': 16, 'date_count': 18, 'marker_col': 12,   'layout': 'multirow_3'},
        'iabg':                {'sheet_signature': 'SHEET1',           'partno_col': 4,  'date_start_col': 13, 'date_count': 22, 'marker_col': None, 'layout': 'flat'},
        'ictbg_ntl7':          {'sheet_signature': 'SHEET1',           'partno_col': 2,  'date_start_col': 13, 'date_count': 57, 'marker_col': 10,   'layout': 'multirow_3'},
        'ictbg_psb9_mrp':      {'sheet_signature': 'PSB9_MRP_DYNAMIC', 'partno_col': 3,  'date_start_col': 15, 'date_count': 31, 'marker_col': 14,   'layout': 'multirow_3'},
        'ictbg_psb9_siriraht': {'sheet_signature': 'SHEET1',           'partno_col': 4,  'date_start_col': 16, 'date_count': 30, 'marker_col': 15,   'layout': 'multirow_3'},
        'svc1pwc1_diode_mos':  {'sheet_signature': 'DIODE_MOS',        'partno_col': 3,  'date_start_col': 9,  'date_count': 29, 'marker_col': None, 'layout': 'flat'},
        'nbq1':                {'sheet_signature': 'PAN_JIT',          'partno_col': 1,  'date_start_col': 16, 'date_count': 29, 'marker_col': None, 'layout': 'flat'},
        'india_iai1':          {'sheet_signature': 'PAN_JIT',          'partno_col': 4,  'date_start_col': 14, 'date_count': 22, 'marker_col': 13,   'layout': 'multirow_3'},
        'psw1_cew1':           {'sheet_signature': 'SHEET1',           'partno_col': 6,  'date_start_col': 14, 'date_count': 29, 'marker_col': 12,   'layout': 'multirow_4'},
        'ketwadee':            {'sheet_signature': 'MRP',              'partno_col': 3,  'date_start_col': 16, 'date_count': 31, 'marker_col': 15,   'layout': 'multirow_3'},
        'weeraya':             {'sheet_signature': 'SHEET1',           'partno_col': 4,  'date_start_col': 14, 'date_count': 31, 'marker_col': 12,   'layout': 'multirow_3'},
        'kanyanat':            {'sheet_signature': 'SHEET1',           'partno_col': 5,  'date_start_col': 25, 'date_count': 31, 'marker_col': 24,   'layout': 'multirow_3'},
        'mwc1ipc1':            {'sheet_signature': 'SHEET1',           'partno_col': 2,  'date_start_col': 9,  'date_count': 45, 'marker_col': 6,    'layout': 'multirow_3'},
        # PSBG 沒有真實樣本, 從 detect_format 程式碼推測 (col 15=filter, 同 SHEET1)
        'psbg':                {'sheet_signature': 'SHEET1',           'partno_col': 4,  'date_start_col': 16, 'date_count': 30, 'marker_col': 15,   'layout': 'multirow_3'},
    }


FINGERPRINTS = _load_fingerprints()
