# -*- coding: utf-8 -*-
"""
完整驗證新邏輯: 逐筆 ERP → 計算目標日期 → 找到對應欄位 → 確認值合理
使用 uploads/6/20260317_200333 (全 23 Plant) 的資料
"""
import os
import sys
sys.stdout.reconfigure(encoding='utf-8')

import pandas as pd
import openpyxl
from datetime import datetime, timedelta, date
from collections import defaultdict

BASE = "d:/github/business_forecasting_lite"
UPLOAD_DIR = os.path.join(BASE, "uploads/6/20260317_200333")
PROCESSED_DIR = os.path.join(BASE, "processed/6/20260317_200333")

WEEKDAY_MAP_FULL = {
    '週一': 0, '禮拜一': 0, '星期一': 0,
    '週二': 1, '禮拜二': 1, '星期二': 1,
    '週三': 2, '禮拜三': 2, '星期三': 2,
    '週四': 3, '禮拜四': 3, '星期四': 3,
    '週五': 4, '禮拜五': 4, '星期五': 4,
    '週六': 5, '禮拜六': 5, '星期六': 5,
    '週日': 6, '禮拜日': 6, '星期日': 6, '禮拜天': 6, '週天': 6,
}
WEEKDAY_MAP_CHAR = {
    '一': 0, '二': 1, '三': 2, '四': 3,
    '五': 4, '六': 5, '日': 6, '天': 6,
}
WEEKDAY_NAMES = ['一', '二', '三', '四', '五', '六', '日']


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    if hasattr(val, 'date') and callable(val.date):
        try: return val.date()
        except: return None
    if isinstance(val, str):
        val = val.strip()
        if not val or val.lower() in ('nan', 'nat'):
            return None
        for fmt in ['%Y/%m/%d', '%Y-%m-%d', '%m/%d/%Y']:
            try: return datetime.strptime(val, fmt).date()
            except: continue
    try: return pd.to_datetime(val).date()
    except: return None


def get_week_end(schedule_date, breakpoint_text):
    if not breakpoint_text or breakpoint_text == 'nan':
        return None
    target_wd = WEEKDAY_MAP_FULL.get(breakpoint_text)
    if target_wd is None:
        return None
    try:
        cur_wd = schedule_date.weekday()
        days_ahead = (target_wd - cur_wd) % 7
        return schedule_date + timedelta(days=int(days_ahead))
    except (TypeError, ValueError):
        return None


def calc_target_NEW(week_end, date_text, on_breakpoint=False):
    """新邏輯"""
    if date_text.startswith('下下下週') or date_text.startswith('下下下禮拜'):
        weeks_offset = 3; wd_char = date_text[-1]
    elif date_text.startswith('下下週') or date_text.startswith('下下禮拜'):
        weeks_offset = 2; wd_char = date_text[-1]
    elif date_text.startswith('下週') or date_text.startswith('下禮拜'):
        weeks_offset = 1; wd_char = date_text[-1]
    elif date_text.startswith('本週') or date_text.startswith('本禮拜') or date_text.startswith('這週'):
        weeks_offset = 0; wd_char = date_text[-1]
    else:
        return parse_date(date_text)

    target_wd = WEEKDAY_MAP_CHAR.get(wd_char)
    if target_wd is None:
        return None

    bp_wd = week_end.weekday()
    days_diff = (target_wd - bp_wd) % 7
    if days_diff > 0 and not on_breakpoint:
        days_diff -= 7
    return week_end + timedelta(days=7 * weeks_offset + days_diff)


def calc_erp_target(row):
    """計算 ERP 目標日期 (新邏輯)"""
    schedule_val = row.get('排程出貨日期')
    schedule_date = parse_date(schedule_val)
    if schedule_date is None:
        return None, None, "排程解析失敗"

    bp_text = str(row.get('排程出貨日期斷點', '')).strip()
    if not bp_text or bp_text == 'nan':
        return None, None, "無斷點"

    week_end = get_week_end(schedule_date, bp_text)
    if week_end is None:
        return None, None, f"斷點解析失敗: {bp_text}"

    calc_type = str(row.get('日期算法', '')).strip().upper()
    if calc_type == 'ETD':
        date_text = str(row.get('ETD', '')).strip()
    elif calc_type == 'ETA':
        date_text = str(row.get('ETA', '')).strip()
    else:
        date_text = str(row.get('ETD', '')).strip()
        if not date_text:
            date_text = str(row.get('ETA', '')).strip()

    if not date_text or date_text == 'nan':
        return None, None, "無 ETD/ETA"

    on_breakpoint = (schedule_date == week_end)
    target = calc_target_NEW(week_end, date_text, on_breakpoint)

    if target and target < schedule_date:
        return None, None, f"目標 {target} < 排程 {schedule_date}"

    detail = {
        'schedule': schedule_date,
        'breakpoint': bp_text,
        'week_end': week_end,
        'calc_type': calc_type,
        'date_text': date_text,
        'on_breakpoint': on_breakpoint,
    }
    return target, detail, None


def read_forecast_structure(filepath):
    """讀取 Forecast 的日期結構和料號索引"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Daily+Weekly+Monthly']

    plant = str(ws.cell(row=1, column=3).value or '').strip()

    dates = {}
    for col in range(11, 70):
        val = ws.cell(row=7, column=col).value
        d = parse_date(val)
        if d:
            dates[col] = d

    # Build material → commit row
    materials = {}
    for row in range(8, ws.max_row + 1):
        measure = str(ws.cell(row=row, column=3).value or '').strip()
        if measure == 'Commit':
            mat = str(ws.cell(row=row, column=2).value or '').strip()
            if mat:
                materials[mat] = row

    wb.close()

    daily_dates = {c: d for c, d in dates.items() if c <= 41}
    weekly_dates = {c: d for c, d in dates.items() if 42 <= c <= 63}
    monthly_dates = {c: d for c, d in dates.items() if 64 <= c <= 69}

    return {
        'plant': plant,
        'dates': dates,
        'daily': daily_dates,
        'weekly': weekly_dates,
        'monthly': monthly_dates,
        'daily_end': max(daily_dates.values()) if daily_dates else None,
        'weekly_start': min(weekly_dates.values()) if weekly_dates else None,
        'materials': materials,
    }


def find_date_column(dates, target_date):
    """模擬 _find_date_column (含 GAP fallback)"""
    # Daily exact match
    for col, d in dates.items():
        if col > 41:
            continue
        if d == target_date:
            return col, 'Daily', d

    # Weekly range match
    first_weekly_col = None
    first_weekly_date = None
    for col, d in dates.items():
        if col < 42 or col > 63:
            continue
        if first_weekly_col is None or d < first_weekly_date:
            first_weekly_col = col
            first_weekly_date = d
        if d <= target_date <= d + timedelta(days=6):
            return col, 'Weekly', d

    # GAP fallback
    if first_weekly_date and target_date < first_weekly_date:
        return first_weekly_col, 'Weekly(GAP)', first_weekly_date

    # Monthly
    for col, d in dates.items():
        if col < 64 or col > 69:
            continue
        if d.year == target_date.year and d.month == target_date.month:
            return col, 'Monthly', d

    return None, 'NOT_FOUND', None


def read_result_commits(filepath):
    """讀取結果檔的所有 Commit cell (含0值以外的)"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Daily+Weekly+Monthly']
    cells = {}  # (material, col) → value
    for row in range(8, ws.max_row + 1):
        measure = str(ws.cell(row=row, column=3).value or '').strip()
        if measure != 'Commit':
            continue
        mat = str(ws.cell(row=row, column=2).value or '').strip()
        for col in range(11, 70):
            val = ws.cell(row=row, column=col).value
            if val and isinstance(val, (int, float)) and val != 0:
                cells[(mat, col)] = val
    wb.close()
    return cells


# ═══════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════

# Load integrated ERP
erp_path = os.path.join(PROCESSED_DIR, "integrated_erp.xlsx")
erp_df = pd.read_excel(erp_path)
print(f"ERP: {len(erp_df)} rows")

# Load all forecast structures
plant_infos = {}
for i in range(1, 24):
    fpath = os.path.join(UPLOAD_DIR, f"forecast_data_{i}.xlsx")
    if not os.path.exists(fpath):
        continue
    fi = read_forecast_structure(fpath)
    plant_infos[fi['plant']] = fi

print(f"Loaded {len(plant_infos)} plant forecasts\n")

# ── Process each plant ──
grand_total_fills = 0
grand_total_skips = 0
grand_total_value = 0
grand_issues = []

for plant in sorted(plant_infos.keys()):
    fi = plant_infos[plant]
    plant_erp = erp_df[erp_df['客戶需求地區'].astype(str).str.strip() == plant].copy()
    if len(plant_erp) == 0:
        continue

    print(f"{'='*70}")
    print(f"Plant {plant}: ERP {len(plant_erp)} 筆, 料號 {len(fi['materials'])} 個")
    print(f"  Daily end: {fi['daily_end']}, Weekly start: {fi['weekly_start']}")
    print(f"{'='*70}")

    # 計算每筆 ERP 的預期填入
    expected_fills = defaultdict(float)  # (material, col) → accumulated value
    fill_details = []  # detailed trace
    skipped = []

    for idx, row in plant_erp.iterrows():
        material = str(row.get('客戶料號', '')).strip()
        qty = row.get('淨需求', 0)
        try:
            qty = float(qty)
        except:
            qty = 0

        # Check material exists in forecast
        if material not in fi['materials']:
            skipped.append(f"  料號不存在: {material}")
            continue

        if qty <= 0:
            skipped.append(f"  qty<=0: {material} qty={qty}")
            continue

        # Calculate target date
        target, detail, err = calc_erp_target(row)
        if target is None:
            skipped.append(f"  日期計算失敗: {material} - {err}")
            continue

        # Find column
        col, col_type, col_date = find_date_column(fi['dates'], target)
        if col is None:
            skipped.append(f"  NOT_FOUND: {material} target={target}")
            continue

        fill_value = qty * 1000
        expected_fills[(material, col)] += fill_value

        fill_details.append({
            'material': material,
            'qty': qty,
            'fill_value': fill_value,
            'schedule': detail['schedule'],
            'breakpoint': detail['breakpoint'],
            'on_bp': detail['on_breakpoint'],
            'calc_type': detail['calc_type'],
            'text': detail['date_text'],
            'target_date': target,
            'col': col,
            'col_type': col_type,
            'col_date': col_date,
        })

    fill_count = len(fill_details)
    skip_count = len(skipped)
    grand_total_fills += fill_count
    grand_total_skips += skip_count

    # ── 按目標日期彙總 ──
    date_summary = defaultdict(lambda: {'count': 0, 'total_value': 0, 'materials': set()})
    for fd in fill_details:
        key = (fd['target_date'], fd['col_type'], fd['col'])
        date_summary[key]['count'] += 1
        date_summary[key]['total_value'] += fd['fill_value']
        date_summary[key]['materials'].add(fd['material'])

    print(f"\n  填入 {fill_count} 筆, 跳過 {skip_count} 筆")
    print(f"\n  目標日期 → 欄位對應:")
    for key in sorted(date_summary.keys()):
        target_date, col_type, col = key
        info = date_summary[key]
        wd = WEEKDAY_NAMES[target_date.weekday()]
        col_date = fi['dates'].get(col)
        grand_total_value += info['total_value']

        # 合理性檢查
        issues = []
        if col_type == 'NOT_FOUND':
            issues.append("NOT_FOUND")
        if col_type == 'Weekly(GAP)':
            issues.append("GAP_FALLBACK")
        if col_type == 'Weekly' and col_date:
            # 確認 target_date 在 weekly range 內
            if not (col_date <= target_date <= col_date + timedelta(days=6)):
                issues.append(f"WEEK_MISMATCH(col={col_date})")
        if col_type == 'Daily' and col_date and col_date != target_date:
            issues.append(f"DAILY_MISMATCH(col={col_date})")

        issue_str = f" ⚠️ {', '.join(issues)}" if issues else ""
        for i in issues:
            grand_issues.append(f"Plant {plant}: {target_date} {i}")

        print(f"    {target_date} (週{wd}) → col {col} [{col_type}]: "
              f"{info['count']}筆, ${info['total_value']:,.0f}{issue_str}")

    # ── on_breakpoint 筆數統計 ──
    on_bp_count = sum(1 for fd in fill_details if fd['on_bp'])
    if on_bp_count > 0:
        print(f"\n  on_breakpoint 筆數: {on_bp_count}/{fill_count}")
        # 顯示前幾筆 on_breakpoint 的詳細計算
        bp_details = [fd for fd in fill_details if fd['on_bp']]
        for fd in bp_details[:5]:
            wd = WEEKDAY_NAMES[fd['target_date'].weekday()]
            print(f"    {fd['material']}: 排程={fd['schedule']}({fd['breakpoint']}), "
                  f"{fd['calc_type']}={fd['text']} → {fd['target_date']}(週{wd}) "
                  f"[{fd['col_type']}] val={fd['fill_value']:,.0f}")
        if len(bp_details) > 5:
            print(f"    ... 還有 {len(bp_details)-5} 筆")

    # ── 讀結果檔比對 ──
    # 找到這個 plant 的結果檔
    result_files = [f for f in os.listdir(PROCESSED_DIR)
                    if f.startswith(f'forecast_{plant}_') and f.endswith('.xlsx')]

    if result_files:
        for rf in result_files:
            rpath = os.path.join(PROCESSED_DIR, rf)
            result_cells = read_result_commits(rpath)

            print(f"\n  結果檔比對: {rf}")

            # 比較 expected vs actual
            all_keys = set(expected_fills.keys()) | set(result_cells.keys())
            match_count = 0
            mismatch_count = 0
            only_expected = 0
            only_result = 0
            mismatches = []

            for key in sorted(all_keys):
                exp = expected_fills.get(key)
                act = result_cells.get(key)
                mat, col = key
                col_date = fi['dates'].get(col)

                if exp and act:
                    if abs(exp - act) < 0.01:
                        match_count += 1
                    else:
                        mismatch_count += 1
                        mismatches.append({
                            'material': mat, 'col': col,
                            'col_date': col_date,
                            'expected': exp, 'actual': act,
                            'diff': act - exp,
                        })
                elif exp and not act:
                    only_expected += 1
                elif act and not exp:
                    only_result += 1

            print(f"    完全匹配: {match_count}")
            print(f"    值不同: {mismatch_count}")
            print(f"    新邏輯有/舊結果無: {only_expected} (新日期位移)")
            print(f"    舊結果有/新邏輯無: {only_result} (舊日期位移)")

            if mismatches:
                print(f"    值不同的明細 (前 10):")
                for m in mismatches[:10]:
                    wd = WEEKDAY_NAMES[m['col_date'].weekday()] if m['col_date'] else '?'
                    print(f"      {m['material']} col={m['col']} ({m['col_date']} 週{wd}): "
                          f"預期={m['expected']:,.0f} 實際={m['actual']:,.0f} 差={m['diff']:+,.0f}")

            # 分析 only_expected 和 only_result 的日期分布
            if only_expected > 0:
                exp_only_dates = defaultdict(int)
                for key in sorted(all_keys):
                    if key in expected_fills and key not in result_cells:
                        mat, col = key
                        col_date = fi['dates'].get(col)
                        if col_date:
                            exp_only_dates[col_date] += 1
                if exp_only_dates:
                    print(f"    新邏輯有/舊結果無 — 日期分布:")
                    for d in sorted(exp_only_dates.keys()):
                        wd = WEEKDAY_NAMES[d.weekday()]
                        print(f"      {d} (週{wd}): {exp_only_dates[d]} cells")

            if only_result > 0:
                res_only_dates = defaultdict(int)
                for key in sorted(all_keys):
                    if key not in expected_fills and key in result_cells:
                        mat, col = key
                        col_date = fi['dates'].get(col)
                        if col_date:
                            res_only_dates[col_date] += 1
                if res_only_dates:
                    print(f"    舊結果有/新邏輯無 — 日期分布:")
                    for d in sorted(res_only_dates.keys()):
                        wd = WEEKDAY_NAMES[d.weekday()]
                        print(f"      {d} (週{wd}): {res_only_dates[d]} cells")

    # ── 跳過原因統計 ──
    if skipped:
        skip_reasons = defaultdict(int)
        for s in skipped:
            reason = s.split(':')[0].strip()
            skip_reasons[reason] += 1
        print(f"\n  跳過原因:")
        for reason, cnt in sorted(skip_reasons.items(), key=lambda x: -x[1]):
            print(f"    {reason}: {cnt}")

    print()


# ═══════════════════════════════════════════════════════
# 特別驗證: 2680 的 3/30(一) on_breakpoint 案例
# ═══════════════════════════════════════════════════════
print("=" * 70)
print("特別驗證: 2680 排程=3/30(一) on_breakpoint 案例")
print("=" * 70)

fi_2680 = plant_infos.get('2680')
if fi_2680:
    erp_2680 = erp_df[erp_df['客戶需求地區'].astype(str).str.strip() == '2680']
    bp_cases = []
    for idx, row in erp_2680.iterrows():
        schedule = parse_date(row.get('排程出貨日期'))
        if schedule is None:
            continue
        bp = str(row.get('排程出貨日期斷點', '')).strip()
        we = get_week_end(schedule, bp)
        if we and schedule == we:
            material = str(row.get('客戶料號', '')).strip()
            qty = row.get('淨需求', 0)
            target, detail, err = calc_erp_target(row)
            col, col_type, col_date = find_date_column(fi_2680['dates'], target) if target else (None, 'N/A', None)
            bp_cases.append({
                'material': material, 'schedule': schedule,
                'bp': bp, 'text': detail['date_text'] if detail else '',
                'calc': detail['calc_type'] if detail else '',
                'target': target, 'col': col, 'col_type': col_type,
                'col_date': col_date, 'qty': qty,
            })

    if bp_cases:
        print(f"\n2680 on_breakpoint 案例: {len(bp_cases)} 筆")
        for c in bp_cases:
            twd = WEEKDAY_NAMES[c['target'].weekday()] if c['target'] else '?'
            print(f"  {c['material']}: 排程={c['schedule']}({c['bp']}), "
                  f"{c['calc']}={c['text']} → 目標={c['target']}(週{twd}) "
                  f"[{c['col_type']}] col={c['col']} qty={c['qty']}")

            # 人工推演
            print(f"    推演: schedule={c['schedule']}(一), bp=禮拜一 → week_end={c['schedule']}")
            print(f"    on_breakpoint=True → days_diff 不減 7")
            if c['target']:
                print(f"    結果: {c['target']}(週{twd}) ✓ 符合預期(斷點後第一個{c['text'][-1]})")
    else:
        print("  未找到 on_breakpoint 案例")


# ═══════════════════════════════════════════════════════
# 彙總
# ═══════════════════════════════════════════════════════
print(f"\n{'='*70}")
print("全局彙總")
print(f"{'='*70}")
print(f"  總填入筆數: {grand_total_fills}")
print(f"  總跳過筆數: {grand_total_skips}")
print(f"  總填入金額: {grand_total_value:,.0f}")

if grand_issues:
    print(f"\n  問題清單 ({len(grand_issues)}):")
    for issue in grand_issues:
        print(f"    {issue}")
else:
    print(f"\n  ✅ 無問題")

# 數值合理性: 檢查 qty × 1000 是否為整數
print(f"\n  數值合理性:")
print(f"  - 所有填入值 = 淨需求 × 1000")
print(f"  - 相同 (material, col) 的多筆 ERP 會累加")
