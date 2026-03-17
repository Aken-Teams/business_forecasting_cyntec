"""
Verification script for Liteon forecast date calculation fix.

Tests the corrected logic where breakpoint is the week START (not END):
  days_diff = (target_weekday - breakpoint_weekday) % 7
  # No subtraction — breakpoint is week START

Example: schedule 3/30 (Mon), breakpoint=禮拜一, ETA=下週四
  Old: 3/30 + (7*1 + (-4)) = 4/2  (WRONG)
  New: 3/30 + (7*1 + 3)   = 4/9  (CORRECT)
"""

import sys
import os
import shutil
import tempfile

sys.stdout.reconfigure(encoding='utf-8')

import pandas as pd
from datetime import datetime, timedelta

# ── 1. Setup paths ──────────────────────────────────────────────
BASE_DIR = r'D:\github\business_forecasting_lite'
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads', '6', '20260317_195848')
ERP_FILE = os.path.join(UPLOAD_DIR, 'erp_data.xlsx')
FORECAST_FILE = os.path.join(UPLOAD_DIR, 'forecast_data.xlsx')

sys.path.insert(0, BASE_DIR)
from liteon_forecast_processor import LiteonForecastProcessor

# ── 2. Get mapping data from database ──────────────────────────
from database import get_customer_mappings_raw

raw_mappings = get_customer_mappings_raw(6)
print(f"Loaded {len(raw_mappings)} mapping records from database")

# Build Liteon lookup tables (same logic as app.py lines 3014-3035)
liteon_lookup_11 = {}
liteon_lookup_32 = {}
for m in raw_mappings:
    cname = str(m.get('customer_name', '')).strip()
    order_type = str(m.get('order_type', '')).strip()
    delivery_loc = str(m.get('delivery_location', '')).strip() if m.get('delivery_location') else ''
    warehouse = str(m.get('warehouse', '')).strip() if m.get('warehouse') else ''

    mapping_values = {
        'region': str(m['region']).strip() if m['region'] else '',
        'schedule_breakpoint': str(m['schedule_breakpoint']).strip() if m['schedule_breakpoint'] else '',
        'etd': str(m['etd']).strip() if m['etd'] else '',
        'eta': str(m['eta']).strip() if m['eta'] else '',
        'date_calc_type': str(m.get('date_calc_type', '')).strip() if m.get('date_calc_type') else ''
    }

    if order_type == '11' and delivery_loc:
        liteon_lookup_11[(cname, delivery_loc)] = mapping_values
    elif order_type == '32' and warehouse:
        liteon_lookup_32[(cname, warehouse)] = mapping_values

print(f"Built {len(liteon_lookup_11)} type-11 lookups + {len(liteon_lookup_32)} type-32 lookups")

# ── 3. Enrich ERP data (replicate app.py logic) ────────────────
erp_df = pd.read_excel(ERP_FILE, engine='openpyxl')
print(f"Loaded ERP: {len(erp_df)} rows")

customer_col = '客戶簡稱'
delivery_col = '送貨地點'
order_type_col = '訂單型態'
warehouse_col = '倉庫'


def get_liteon_mapping(row, field):
    customer = str(row[customer_col]).strip() if pd.notna(row[customer_col]) else ''
    order_type_val = str(row[order_type_col]).strip() if pd.notna(row[order_type_col]) else ''
    ot_prefix = order_type_val[:2] if len(order_type_val) >= 2 else order_type_val

    if ot_prefix == '11':
        delivery = str(row[delivery_col]).strip() if pd.notna(row[delivery_col]) else ''
        key = (customer, delivery)
        mapping = liteon_lookup_11.get(key, {})
    elif ot_prefix == '32':
        wh = str(row[warehouse_col]).strip() if warehouse_col and pd.notna(row[warehouse_col]) else ''
        key = (customer, wh)
        mapping = liteon_lookup_32.get(key, {})
    else:
        mapping = {}

    return mapping.get(field, '') if mapping else ''


erp_df['客戶需求地區'] = erp_df.apply(lambda row: get_liteon_mapping(row, 'region'), axis=1)
erp_df['排程出貨日期斷點'] = erp_df.apply(lambda row: get_liteon_mapping(row, 'schedule_breakpoint'), axis=1)
erp_df['ETD'] = erp_df.apply(lambda row: get_liteon_mapping(row, 'etd'), axis=1)
erp_df['ETA'] = erp_df.apply(lambda row: get_liteon_mapping(row, 'eta'), axis=1)
erp_df['日期算法'] = erp_df.apply(lambda row: get_liteon_mapping(row, 'date_calc_type'), axis=1)

matched_count = (erp_df['客戶需求地區'] != '').sum()
print(f"Mapping matched: {matched_count}/{len(erp_df)}")
print(f"  客戶需求地區 values: {erp_df['客戶需求地區'].unique().tolist()}")
print(f"  排程出貨日期斷點 values: {erp_df['排程出貨日期斷點'].unique().tolist()}")
print(f"  日期算法 values: {erp_df['日期算法'].unique().tolist()}")
print(f"  ETD values: {erp_df['ETD'].unique().tolist()}")
print(f"  ETA values: {erp_df['ETA'].unique().tolist()}")

# ── 4. Manual date calculation verification ─────────────────────
print("\n" + "=" * 90)
print("MANUAL DATE CALCULATION VERIFICATION")
print("=" * 90)

# Use the processor's internal methods for date calculation
proc_temp = LiteonForecastProcessor.__new__(LiteonForecastProcessor)

WEEKDAY_NAMES = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']


def parse_date(val):
    """Parse date from string or datetime"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, datetime):
        return val.date()
    if hasattr(val, 'date') and callable(val.date):
        return val.date()
    if isinstance(val, str):
        val = val.strip()
        for fmt in ['%Y/%m/%d', '%Y-%m-%d', '%m/%d/%Y']:
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    try:
        return pd.to_datetime(val).date()
    except:
        return None


print(f"\n{'Idx':<5} {'Part#':<14} {'Schedule':<12} {'Breakpoint':<10} {'WeekEnd':<12} {'算法':<6} {'Text':<10} {'Target':<12} {'Weekday':<10} {'Qty':<8}")
print("-" * 100)

sample_count = 0
for idx, row in erp_df.iterrows():
    part_number = str(row.get('客戶料號', '')).strip()
    schedule_str = row.get('排程出貨日期', '')
    breakpoint_text = str(row.get('排程出貨日期斷點', '')).strip()
    etd_text = str(row.get('ETD', '')).strip()
    eta_text = str(row.get('ETA', '')).strip()
    date_calc_type = str(row.get('日期算法', '')).strip().upper()
    qty = row.get('淨需求', 0)

    schedule_date = parse_date(schedule_str)
    if schedule_date is None:
        continue

    # Calculate week_end using the processor's method
    week_end = proc_temp._get_week_end_by_breakpoint(schedule_date, breakpoint_text)
    if week_end is None:
        continue

    # Determine which text to use
    if date_calc_type == 'ETD':
        date_text = etd_text
    elif date_calc_type == 'ETA':
        date_text = eta_text
    else:
        date_text = etd_text or eta_text

    if not date_text:
        continue

    # Calculate target date
    target_date = proc_temp._calculate_target_from_text(week_end, date_text)

    # Safety guard: target must not be before schedule
    guarded = ''
    if target_date is not None and target_date < schedule_date:
        guarded = ' [GUARDED->None]'
        target_date = None

    target_str = str(target_date) if target_date else 'None'
    weekday_str = WEEKDAY_NAMES[target_date.weekday()] if target_date else '-'

    if sample_count < 30:
        print(f"{idx:<5} {part_number:<14} {str(schedule_date):<12} {breakpoint_text:<10} "
              f"{str(week_end):<12} {date_calc_type:<6} {date_text:<10} "
              f"{target_str:<12} {weekday_str:<10} {qty:<8}{guarded}")
    sample_count += 1

print(f"\n(Showed first 30 of {sample_count} calculable rows)")

# ── 5. Verify specific example from the bug report ──────────────
print("\n" + "=" * 90)
print("BUG FIX VERIFICATION: breakpoint as week START")
print("=" * 90)


def verify_example(schedule_str, breakpoint_text, date_text, expected_str):
    schedule_date = datetime.strptime(schedule_str, '%Y-%m-%d').date()
    week_end = proc_temp._get_week_end_by_breakpoint(schedule_date, breakpoint_text)
    target = proc_temp._calculate_target_from_text(week_end, date_text)

    weekday_map_rev = {0: '一', 1: '二', 2: '三', 3: '四', 4: '五', 5: '六', 6: '日'}

    print(f"\n  Schedule: {schedule_date} ({WEEKDAY_NAMES[schedule_date.weekday()]})")
    print(f"  Breakpoint: {breakpoint_text}")
    print(f"  Week-end (anchor): {week_end} ({WEEKDAY_NAMES[week_end.weekday()]})")
    print(f"  Date text: {date_text}")
    print(f"  Calculated target: {target} ({WEEKDAY_NAMES[target.weekday()] if target else 'N/A'})")
    print(f"  Expected: {expected_str}")

    bp_wd = week_end.weekday()
    tgt_wd_char = date_text[-1]
    tgt_wd_map = {'一': 0, '二': 1, '三': 2, '四': 3, '五': 4, '六': 5, '日': 6, '天': 6}
    tgt_wd = tgt_wd_map.get(tgt_wd_char, -1)
    days_diff = (tgt_wd - bp_wd) % 7

    # Figure out weeks_offset from text
    if date_text.startswith('下下下'):
        weeks_offset = 3
    elif date_text.startswith('下下'):
        weeks_offset = 2
    elif date_text.startswith('下'):
        weeks_offset = 1
    else:
        weeks_offset = 0

    old_days_diff = days_diff - 7 if days_diff > 0 else days_diff
    old_target = week_end + timedelta(days=7 * weeks_offset + old_days_diff)
    new_target = week_end + timedelta(days=7 * weeks_offset + days_diff)

    print(f"\n  MATH BREAKDOWN:")
    print(f"    breakpoint_weekday = {bp_wd} ({WEEKDAY_NAMES[bp_wd]})")
    print(f"    target_weekday     = {tgt_wd} ({WEEKDAY_NAMES[tgt_wd]})")
    print(f"    days_diff = ({tgt_wd} - {bp_wd}) % 7 = {days_diff}")
    print(f"    weeks_offset = {weeks_offset}")
    print(f"    OLD formula: {week_end} + (7*{weeks_offset} + ({days_diff}-7)) = {week_end} + {7*weeks_offset + old_days_diff} = {old_target}")
    print(f"    NEW formula: {week_end} + (7*{weeks_offset} + {days_diff})     = {week_end} + {7*weeks_offset + days_diff} = {new_target}")

    expected_date = datetime.strptime(expected_str, '%Y-%m-%d').date()
    status = "PASS" if target == expected_date else "FAIL"
    print(f"\n  Result: {status} (target={target}, expected={expected_date})")
    return status == "PASS"


# Example from bug report
print("\nExample 1: schedule=3/30(Mon), breakpoint=禮拜一, ETA=下週四")
ok1 = verify_example('2026-03-30', '禮拜一', '下週四', '2026-04-09')

# More examples using actual data patterns
print("\nExample 2: schedule=3/17(Tue), breakpoint=禮拜一, ETA=下週四 (real data)")
# 3/17(Tue) -> breakpoint Mon -> anchor=3/23(Mon), 下週四 = 3/23 + 7 + 3 = 4/2
ok2 = verify_example('2026-03-17', '禮拜一', '下週四', '2026-04-02')

print("\nExample 3: schedule=3/17(Tue), breakpoint=禮拜一, ETD=本週五")
# 3/17(Tue) -> breakpoint Mon -> anchor=3/23(Mon), 本週五 = 3/23 + 0 + 4 = 3/27
ok3 = verify_example('2026-03-17', '禮拜一', '本週五', '2026-03-27')

print("\nExample 4: schedule=3/23(Mon), breakpoint=禮拜一, ETA=下週四")
ok4 = verify_example('2026-03-23', '禮拜一', '下週四', '2026-04-02')

print("\nExample 5: schedule=3/20(Fri), breakpoint=禮拜一, ETA=下週四")
# 3/20 is Friday. Breakpoint=Monday → _get_week_end_by_breakpoint moves to next Monday = 3/23.
# Then 下週四 = 3/23 + 7*1 + 3 = 4/2
ok5 = verify_example('2026-03-20', '禮拜一', '下週四', '2026-04-02')

all_pass = all([ok1, ok2, ok3, ok4, ok5])
print(f"\n{'='*60}")
print(f"Manual verification: {'ALL PASSED' if all_pass else 'SOME FAILED'}")
print(f"{'='*60}")

# ── 6. Create temp copy for processing (don't modify originals) ─
print("\n" + "=" * 90)
print("RUNNING FULL PROCESSOR")
print("=" * 90)

tmp_dir = tempfile.mkdtemp(prefix='liteon_verify_')
print(f"Temp directory: {tmp_dir}")

# Copy forecast file (we need to zero out existing Commit values for single-file mode)
tmp_forecast = os.path.join(tmp_dir, 'forecast_data.xlsx')
shutil.copy2(FORECAST_FILE, tmp_forecast)

# Zero out existing Commit values in the temp forecast copy
import openpyxl
wb_clean = openpyxl.load_workbook(tmp_forecast)
ws_clean = wb_clean['Daily+Weekly+Monthly']
commit_rows_cleaned = 0
for row_num in range(8, ws_clean.max_row + 1):
    measure = str(ws_clean.cell(row=row_num, column=3).value or '').strip()
    if measure == 'Commit':
        for col in range(11, 70):  # K through BQ
            cell = ws_clean.cell(row=row_num, column=col)
            if cell.value and cell.value != 0:
                cell.value = 0
                commit_rows_cleaned += 1
wb_clean.save(tmp_forecast)
print(f"Cleaned {commit_rows_cleaned} non-zero Commit cells in temp forecast")

# Save enriched ERP to temp
tmp_erp = os.path.join(tmp_dir, 'erp_data.xlsx')
erp_df['已分配'] = ''
erp_df.to_excel(tmp_erp, index=False)
print(f"Saved enriched ERP ({len(erp_df)} rows) to temp")

# Run processor
processor = LiteonForecastProcessor(
    forecast_file=tmp_forecast,
    erp_file=tmp_erp,
    transit_file=None,
    output_folder=tmp_dir,
    output_filename='forecast_result.xlsx',
    merged_mode=False
)

success = processor.process_all_blocks()

print(f"\nProcessor result: {'SUCCESS' if success else 'FAILURE'}")
print(f"ERP filled: {processor.total_filled}")
print(f"ERP skipped: {processor.total_skipped}")
print(f"Pending changes: {len(processor.pending_changes)}")

# ── 7. Show the pending changes (what was filled where) ─────────
print("\n" + "=" * 90)
print("FILLED CELLS DETAIL")
print("=" * 90)

# Build reverse date map
date_map = processor.date_map
rev_date_map = {col: date_obj for col, date_obj in date_map.items()}

# Build reverse material map
mat_map = processor.material_commit_rows
rev_mat_map = {v: k for k, v in mat_map.items()}

print(f"\n{'Material':<14} {'Row':<6} {'Col':<6} {'Date':<12} {'Value(x1000)':<14} {'Zone':<8}")
print("-" * 70)

changes_sorted = sorted(processor.pending_changes, key=lambda c: (c['row'], c['col']))
for change in changes_sorted[:50]:
    row = change['row']
    col = change['col']
    value = change['value']
    material = rev_mat_map.get(row, '?')
    date_obj = rev_date_map.get(col, None)
    date_str = str(date_obj) if date_obj else f'col{col}'

    # Determine zone
    if col <= processor.daily_end_col:
        zone = 'Daily'
    elif col <= processor.weekly_end_col:
        zone = 'Weekly'
    else:
        zone = 'Monthly'

    print(f"{material:<14} {row:<6} {col:<6} {date_str:<12} {value:<14.0f} {zone:<8}")

if len(changes_sorted) > 50:
    print(f"... ({len(changes_sorted) - 50} more changes)")

# ── 8. Summary stats ────────────────────────────────────────────
print("\n" + "=" * 90)
print("SUMMARY")
print("=" * 90)

daily_count = sum(1 for c in changes_sorted if c['col'] <= processor.daily_end_col)
weekly_count = sum(1 for c in changes_sorted if processor.daily_end_col < c['col'] <= processor.weekly_end_col)
monthly_count = sum(1 for c in changes_sorted if c['col'] > processor.weekly_end_col)

print(f"Total changes: {len(changes_sorted)}")
print(f"  Daily columns: {daily_count}")
print(f"  Weekly columns: {weekly_count}")
print(f"  Monthly columns: {monthly_count}")
print(f"Total quantity filled: {sum(c['value'] for c in changes_sorted):,.0f}")

# Cleanup
print(f"\nTemp files at: {tmp_dir}")
print("(not auto-deleted so you can inspect forecast_result.xlsx)")
