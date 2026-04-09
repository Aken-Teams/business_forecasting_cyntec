"""
台達 Forecast 匯總驗證腳本
=============================
驗證 3 個原始 Forecast 檔案能否整理成 0407匯總格式.xlsx

原始檔案:
  1. PSBG PSB5- Ketwadee0406(完成).xlsx  → Buyer=Ketwadee, Plant=PSB5
  2. PSBG PSB7_Kanyanat.S0406(完成).xlsx → Buyer=Kanyanat, Plant=PSB7
  3. PSBG PSB7-Weeraya0406(完成).xlsx    → Buyer=Weeraya,  Plant=PSB7

匯總格式欄位:
  A=Buyer, B=PLANT, C=ERP客戶簡稱, D=ERP送貨地點,
  E=PARTNO, F=VENDOR PARTNO, G=STOCK, H=ON-WAY,
  I=Date(Demand/Supply/Balance),
  J=PASSDUE, K~Z=週別日期, AA~AI=月份(JUL~MAR)

轉換規則:
  Ketwadee (PSB5): Demand/Supply 直接複製, Balance 重新計算
    - col O(Filter): Demand→Demand, Supply→Supply, Net→(不使用,重算)
    - STOCK = col H
    - Balance[PASSDUE] = STOCK - Demand[PASSDUE]
    - Balance[n] = Balance[n-1] - Demand[n] + Supply[n-1]

  Kanyanat (PSB7): 4-row groups → 3-row groups (REMARK 刪除)
    - col X(TYPE): A-Demand→Demand, B-Supply→Supply, D-Net Demand→(不使用,重算)
    - STOCK 不使用 (None), Balance 從 0 開始
    - Balance 公式同上

  Weeraya (PSB7): 5-row groups → 3-row groups (Firmed order, Remark 刪除)
    - col L(TYPE): Demand→Demand, Forecast Conf→Supply, Net Demand→Balance
    - STOCK = col M (Total stock)
    - Balance 直接從 Net Demand 複製 (因為源頭已用相同公式計算)
"""

import openpyxl
import sys
import os
from datetime import datetime
from collections import OrderedDict

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = r'C:\Users\petty\Desktop\客戶相關資料\01.強茂\台達業務'

# 匯總格式的日期欄位 (col J ~ col AI)
CONSOLIDATED_DATE_COLS = [
    'PASSDUE', '20260406', '20260413', '20260420', '20260427',
    '20260504', '20260511', '20260518', '20260525',
    '20260601', '20260608', '20260615', '20260622', '20260629',
    '20260706', '20260713', '20260720',
    'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC', 'JAN', 'FEB', 'MAR'
]


def normalize_date_header(val):
    """將不同源頭的日期格式統一為匯總格式"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y%m%d')
    s = str(val).strip()
    # "PAST DUE" or "PASSDUE" → "PASSDUE"
    if 'PAST' in s.upper() or 'PASSDUE' in s.upper():
        return 'PASSDUE'
    # "OVER DUE" → "PASSDUE"
    if 'OVER DUE' in s.upper():
        return 'PASSDUE'
    # "20260406" format
    if s.isdigit() and len(s) == 8:
        return s
    # "2026-JUL" → "JUL", "2027-JAN" → "JAN"
    # 但排除超過 2027-MAR 的月份 (2027-APR ~ 2027-AUG 不在匯總中)
    if '-' in s and len(s) >= 8:
        parts = s.split('-')
        if len(parts) == 2 and len(parts[1]) == 3:
            month = parts[1].upper()
            year = parts[0]
            # 匯總只包含: 2026-JUL~DEC, 2027-JAN~MAR
            if year == '2026' and month in ('JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC'):
                return month
            elif year == '2027' and month in ('JAN', 'FEB', 'MAR'):
                return month
            else:
                return None  # 2027-APR以後不在匯總格式中
    return s.upper()


def read_ketwadee(filepath):
    """讀取 PSB5 Ketwadee 資料"""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['MRP']

    # 讀取表頭, 建立日期欄位對應
    headers = {}
    date_col_map = {}  # col_idx → normalized_date
    for cell in ws[1]:
        if cell.value is not None:
            headers[cell.column] = cell.value
            if cell.column >= 16:  # P onwards = date columns
                norm = normalize_date_header(cell.value)
                if norm in CONSOLIDATED_DATE_COLS:
                    date_col_map[cell.column] = norm

    results = []
    max_col = ws.max_column

    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row,
                             min_col=1, max_col=max_col, values_only=False))
    i = 0
    while i < len(rows):
        row = rows[i]
        filter_val = row[14].value if len(row) > 14 else None  # col O

        if filter_val == 'Demand':
            part_no = row[2].value   # C = Part No
            vendor_part = row[3].value  # D = Vendor Part
            stock = row[7].value or 0   # H = STOCK

            # 讀取 Demand 數據
            demand = {}
            for col_idx, date_key in date_col_map.items():
                if col_idx - 1 < len(row):
                    v = row[col_idx - 1].value
                    demand[date_key] = v if v is not None else 0
                else:
                    demand[date_key] = 0

            # Supply (next row)
            supply = {}
            if i + 1 < len(rows):
                supply_row = rows[i + 1]
                for col_idx, date_key in date_col_map.items():
                    if col_idx - 1 < len(supply_row):
                        v = supply_row[col_idx - 1].value
                        supply[date_key] = v if v is not None else 0
                    else:
                        supply[date_key] = 0

            results.append({
                'buyer': 'Ketwadee',
                'plant': 'PSB5',
                'part_no': str(part_no),
                'vendor_part': str(vendor_part) if vendor_part else '',
                'stock': stock,
                'on_way': None,
                'demand': demand,
                'supply': supply,
            })
            i += 3  # Skip Demand + Supply + Net
        else:
            i += 1

    wb.close()
    return results


def read_kanyanat(filepath):
    """讀取 PSB7 Kanyanat 資料"""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['Sheet1']

    headers = {}
    date_col_map = {}
    for cell in ws[1]:
        if cell.value is not None:
            headers[cell.column] = cell.value
            if cell.column >= 25:  # Y onwards
                norm = normalize_date_header(cell.value)
                if norm in CONSOLIDATED_DATE_COLS:
                    date_col_map[cell.column] = norm

    results = []
    max_col = ws.max_column
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row,
                             min_col=1, max_col=max_col, values_only=False))
    i = 0
    while i < len(rows):
        row = rows[i]
        type_val = row[23].value if len(row) > 23 else None  # X = TYPE

        if type_val == 'A-Demand':
            part_no = row[4].value   # E = Raw Material(P/N)
            vendor_part = row[5].value  # F = Vendor Part

            # Demand 數據
            demand = {}
            for col_idx, date_key in date_col_map.items():
                if col_idx - 1 < len(row):
                    v = row[col_idx - 1].value
                    demand[date_key] = v if v is not None else 0
                else:
                    demand[date_key] = 0

            # Supply (next row: B-Supply)
            supply = {}
            if i + 1 < len(rows):
                supply_row = rows[i + 1]
                for col_idx, date_key in date_col_map.items():
                    if col_idx - 1 < len(supply_row):
                        v = supply_row[col_idx - 1].value
                        supply[date_key] = v if v is not None else 0
                    else:
                        supply[date_key] = 0

            results.append({
                'buyer': 'Kanyanat',
                'plant': 'PSB7',
                'part_no': str(part_no),
                'vendor_part': str(vendor_part) if vendor_part else '',
                'stock': None,  # Kanyanat 不使用 STOCK
                'on_way': None,
                'demand': demand,
                'supply': supply,
            })
            i += 4  # Skip A-Demand + B-Supply + D-Net Demand + E-REMARK
        else:
            i += 1

    wb.close()
    return results


def read_weeraya(filepath):
    """讀取 PSB7 Weeraya 資料"""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['Sheet1']

    headers = {}
    date_col_map = {}
    for cell in ws[1]:
        if cell.value is not None:
            headers[cell.column] = cell.value
            if cell.column >= 14:  # N onwards
                norm = normalize_date_header(cell.value)
                if norm in CONSOLIDATED_DATE_COLS:
                    date_col_map[cell.column] = norm

    results = []
    max_col = ws.max_column
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row,
                             min_col=1, max_col=max_col, values_only=False))
    i = 0
    while i < len(rows):
        row = rows[i]
        type_val = row[11].value if len(row) > 11 else None  # L = TYPE

        if type_val == 'Demand':
            part_no = row[3].value    # D = Raw Material(P/N)
            vendor_part = row[4].value  # E = Vendor Part
            stock = row[12].value or 0  # M = Total stock

            def read_date_values(r):
                data = {}
                for col_idx, date_key in date_col_map.items():
                    if col_idx - 1 < len(r):
                        v = r[col_idx - 1].value
                        data[date_key] = v if v is not None else 0
                    else:
                        data[date_key] = 0
                return data

            demand = read_date_values(row)

            # Supply = Forecast Conf (row i+2)
            supply = {}
            if i + 2 < len(rows):
                supply = read_date_values(rows[i + 2])

            # Balance = Net Demand (row i+3) - 直接複製
            balance_data = {}
            if i + 3 < len(rows):
                balance_data = read_date_values(rows[i + 3])

            results.append({
                'buyer': 'Weeraya',
                'plant': 'PSB7',
                'part_no': str(part_no),
                'vendor_part': str(vendor_part) if vendor_part else '',
                'stock': stock,
                'on_way': None,
                'demand': demand,
                'supply': supply,
                'balance_override': balance_data,  # Weeraya 直接用 Net Demand
            })
            i += 5  # Demand + Firmed order + Forecast Conf + Net Demand + Remark
        else:
            i += 1

    wb.close()
    return results


def calculate_balance(demand, supply, stock):
    """
    計算 Balance:
      Balance[PASSDUE] = STOCK - Demand[PASSDUE]
      Balance[n] = Balance[n-1] - Demand[n] + Supply[n-1]
    Supply 在同一週不計入, 延後一期生效
    """
    balance = {}
    prev_balance = 0
    prev_supply = 0
    initial_stock = stock if stock else 0

    for idx, date_key in enumerate(CONSOLIDATED_DATE_COLS):
        d = demand.get(date_key, 0) or 0
        s = supply.get(date_key, 0) or 0

        if idx == 0:  # PASSDUE
            balance[date_key] = initial_stock - d
            prev_balance = balance[date_key]
            prev_supply = s
        else:
            balance[date_key] = prev_balance - d + prev_supply
            prev_balance = balance[date_key]
            prev_supply = s

    return balance


def read_consolidated(filepath):
    """讀取匯總格式"""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['工作表1']

    con_headers = [cell.value for cell in ws[1]]

    # 建立日期欄位對應
    date_col_map = {}
    for i, h in enumerate(con_headers):
        if h is not None and i >= 9:  # J onwards
            norm = normalize_date_header(h)
            if norm in CONSOLIDATED_DATE_COLS:
                date_col_map[i] = norm

    results = {}  # key=(buyer, part_no, date_type) → {date_key: value}
    metadata = {}  # key=(buyer, part_no) → {stock, on_way, ...}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False), 2):
        buyer = row[0].value
        plant = row[1].value
        part_no = str(row[4].value) if row[4].value else ''
        vendor_part = str(row[5].value) if row[5].value else ''
        stock = row[6].value   # G
        on_way = row[7].value  # H
        date_type = row[8].value  # I

        if not buyer or not date_type:
            continue

        data = {}
        for col_i, date_key in date_col_map.items():
            v = row[col_i].value
            data[date_key] = v if v is not None else 0

        key = (buyer, part_no, date_type)
        results[key] = data

        if date_type == 'Demand':
            metadata[(buyer, part_no)] = {
                'plant': plant,
                'vendor_part': vendor_part,
                'stock': stock,
                'on_way': on_way,
            }

    wb.close()
    return results, metadata


def compare_values(expected, actual, tolerance=0.01):
    """比較兩個數值, 允許浮點誤差"""
    e = expected or 0
    a = actual or 0
    if isinstance(e, (int, float)) and isinstance(a, (int, float)):
        return abs(e - a) < tolerance
    return str(e) == str(a)


def generate_consolidated_excel(all_source, output_path, reference_path):
    """
    根據原始資料產出匯總格式 Excel
    使用 reference_path 的原始檔案作為模板 (保留格式/公式/樣式)
    完整複製: 字體、填色、框線、對齊、數值格式、凍結窗格、自動篩選
    """
    from copy import copy
    from openpyxl.styles import (Font, Alignment, PatternFill, Border, Side,
                                  numbers)
    from openpyxl.styles.colors import Color

    # 讀取原始匯總檔作為模板
    wb_ref = openpyxl.load_workbook(reference_path)
    ws_ref = wb_ref['工作表1']

    # 建立新的 workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '工作表1'

    # === 複製表頭 (row 1) - 保留所有格式 ===
    for col in range(1, ws_ref.max_column + 1):
        src = ws_ref.cell(row=1, column=col)
        dst = ws.cell(row=1, column=col, value=src.value)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.border = copy(src.border)

    # === 複製欄寬 ===
    for col_letter, dim in ws_ref.column_dimensions.items():
        ws.column_dimensions[col_letter].width = dim.width

    # === 定義資料列的樣式 ===
    # 字體
    font_mingliu_12 = Font(name='新細明體', size=12)
    font_arial_9 = Font(name='Arial', size=9)
    font_arial_10 = Font(name='Arial', size=10)

    # 數值格式
    nf_data = '#,##0_);[Red]\\(#,##0\\)'
    nf_partno = '000'
    nf_stock = '#,##0'

    # 對齊
    align_center = Alignment(vertical='center')
    align_top_left = Alignment(horizontal='left', vertical='top')
    align_top = Alignment(vertical='top')

    # 填色 (背景色)
    fill_cd = PatternFill(patternType='solid',
                          fgColor=Color(theme=4, tint=0.7999816888943144))
    fill_i = PatternFill(patternType='solid',
                         fgColor=Color(theme=0, tint=0.0))
    fill_supply = PatternFill(patternType='solid',
                              fgColor=Color(theme=9, tint=0.7999816888943144))
    fill_balance = PatternFill(patternType='solid',
                               fgColor=Color(theme=7, tint=0.7999816888943144))

    # 框線
    thin_side = Side(style='thin')
    medium_side = Side(style='medium')
    border_normal = Border(left=thin_side, right=thin_side,
                           top=thin_side, bottom=thin_side)
    border_balance = Border(left=thin_side, right=thin_side,
                            top=thin_side, bottom=medium_side)

    # 對照表: Plant → (ERP客戶簡稱, ERP送貨地點)
    erp_mapping = {
        'PSB5': ('台達泰國', '台達PSB5SH'),
        'PSB7': ('台達泰國', '台達PSB7SH'),
    }

    def apply_row_style(row_num, date_type):
        """套用整列的框線、對齊、填色"""
        is_balance = (date_type == 'Balance')
        bdr = border_balance if is_balance else border_normal

        # A-B: 新細明體/12, center, 框線
        for col in [1, 2]:
            c = ws.cell(row=row_num, column=col)
            c.alignment = align_center
            c.border = bdr

        # C-D: 新細明體/12, center, 框線, 淺藍背景
        for col in [3, 4]:
            c = ws.cell(row=row_num, column=col)
            c.alignment = align_center
            c.border = bdr
            c.fill = fill_cd

        # E-F: Arial/9, top-left, 框線
        for col in [5, 6]:
            c = ws.cell(row=row_num, column=col)
            c.alignment = align_top_left
            c.border = bdr

        # G: Arial/9, top-left, 框線
        c = ws.cell(row=row_num, column=7)
        c.font = font_arial_9
        c.alignment = align_top_left
        c.border = bdr

        # H: 新細明體/12, center, 框線
        c = ws.cell(row=row_num, column=8)
        c.font = font_mingliu_12
        c.alignment = align_center
        c.border = bdr

        # I: Arial/10, top-left, 框線, 白底
        c = ws.cell(row=row_num, column=9)
        c.alignment = align_top_left
        c.border = bdr
        c.fill = fill_i

        # J~AI (col 10~35): 新細明體/12, center, 框線
        # Supply行: 綠底 (theme9), Balance行: 橄欖底 (theme7)
        data_fill = None
        if date_type == 'Supply':
            data_fill = fill_supply
        elif date_type == 'Balance':
            data_fill = fill_balance

        for col in range(10, 36):
            c = ws.cell(row=row_num, column=col)
            c.alignment = align_center
            c.border = bdr
            if data_fill:
                c.fill = data_fill

    # 寫入資料
    row_num = 2
    for item in all_source:
        buyer = item['buyer']
        plant = item['plant']
        part_no = item['part_no']
        vendor_part = item['vendor_part']
        stock = item['stock']
        erp_name, erp_location = erp_mapping.get(plant, ('', ''))

        # === Demand 行 ===
        demand_row = row_num
        # A~F 資訊
        ws.cell(row=row_num, column=1, value=buyer).font = font_mingliu_12
        ws.cell(row=row_num, column=2, value=plant).font = font_mingliu_12
        ws.cell(row=row_num, column=3, value=erp_name).font = font_mingliu_12
        ws.cell(row=row_num, column=4, value=erp_location).font = font_mingliu_12
        c = ws.cell(row=row_num, column=5, value=part_no)
        c.font = font_arial_9
        c.number_format = nf_partno
        ws.cell(row=row_num, column=6, value=vendor_part).font = font_arial_9
        if stock is not None:
            c = ws.cell(row=row_num, column=7, value=stock)
            c.font = font_arial_9
            c.number_format = nf_stock
        ws.cell(row=row_num, column=9, value='Demand').font = font_arial_10
        # J~AI 資料
        for col_offset, date_key in enumerate(CONSOLIDATED_DATE_COLS):
            v = item['demand'].get(date_key, 0) or 0
            c = ws.cell(row=row_num, column=10 + col_offset, value=v)
            c.font = font_mingliu_12
            c.number_format = nf_data
        apply_row_style(row_num, 'Demand')
        row_num += 1

        # === Supply 行 ===
        supply_row = row_num
        ws.cell(row=row_num, column=1, value=buyer).font = font_mingliu_12
        ws.cell(row=row_num, column=2, value=plant).font = font_mingliu_12
        ws.cell(row=row_num, column=3, value=erp_name).font = font_mingliu_12
        ws.cell(row=row_num, column=4, value=erp_location).font = font_mingliu_12
        c = ws.cell(row=row_num, column=5, value=part_no)
        c.font = font_arial_9
        c.number_format = nf_partno
        ws.cell(row=row_num, column=6, value=vendor_part).font = font_arial_9
        ws.cell(row=row_num, column=9, value='Supply').font = font_arial_10
        for col_offset, date_key in enumerate(CONSOLIDATED_DATE_COLS):
            v = item['supply'].get(date_key, 0) or 0
            col_num = 10 + col_offset
            c = ws.cell(row=row_num, column=col_num,
                        value=v if v != 0 else None)
            c.font = font_mingliu_12
            c.number_format = nf_data
        apply_row_style(row_num, 'Supply')
        row_num += 1

        # === Balance 行 (用公式) ===
        balance_row = row_num
        ws.cell(row=row_num, column=1, value=buyer).font = font_mingliu_12
        ws.cell(row=row_num, column=2, value=plant).font = font_mingliu_12
        ws.cell(row=row_num, column=3, value=erp_name).font = font_mingliu_12
        ws.cell(row=row_num, column=4, value=erp_location).font = font_mingliu_12
        c = ws.cell(row=row_num, column=5, value=part_no)
        c.font = font_arial_9
        c.number_format = nf_partno
        ws.cell(row=row_num, column=6, value=vendor_part).font = font_arial_9
        ws.cell(row=row_num, column=9, value='Balance').font = font_arial_10
        for col_offset, date_key in enumerate(CONSOLIDATED_DATE_COLS):
            col_num = 10 + col_offset
            col_letter = openpyxl.utils.get_column_letter(col_num)
            if col_offset == 0:
                formula = f'=G{demand_row}+H{demand_row}-{col_letter}{demand_row}'
            elif col_offset == 1:
                prev_col = openpyxl.utils.get_column_letter(col_num - 1)
                formula = (f'={prev_col}{supply_row}+{prev_col}{balance_row}'
                           f'-{col_letter}{demand_row}')
            else:
                prev_col = openpyxl.utils.get_column_letter(col_num - 1)
                formula = (f'={prev_col}{balance_row}+{prev_col}{supply_row}'
                           f'-{col_letter}{demand_row}')
            c = ws.cell(row=balance_row, column=col_num, value=formula)
            c.font = font_mingliu_12
            c.number_format = nf_data
        apply_row_style(row_num, 'Balance')
        row_num += 1

    # === 凍結窗格: J2 ===
    ws.freeze_panes = 'J2'

    # === 自動篩選 ===
    last_col_letter = openpyxl.utils.get_column_letter(ws_ref.max_column)
    ws.auto_filter.ref = f'A1:{last_col_letter}{row_num - 1}'

    # === 條件式格式設定 (從模板複製) ===
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles.differential import DifferentialStyle
    for cf in ws_ref.conditional_formatting:
        for rule in cf.rules:
            ws.conditional_formatting.add(str(cf.cells), rule)

    # === 複製工作表5 (含格式) ===
    ws5_ref = wb_ref['工作表5']
    ws5 = wb.create_sheet('工作表5')
    for row in ws5_ref.iter_rows(min_row=1, max_row=ws5_ref.max_row,
                                  max_col=ws5_ref.max_column, values_only=False):
        for cell in row:
            dst = ws5.cell(row=cell.row, column=cell.column, value=cell.value)
            dst.font = copy(cell.font)
            dst.fill = copy(cell.fill)
            dst.alignment = copy(cell.alignment)
            dst.number_format = cell.number_format
            dst.border = copy(cell.border)
    for col_letter, dim in ws5_ref.column_dimensions.items():
        ws5.column_dimensions[col_letter].width = dim.width

    wb_ref.close()
    wb.save(output_path)
    return row_num - 2  # 總資料行數


def verify():
    """主驗證流程"""
    print('=' * 80)
    print('台達 Forecast 匯總格式驗證')
    print('=' * 80)

    # 讀取 3 個原始檔案
    ketwadee_file = None
    kanyanat_file = None
    weeraya_file = None

    for f in os.listdir(BASE_DIR):
        if 'Ketwadee' in f and f.endswith('.xlsx'):
            ketwadee_file = os.path.join(BASE_DIR, f)
        elif 'Kanyanat' in f and f.endswith('.xlsx'):
            kanyanat_file = os.path.join(BASE_DIR, f)
        elif 'Weeraya' in f and f.endswith('.xlsx'):
            weeraya_file = os.path.join(BASE_DIR, f)

    print(f'\n[1] 讀取原始檔案...')
    print(f'  Ketwadee: {os.path.basename(ketwadee_file)}')
    ketwadee_data = read_ketwadee(ketwadee_file)
    print(f'    -> {len(ketwadee_data)} 個料號')

    print(f'  Kanyanat: {os.path.basename(kanyanat_file)}')
    kanyanat_data = read_kanyanat(kanyanat_file)
    print(f'    -> {len(kanyanat_data)} 個料號')

    print(f'  Weeraya: {os.path.basename(weeraya_file)}')
    weeraya_data = read_weeraya(weeraya_file)
    print(f'    -> {len(weeraya_data)} 個料號')

    # 讀取匯總格式
    con_file = os.path.join(BASE_DIR, '0407匯總格式.xlsx')
    print(f'\n[2] 讀取匯總格式: {os.path.basename(con_file)}')
    con_data, con_meta = read_consolidated(con_file)

    # 統計匯總格式各 buyer 的料號數
    con_buyers = {}
    for (buyer, part, dtype) in con_data:
        if dtype == 'Demand':
            con_buyers.setdefault(buyer, set()).add(part)
    for b, parts in con_buyers.items():
        print(f'  {b}: {len(parts)} 個料號')

    # 合併所有原始資料
    all_source = ketwadee_data + kanyanat_data + weeraya_data

    # 產出匯總格式 Excel
    output_path = os.path.join(BASE_DIR, '0407匯總格式_程式產出.xlsx')
    print(f'\n[3] 產出匯總格式 Excel...')
    total_rows = generate_consolidated_excel(all_source, output_path, con_file)
    print(f'  已產出: {os.path.basename(output_path)}')
    print(f'  資料行數: {total_rows} (含 Demand/Supply/Balance)')
    print(f'  Balance 行使用公式, 可在 Excel 中驗證')

    # 驗證
    print(f'\n[4] 開始比對...')
    total_checks = 0
    total_match = 0
    total_mismatch = 0
    mismatch_details = []

    for item in all_source:
        buyer = item['buyer']
        part = item['part_no']

        # 檢查這個料號是否存在於匯總格式中
        con_demand_key = (buyer, part, 'Demand')
        con_supply_key = (buyer, part, 'Supply')
        con_balance_key = (buyer, part, 'Balance')

        if con_demand_key not in con_data:
            continue  # 這個料號不在匯總中 (可能被篩選掉)

        # === 比對 Demand ===
        for date_key in CONSOLIDATED_DATE_COLS:
            src_val = item['demand'].get(date_key, 0) or 0
            con_val = con_data[con_demand_key].get(date_key, 0) or 0
            total_checks += 1
            if compare_values(src_val, con_val):
                total_match += 1
            else:
                total_mismatch += 1
                mismatch_details.append(
                    f'  {buyer}/{part} Demand [{date_key}]: '
                    f'source={src_val}, consolidated={con_val}'
                )

        # === 比對 Supply ===
        if con_supply_key in con_data:
            for date_key in CONSOLIDATED_DATE_COLS:
                src_val = item['supply'].get(date_key, 0) or 0
                con_val = con_data[con_supply_key].get(date_key, 0) or 0
                total_checks += 1
                if compare_values(src_val, con_val):
                    total_match += 1
                else:
                    total_mismatch += 1
                    mismatch_details.append(
                        f'  {buyer}/{part} Supply [{date_key}]: '
                        f'source={src_val}, consolidated={con_val}'
                    )

        # === 比對 Balance ===
        if con_balance_key in con_data:
            # 決定 Balance 來源
            if 'balance_override' in item:
                # Weeraya: 直接用 Net Demand
                calc_balance = item['balance_override']
            else:
                # Ketwadee / Kanyanat: 重新計算
                stock = item['stock'] if item['stock'] else 0
                calc_balance = calculate_balance(
                    item['demand'], item['supply'], stock
                )

            for date_key in CONSOLIDATED_DATE_COLS:
                calc_val = calc_balance.get(date_key, 0) or 0
                con_val = con_data[con_balance_key].get(date_key, 0) or 0
                total_checks += 1
                if compare_values(calc_val, con_val, tolerance=1.0):
                    total_match += 1
                else:
                    total_mismatch += 1
                    mismatch_details.append(
                        f'  {buyer}/{part} Balance [{date_key}]: '
                        f'calculated={calc_val}, consolidated={con_val}'
                    )

        # === 比對 STOCK ===
        meta = con_meta.get((buyer, part), {})
        con_stock = meta.get('stock')
        if item['stock'] is not None and con_stock is not None:
            total_checks += 1
            if compare_values(item['stock'], con_stock):
                total_match += 1
            else:
                total_mismatch += 1
                mismatch_details.append(
                    f'  {buyer}/{part} STOCK: source={item["stock"]}, '
                    f'consolidated={con_stock}'
                )

    # 報告
    print(f'\n{"=" * 80}')
    print(f'驗證結果')
    print(f'{"=" * 80}')

    # 匯總中存在但原始中找不到的料號
    source_keys = {(item['buyer'], item['part_no']) for item in all_source}
    con_only = set()
    for (buyer, part, dtype) in con_data:
        if dtype == 'Demand' and (buyer, part) not in source_keys:
            con_only.add((buyer, part))

    # 原始中存在但匯總中沒有的料號
    source_only = {}
    for item in all_source:
        key = (item['buyer'], item['part_no'])
        if (item['buyer'], item['part_no'], 'Demand') not in con_data:
            source_only.setdefault(item['buyer'], set()).add(item['part_no'])

    print(f'\n--- 料號涵蓋率 ---')
    for buyer_name in ['Ketwadee', 'Kanyanat', 'Weeraya']:
        src_count = len([x for x in all_source if x['buyer'] == buyer_name])
        con_count = len(con_buyers.get(buyer_name, set()))
        matched = len([x for x in all_source
                       if x['buyer'] == buyer_name
                       and (x['buyer'], x['part_no'], 'Demand') in con_data])
        filtered = src_count - matched
        print(f'  {buyer_name}: 原始={src_count}, 匯總={con_count}, '
              f'匹配={matched}, 被篩選={filtered}')

    if source_only:
        print(f'\n--- 原始有但匯總沒有的料號 (被篩選) ---')
        for buyer_name, parts in source_only.items():
            print(f'  {buyer_name}: {len(parts)} 個料號被篩選')

    print(f'\n--- 數值比對 ---')
    print(f'  總比對數: {total_checks}')
    print(f'  匹配數:   {total_match}')
    print(f'  不匹配數: {total_mismatch}')

    if total_checks > 0:
        accuracy = total_match / total_checks * 100
        print(f'  準確率:   {accuracy:.2f}%')

    if total_mismatch == 0:
        print(f'\n*** 驗證通過! 所有數值完全匹配! ***')
    else:
        print(f'\n*** 發現 {total_mismatch} 個不匹配 ***')
        print(f'\n--- 不匹配詳情 (前 50 筆) ---')
        for detail in mismatch_details[:50]:
            print(detail)
        if len(mismatch_details) > 50:
            print(f'  ... 還有 {len(mismatch_details) - 50} 筆')

    return total_mismatch == 0


if __name__ == '__main__':
    success = verify()
    sys.exit(0 if success else 1)
