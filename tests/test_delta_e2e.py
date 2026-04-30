# -*- coding: utf-8 -*-
"""
tests/test_delta_e2e.py
Delta 27 個 Buyer 檔案基線 + 6 維度 Mutation 回歸測試

用法:
  pytest tests/test_delta_e2e.py -v
  pytest tests/test_delta_e2e.py -v -k "baseline"
  pytest tests/test_delta_e2e.py -v -k "mutation"
  pytest tests/test_delta_e2e.py -v -k "eibg or ketwadee"

測試項目:
  TestBaseline   - 23 個實際檔案: format detection / read / backfill
  TestMutation   - 14 格式 × 6 漂移維度 × ~3 變體 ≈ 80+ cases
"""

import sys
import os
import shutil
import json
from datetime import datetime
from pathlib import Path

import pytest
import openpyxl

# ── 專案路徑 ────────────────────────────────────────────────────────
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from delta_forecast_processor import detect_format
from delta_format_fingerprint import match_known_format_fingerprint
from delta_unified_reader import (
    read_buyer_file, find_valid_sheets, scan_headers,
    collect_date_cols, find_first_date_col,
)
from delta_original_backfill import backfill_one_file

# ── 路徑常數 ─────────────────────────────────────────────────────────
DESKTOP_DIR = Path(r"C:/Users/petty/Desktop/客戶相關資料/01.強茂/台達業務")

# 優先用從真實買方檔案生成的測試 forecast（含實際 supply 值）
# 若不存在，退回到生產用的 forecast_result（supply 全為空，只能測試不崩潰）
_FIXTURE_FORECAST = ROOT / "tests" / "fixtures" / "forecast_result_test.xlsx"
_PROD_FORECAST    = ROOT / "processed/7/20260430_155536/forecast_result.xlsx"
FORECAST_RESULT = _FIXTURE_FORECAST if _FIXTURE_FORECAST.exists() else _PROD_FORECAST

# ── 27 個基線檔案清單 ─────────────────────────────────────────────────
# (fname_relative_to_DESKTOP_DIR, expected_fmt, plant_codes)
# fmt=None → 跳過 format 偵測 (彙總/參考檔)
BASELINE_FILES = [
    # ── EIBG/EISBG 系列 ──
    ("EIBG-TPW1\u2014Lydia--0427.xlsx",                           "eibg_eisbg",           ["TPW1"]),
    ("EIBG-TPW1-PAN JIT MRP Lydia TPW1--0420_backfilled.xlsx",    "eibg_eisbg",           ["TPW1"]),
    ("EIBG-UPW1 PANJIT 0413.xlsx",                                "eibg_eisbg",           ["UPW1"]),
    ("EISBG-.xlsx",                                               "eibg_eisbg",           ["UPW1"]),
    # ── ICTBG 系列 ──
    # 已回填版本：結構已被修改，format detection 行為不同，不驗證 format
    ("DNI-NTL7-AMY-100109-PAN JIT-BNV_20260417_backfilled.xlsx",  None,                   ["NTL7"]),
    ("ICTBG(DNI)-NTL7  4.13 MRP CFM.xlsx",                       "ictbg_ntl7",           ["NTL7"]),
    ("ICTBG-PSB9-Kaewarin_20260413.xlsx",                         "ictbg_psb9_mrp",       ["PSB9"]),
    ("ICTBG-PSB9-Siriraht_20260411.xlsx",                         "ictbg_psb9_siriraht",  ["PSB9"]),
    # ── FMBG / IABG ──
    ("FMBG-MRP(TPC5)-100109-2026-4-15.xlsx",                      "fmbg",                 ["TPC5"]),
    ("IABG-IMW1-\u9648\u59ff\u5bb9_20260413.xlsx",                "iabg",                 ["IMW1"]),
    # ── SVC1PWC1 Diode&MOS ──
    ("MRP(SVC1PWC1 DIODE&MOS)-100109-2026-3-30.xlsx",             "svc1pwc1_diode_mos",   ["SVC1", "PWC1"]),
    # ── NBQ1 ──
    ("NBQ1.xlsx",                                                 "nbq1",                 ["NBQ1"]),
    # ── India IAI1 ──
    ("PSBG (India IAI1&UPI2&DFI1 DIODES)-Jack0401.xlsx",          "india_iai1",           ["IAI1", "UPI2", "DFI1"]),
    # ── PSW1+CEW1 ──
    ("PSBG PSW1+CEW1- \u694a\u6d0b\u5f59\u6574 0330(\u5b8c\u6210) (002).xlsx", "psw1_cew1", ["PSW1", "CEW1"]),
    ("PSBG PSW1+CEW1\u5408\u4f75-Aviva_20260416.xlsx",            "psw1_cew1",            ["PSW1", "CEW1"]),
    # ── Kanyanat ──
    ("PSBG-PSB7PAN JIT YTMDS APR 20 2026 Kanyanat.xlsx",          "kanyanat",             ["PSB7"]),
    ("PSBG\xa0PSB7_Kanyanat.S0406(\u5b8c\u6210).xlsx",            "kanyanat",             ["PSB7"]),
    # ── Ketwadee ──
    ("PSBG\xa0PSB5-\xa0Ketwadee0406(\u5b8c\u6210).xlsx",          "ketwadee",             ["PSB5"]),
    # ── Weeraya ──
    ("PSBG\xa0PSB7-Weeraya0406(\u5b8c\u6210).xlsx",               "weeraya",              ["PSB7"]),
    # ── MWC1+IPC1 系列 ──
    ("W4-PSBG DNI-MWC1&IPC1 MRP 04.20.2026\u738b\u8ff0\u9023.xlsx",            "mwc1ipc1", ["MWC1", "IPC1"]),
    ("W4-PSBG DNI-MWC1+IPC1 \u5f37\u8302 MRP+SHIP 2026-4-17\u5468\u6843\u6625.xlsx", "mwc1ipc1", ["MWC1", "IPC1"]),
    ("W4-PSBG DNI-MWC1-IPC1-MWT-IPT-100109-\u5f37\u8302-0420MRP \u7f85\u5a1f.xlsx", "mwc1ipc1", ["MWC1", "IPC1"]),
    ("\u5f37\u8302 MWC1IPC1 MRP 03.30.2026.xlsx",                 "mwc1ipc1",             ["MWC1", "IPC1"]),
    # ── 邏輯/ 目錄 (彙總檔，跳過格式偵測) ──
    ("\u908f\u8f2f/0407\u532f\u7e3d\u683c\u5f0f.xlsx",                                          None, []),
    ("\u908f\u8f2f/0407\u532f\u7e3d\u683c\u5f0f_\u7a0b\u5f0f\u7522\u51fav5_\u51688\u6a94.xlsx", None, []),
    ("\u908f\u8f2f/0408-\u4e0a\u5348\u6de8\u9700\u6c42\xa0(\u53f0\u9054).xlsx",                 None, []),
    ("\u908f\u8f2f/\u5c0d\u7167\u8868.xlsx",                                                     None, []),
]

# ── 工具函式 ──────────────────────────────────────────────────────────

def _discover_structure(filepath):
    """探測檔案結構，回傳 sheet/header_row/found_cols/date_map 等資訊"""
    wb = openpyxl.load_workbook(str(filepath), read_only=False)
    specs = find_valid_sheets(wb)
    if not specs:
        wb.close()
        return None
    sname, hr = specs[0]
    ws = wb[sname]
    max_col = ws.max_column or 40
    hv = [ws.cell(hr, c).value for c in range(1, max_col + 1)]
    col_map = {c + 1: v for c, v in enumerate(hv)}
    found, _ = scan_headers(hv)
    ds = find_first_date_col(col_map, hv)
    dcm = collect_date_cols(ds, hv) if ds else {}
    # 找 marker col 的值（抽樣前 60 列）
    marker_col = None
    marker_vals = {}  # col → set of unique values
    for r in range(hr + 1, min((ws.max_row or hr + 60), hr + 60) + 1):
        for c in range(1, min(max_col, 20) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and any(
                kw in v.lower() for kw in ['demand', 'supply', 'balance', '需求', '供應', '結餘']
            ):
                marker_vals.setdefault(c, set()).add(v.strip())
    if marker_vals:
        # 最多唯一值的欄
        marker_col = max(marker_vals, key=lambda c: len(marker_vals[c]))
    wb.close()
    return {
        'sheet': sname,
        'header_row': hr,
        'found': found,
        'date_start': ds,
        'date_col_map': dcm,
        'header_values': hv,
        'marker_col': marker_col,
        'marker_vals': marker_vals.get(marker_col, set()) if marker_col else set(),
    }


def _make_mutation(src_path, dst_path, mutation_type):
    """
    複製 src_path → dst_path，套用 mutation_type 指定的漂移。
    成功回傳 True，無法套用（欄不存在）回傳 False（測試標記 xfail）。

    mutation_type 支援:
      date_mmdd          YYYYMMDD → MMDD (4 位)
      date_slash         YYYYMMDD → M/D
      date_month_year    YYYYMMDD → APR-2026
      partno_case        PARTNO → PartNo
      partno_dotspace    PARTNO → PART NO.
      partno_chinese     PARTNO → 料號
      plant_case         PLANT → Plant
      plant_synonym      PLANT → WAREHOUSE
      plant_value_spaces plant value " PSB5 " (加空格)
      marker_upper       demand/supply/balance → UPPER
      marker_chinese     demand/supply/balance → 中文
      marker_prefix      A-Demand → 1.Demand
      stock_synonym      STOCK → Inventory
      onway_synonym      ON WAY → In-Transit
      stock_onway_both   同時替換 STOCK + ON WAY
      sheet_upper        Sheet1 → SHEET1
      sheet_suffix       Sheet1 → Sheet1_Apr
      sheet_rename       Sheet1 → Forecast (完全改名)
    """
    shutil.copy2(str(src_path), str(dst_path))
    struct = _discover_structure(src_path)
    if not struct:
        return False

    wb = openpyxl.load_workbook(str(dst_path))
    ws = wb[struct['sheet']]
    hr = struct['header_row']
    found = struct['found']
    dcm = struct['date_col_map']
    hv = struct['header_values']

    applied = False

    if mutation_type == 'date_mmdd':
        for col, date_str in dcm.items():
            if len(str(date_str)) == 8 and str(date_str).isdigit():
                ws.cell(hr, col).value = str(date_str)[4:]  # YYYYMMDD → MMDD
                applied = True

    elif mutation_type == 'date_slash':
        for col, date_str in dcm.items():
            if len(str(date_str)) == 8 and str(date_str).isdigit():
                mm = int(str(date_str)[4:6])
                dd = int(str(date_str)[6:8])
                ws.cell(hr, col).value = f"{mm}/{dd}"
                applied = True

    elif mutation_type == 'date_month_year':
        MONTH_ABBR = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN',
                      'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
        seen_months = set()
        for col, date_str in sorted(dcm.items()):
            if len(str(date_str)) == 8 and str(date_str).isdigit():
                yr = str(date_str)[:4]
                mm = int(str(date_str)[4:6])
                mon = MONTH_ABBR[mm - 1]
                key = f"{mon}-{yr}"
                if key not in seen_months:
                    ws.cell(hr, col).value = key
                    seen_months.add(key)
                    applied = True
                else:
                    # 同月第二週 → 跳過（保持原值避免重複）
                    pass

    elif mutation_type == 'partno_case':
        col = found.get('partno')
        if col:
            ws.cell(hr, col).value = 'PartNo'
            applied = True

    elif mutation_type == 'partno_dotspace':
        col = found.get('partno')
        if col:
            ws.cell(hr, col).value = 'PART NO.'
            applied = True

    elif mutation_type == 'partno_chinese':
        col = found.get('partno')
        if col:
            ws.cell(hr, col).value = '料號'
            applied = True

    elif mutation_type == 'plant_case':
        col = found.get('plant')
        if col:
            ws.cell(hr, col).value = 'Plant'
            applied = True

    elif mutation_type == 'plant_synonym':
        col = found.get('plant')
        if col:
            ws.cell(hr, col).value = 'WAREHOUSE'
            applied = True

    elif mutation_type == 'plant_value_spaces':
        col = found.get('plant')
        if col:
            max_row = ws.max_row or (hr + 200)
            for r in range(hr + 1, max_row + 1):
                v = ws.cell(r, col).value
                if isinstance(v, str) and v.strip():
                    ws.cell(r, col).value = f' {v.strip()} '
                    applied = True

    elif mutation_type == 'marker_upper':
        mc = struct['marker_col']
        if mc:
            max_row = ws.max_row or (hr + 200)
            for r in range(hr + 1, max_row + 1):
                v = ws.cell(r, mc).value
                if isinstance(v, str):
                    ws.cell(r, mc).value = v.upper()
                    applied = True

    elif mutation_type == 'marker_chinese':
        mc = struct['marker_col']
        MARKER_ZH = {
            'demand': '需求量', 'supply': '供應量', 'balance': '結餘',
            'net demand': '淨需求', 'net supply': '淨供應',
            'a-demand': '需求量', 'b-supply': '供應量',
        }
        if mc:
            max_row = ws.max_row or (hr + 200)
            for r in range(hr + 1, max_row + 1):
                v = ws.cell(r, mc).value
                if isinstance(v, str):
                    lower_v = v.strip().lower()
                    replacement = None
                    for k, zh in MARKER_ZH.items():
                        if k in lower_v:
                            replacement = zh
                            break
                    if replacement:
                        ws.cell(r, mc).value = replacement
                        applied = True

    elif mutation_type == 'marker_prefix':
        # A-Demand → 1.Demand, B-Supply → 2.Supply (保留 dash 前綴變數字)
        mc = struct['marker_col']
        if mc:
            max_row = ws.max_row or (hr + 200)
            prefix_map = {'a': '1', 'b': '2', 'c': '3', 'd': '4'}
            for r in range(hr + 1, max_row + 1):
                v = ws.cell(r, mc).value
                if isinstance(v, str) and '-' in v:
                    parts = v.split('-', 1)
                    if parts[0].strip().lower() in prefix_map:
                        new_prefix = prefix_map[parts[0].strip().lower()]
                        ws.cell(r, mc).value = f'{new_prefix}.{parts[1]}'
                        applied = True

    elif mutation_type == 'stock_synonym':
        col = found.get('stock')
        if col:
            ws.cell(hr, col).value = 'Inventory'
            applied = True

    elif mutation_type == 'onway_synonym':
        col = found.get('on_way')
        if col:
            ws.cell(hr, col).value = 'In-Transit'
            applied = True

    elif mutation_type == 'stock_onway_both':
        for key, new_name in [('stock', 'Inventory'), ('on_way', 'In-Transit')]:
            col = found.get(key)
            if col:
                ws.cell(hr, col).value = new_name
                applied = True

    elif mutation_type == 'sheet_upper':
        old = struct['sheet']
        if old.upper() != old:
            ws.title = old.upper()
            applied = True

    elif mutation_type == 'sheet_suffix':
        ws.title = struct['sheet'] + '_Apr'
        applied = True

    elif mutation_type == 'sheet_rename':
        ws.title = 'Forecast'
        applied = True

    # ── 結構性 mutations (rule-based 應能處理) ─────────────────────────

    elif mutation_type == 'extra_title_row':
        # 在表頭列前插入一列標題，讓 header 往下移一列
        ws.insert_rows(hr)
        ws.cell(hr, 1).value = 'Delta Electronics MRP Forecast Report'
        applied = True

    elif mutation_type == 'header_pushed_down_3':
        # 插入 3 列空白列，表頭從 row 1 移到 row 4
        ws.insert_rows(hr, amount=3)
        applied = True

    elif mutation_type == 'add_decoy_sheet':
        # 在 workbook 最前面插入無資料的 Cover sheet
        cover = wb.create_sheet('Cover', 0)
        cover.cell(1, 1).value = 'Delta Electronics'
        cover.cell(2, 1).value = 'MRP Forecast Summary'
        applied = True

    elif mutation_type == 'extra_col_seq_no':
        # 在所有欄前插入 "No." 序號欄
        ws.insert_cols(1)
        ws.cell(hr, 1).value = 'No.'
        max_row = ws.max_row or (hr + 200)
        for r in range(hr + 1, max_row + 1):
            if ws.cell(r, 2).value is not None:
                ws.cell(r, 1).value = r - hr
                applied = True

    elif mutation_type == 'extra_blank_rows':
        # 在資料區散佈 5 列空白列（模擬人工加入的分隔列）
        max_row = ws.max_row or (hr + 100)
        insert_at = sorted(
            [hr + 5, hr + 12, hr + 20, hr + 30, hr + 45],
            reverse=True
        )
        for pos in insert_at:
            if pos < max_row:
                ws.insert_rows(pos)
                applied = True

    elif mutation_type == 'date_iso_hyphen':
        # YYYYMMDD → 2026-04-27 (ISO 格式，DATE_PAT_HYPHEN 應能識別)
        for col, date_str in dcm.items():
            if len(str(date_str)) == 8 and str(date_str).isdigit():
                yr = str(date_str)[:4]
                mm = str(date_str)[4:6]
                dd = str(date_str)[6:8]
                ws.cell(hr, col).value = f'{yr}-{mm}-{dd}'
                applied = True

    elif mutation_type == 'cols_swap_partno_plant':
        # 互換 PARTNO 欄和 PLANT 欄的所有資料（測試位置獨立性）
        partno_col = found.get('partno')
        plant_col = found.get('plant')
        if partno_col and plant_col:
            max_row_d = ws.max_row or (hr + 200)
            for r in range(1, max_row_d + 1):
                pn_val = ws.cell(r, partno_col).value
                pl_val = ws.cell(r, plant_col).value
                ws.cell(r, partno_col).value = pl_val
                ws.cell(r, plant_col).value = pn_val
            applied = True

    elif mutation_type == 'cols_all_shuffled':
        # 所有欄完全亂序（含 PARTNO / PLANT / 日期欄），固定 seed=42 確保可重現
        import random as _rnd
        max_col_s = ws.max_column or 30
        max_row_s = ws.max_row or (hr + 200)
        all_data_s = []
        for r in range(1, max_row_s + 1):
            all_data_s.append([ws.cell(r, c).value for c in range(1, max_col_s + 1)])
        col_order_s = list(range(max_col_s))
        _rnd.Random(42).shuffle(col_order_s)
        for r in range(1, max_row_s + 1):
            orig_s = all_data_s[r - 1]
            for new_c, old_c in enumerate(col_order_s, 1):
                ws.cell(r, new_c).value = orig_s[old_c] if old_c < len(orig_s) else None
        applied = True

    elif mutation_type == 'date_cols_pushed_right':
        # 在日期欄前插入 5 個雜訊欄（日期從更右邊開始）
        if dcm:
            first_date_col = min(dcm.keys())
            ws.insert_cols(first_date_col, amount=5)
            for c in range(first_date_col, first_date_col + 5):
                ws.cell(hr, c).value = f'Extra_{c - first_date_col + 1}'
            applied = True

    # ── AI 深度 mutations (rule-based 無法處理，需 AI fallback) ──────────

    elif mutation_type == 'partno_jp':
        # PARTNO → 品番 (日文料號，不在 keyword 清單中)
        col = found.get('partno')
        if col:
            ws.cell(hr, col).value = '品番'
            applied = True

    elif mutation_type == 'partno_item_code':
        # PARTNO → Item Code (不在 keyword 清單中)
        col = found.get('partno')
        if col:
            ws.cell(hr, col).value = 'Item Code'
            applied = True

    elif mutation_type == 'plant_factory':
        # PLANT → Factory (不在 keyword 清單中)
        col = found.get('plant')
        if col:
            ws.cell(hr, col).value = 'Factory'
            applied = True

    elif mutation_type == 'plant_location':
        # PLANT → Location (不在 keyword 清單中)
        col = found.get('plant')
        if col:
            ws.cell(hr, col).value = 'Location'
            applied = True

    elif mutation_type == 'marker_forecast_qty':
        # demand → "Forecast Qty", supply → "Order Qty", balance → "Net Qty"
        # 完全不在 MARKER_PATTERNS 中
        mc = struct['marker_col']
        FORECAST_MAP = {
            'demand': 'Forecast Qty', 'a-demand': 'Forecast Qty',
            'supply': 'Order Qty', 'b-supply': 'Order Qty', 'b-cfm': 'Order Qty',
            'balance': 'Net Qty', 'c-balance': 'Net Qty', 'net demand': 'Net Qty',
            'request': 'Forecast Qty',
        }
        if mc:
            max_row = ws.max_row or (hr + 200)
            for r in range(hr + 1, max_row + 1):
                v = ws.cell(r, mc).value
                if isinstance(v, str):
                    lower_v = v.strip().lower()
                    for k, new_v in FORECAST_MAP.items():
                        if k in lower_v:
                            ws.cell(r, mc).value = new_v
                            applied = True
                            break

    elif mutation_type == 'marker_jp':
        # demand → 予測量, supply → 発注量, balance → 在庫量 (日文，完全陌生)
        mc = struct['marker_col']
        JP_MAP = {
            'demand': '予測量', 'a-demand': '予測量',
            'supply': '発注量', 'b-supply': '発注量', 'b-cfm': '発注量',
            'balance': '在庫量', 'c-balance': '在庫量', 'net demand': '在庫量',
            'request': '予測量',
        }
        if mc:
            max_row = ws.max_row or (hr + 200)
            for r in range(hr + 1, max_row + 1):
                v = ws.cell(r, mc).value
                if isinstance(v, str):
                    lower_v = v.strip().lower()
                    for k, new_v in JP_MAP.items():
                        if k in lower_v:
                            ws.cell(r, mc).value = new_v
                            applied = True
                            break

    elif mutation_type == 'date_long_en':
        # 20260427 → "April 27, 2026"（長格式英文，不符合任何 date pattern）
        MONTH_FULL = ['January', 'February', 'March', 'April', 'May', 'June',
                      'July', 'August', 'September', 'October', 'November', 'December']
        seen_months = set()
        for col, date_str in sorted(dcm.items()):
            if len(str(date_str)) == 8 and str(date_str).isdigit():
                yr = int(str(date_str)[:4])
                mm = int(str(date_str)[4:6])
                dd = int(str(date_str)[6:8])
                key = f"{MONTH_FULL[mm - 1]}-{yr}"
                if key not in seen_months:
                    ws.cell(hr, col).value = f"{MONTH_FULL[mm - 1]} {dd}, {yr}"
                    seen_months.add(key)
                    applied = True
                else:
                    ws.cell(hr, col).value = None  # 同月只保留第一週

    elif mutation_type == 'date_excel_serial':
        # 將日期欄改為 Excel 序號整數（非字串），is_date_header() 無法辨識
        from datetime import date as _date
        _excel_epoch = _date(1899, 12, 30)
        for col, date_str in dcm.items():
            if len(str(date_str)) == 8 and str(date_str).isdigit():
                try:
                    yr = int(str(date_str)[:4])
                    mm = int(str(date_str)[4:6])
                    dd = int(str(date_str)[6:8])
                    serial = (_date(yr, mm, dd) - _excel_epoch).days
                    ws.cell(hr, col).value = serial
                    applied = True
                except ValueError:
                    pass

    elif mutation_type == 'all_headers_jp':
        # 所有關鍵欄名全部改為日文（最高難度 AI 測試）
        JP_HEADERS = {
            'partno': '品番', 'plant': '工場',
            'stock': '現在庫', 'on_way': '輸送中',
        }
        for field, jp_name in JP_HEADERS.items():
            col = found.get(field)
            if col:
                ws.cell(hr, col).value = jp_name
                applied = True

    wb.save(str(dst_path))
    wb.close()
    return applied


# ── Mutation 定義表 ───────────────────────────────────────────────────
# (fmt, src_filename, plant_codes, m_id, label, mutation_type)
# 每個格式選一個代表檔，套用相關維度的變體

_MUTATION_DEFS_RAW = [
    # eibg_eisbg: MMDD 日期已是特殊格式，重點測試欄名漂移
    ('eibg_eisbg', 'EIBG-TPW1\u2014Lydia--0427.xlsx', ['TPW1'], [
        ('m2a', 'PARTNO→PartNo',          'partno_case'),
        ('m2b', 'PARTNO→PART NO.',        'partno_dotspace'),
        ('m2c', 'PARTNO→料號',             'partno_chinese'),
        ('m3a', 'PLANT→Plant',            'plant_case'),
        ('m3b', 'PLANT→WAREHOUSE',        'plant_synonym'),
        ('m3c', 'plant值加空格',            'plant_value_spaces'),
        ('m5a', 'STOCK→Inventory',        'stock_synonym'),
        ('m5b', 'ON WAY→In-Transit',      'onway_synonym'),
        ('m5c', 'STOCK+ONWAY同時替換',     'stock_onway_both'),
    ]),
    # ketwadee: YYYYMMDD 日期，有 Demand/Supply/Balance marker
    ('ketwadee', 'PSBG\xa0PSB5-\xa0Ketwadee0406(\u5b8c\u6210).xlsx', ['PSB5'], [
        ('m1a', 'date→MMDD',              'date_mmdd'),
        ('m1b', 'date→M/D斜線',           'date_slash'),
        ('m1c', 'date→MON-YYYY月份',      'date_month_year'),
        ('m2a', 'PARTNO→PartNo',          'partno_case'),
        ('m4a', 'marker→UPPERCASE',       'marker_upper'),
        ('m4b', 'marker→中文',             'marker_chinese'),
        ('m6a', 'Sheet→SHEET(大寫)',       'sheet_upper'),
        ('m6b', 'Sheet→Sheet_Apr後綴',    'sheet_suffix'),
    ]),
    # kanyanat: YYYYMMDD 日期，col24=TYPE marker
    ('kanyanat', 'PSBG\xa0PSB7_Kanyanat.S0406(\u5b8c\u6210).xlsx', ['PSB7'], [
        ('m1a', 'date→MMDD',              'date_mmdd'),
        ('m1b', 'date→M/D斜線',           'date_slash'),
        ('m2a', 'PARTNO→PartNo',          'partno_case'),
        ('m4a', 'marker→UPPERCASE',       'marker_upper'),
        ('m4b', 'marker→中文',             'marker_chinese'),
        ('m6b', 'Sheet→Sheet_Apr後綴',    'sheet_suffix'),
    ]),
    # weeraya: col12=TYPE marker
    ('weeraya', 'PSBG\xa0PSB7-Weeraya0406(\u5b8c\u6210).xlsx', ['PSB7'], [
        ('m1a', 'date→MMDD',              'date_mmdd'),
        ('m1b', 'date→M/D斜線',           'date_slash'),
        ('m2a', 'PARTNO→PartNo',          'partno_case'),
        ('m4a', 'marker→UPPERCASE',       'marker_upper'),
        ('m4c', 'marker前綴改數字',        'marker_prefix'),
    ]),
    # mwc1ipc1: 多 PLANT, col6=REQUEST ITEM marker
    ('mwc1ipc1', 'W4-PSBG DNI-MWC1&IPC1 MRP 04.20.2026\u738b\u8ff0\u9023.xlsx', ['MWC1', 'IPC1'], [
        ('m1a', 'date→MMDD',              'date_mmdd'),
        ('m1b', 'date→M/D斜線',           'date_slash'),
        ('m2a', 'PARTNO→PartNo',          'partno_case'),
        ('m2c', 'PARTNO→料號',             'partno_chinese'),
        ('m3a', 'PLANT→Plant',            'plant_case'),
        ('m3c', 'plant值加空格',            'plant_value_spaces'),
        ('m4a', 'marker→UPPERCASE',       'marker_upper'),
    ]),
    # india_iai1: 多 PLANT
    ('india_iai1', 'PSBG (India IAI1&UPI2&DFI1 DIODES)-Jack0401.xlsx', ['IAI1', 'UPI2', 'DFI1'], [
        ('m1a', 'date→MMDD',              'date_mmdd'),
        ('m1b', 'date→M/D斜線',           'date_slash'),
        ('m3a', 'PLANT→Plant',            'plant_case'),
        ('m3c', 'plant值加空格',            'plant_value_spaces'),
        ('m6b', 'Sheet→Sheet_Apr後綴',    'sheet_suffix'),
    ]),
    # nbq1: flat/single row
    ('nbq1', 'NBQ1.xlsx', ['NBQ1'], [
        ('m1a', 'date→MMDD',              'date_mmdd'),
        ('m1b', 'date→M/D斜線',           'date_slash'),
        ('m2a', 'PARTNO→PartNo',          'partno_case'),
        ('m2b', 'PARTNO→PART NO.',        'partno_dotspace'),
    ]),
    # psw1_cew1: col12=Status marker
    ('psw1_cew1', 'PSBG PSW1+CEW1- \u694a\u6d0b\u5f59\u6574 0330(\u5b8c\u6210) (002).xlsx', ['PSW1', 'CEW1'], [
        ('m1a', 'date→MMDD',              'date_mmdd'),
        ('m1b', 'date→M/D斜線',           'date_slash'),
        ('m4a', 'marker→UPPERCASE',       'marker_upper'),
        ('m4b', 'marker→中文',             'marker_chinese'),
    ]),
    # iabg: flat (no marker)
    ('iabg', 'IABG-IMW1-\u9648\u59ff\u5bb9_20260413.xlsx', ['IMW1'], [
        ('m1a', 'date→MMDD',              'date_mmdd'),
        ('m1b', 'date→M/D斜線',           'date_slash'),
        ('m2a', 'PARTNO→PartNo',          'partno_case'),
        ('m5a', 'STOCK→Inventory',        'stock_synonym'),
        ('m5b', 'ON WAY→In-Transit',      'onway_synonym'),
    ]),
    # fmbg: col12=REQUEST ITEM marker
    ('fmbg', 'FMBG-MRP(TPC5)-100109-2026-4-15.xlsx', ['TPC5'], [
        ('m1a', 'date→MMDD',              'date_mmdd'),
        ('m1b', 'date→M/D斜線',           'date_slash'),
        ('m4a', 'marker→UPPERCASE',       'marker_upper'),
        ('m4b', 'marker→中文',             'marker_chinese'),
    ]),
    # ictbg_ntl7: col10=REQUEST ITEM marker
    ('ictbg_ntl7', 'ICTBG(DNI)-NTL7  4.13 MRP CFM.xlsx', ['NTL7'], [
        ('m1a', 'date→MMDD',              'date_mmdd'),
        ('m1b', 'date→M/D斜線',           'date_slash'),
        ('m4a', 'marker→UPPERCASE',       'marker_upper'),
        ('m6a', 'Sheet→SHEET(大寫)',       'sheet_upper'),
        ('m6b', 'Sheet→Sheet_Apr後綴',    'sheet_suffix'),
    ]),
    # ictbg_psb9_mrp: PSB9_MRP* sheet
    ('ictbg_psb9_mrp', 'ICTBG-PSB9-Kaewarin_20260413.xlsx', ['PSB9'], [
        ('m1a', 'date→MMDD',              'date_mmdd'),
        ('m4a', 'marker→UPPERCASE',       'marker_upper'),
        ('m6b', 'Sheet→Sheet_Apr後綴',    'sheet_suffix'),
    ]),
    # ictbg_psb9_siriraht
    ('ictbg_psb9_siriraht', 'ICTBG-PSB9-Siriraht_20260411.xlsx', ['PSB9'], [
        ('m1a', 'date→MMDD',              'date_mmdd'),
        ('m4a', 'marker→UPPERCASE',       'marker_upper'),
    ]),
    # svc1pwc1_diode_mos: 兩 sheet 組合
    ('svc1pwc1_diode_mos', 'MRP(SVC1PWC1 DIODE&MOS)-100109-2026-3-30.xlsx', ['SVC1', 'PWC1'], [
        ('m1a', 'date→MMDD',              'date_mmdd'),
        ('m1b', 'date→M/D斜線',           'date_slash'),
        ('m6b', 'Sheet→Sheet_Apr後綴',    'sheet_suffix'),
    ]),
]

# 展開 MUTATION_CASES list
MUTATION_CASES = []
for fmt_name, src_file, plant_codes, mutations_list in _MUTATION_DEFS_RAW:
    for m_id, label, mutation_type in mutations_list:
        MUTATION_CASES.append((fmt_name, src_file, plant_codes, m_id, label, mutation_type))


# ── 結構性 Mutation 定義 ───────────────────────────────────────────────
# 測試 rule-based 面對表頭位移、插入欄列、多 Sheet 等結構變化的韌性

_MUTATION_DEFS_STRUCTURAL = [
    # multirow_3 代表: ketwadee (有 Demand/Supply/Balance)
    ('ketwadee', 'PSBG\xa0PSB5-\xa0Ketwadee0406(\u5b8c\u6210).xlsx', ['PSB5'], [
        ('s1', '標題列插在表頭前(header→row2)',        'extra_title_row'),
        ('s2', '表頭往下推3列(header→row4)',            'header_pushed_down_3'),
        ('s3', '前面加 Cover Sheet',                    'add_decoy_sheet'),
        ('s4', '插入 No. 序號欄在最前面',               'extra_col_seq_no'),
        ('s5', '資料區散佈5列空白列',                    'extra_blank_rows'),
        ('s6', 'date→ISO 2026-04-27格式',               'date_iso_hyphen'),
        ('s7', 'PARTNO↔PLANT 欄互換',                   'cols_swap_partno_plant'),
        ('s8', '所有欄完全亂序(seed=42)',                'cols_all_shuffled'),
        ('s9', '日期欄前插入5個雜訊欄',                  'date_cols_pushed_right'),
    ]),
    # flat 代表: eibg_eisbg (無 marker 欄)
    ('eibg_eisbg', 'EIBG-TPW1\u2014Lydia--0427.xlsx', ['TPW1'], [
        ('s1', '標題列插在表頭前(header→row2)',        'extra_title_row'),
        ('s2', '表頭往下推3列(header→row4)',            'header_pushed_down_3'),
        ('s3', '前面加 Cover Sheet',                    'add_decoy_sheet'),
        ('s4', '插入 No. 序號欄在最前面',               'extra_col_seq_no'),
        ('s5', '資料區散佈5列空白列',                    'extra_blank_rows'),
        ('s7', 'PARTNO↔PLANT 欄互換',                   'cols_swap_partno_plant'),
        ('s8', '所有欄完全亂序(seed=42)',                'cols_all_shuffled'),
        ('s9', '日期欄前插入5個雜訊欄',                  'date_cols_pushed_right'),
    ]),
    # multi-plant 代表: mwc1ipc1
    ('mwc1ipc1', 'W4-PSBG DNI-MWC1&IPC1 MRP 04.20.2026\u738b\u8ff0\u9023.xlsx', ['MWC1', 'IPC1'], [
        ('s1', '標題列插在表頭前(header→row2)',        'extra_title_row'),
        ('s2', '表頭往下推3列(header→row4)',            'header_pushed_down_3'),
        ('s3', '前面加 Cover Sheet',                    'add_decoy_sheet'),
        ('s4', '插入 No. 序號欄在最前面',               'extra_col_seq_no'),
        ('s5', '資料區散佈5列空白列',                    'extra_blank_rows'),
        ('s6', 'date→ISO 2026-04-27格式',               'date_iso_hyphen'),
        ('s7', 'PARTNO↔PLANT 欄互換',                   'cols_swap_partno_plant'),
        ('s8', '所有欄完全亂序(seed=42)',                'cols_all_shuffled'),
        ('s9', '日期欄前插入5個雜訊欄',                  'date_cols_pushed_right'),
    ]),
    # 含 PSB9 雙 sheet 的格式
    ('ictbg_psb9_mrp', 'ICTBG-PSB9-Kaewarin_20260413.xlsx', ['PSB9'], [
        ('s1', '標題列插在表頭前(header→row2)',        'extra_title_row'),
        ('s2', '表頭往下推3列(header→row4)',            'header_pushed_down_3'),
        ('s4', '插入 No. 序號欄在最前面',               'extra_col_seq_no'),
        ('s5', '資料區散佈5列空白列',                    'extra_blank_rows'),
        ('s7', 'PARTNO↔PLANT 欄互換',                   'cols_swap_partno_plant'),
        ('s8', '所有欄完全亂序(seed=42)',                'cols_all_shuffled'),
        ('s9', '日期欄前插入5個雜訊欄',                  'date_cols_pushed_right'),
    ]),
]

STRUCTURAL_CASES = []
for fmt_name, src_file, plant_codes, mutations_list in _MUTATION_DEFS_STRUCTURAL:
    for m_id, label, mutation_type in mutations_list:
        STRUCTURAL_CASES.append((fmt_name, src_file, plant_codes, m_id, label, mutation_type))


# ── AI 深度 Mutation 定義 ─────────────────────────────────────────────
# 使用 rule-based 完全不認識的欄名/格式，驗證 AI fallback 能正確識別並讀取
# ⚠️ 需要 DEEPSEEK_API_KEY 環境變數

_MUTATION_DEFS_AI = [
    # multirow_3 + 有 marker: ketwadee
    ('ketwadee', 'PSBG\xa0PSB5-\xa0Ketwadee0406(\u5b8c\u6210).xlsx', ['PSB5'], [
        ('ai1',  'PARTNO→品番(日文，不在keyword清單)',        'partno_jp'),
        ('ai2',  'PARTNO→Item Code(不在keyword清單)',        'partno_item_code'),
        ('ai3',  'PLANT→Factory(不在keyword清單)',           'plant_factory'),
        ('ai3b', 'PLANT→Location(不在keyword清單)',          'plant_location'),
        ('ai4',  'marker→Forecast Qty/Order Qty/Net Qty',   'marker_forecast_qty'),
        ('ai5',  'marker→日文(予測量/発注量/在庫量)',         'marker_jp'),
        ('ai6',  'date→April 27 2026 英文長格式',            'date_long_en'),
        ('ai7',  'date→Excel序號整數(非字串)',                'date_excel_serial'),
    ]),
    # flat + 無 marker: eibg_eisbg
    ('eibg_eisbg', 'EIBG-TPW1\u2014Lydia--0427.xlsx', ['TPW1'], [
        ('ai1',  'PARTNO→品番(日文)',                        'partno_jp'),
        ('ai2',  'PARTNO→Item Code',                        'partno_item_code'),
        ('ai3',  'PLANT→Factory',                           'plant_factory'),
        ('ai8',  '所有欄改日文(品番/工場/現在庫)',             'all_headers_jp'),
    ]),
    # multi-plant + marker: mwc1ipc1
    ('mwc1ipc1', 'W4-PSBG DNI-MWC1&IPC1 MRP 04.20.2026\u738b\u8ff0\u9023.xlsx', ['MWC1', 'IPC1'], [
        ('ai1',  'PARTNO→品番(日文)',                        'partno_jp'),
        ('ai3',  'PLANT→Factory',                           'plant_factory'),
        ('ai4',  'marker→Forecast Qty/Order Qty/Net Qty',   'marker_forecast_qty'),
        ('ai6',  'date→April 27 2026 英文長格式',            'date_long_en'),
        ('ai7',  'date→Excel序號整數',                       'date_excel_serial'),
        ('ai8',  '所有關鍵欄改日文',                          'all_headers_jp'),
    ]),
    # 有 marker + 單 plant: kanyanat
    ('kanyanat', 'PSBG\xa0PSB7_Kanyanat.S0406(\u5b8c\u6210).xlsx', ['PSB7'], [
        ('ai1',  'PARTNO→品番(日文)',                        'partno_jp'),
        ('ai4',  'marker→Forecast Qty/Order Qty/Net Qty',   'marker_forecast_qty'),
        ('ai5',  'marker→日文(予測量/発注量/在庫量)',         'marker_jp'),
        ('ai7',  'date→Excel序號整數',                       'date_excel_serial'),
    ]),
    # india_iai1: 多 plant，無 marker，有 STOCK/ONWAY
    ('india_iai1', 'PSBG (India IAI1&UPI2&DFI1 DIODES)-Jack0401.xlsx', ['IAI1', 'UPI2', 'DFI1'], [
        ('ai1',  'PARTNO→品番(日文)',                        'partno_jp'),
        ('ai3',  'PLANT→Factory',                           'plant_factory'),
        ('ai7',  'date→Excel序號整數',                       'date_excel_serial'),
    ]),
]

AI_CASES = []
for fmt_name, src_file, plant_codes, mutations_list in _MUTATION_DEFS_AI:
    for m_id, label, mutation_type in mutations_list:
        AI_CASES.append((fmt_name, src_file, plant_codes, m_id, label, mutation_type))

# AI 是否可用（需要 deepseek_api_key 環境變數）
# 先呼叫 load_dotenv，確保 .env 中的 key 已載入
try:
    from dotenv import load_dotenv as _load_dotenv
    _load_dotenv()
except ImportError:
    pass
_AI_AVAILABLE = bool(os.getenv('deepseek_api_key'))


# ── 輔助: 格式偵測 ────────────────────────────────────────────────────
def _detect_fmt(filepath):
    """嘗試 detect_format → fingerprint，回傳 fmt 字串或 None"""
    fmt = detect_format(str(filepath))
    if not fmt:
        fmt, score = match_known_format_fingerprint(str(filepath))
    return fmt


# ── 輔助: 取得測試 ID ─────────────────────────────────────────────────
def _baseline_id(val):
    if isinstance(val, str) and '.xlsx' in val:
        return Path(val).stem[:40]
    return str(val)

def _mutation_id(val):
    if isinstance(val, tuple) and len(val) >= 4:
        fmt_name, src, plants, m_id = val[0], val[1], val[2], val[3]
        return f"{fmt_name}-{m_id}"
    return str(val)


# ══════════════════════════════════════════════════════════════════════
# TestBaseline: 23 個實際檔案基線測試
# ══════════════════════════════════════════════════════════════════════

@pytest.mark.baseline
class TestBaseline:
    """實際 Buyer 檔案基線測試：format detection / read / backfill"""

    @pytest.mark.parametrize("fname,expected_fmt,plant_codes", BASELINE_FILES,
                             ids=[_baseline_id(f[0]) for f in BASELINE_FILES])
    def test_file_exists(self, fname, expected_fmt, plant_codes):
        """測試檔案必須存在"""
        path = DESKTOP_DIR / fname
        assert path.exists(), f"File not found: {path}"

    @pytest.mark.parametrize("fname,expected_fmt,plant_codes", BASELINE_FILES,
                             ids=[_baseline_id(f[0]) for f in BASELINE_FILES])
    def test_format_detection(self, fname, expected_fmt, plant_codes):
        """detect_format() / fingerprint 能識別正確格式"""
        if expected_fmt is None:
            pytest.skip("彙總/參考檔，跳過格式偵測")
        path = DESKTOP_DIR / fname
        if not path.exists():
            pytest.skip(f"檔案不存在: {fname}")
        fmt = _detect_fmt(path)
        assert fmt == expected_fmt, (
            f"格式偵測錯誤: expected={expected_fmt!r}, got={fmt!r}\n"
            f"  File: {fname}"
        )

    @pytest.mark.parametrize("fname,expected_fmt,plant_codes", BASELINE_FILES,
                             ids=[_baseline_id(f[0]) for f in BASELINE_FILES])
    def test_read_returns_data(self, fname, expected_fmt, plant_codes):
        """read_buyer_file() 回傳 ≥1 筆，partno 非空"""
        if expected_fmt is None:
            pytest.skip("彙總/參考檔，跳過讀取測試")
        path = DESKTOP_DIR / fname
        if not path.exists():
            pytest.skip(f"檔案不存在: {fname}")
        rows = read_buyer_file(str(path), plant_codes=plant_codes or None)
        assert len(rows) >= 1, (
            f"read_buyer_file 回傳 0 筆資料\n  File: {fname}"
        )
        empty_partno = [r for r in rows if not r.get('part_no')]
        assert len(empty_partno) == 0, (
            f"{len(empty_partno)}/{len(rows)} 筆 partno 為空\n"
            f"  File: {fname}\n"
            f"  範例: {empty_partno[:2]}"
        )

    @pytest.mark.parametrize("fname,expected_fmt,plant_codes", BASELINE_FILES,
                             ids=[_baseline_id(f[0]) for f in BASELINE_FILES])
    def test_backfill_success(self, fname, expected_fmt, plant_codes, tmp_path):
        """backfill_one_file() success=True，對 flat 格式 n_cells_written > 0"""
        if expected_fmt is None:
            pytest.skip("彙總/參考檔，跳過回填測試")
        if not FORECAST_RESULT.exists():
            pytest.skip(f"forecast_result.xlsx 不存在: {FORECAST_RESULT}")
        path = DESKTOP_DIR / fname
        if not path.exists():
            pytest.skip(f"檔案不存在: {fname}")
        out = tmp_path / Path(fname).name
        out.parent.mkdir(parents=True, exist_ok=True)
        result = backfill_one_file(
            str(path), str(FORECAST_RESULT), str(out),
            plant_codes=plant_codes or None,
            file_label=fname,
        )
        assert result['success'], (
            f"backfill 失敗: {result.get('message', '')}\n"
            f"  skip_reason: {result.get('skip_reason')}\n"
            f"  File: {fname}"
        )
        is_already_backfilled = '_backfilled' in fname
        using_real_fixture = (FORECAST_RESULT == _FIXTURE_FORECAST)

        # Flat 格式: Balance 公式無條件產生
        FLAT_FORMATS = {'eibg_eisbg', 'iabg', 'nbq1'}
        if expected_fmt in FLAT_FORMATS and not is_already_backfilled:
            assert result['n_cells_written'] > 0, (
                f"Flat 格式 n_cells_written=0，應有 Balance 公式\n"
                f"  File: {fname}, format={expected_fmt}"
            )

        # Multirow_3 格式: 使用真實 fixture 時，supply lookup 有資料，應有值寫入
        MULTIROW_FORMATS = {
            'ketwadee', 'kanyanat', 'weeraya', 'mwc1ipc1',
            'india_iai1', 'fmbg', 'psw1_cew1', 'ictbg_ntl7',
            'ictbg_psb9_mrp', 'ictbg_psb9_siriraht', 'svc1pwc1_diode_mos',
        }
        if (using_real_fixture and expected_fmt in MULTIROW_FORMATS
                and not is_already_backfilled):
            assert result['n_cells_written'] > 0, (
                f"Multirow_3 格式使用真實 fixture 時 n_cells_written=0\n"
                f"  File: {fname}, format={expected_fmt}\n"
                f"  n_partno_matched={result.get('n_partno_matched', 0)}\n"
                f"  message: {result.get('message', '')}"
            )


# ══════════════════════════════════════════════════════════════════════
# TestMutation: 6 維度 × 14 格式 Mutation 測試
# ══════════════════════════════════════════════════════════════════════

@pytest.mark.mutation
class TestMutation:
    """Mutation 測試：漂移欄位後系統仍能正確讀取與回填"""

    @pytest.mark.parametrize("fmt,src,plants,m_id,label,mutation_type", MUTATION_CASES,
                             ids=[f"{c[0]}-{c[3]}" for c in MUTATION_CASES])
    def test_mutation_read(self, fmt, src, plants, m_id, label, mutation_type, tmp_path):
        """mutated 檔案能被 read_buyer_file() 正確解析 (partno 非空)"""
        src_path = DESKTOP_DIR / src
        if not src_path.exists():
            pytest.skip(f"來源檔不存在: {src}")

        # 安全的目標檔名（避免特殊字元）
        safe_name = f"mut_{m_id}_{fmt}.xlsx"
        mut_path = tmp_path / safe_name

        applied = _make_mutation(src_path, mut_path, mutation_type)
        if not applied:
            pytest.xfail(f"[{fmt}-{m_id}] mutation 無法套用（欄位不存在於此檔案）: {label}")

        rows = read_buyer_file(str(mut_path), plant_codes=plants or None,
                               file_label=safe_name)
        assert len(rows) >= 1, (
            f"[{fmt}-{label}] read_buyer_file 回傳 0 筆\n"
            f"  Mutation: {mutation_type}\n"
            f"  Source: {src}"
        )
        empty_partno = [r for r in rows if not r.get('part_no')]
        assert len(empty_partno) == 0, (
            f"[{fmt}-{label}] {len(empty_partno)}/{len(rows)} 筆 partno 為空\n"
            f"  Mutation: {mutation_type}"
        )

    @pytest.mark.parametrize("fmt,src,plants,m_id,label,mutation_type", MUTATION_CASES,
                             ids=[f"{c[0]}-{c[3]}" for c in MUTATION_CASES])
    def test_mutation_backfill(self, fmt, src, plants, m_id, label, mutation_type, tmp_path):
        """mutated 檔案能被 backfill_one_file() 成功回填"""
        if not FORECAST_RESULT.exists():
            pytest.skip(f"forecast_result.xlsx 不存在")
        src_path = DESKTOP_DIR / src
        if not src_path.exists():
            pytest.skip(f"來源檔不存在: {src}")

        safe_name = f"mut_{m_id}_{fmt}.xlsx"
        mut_path = tmp_path / safe_name
        applied = _make_mutation(src_path, mut_path, mutation_type)
        if not applied:
            pytest.xfail(f"[{fmt}-{m_id}] mutation 無法套用: {label}")

        out_path = tmp_path / f"out_{safe_name}"
        result = backfill_one_file(
            str(mut_path), str(FORECAST_RESULT), str(out_path),
            plant_codes=plants or None,
            file_label=safe_name,
        )
        assert result['success'], (
            f"[{fmt}-{label}] backfill 失敗: {result.get('message', '')}\n"
            f"  skip_reason: {result.get('skip_reason')}\n"
            f"  Mutation: {mutation_type}"
        )
        # flat 格式一定有 Balance 公式
        FLAT_FORMATS = {'eibg_eisbg', 'iabg', 'nbq1'}
        if fmt in FLAT_FORMATS:
            assert result['n_cells_written'] > 0, (
                f"[{fmt}-{label}] Flat 格式 n_cells_written=0 (應有 Balance 公式)"
            )


# ══════════════════════════════════════════════════════════════════════
# TestMutationStructural: 結構性漂移測試 (rule-based 應能處理)
# ══════════════════════════════════════════════════════════════════════

@pytest.mark.structural
class TestMutationStructural:
    """結構性 Mutation：表頭位移、插入欄/列、多 Sheet 順序等。

    這些變化應由 rule-based 層自行處理（find_header_row 掃前10列，
    find_valid_sheets 跳過無 PARTNO 的 sheet）。不需要 AI。
    """

    @pytest.mark.parametrize("fmt,src,plants,m_id,label,mutation_type", STRUCTURAL_CASES,
                             ids=[f"{c[0]}-{c[3]}" for c in STRUCTURAL_CASES])
    def test_structural_read(self, fmt, src, plants, m_id, label, mutation_type, tmp_path):
        """結構性漂移後 read_buyer_file() 仍能正確解析"""
        src_path = DESKTOP_DIR / src
        if not src_path.exists():
            pytest.skip(f"來源檔不存在: {src}")
        safe_name = f"struct_{m_id}_{fmt}.xlsx"
        mut_path = tmp_path / safe_name
        applied = _make_mutation(src_path, mut_path, mutation_type)
        if not applied:
            pytest.xfail(f"[{fmt}-{m_id}] structural mutation 無法套用: {label}")
        rows = read_buyer_file(str(mut_path), plant_codes=plants or None, file_label=safe_name)
        assert len(rows) >= 1, (
            f"[{fmt}-{label}] read_buyer_file 回傳 0 筆\n"
            f"  mutation: {mutation_type}, source: {src}"
        )
        empty_partno = [r for r in rows if not r.get('part_no')]
        assert len(empty_partno) == 0, (
            f"[{fmt}-{label}] {len(empty_partno)}/{len(rows)} 筆 partno 為空"
        )

    @pytest.mark.parametrize("fmt,src,plants,m_id,label,mutation_type", STRUCTURAL_CASES,
                             ids=[f"{c[0]}-{c[3]}" for c in STRUCTURAL_CASES])
    def test_structural_backfill(self, fmt, src, plants, m_id, label, mutation_type, tmp_path):
        """結構性漂移後 backfill_one_file() 仍能成功回填"""
        if not FORECAST_RESULT.exists():
            pytest.skip("forecast_result.xlsx 不存在")
        src_path = DESKTOP_DIR / src
        if not src_path.exists():
            pytest.skip(f"來源檔不存在: {src}")
        safe_name = f"struct_{m_id}_{fmt}.xlsx"
        mut_path = tmp_path / safe_name
        applied = _make_mutation(src_path, mut_path, mutation_type)
        if not applied:
            pytest.xfail(f"[{fmt}-{m_id}] structural mutation 無法套用: {label}")
        out_path = tmp_path / f"out_{safe_name}"
        result = backfill_one_file(
            str(mut_path), str(FORECAST_RESULT), str(out_path),
            plant_codes=plants or None, file_label=safe_name,
        )
        assert result['success'], (
            f"[{fmt}-{label}] backfill 失敗: {result.get('message', '')}\n"
            f"  skip_reason: {result.get('skip_reason')}"
        )


# ══════════════════════════════════════════════════════════════════════
# TestMutationAI: AI 深度 Mutation 測試 (rule-based 無法處理)
# ══════════════════════════════════════════════════════════════════════

@pytest.mark.ai
class TestMutationAI:
    """AI 深度 Mutation：完全陌生的欄名/日期格式。

    這些 mutation 使用 rule-based keyword 清單中不存在的詞彙：
    - 品番/工場 (日文)、Factory/Location/Item Code (英文新詞)
    - Forecast Qty/Order Qty (全新 marker 詞彙)
    - 長格式英文日期 / Excel 序號整數

    rule-based 層完全失敗 → 觸發 AI fallback (delta_ai_helper.py)
    → 驗證 AI 能正確識別後系統仍可讀取+回填。

    ⚠️  需要 DEEPSEEK_API_KEY 環境變數，否則整個 class 跳過。
    """

    @pytest.mark.parametrize("fmt,src,plants,m_id,label,mutation_type", AI_CASES,
                             ids=[f"{c[0]}-{c[3]}" for c in AI_CASES])
    def test_ai_read(self, fmt, src, plants, m_id, label, mutation_type, tmp_path):
        """AI 辨識後 read_buyer_file() 能正確解析 (partno 非空)"""
        if not _AI_AVAILABLE:
            pytest.skip("DEEPSEEK_API_KEY 未設定，跳過 AI fallback 測試")
        src_path = DESKTOP_DIR / src
        if not src_path.exists():
            pytest.skip(f"來源檔不存在: {src}")
        safe_name = f"ai_{m_id}_{fmt}.xlsx"
        mut_path = tmp_path / safe_name
        applied = _make_mutation(src_path, mut_path, mutation_type)
        if not applied:
            pytest.xfail(f"[{fmt}-{m_id}] AI mutation 無法套用（欄不存在）: {label}")
        rows = read_buyer_file(str(mut_path), plant_codes=plants or None, file_label=safe_name)
        assert len(rows) >= 1, (
            f"[{fmt}-{label}] AI fallback 後 read_buyer_file 仍回傳 0 筆\n"
            f"  → AI 未能識別欄位，或識別結果不正確\n"
            f"  mutation: {mutation_type}"
        )
        empty_partno = [r for r in rows if not r.get('part_no')]
        assert len(empty_partno) == 0, (
            f"[{fmt}-{label}] AI fallback 後 {len(empty_partno)}/{len(rows)} 筆 partno 為空\n"
            f"  → AI 可能識別了錯誤的 partno 欄位"
        )

    @pytest.mark.parametrize("fmt,src,plants,m_id,label,mutation_type", AI_CASES,
                             ids=[f"{c[0]}-{c[3]}" for c in AI_CASES])
    def test_ai_backfill(self, fmt, src, plants, m_id, label, mutation_type, tmp_path):
        """AI 辨識後 backfill_one_file() 能成功回填"""
        if not _AI_AVAILABLE:
            pytest.skip("DEEPSEEK_API_KEY 未設定，跳過 AI fallback 測試")
        if not FORECAST_RESULT.exists():
            pytest.skip("forecast_result.xlsx 不存在")
        src_path = DESKTOP_DIR / src
        if not src_path.exists():
            pytest.skip(f"來源檔不存在: {src}")
        safe_name = f"ai_{m_id}_{fmt}.xlsx"
        mut_path = tmp_path / safe_name
        applied = _make_mutation(src_path, mut_path, mutation_type)
        if not applied:
            pytest.xfail(f"[{fmt}-{m_id}] AI mutation 無法套用: {label}")
        out_path = tmp_path / f"out_{safe_name}"
        result = backfill_one_file(
            str(mut_path), str(FORECAST_RESULT), str(out_path),
            plant_codes=plants or None, file_label=safe_name,
        )
        assert result['success'], (
            f"[{fmt}-{label}] AI fallback 後 backfill 仍失敗\n"
            f"  message: {result.get('message', '')}\n"
            f"  skip_reason: {result.get('skip_reason')}\n"
            f"  mutation: {mutation_type}"
        )


# ══════════════════════════════════════════════════════════════════════
# TestMultirow3Supply: 合成 forecast_result 驗證 multirow_3 供給值寫入
# ══════════════════════════════════════════════════════════════════════

def _make_synthetic_forecast_result(tmp_path, plant, part_supply_map, date_keys):
    """建立合成的 forecast_result.xlsx，讓 multirow_3 回填測試能實際寫入資料。

    Args:
        tmp_path: pathlib.Path — 輸出目錄
        plant: str — PLANT 代碼 (如 'PSB5')
        part_supply_map: {partno: {date_key: value}} — 供給資料
        date_keys: list[str] — 日期欄清單 (如 ['20260427', '20260504', ...])

    Returns:
        Path to the synthetic forecast_result.xlsx
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # forecast_result.xlsx 固定欄位:
    # A=Buyer, B=PLANT, C=ERP客戶, D=ERP廠址, E=PARTNO, F=VENDOR PARTNO,
    # G=STOCK, H=ON-WAY, I=Date, J=PASSDUE, K+=日期欄
    header = ['Buyer', 'PLANT', 'ERP客戶', 'ERP廠址', 'PARTNO',
              'VENDOR PARTNO', 'STOCK', 'ON-WAY', 'Date', 'PASSDUE'] + date_keys
    ws.append(header)

    for partno, supply_dict in part_supply_map.items():
        # Demand row (必要, 不影響 lookup)
        demand_row = ['SyntheticBuyer', plant, '測試', '測試廠', partno,
                      '', 0, 0, 'Demand', 0] + [0] * len(date_keys)
        ws.append(demand_row)

        # Supply row
        supply_vals = [supply_dict.get(dk, 0) for dk in date_keys]
        supply_row = ['SyntheticBuyer', plant, '測試', '測試廠', partno,
                      '', None, None, 'Supply', None] + supply_vals
        ws.append(supply_row)

    out = tmp_path / "synthetic_forecast_result.xlsx"
    wb.save(str(out))
    return out


# 用於 multirow_3 供給值寫入測試的資料 (plant, 部分 partno, 日期值)
# partno 來自 read_buyer_file() 的實際回傳，確保能比對到 lookup
_MULTIROW3_SUPPLY_CASES = [
    (
        "ketwadee",
        "PSBG\xa0PSB5-\xa0Ketwadee0406(\u5b8c\u6210).xlsx",
        ["PSB5"],
        "PSB5",
        {
            "203318830134": {"20260427": 1000, "20260504": 2000},
            "203318870079": {"20260427": 500,  "20260504": 1500},
        },
        ["20260427", "20260504", "20260511"],
    ),
    (
        "kanyanat",
        "PSBG\xa0PSB7_Kanyanat.S0406(\u5b8c\u6210).xlsx",  # \xa0 non-breaking space
        ["PSB7"],
        "PSB7",
        {
            "242953653306": {"20260427": 800,  "20260504": 1600},
            "243690600736": {"20260427": 400,  "20260504": 800},
        },
        ["20260427", "20260504", "20260511"],
    ),
    (
        "mwc1ipc1",
        "W4-PSBG DNI-MWC1&IPC1 MRP 04.20.2026\u738b\u8ff0\u9023.xlsx",
        ["MWC1", "IPC1"],
        "MWC1",
        {
            "203823280034": {"20260427": 600, "20260504": 1200},
            "203852600034": {"20260427": 300, "20260504": 900},
        },
        ["20260427", "20260504", "20260511"],
    ),
]


@pytest.mark.structural
class TestMultirow3Supply:
    """驗證 multirow_3 格式在有合成 Supply 資料時能實際寫入儲存格。

    使用 _make_synthetic_forecast_result 建立含已知供給量的 forecast_result，
    確保 backfill_one_file 能真正寫入 n_cells_written > 0。
    """

    @pytest.mark.parametrize(
        "fmt,src,plants,plant,supply_map,date_keys",
        _MULTIROW3_SUPPLY_CASES,
        ids=[c[0] for c in _MULTIROW3_SUPPLY_CASES],
    )
    def test_supply_cells_written(self, fmt, src, plants, plant,
                                  supply_map, date_keys, tmp_path):
        """multirow_3 回填有 supply lookup 時 n_cells_written > 0"""
        src_path = DESKTOP_DIR / src
        if not src_path.exists():
            pytest.skip(f"來源檔不存在: {src}")

        # 建合成 forecast_result
        fr_path = _make_synthetic_forecast_result(tmp_path, plant, supply_map, date_keys)
        out_path = tmp_path / f"out_{fmt}.xlsx"

        result = backfill_one_file(
            str(src_path), str(fr_path), str(out_path),
            plant_codes=plants or None,
            file_label=src,
        )

        assert result['success'], (
            f"[{fmt}] multirow_3 回填失敗: {result.get('message', '')}\n"
            f"  skip_reason: {result.get('skip_reason')}"
        )
        assert result['n_cells_written'] > 0, (
            f"[{fmt}] 合成 supply 資料後 n_cells_written=0\n"
            f"  n_partno_matched={result.get('n_partno_matched', 0)}\n"
            f"  lookup plant={plant!r}, partno keys={list(supply_map.keys())[:3]}\n"
            f"  message: {result.get('message', '')}"
        )

    @pytest.mark.parametrize(
        "fmt,src,plants,plant,supply_map,date_keys",
        _MULTIROW3_SUPPLY_CASES,
        ids=[c[0] for c in _MULTIROW3_SUPPLY_CASES],
    )
    def test_supply_values_correct(self, fmt, src, plants, plant,
                                   supply_map, date_keys, tmp_path):
        """驗證回填後輸出檔的 supply 值非零（用 read_buyer_file 重新解析確認）"""
        src_path = DESKTOP_DIR / src
        if not src_path.exists():
            pytest.skip(f"來源檔不存在: {src}")

        fr_path = _make_synthetic_forecast_result(tmp_path, plant, supply_map, date_keys)
        out_path = tmp_path / f"out_vals_{fmt}.xlsx"

        result = backfill_one_file(
            str(src_path), str(fr_path), str(out_path),
            plant_codes=plants or None,
            file_label=src,
        )

        if not result['success'] or result['n_cells_written'] == 0:
            pytest.skip(f"[{fmt}] 未寫入任何儲存格，略過值驗證")

        # 用 read_buyer_file 重新解析輸出檔，確認 supply 值已寫入
        out_rows = read_buyer_file(str(out_path), plant_codes=plants or None,
                                   file_label=str(out_path.name))
        assert len(out_rows) >= 1, f"[{fmt}] 輸出檔 read_buyer_file 回傳 0 筆"

        # 找到對應 partno 的 supply 列，確認有非零值
        first_partno = list(supply_map.keys())[0]
        matched_rows = [r for r in out_rows if r.get('part_no') == first_partno]
        assert matched_rows, (
            f"[{fmt}] 輸出檔找不到 partno={first_partno}，"
            f"有的 partno: {[r.get('part_no') for r in out_rows[:3]]}"
        )
        row0 = matched_rows[0]
        supply_dict = row0.get('supply', {})
        nonzero_supply = {k: v for k, v in supply_dict.items() if v}
        assert nonzero_supply, (
            f"[{fmt}] partno={first_partno} supply 全為零\n"
            f"  supply dict: {supply_dict}"
        )


# ══════════════════════════════════════════════════════════════════════
# conftest hook: session-end 報告
# （掛在本檔，在 conftest.py 也可設）
# ══════════════════════════════════════════════════════════════════════

def pytest_sessionfinish(session, exitstatus):
    """Session 結束時輸出 Markdown 測試報告"""
    report_path = ROOT / "test_delta_report.md"
    terminalreporter = session.config.pluginmanager.get_plugin("terminalreporter")
    if not terminalreporter:
        return

    passed = terminalreporter.stats.get("passed", [])
    failed = terminalreporter.stats.get("failed", [])
    skipped = terminalreporter.stats.get("skipped", [])
    xfailed = terminalreporter.stats.get("xfailed", [])

    lines = [
        "# Delta 格式測試報告",
        f"產出時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        f"## 統計摘要",
        f"| 項目 | 數量 |",
        f"|------|------|",
        f"| PASSED | {len(passed)} |",
        f"| FAILED | {len(failed)} |",
        f"| SKIPPED | {len(skipped)} |",
        f"| XFAILED | {len(xfailed)} |",
        "",
    ]

    if failed:
        lines += [
            "## 失敗案例",
            "| 測試 | 錯誤訊息 |",
            "|------|---------|",
        ]
        for rep in failed:
            node = rep.nodeid
            msg = str(rep.longreprtext if hasattr(rep, 'longreprtext') else rep.longrepr)
            # 取第一行
            first_line = msg.split('\n')[0][:120] if msg else ''
            lines.append(f"| `{node}` | {first_line} |")
        lines.append("")

    if xfailed:
        lines += [
            "## XFAILED (mutation 欄位不存在，預期失敗)",
            "| 測試 |",
            "|------|",
        ]
        for rep in xfailed:
            lines.append(f"| `{rep.nodeid}` |")
        lines.append("")

    try:
        report_path.write_text('\n'.join(lines), encoding='utf-8')
    except Exception:
        pass
