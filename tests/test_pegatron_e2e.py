# -*- coding: utf-8 -*-
"""
Pegatron Forecast Processor - End-to-End Test
完整流程: 讀取 → 匹配 → 填入 ETA QTY → 公式重算 → 驗證 Commit
"""
import pytest
import os
import shutil
from datetime import datetime, date
from unittest.mock import patch
from openpyxl import Workbook, load_workbook
import pandas as pd

from pegatron_forecast_processor import PegatronForecastProcessor


# ============================================================
# Test Data Builders
# ============================================================

def _write_group(ws, start_row, f_val, g_val, ordered_item,
                 balance1_rw, balance2_rw):
    """
    寫入一個 8-row Pegatron 群組到 worksheet。
    balance1_rw / balance2_rw: 長度 6 的 list，對應 R~W (cols 18-23) 的值。
    """
    r = start_row
    # Row 1: WEEK#
    ws.cell(row=r, column=6, value=f_val)
    ws.cell(row=r, column=7, value=g_val)
    ws.cell(row=r, column=13, value="WEEK#")
    # Row 2: FORECAST
    ws.cell(row=r + 1, column=9, value=ordered_item)
    ws.cell(row=r + 1, column=13, value="FORECAST")
    # Row 3: OTW QTY
    ws.cell(row=r + 2, column=13, value="OTW QTY")
    # Row 4: ETA DATE
    ws.cell(row=r + 3, column=13, value="ETA DATE")
    # Row 5: ETA QTY (初始 0)
    ws.cell(row=r + 4, column=13, value="ETA QTY")
    for col in range(15, 24):
        ws.cell(row=r + 4, column=col, value=0)
    # Row 6: Balance 第一列 (offset +5)
    ws.cell(row=r + 5, column=13, value="Balance(VMI)")
    for i, val in enumerate(balance1_rw):
        ws.cell(row=r + 5, column=18 + i, value=val)
    # Row 7: Balance 第二列 (offset +6)
    ws.cell(row=r + 6, column=13, value="Balance(REAL)")
    for i, val in enumerate(balance2_rw):
        ws.cell(row=r + 6, column=18 + i, value=val)
    # Row 8: WOS / DSI
    ws.cell(row=r + 7, column=13, value="WOS")


def create_test_forecast(path):
    """
    建立測試用 Forecast (3 groups):
      Group 1: Balance1 全正 → Commit = Y
      Group 2: Balance1/2 都有負值 → Commit = N
      Group 3: Balance1 有負, Balance2 全正 → Commit = Y
    """
    wb = Workbook()
    ws = wb.active

    # Row 1: 檔案標題
    ws.cell(row=1, column=1, value="SH1")

    # Row 2: 欄位標題 + 日期欄位
    ws.cell(row=2, column=6, value="Plant")
    ws.cell(row=2, column=7, value="MRP ID")
    ws.cell(row=2, column=9, value="Description/PN")
    ws.cell(row=2, column=12, value="Commit")
    ws.cell(row=2, column=13, value="Date")
    ws.cell(row=2, column=14, value="Pass due")
    # 日期欄位 (cols 15~23) - 週一日期
    week_dates = [
        datetime(2026, 3, 30),  # col 15 (O)
        datetime(2026, 4, 6),   # col 16 (P)
        datetime(2026, 4, 13),  # col 17 (Q)
        datetime(2026, 4, 20),  # col 18 (R)
        datetime(2026, 4, 27),  # col 19 (S)
        datetime(2026, 5, 4),   # col 20 (T)
        datetime(2026, 5, 11),  # col 21 (U)
        datetime(2026, 5, 18),  # col 22 (V)
        datetime(2026, 5, 25),  # col 23 (W)
    ]
    for i, d in enumerate(week_dates):
        ws.cell(row=2, column=15 + i, value=d)

    # Group 1 (rows 3-10): Balance1 R~W 全正 → Y
    _write_group(ws, 3, "3A33", "A00Y", "PART001",
                 balance1_rw=[1000, 2000, 3000, 500, 800, 1200],
                 balance2_rw=[1000, -500, 3000, 500, 800, 1200])

    # Group 2 (rows 11-18): Balance1/2 都有負 → N
    _write_group(ws, 11, "3A33", "B00Y", "PART002",
                 balance1_rw=[-100, 2000, -300, 500, 800, 1200],
                 balance2_rw=[1000, -500, 3000, -500, 800, -1200])

    # Group 3 (rows 19-26): Balance1 有負, Balance2 全正 → Y
    _write_group(ws, 19, "3A33", "C00Y", "PART003",
                 balance1_rw=[1000, -200, 3000, 500, 800, 1200],
                 balance2_rw=[500, 600, 700, 800, 900, 100])

    wb.save(path)
    wb.close()


def create_test_erp(path):
    """
    建立測試用 ERP (2 records):
      Record 1: 匹配 Group 1 (3A33-A00Y / PART001)
      Record 2: 匹配 Group 3 (3A33-C00Y / PART003)
    """
    data = {
        '客戶需求地區': ['TW', 'TW'],
        'Line 客戶採購單號': ['3A33-A00Y', '3A33-C00Y'],
        '客戶料號': ['PART001', 'PART003'],
        '淨需求': [5, 3],
        '排程出貨日期': [datetime(2026, 3, 30), datetime(2026, 4, 6)],
        '排程出貨日期斷點': ['週三', '週三'],
        'ETA': ['本週五', '本週四'],
    }
    df = pd.DataFrame(data)
    df.to_excel(path, index=False)


def _mock_recalc(file_path, output_path=None):
    """Mock recalculate: Balance 是數值不是公式，直接複製即可"""
    target = output_path or file_path
    if os.path.abspath(file_path) != os.path.abspath(target):
        shutil.copy2(file_path, target)
    return True


# ============================================================
# Tests
# ============================================================

class TestPegatronEndToEnd:
    """Pegatron 完整流程測試: ERP 匹配 → ETA QTY → Commit"""

    @pytest.fixture
    def test_env(self, tmp_path):
        """建立完整測試環境"""
        forecast_path = tmp_path / "forecast.xlsx"
        erp_path = tmp_path / "erp.xlsx"
        output_dir = tmp_path / "output"
        output_dir.mkdir()

        create_test_forecast(str(forecast_path))
        create_test_erp(str(erp_path))

        return {
            'forecast': str(forecast_path),
            'erp': str(erp_path),
            'output_dir': str(output_dir),
            'output_file': str(output_dir / "result.xlsx"),
        }

    def _run_processor(self, test_env, is_merged=False):
        """建立並執行 processor"""
        processor = PegatronForecastProcessor(
            forecast_file=test_env['forecast'],
            erp_file=test_env['erp'],
            output_folder=test_env['output_dir'],
            output_filename='result.xlsx',
            is_merged=is_merged,
        )
        processor.process_all_blocks()
        return processor

    @patch('libreoffice_utils.recalculate_xlsx', side_effect=_mock_recalc)
    def test_full_flow_commit(self, mock_recalc, test_env):
        """
        E2E: 非合併模式 → 不插入 A 欄，欄位位置不變
        Group 1: Balance1 全正 → Y
        Group 2: 都有負 → N
        Group 3: Balance2 全正 → Y
        """
        self._run_processor(test_env, is_merged=False)

        wb = load_workbook(test_env['output_file'])
        ws = wb.active

        # --- Commit 驗證 (非合併，col 12 不變) ---
        assert ws.cell(row=3, column=12).value == "Y", "Group 1: Balance1 全正 → Y"
        assert ws.cell(row=11, column=12).value == "N", "Group 2: 都有負值 → N"
        assert ws.cell(row=19, column=12).value == "Y", "Group 3: Balance2 全正 → Y"

        # --- ETA QTY 驗證 ---
        # Group 1: qty=5, *1000=5000, target=2026-04-03 (Fri) → week of 3/30 → col 15
        assert ws.cell(row=7, column=15).value == 5000, "Group 1 ETA QTY"
        # Group 3: qty=3, *1000=3000, target=2026-04-09 (Thu) → week of 4/6 → col 16
        assert ws.cell(row=23, column=16).value == 3000, "Group 3 ETA QTY"

        # Group 2: 沒有 ERP 匹配 → ETA QTY 維持 0
        for col in range(15, 24):
            val = ws.cell(row=15, column=col).value
            assert val == 0 or val is None, f"Group 2 ETA QTY col {col} should be 0"

        wb.close()

    @patch('libreoffice_utils.recalculate_xlsx', side_effect=_mock_recalc)
    def test_commit_all_negative_is_n(self, mock_recalc, test_env):
        """單獨驗證: Balance1/2 都有負值 → Commit = N"""
        self._run_processor(test_env)

        wb = load_workbook(test_env['output_file'])
        ws = wb.active
        assert ws.cell(row=11, column=12).value == "N"
        wb.close()

    @patch('libreoffice_utils.recalculate_xlsx', side_effect=_mock_recalc)
    def test_erp_allocation_marked(self, mock_recalc, test_env):
        """驗證 ERP 已分配狀態被正確標記"""
        self._run_processor(test_env)

        erp_df = pd.read_excel(test_env['erp'])
        assert (erp_df['已分配'] == '✓').sum() == 2, "2 筆 ERP 應被標記已分配"

    @patch('libreoffice_utils.recalculate_xlsx', side_effect=_mock_recalc)
    def test_stats_correct(self, mock_recalc, test_env):
        """驗證統計數字"""
        processor = self._run_processor(test_env)

        assert processor.total_filled == 2, "ERP 應有 2 筆更新"
        assert processor.total_transit_filled == 0, "無 Transit"

    @patch('pegatron_forecast_processor.PegatronForecastProcessor._fill_part_number_with_excel_com',
           side_effect=Exception("skip Excel COM in test"))
    @patch('libreoffice_utils.recalculate_xlsx', side_effect=_mock_recalc)
    def test_merged_has_part_number_in_column_a(self, mock_recalc, mock_excel_com, test_env):
        """
        驗證: 合併模式 → 插入新 A 欄，每個群組 8 列都填入料號
        """
        self._run_processor(test_env, is_merged=True)

        wb = load_workbook(test_env['output_file'])
        ws = wb.active

        # 標題列
        assert ws.cell(row=2, column=1).value == "PN Model", "A2 應為 PN Model 標題"

        # 每個群組的 8 列都應有對應的客戶料號（不合併）
        for r in range(3, 11):
            assert ws.cell(row=r, column=1).value == "PART001", f"Group 1 row {r} A欄=PART001"
        for r in range(11, 19):
            assert ws.cell(row=r, column=1).value == "PART002", f"Group 2 row {r} A欄=PART002"
        for r in range(19, 27):
            assert ws.cell(row=r, column=1).value == "PART003", f"Group 3 row {r} A欄=PART003"

        # 原本的資料應該右移一欄（原 col 6 Plant → col 7）
        assert ws.cell(row=2, column=7).value == "Plant", "原 Plant 欄應右移到 G"

        wb.close()

    @patch('libreoffice_utils.recalculate_xlsx', side_effect=_mock_recalc)
    def test_non_merged_no_column_a(self, mock_recalc, test_env):
        """驗證: 非合併模式 → 不插入 A 欄，A 欄維持原本資料"""
        self._run_processor(test_env, is_merged=False)

        wb = load_workbook(test_env['output_file'])
        ws = wb.active

        # A 欄不應有 PN Model 標題
        assert ws.cell(row=2, column=1).value != "PN Model", "非合併模式不應有 PN Model"
        # Plant 維持在原本的 col 6
        assert ws.cell(row=2, column=6).value == "Plant", "Plant 應維持在 col 6"

        wb.close()
