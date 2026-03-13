"""
光寶(Liteon) Forecast Processor
處理 Liteon 的 Forecast 填入邏輯

支援兩種模式:
1. 單檔模式 (merged_mode=False): 每個 Forecast 檔案獨立處理
2. 合併模式 (merged_mode=True): 多個檔案合併後一次處理

單檔 Forecast 結構 (Daily+Weekly+Monthly sheet):
- C1: Plant code (e.g. "15K0")
- Row 7: 日期 headers (K~AO=BY天, AP~BK=BY周, BL~BQ=BY月)
- Row 8+: 每個料號 3 行一組 (Demand / Commit / Accumulate Shortage)
- Column B: Material (料號)
- Column C: Data Measures ("Demand" / "Commit" / "Accumulate Shortage")

合併 Forecast 結構:
- Row 1: [Plant, Buyer Code, 原 Row 7 headers] ← 欄位右移 2
- Row 2+: [plant_val, buyer_val, 原 Row 8+ data]
- Column A: Plant code (per row)
- Column D: Material (原 Column B + 2)
- Column E: Data Measures (原 Column C + 2)

Transit/ERP 填入邏輯:
- 客戶需求地區 == Plant code (單檔: C1, 合併: Column A)
- 料號匹配 + 日期計算 → 填入 Commit row
"""

import os
import pandas as pd
import openpyxl
from datetime import datetime, timedelta


class LiteonForecastProcessor:
    """Liteon Forecast Processor - 處理光寶 Forecast 填入"""

    # Forecast sheet layout constants
    SHEET_NAME = 'Daily+Weekly+Monthly'
    PLANT_CELL_ROW = 1      # C1 = Plant code
    PLANT_CELL_COL = 3       # Column C
    DATE_HEADER_ROW = 7      # Row 7 = date headers
    DATA_START_ROW = 8       # Row 8+ = data rows
    MATERIAL_COL = 2         # Column B = Material
    DATA_MEASURES_COL = 3    # Column C = "Demand" / "Commit" / "Accumulate Shortage"

    # Date column ranges (1-based column indices)
    DAILY_START_COL = 11     # Column K
    DAILY_END_COL = 41       # Column AO
    WEEKLY_START_COL = 42    # Column AP
    WEEKLY_END_COL = 63      # Column BK
    MONTHLY_START_COL = 64   # Column BL
    MONTHLY_END_COL = 69     # Column BQ

    def __init__(self, forecast_file, erp_file, transit_file=None,
                 output_folder=None, output_filename=None, merged_mode=False):
        self.forecast_file = forecast_file
        self.erp_file = erp_file
        self.transit_file = transit_file
        self.output_folder = output_folder or os.path.dirname(forecast_file)
        self.output_filename = output_filename or 'forecast_result.xlsx'
        self.merged_mode = merged_mode

        # 合併模式: 所有欄位右移 2 (Plant + Buyer Code 佔 A, B)
        if merged_mode:
            self.col_offset = 2
            self.plant_col = 1           # Column A = Plant (per row)
            self.buyer_col = 2           # Column B = Buyer Code (per row)
            self.date_header_row = 1     # Row 1 = headers
            self.data_start_row = 2      # Row 2+ = data
        else:
            self.col_offset = 0
            self.date_header_row = self.DATE_HEADER_ROW
            self.data_start_row = self.DATA_START_ROW

        # Dynamic column positions (apply offset)
        self.material_col = self.MATERIAL_COL + self.col_offset
        self.data_measures_col = self.DATA_MEASURES_COL + self.col_offset
        self.daily_start_col = self.DAILY_START_COL + self.col_offset
        self.daily_end_col = self.DAILY_END_COL + self.col_offset
        self.weekly_start_col = self.WEEKLY_START_COL + self.col_offset
        self.weekly_end_col = self.WEEKLY_END_COL + self.col_offset
        self.monthly_start_col = self.MONTHLY_START_COL + self.col_offset
        self.monthly_end_col = self.MONTHLY_END_COL + self.col_offset

        # Statistics
        self.total_filled = 0
        self.total_skipped = 0
        self.total_transit_filled = 0
        self.total_transit_skipped = 0

        # Internal state
        self.wb = None
        self.ws = None
        self.plant_code = None
        self.date_map = {}          # col_index -> date object
        # material_commit_rows:
        #   單檔模式: material -> commit row number
        #   合併模式: (plant, material) -> commit row number
        self.material_commit_rows = {}
        self.pending_changes = []   # list of {row, col, value}
        self.erp_df = None
        self.transit_df = None

    def process_all_blocks(self):
        """主入口：處理整個 Forecast 檔案"""
        try:
            print(f"\n{'='*60}")
            print(f"[Liteon] 開始處理 Forecast: {os.path.basename(self.forecast_file)}")
            print(f"{'='*60}")

            # 1. Load forecast workbook
            self._load_forecast()

            # 2. Load ERP & Transit data
            self._load_data()

            # 3. Process Transit first (if available)
            if self.transit_df is not None and len(self.transit_df) > 0:
                self._process_transit()

            # 4. Process ERP
            if self.erp_df is not None and len(self.erp_df) > 0:
                self._process_erp()

            # 5. Apply all changes
            self._apply_changes()

            # 6. Save output
            self._save_file()

            # 7. Save allocation status back to source files
            self._save_allocation_status()

            print(f"\n[Liteon] 處理完成:")
            print(f"  Transit 填入: {self.total_transit_filled}, 跳過: {self.total_transit_skipped}")
            print(f"  ERP 填入: {self.total_filled}, 跳過: {self.total_skipped}")
            return True

        except Exception as e:
            print(f"[Liteon] 處理失敗: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    def _load_forecast(self):
        """Load forecast workbook and parse structure"""
        print(f"[Liteon] 載入 Forecast: {os.path.basename(self.forecast_file)}")
        self.wb = openpyxl.load_workbook(self.forecast_file)

        if self.SHEET_NAME not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{self.SHEET_NAME}' not found in {self.forecast_file}")

        self.ws = self.wb[self.SHEET_NAME]

        # Get Plant code
        if self.merged_mode:
            self.plant_code = None  # 合併模式: Plant 在每行 Column A
            print(f"[Liteon] 合併模式: Plant 從每行 Column A 讀取")
        else:
            self.plant_code = str(self.ws.cell(row=self.PLANT_CELL_ROW, column=self.PLANT_CELL_COL).value or '').strip()
            print(f"[Liteon] Plant: {self.plant_code}")

        # Parse date headers from row 7
        self._parse_date_headers()

        # Build material -> commit row mapping
        self._build_material_index()

    def _parse_date_headers(self):
        """Parse date columns from header row"""
        self.date_map = {}  # col_index (1-based) -> date object

        for col in range(self.daily_start_col, self.monthly_end_col + 1):
            cell_val = self.ws.cell(row=self.date_header_row, column=col).value
            if cell_val is None:
                continue

            date_obj = self._parse_date_value(cell_val)
            if date_obj:
                self.date_map[col] = date_obj

        print(f"[Liteon] 日期欄位: {len(self.date_map)} columns "
              f"(offset={self.col_offset})")

    def _parse_date_value(self, val):
        """Parse a cell value to a date object"""
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return None
        if isinstance(val, datetime):
            return val.date()
        if hasattr(val, 'date') and callable(val.date):
            try:
                return val.date()
            except:
                return None
        if isinstance(val, str):
            val = val.strip()
            if not val or val.lower() == 'nan' or val.lower() == 'nat':
                return None
            for fmt in ['%Y/%m/%d', '%Y-%m-%d', '%m/%d/%Y']:
                try:
                    return datetime.strptime(val, fmt).date()
                except ValueError:
                    continue
        return None

    def _build_material_index(self):
        """Build mapping: material number -> commit row number
        單檔模式: material -> row
        合併模式: (plant, material) -> row
        """
        self.material_commit_rows = {}

        for row in range(self.data_start_row, self.ws.max_row + 1):
            measure = str(self.ws.cell(row=row, column=self.data_measures_col).value or '').strip()
            if measure == 'Commit':
                material = str(self.ws.cell(row=row, column=self.material_col).value or '').strip()
                if not material:
                    continue
                if self.merged_mode:
                    plant = str(self.ws.cell(row=row, column=self.plant_col).value or '').strip()
                    if plant:
                        self.material_commit_rows[(plant, material)] = row
                else:
                    self.material_commit_rows[material] = row

        print(f"[Liteon] 料號數: {len(self.material_commit_rows)}")

    def _load_data(self):
        """Load ERP and Transit data"""
        # Load ERP
        if os.path.exists(self.erp_file):
            self.erp_df = pd.read_excel(self.erp_file)
            if '已分配' not in self.erp_df.columns:
                self.erp_df['已分配'] = ''
            self.erp_df['已分配'] = self.erp_df['已分配'].astype(str).replace('nan', '')
            print(f"[Liteon] ERP: {len(self.erp_df)} rows")
        else:
            self.erp_df = pd.DataFrame()

        # Load Transit
        if self.transit_file and os.path.exists(self.transit_file):
            self.transit_df = pd.read_excel(self.transit_file)
            if '已分配' not in self.transit_df.columns:
                self.transit_df['已分配'] = ''
            self.transit_df['已分配'] = self.transit_df['已分配'].astype(str).replace('nan', '')
            print(f"[Liteon] Transit: {len(self.transit_df)} rows")
        else:
            self.transit_df = None

    def _find_date_column(self, target_date):
        """
        找到目標日期對應的欄位。
        策略: 先找精確天 → 再找所屬周 → 再找所屬月
        """
        if target_date is None:
            return None

        # Step 1: 精確匹配日期 (BY天)
        for col, date_obj in self.date_map.items():
            if col > self.daily_end_col:
                continue
            if date_obj == target_date:
                return col

        # Step 2: 找所屬的周 (BY周)
        # Weekly columns represent week start (Monday), find the week that contains target_date
        for col, date_obj in self.date_map.items():
            if col < self.weekly_start_col or col > self.weekly_end_col:
                continue
            # Each weekly column covers a 7-day range starting from date_obj
            week_start = date_obj
            week_end = date_obj + timedelta(days=6)
            if week_start <= target_date <= week_end:
                return col

        # Step 3: 找所屬的月 (BY月)
        for col, date_obj in self.date_map.items():
            if col < self.monthly_start_col or col > self.monthly_end_col:
                continue
            # Monthly column: same year+month
            if date_obj.year == target_date.year and date_obj.month == target_date.month:
                return col

        return None

    def _add_change(self, row, col, value):
        """Add a pending change, accumulating values for the same cell"""
        for change in self.pending_changes:
            if change['row'] == row and change['col'] == col:
                change['value'] += value
                return
        self.pending_changes.append({'row': row, 'col': col, 'value': value})

    def _process_transit(self):
        """Process transit data → fill into forecast"""
        print(f"\n[Liteon] === 處理 Transit 填入 ===")

        # Find the region column (客戶需求地區)
        region_col = '客戶需求地區'
        item_col = 'Ordered Item'
        qty_col = 'Qty'
        eta_col = 'ETA'

        if region_col not in self.transit_df.columns:
            print(f"[Liteon] Transit 缺少 '{region_col}' 欄位，跳過")
            return

        for idx, row in self.transit_df.iterrows():
            # Skip already allocated
            if str(row.get('已分配', '')).strip() == '✓':
                self.total_transit_skipped += 1
                continue

            # Match region and material
            region = str(row.get(region_col, '')).strip()
            material = str(row.get(item_col, '')).strip()

            if self.merged_mode:
                # 合併模式: 用 (region, material) 查 commit row
                key = (region, material)
                if not material or key not in self.material_commit_rows:
                    self.total_transit_skipped += 1
                    continue
                commit_row = self.material_commit_rows[key]
            else:
                # 單檔模式: 先比對 plant_code，再查 material
                if region != self.plant_code:
                    continue
                if not material or material not in self.material_commit_rows:
                    self.total_transit_skipped += 1
                    continue
                commit_row = self.material_commit_rows[material]

            # Parse ETA date
            eta_val = row.get(eta_col)
            eta_date = self._parse_date_value(eta_val)
            if eta_date is None:
                try:
                    eta_date = pd.to_datetime(eta_val).date()
                except:
                    self.total_transit_skipped += 1
                    continue

            # Find target column (day → week → month fallback)
            target_col = self._find_date_column(eta_date)
            if target_col is None:
                self.total_transit_skipped += 1
                continue

            # Get quantity and multiply by 1000
            try:
                qty = float(row.get(qty_col, 0))
            except (ValueError, TypeError):
                self.total_transit_skipped += 1
                continue

            fill_value = qty * 1000

            self._add_change(commit_row, target_col, fill_value)
            self.transit_df.at[idx, '已分配'] = '✓'
            self.total_transit_filled += 1

        print(f"[Liteon] Transit: {self.total_transit_filled} 填入, "
              f"{self.total_transit_skipped} 跳過")

    def _process_erp(self):
        """Process ERP data → fill into forecast"""
        print(f"\n[Liteon] === 處理 ERP 填入 ===")

        # Detect required columns
        region_col = '客戶需求地區'
        part_col = '客戶料號'

        # Check if required columns exist
        required = [region_col, part_col]
        for col_name in required:
            if col_name not in self.erp_df.columns:
                print(f"[Liteon] ERP 缺少 '{col_name}' 欄位，跳過")
                return

        # Date calculation columns
        date_calc_col = '日期算法'        # ETD or ETA
        schedule_col = '排程出貨日期'      # Schedule date
        breakpoint_col = '排程出貨日期斷點'  # Breakpoint (weekday)
        etd_col = 'ETD'
        eta_col = 'ETA'
        qty_col = '淨需求'                 # Net demand quantity

        for idx, row in self.erp_df.iterrows():
            try:
                # Skip already allocated
                if str(row.get('已分配', '')).strip() == '✓':
                    self.total_skipped += 1
                    continue

                # Match region and material
                region = str(row.get(region_col, '')).strip()
                part_number = str(row.get(part_col, '')).strip()

                if self.merged_mode:
                    # 合併模式: 用 (region, material) 查 commit row
                    key = (region, part_number)
                    if not part_number or key not in self.material_commit_rows:
                        self.total_skipped += 1
                        continue
                    commit_row = self.material_commit_rows[key]
                else:
                    # 單檔模式: 先比對 plant_code，再查 material
                    if region != self.plant_code:
                        continue
                    if not part_number or part_number not in self.material_commit_rows:
                        self.total_skipped += 1
                        continue
                    commit_row = self.material_commit_rows[part_number]

                # Calculate target date
                target_date = self._calculate_erp_target_date(row, date_calc_col,
                                                              schedule_col, breakpoint_col,
                                                              etd_col, eta_col)
                if target_date is None:
                    self.total_skipped += 1
                    continue

                # Find target column (day → week → month fallback)
                target_col = self._find_date_column(target_date)
                if target_col is None:
                    self.total_skipped += 1
                    continue

                # Get quantity and multiply by 1000
                try:
                    qty = float(row.get(qty_col, 0))
                except (ValueError, TypeError):
                    self.total_skipped += 1
                    continue

                if qty <= 0:
                    self.total_skipped += 1
                    continue

                fill_value = qty * 1000

                self._add_change(commit_row, target_col, fill_value)
                self.erp_df.at[idx, '已分配'] = '✓'
                self.total_filled += 1

            except Exception as e:
                self.total_skipped += 1
                continue

        print(f"[Liteon] ERP: {self.total_filled} 填入, "
              f"{self.total_skipped} 跳過")

    def _calculate_erp_target_date(self, row, date_calc_col, schedule_col,
                                    breakpoint_col, etd_col, eta_col):
        """
        計算 ERP 的目標填入日期

        日期算法:
        1. 讀取 排程出貨日期 (schedule date)
        2. 讀取 排程出貨日期斷點 (weekday breakpoint)
        3. 計算 week_end (breakpoint 當天)
        4. 根據 日期算法 (ETD/ETA):
           - ETD: 讀取 ETD mapping (e.g. "本週三", "下週一")
           - ETA: 讀取 ETA mapping (e.g. "下下週二")
        5. 從 week_end 計算實際目標日期
        """
        # Get schedule date
        schedule_val = row.get(schedule_col)
        schedule_date = self._parse_date_value(schedule_val)
        if schedule_date is None:
            try:
                schedule_date = pd.to_datetime(schedule_val).date()
            except:
                return None

        # Get breakpoint weekday
        breakpoint_val = str(row.get(breakpoint_col, '')).strip()
        if not breakpoint_val:
            return None

        # Calculate week end from breakpoint
        week_end = self._get_week_end_by_breakpoint(schedule_date, breakpoint_val)
        if week_end is None:
            return None

        # Determine which date mapping to use
        calc_type = str(row.get(date_calc_col, '')).strip().upper()

        if calc_type == 'ETD':
            date_text = str(row.get(etd_col, '')).strip()
        elif calc_type == 'ETA':
            date_text = str(row.get(eta_col, '')).strip()
        else:
            # Default: try ETD first, then ETA
            date_text = str(row.get(etd_col, '')).strip()
            if not date_text:
                date_text = str(row.get(eta_col, '')).strip()

        if not date_text:
            return None

        target_date = self._calculate_target_from_text(week_end, date_text)

        # 防護：目標日期不可早於排程出貨日期
        if target_date is not None and target_date < schedule_date:
            return None

        return target_date

    def _get_week_end_by_breakpoint(self, schedule_date, breakpoint_text):
        """
        根據排程出貨日期和斷點，計算 week end 日期
        斷點 = 每周的截止日 (例如 "週三" 表示本周三)
        """
        weekday_map = {
            '週一': 0, '禮拜一': 0, '星期一': 0,
            '週二': 1, '禮拜二': 1, '星期二': 1,
            '週三': 2, '禮拜三': 2, '星期三': 2,
            '週四': 3, '禮拜四': 3, '星期四': 3,
            '週五': 4, '禮拜五': 4, '星期五': 4,
            '週六': 5, '禮拜六': 5, '星期六': 5,
            '週日': 6, '禮拜日': 6, '星期日': 6, '禮拜天': 6, '週天': 6,
        }

        target_weekday = weekday_map.get(breakpoint_text)
        if target_weekday is None:
            return None

        try:
            current_weekday = schedule_date.weekday()
            days_ahead = (target_weekday - current_weekday) % 7
            return schedule_date + timedelta(days=days_ahead)
        except (TypeError, ValueError):
            return None

    def _calculate_target_from_text(self, week_end, date_text):
        """
        從文字描述計算目標日期
        例如: "本週三" → week_end + 0 weeks + offset to Wednesday
              "下週一" → week_end + 1 week + offset to Monday
              "下下週二" → week_end + 2 weeks + offset to Tuesday
              "下下下週五" → week_end + 3 weeks + offset to Friday
        """
        weekday_map = {
            '一': 0, '二': 1, '三': 2, '四': 3,
            '五': 4, '六': 5, '日': 6, '天': 6,
        }

        # Parse weeks offset
        if date_text.startswith('下下下週') or date_text.startswith('下下下禮拜'):
            weeks_offset = 3
            weekday_char = date_text[-1]
        elif date_text.startswith('下下週') or date_text.startswith('下下禮拜'):
            weeks_offset = 2
            weekday_char = date_text[-1]
        elif date_text.startswith('下週') or date_text.startswith('下禮拜'):
            weeks_offset = 1
            weekday_char = date_text[-1]
        elif date_text.startswith('本週') or date_text.startswith('本禮拜') or date_text.startswith('這週'):
            weeks_offset = 0
            weekday_char = date_text[-1]
        else:
            # Try to parse as a direct date
            parsed = self._parse_date_value(date_text)
            if parsed:
                return parsed
            return None

        target_weekday = weekday_map.get(weekday_char)
        if target_weekday is None:
            return None

        # 以 week_end (斷點日) 為錨點計算
        # 斷點日是一周的最後一天，目標 weekday 在同一周內（斷點日或之前）
        # 例: 斷點=週一(0), 目標=週四(3) → days_diff = (3-0)%7 = 3, 3>0 → 3-7 = -4
        #     表示目標在斷點日的前 4 天
        breakpoint_weekday = week_end.weekday()
        days_diff = (target_weekday - breakpoint_weekday) % 7
        if days_diff > 0:
            days_diff -= 7  # 目標在斷點日之前（同一周內）
        # days_diff == 0 表示目標就是斷點日本身
        return week_end + timedelta(days=7 * weeks_offset + days_diff)

    def _apply_changes(self):
        """Apply all pending changes to the worksheet"""
        print(f"\n[Liteon] 套用 {len(self.pending_changes)} 筆變更...")

        for change in self.pending_changes:
            cell = self.ws.cell(row=change['row'], column=change['col'])
            current = cell.value
            if current and isinstance(current, (int, float)):
                cell.value = current + change['value']
            else:
                cell.value = change['value']

    def _save_file(self):
        """Save the modified forecast workbook"""
        output_path = os.path.join(self.output_folder, self.output_filename)
        self.wb.save(output_path)
        print(f"[Liteon] 已儲存: {self.output_filename}")

    def _save_allocation_status(self):
        """Save allocation status back to ERP and Transit files"""
        try:
            if self.erp_df is not None and len(self.erp_df) > 0:
                self.erp_df.to_excel(self.erp_file, index=False)
                print(f"[Liteon] ERP 已分配狀態已更新")
        except Exception as e:
            print(f"[Liteon] 警告: ERP 已分配狀態更新失敗: {e}")

        try:
            if self.transit_df is not None and len(self.transit_df) > 0 and self.transit_file:
                self.transit_df.to_excel(self.transit_file, index=False)
                print(f"[Liteon] Transit 已分配狀態已更新")
        except Exception as e:
            print(f"[Liteon] 警告: Transit 已分配狀態更新失敗: {e}")
