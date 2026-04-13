"""
Delta Step 4: Forecast Fill (Transit + ERP)

填寫邏輯：
1. Transit 填入 (先做)
   - 匹配: Forecast[B]=Plant ↔ Transit['客戶需求地區']
            Forecast[C]=客戶簡稱 ↔ Transit[K]=客戶簡稱
            Forecast[D]=送貨地點 ↔ Transit[D]=Location
            Forecast[E]=PARTNO ↔ Transit[E]=Ordered Item
   - 位置: Transit[L]=ETA 日期 → 找對應欄位 (週/月)
   - 值: Transit[G]=Qty (原值，不乘 1000)

2. ERP 填入 (後做)
   - 匹配: Forecast[B]=Plant ↔ ERP[客戶需求地區]
            Forecast[C]=客戶簡稱 ↔ ERP[客戶簡稱]
            Forecast[D]=送貨地點 ↔ ERP[送貨地點]
            Forecast[E]=PARTNO ↔ ERP[客戶料號]
   - 排程日期: ERP[排程出貨日期] 依 [排程出貨日期斷點] 計算該週範圍
   - 目標日期: 依 [ETA] 文字 (本週/下週/下下週 + 星期X) 推算填入的週別
   - 值: ERP[淨需求] × 1000

日期欄位匹配規則:
- 優先精確匹配 YYYYMMDD
- 其次依「週一起算 7 天」範圍匹配 (週日期通常是週一)
- 最後退回月份標籤 (JAN-DEC) 以月份數字匹配
"""

import os
import pandas as pd
import openpyxl
from datetime import datetime, timedelta


MONTH_NAMES = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN',
               'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']

WEEKDAY_MAP_BREAKPOINT = {
    '禮拜一': 0, '禮拜二': 1, '禮拜三': 2, '禮拜四': 3,
    '禮拜五': 4, '禮拜六': 5, '禮拜日': 6, '星期日': 6,
}
WEEKDAY_MAP_ETA = {
    '一': 0, '二': 1, '三': 2, '四': 3,
    '五': 4, '六': 5, '日': 6, '天': 6,
}


def normalize_date(val):
    """任意型別轉 datetime。無法解析回 None。"""
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(val, datetime):
        return val
    if isinstance(val, pd.Timestamp):
        return val.to_pydatetime()
    if isinstance(val, str):
        s = val.strip()
        if not s or s.lower() in ('nan', 'none', 'nat', ''):
            return None
        for fmt in ('%Y/%m/%d', '%Y-%m-%d', '%Y%m%d',
                    '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d %H:%M:%S'):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        try:
            return pd.to_datetime(s).to_pydatetime()
        except Exception:
            return None
    try:
        return pd.to_datetime(val).to_pydatetime()
    except Exception:
        return None


def build_date_col_map(ws, start_col=10):
    """
    掃描 forecast worksheet 第 1 列的日期欄位。

    Args:
        ws: openpyxl worksheet
        start_col: 開始掃描的欄位 (1-based, J=10, K=11)

    Returns:
        list of (col_idx, kind, value)
          kind='passdue' → value=None
          kind='weekly'  → value=datetime
          kind='monthly' → value=int(月份 1-12)
    """
    cols = []
    for cell in ws[1]:
        if cell.column < start_col:
            continue
        v = cell.value
        if v is None:
            continue
        s = str(v).strip().upper()
        if s == 'PASSDUE':
            cols.append((cell.column, 'passdue', None))
        elif s in MONTH_NAMES:
            cols.append((cell.column, 'monthly', MONTH_NAMES.index(s) + 1))
        elif s.isdigit() and len(s) == 8:
            try:
                d = datetime.strptime(s, '%Y%m%d')
                cols.append((cell.column, 'weekly', d))
            except ValueError:
                continue
        else:
            d = normalize_date(v)
            if d:
                cols.append((cell.column, 'weekly', d))
    return cols


def find_fill_col(date_col_map, target_date):
    """
    依 target_date 找到要填入的 forecast 欄位索引 (1-based)。

    匹配優先序:
      1) 精確 YYYYMMDD 匹配
      2) 週一 7 天範圍匹配 (只考慮 weekday()==0 的欄位)
      3) 月份匹配 (以月份數字)
    """
    if target_date is None:
        return None

    target_ymd = target_date.strftime('%Y%m%d')
    weekly = [(c, d) for c, k, d in date_col_map if k == 'weekly']
    monthly = [(c, m) for c, k, m in date_col_map if k == 'monthly']

    # 1. 精確匹配
    for col, d in weekly:
        if d.strftime('%Y%m%d') == target_ymd:
            return col

    # 2. 週範圍 (只接受「週一」作為 bucket 起點，避免月末插入日吃掉週範圍)
    for col, d in weekly:
        if d.weekday() == 0:  # Monday
            if d <= target_date < d + timedelta(days=7):
                return col

    # 3. 若週一邏輯沒命中，放寬條件: 任何 weekly col 的 7 天內
    for col, d in weekly:
        if d <= target_date < d + timedelta(days=7):
            return col

    # 4. 月份匹配
    target_month = target_date.month
    for col, m in monthly:
        if m == target_month:
            return col

    return None


def _find_col_by_keywords(df, keywords, exact=False):
    """在 df.columns 中找第一個含關鍵字的欄位。"""
    if isinstance(keywords, str):
        keywords = [keywords]
    for col in df.columns:
        col_str = str(col)
        if exact:
            if col_str in keywords:
                return col
        else:
            for kw in keywords:
                if kw.lower() in col_str.lower():
                    return col
    return None


def fill_transit_into_forecast(ws, date_col_map, transit_file):
    """
    把 Transit 的 Qty 按 ETA 日期填入 forecast。

    Returns:
        (filled_count, skipped_count, matched_rows)
    """
    if not transit_file or not os.path.exists(transit_file):
        print("  ⏭️ 無 Transit 檔案，跳過 Transit 填入")
        return 0, 0, 0

    try:
        transit_df = pd.read_excel(transit_file)
    except Exception as e:
        print(f"  ⚠️ Transit 讀取失敗: {e}")
        return 0, 0, 0

    if len(transit_df) == 0:
        print("  ⏭️ Transit 檔案為空，跳過")
        return 0, 0, 0

    # 尋找 Transit 欄位
    region_col = _find_col_by_keywords(transit_df, ['客戶需求地區', '需求地區'])
    customer_col = _find_col_by_keywords(transit_df, ['客戶簡稱', '簡稱'])
    location_col = _find_col_by_keywords(transit_df, ['Location', '送貨地點'])
    partno_col = _find_col_by_keywords(transit_df, ['Ordered Item', 'ordered item'])
    qty_col = _find_col_by_keywords(transit_df, ['Qty'], exact=True) \
              or _find_col_by_keywords(transit_df, ['qty'])
    eta_col = _find_col_by_keywords(transit_df, ['ETA'], exact=True) \
              or _find_col_by_keywords(transit_df, ['ETA'])

    missing = []
    if not region_col: missing.append('客戶需求地區')
    if not customer_col: missing.append('客戶簡稱')
    if not location_col: missing.append('Location')
    if not partno_col: missing.append('Ordered Item')
    if not qty_col: missing.append('Qty')
    if not eta_col: missing.append('ETA')
    if missing:
        print(f"  ⚠️ Transit 缺少欄位: {missing}")
        return 0, 0, 0

    # 建立 lookup: (plant, customer, location, partno) → [(eta, qty), ...]
    lookup = {}
    for _, row in transit_df.iterrows():
        plant = str(row[region_col]).strip() if pd.notna(row[region_col]) else ''
        customer = str(row[customer_col]).strip() if pd.notna(row[customer_col]) else ''
        location = str(row[location_col]).strip() if pd.notna(row[location_col]) else ''
        partno = str(row[partno_col]).strip() if pd.notna(row[partno_col]) else ''
        eta = normalize_date(row[eta_col])
        qty = row[qty_col] if pd.notna(row[qty_col]) else 0

        if not plant or not partno or eta is None:
            continue

        key = (plant, customer, location, partno)
        lookup.setdefault(key, []).append((eta, qty))

    print(f"  Transit lookup: {len(lookup)} 筆 unique keys, {sum(len(v) for v in lookup.values())} 筆記錄")

    filled = 0
    skipped = 0
    matched_rows = 0

    # 只掃描 Supply 列 (column I == "Supply")
    for r in range(2, ws.max_row + 1):
        row_type = ws.cell(row=r, column=9).value  # I 欄 = row type
        if row_type is None or str(row_type).strip() != 'Supply':
            continue

        plant_v = ws.cell(row=r, column=2).value
        customer_v = ws.cell(row=r, column=3).value
        location_v = ws.cell(row=r, column=4).value
        partno_v = ws.cell(row=r, column=5).value

        if plant_v is None or partno_v is None:
            continue

        key = (
            str(plant_v).strip(),
            str(customer_v).strip() if customer_v is not None else '',
            str(location_v).strip() if location_v is not None else '',
            str(partno_v).strip(),
        )

        if key not in lookup:
            continue

        matched_rows += 1
        for eta, qty in lookup[key]:
            col = find_fill_col(date_col_map, eta)
            if col is None:
                skipped += 1
                continue
            cell = ws.cell(row=r, column=col)
            old = cell.value if isinstance(cell.value, (int, float)) else 0
            try:
                cell.value = (old or 0) + float(qty)
            except (ValueError, TypeError):
                cell.value = qty
            filled += 1

    print(f"  ✅ Transit 填入 Supply: {filled} 筆 (跳過 {skipped}, 匹配 forecast 列數 {matched_rows})")
    return filled, skipped, matched_rows


def get_week_end_day(breakpoint_text):
    """'禮拜四' → 3 (0=Mon)"""
    return WEEKDAY_MAP_BREAKPOINT.get(str(breakpoint_text).strip() if breakpoint_text else '', 3)


def get_eta_weekday(weekday_text):
    """'二' → 1"""
    return WEEKDAY_MAP_ETA.get(str(weekday_text).strip() if weekday_text else '', 1)


def calculate_eta_target_date(schedule_date, breakpoint_text, eta_text):
    """
    依 ERP 的排程出貨日期、排程斷點、ETA 文字，計算目標填入日期。

    與 UltraFastForecastProcessor.calculate_eta_date 相同的演算法:
    - 依斷點找出該排程日所在的「斷點週」的 week_end
    - 從 week_end 往後找到最近的週六, 反推對應的週日
    - 依 ETA 文字 (本週/下週/下下週 + 星期X) 計算最終日期
    """
    if schedule_date is None:
        return None

    date_obj = normalize_date(schedule_date)
    if date_obj is None:
        return None

    # 1. 依斷點計算該週結束日
    week_end_day = get_week_end_day(breakpoint_text)
    days_to_week_end = (week_end_day - date_obj.weekday()) % 7
    week_end_date = date_obj + timedelta(days=days_to_week_end)

    # 2. 找到對應「標準週」(週日~週六) 的週日
    days_to_saturday = (5 - week_end_date.weekday()) % 7
    current_saturday = week_end_date + timedelta(days=days_to_saturday)
    current_sunday = current_saturday - timedelta(days=6)

    eta = str(eta_text).strip() if eta_text else ''
    if not eta:
        return None

    try:
        if '下下週' in eta:
            target_wd = get_eta_weekday(eta.replace('下下週', ''))
            base = current_sunday + timedelta(days=14)
            days_to_target = (target_wd - base.weekday()) % 7
            return base + timedelta(days=days_to_target)
        elif '下週' in eta:
            target_wd = get_eta_weekday(eta.replace('下週', ''))
            base = current_sunday + timedelta(days=7)
            days_to_target = (target_wd - base.weekday()) % 7
            return base + timedelta(days=days_to_target)
        elif '本週' in eta:
            target_wd = get_eta_weekday(eta.replace('本週', ''))
            days_to_target = (target_wd - current_sunday.weekday()) % 7
            return current_sunday + timedelta(days=days_to_target)
    except Exception:
        return None

    return None


def fill_erp_into_forecast(ws, date_col_map, erp_file):
    """
    把 ERP 的淨需求依 (排程斷點 + ETA) 推算的目標日期填入 forecast 的 Supply 列。
    值 = 淨需求 × 1000
    填完後回寫 ERP「已分配」欄位。

    Returns:
        (filled_count, skipped_count, matched_rows)
    """
    if not erp_file or not os.path.exists(erp_file):
        print("  ⚠️ 無 ERP 檔案")
        return 0, 0, 0

    try:
        erp_df = pd.read_excel(erp_file)
    except Exception as e:
        print(f"  ⚠️ ERP 讀取失敗: {e}")
        return 0, 0, 0

    required = ['客戶簡稱', '送貨地點', '客戶料號', '客戶需求地區',
                '淨需求', '排程出貨日期', '排程出貨日期斷點', 'ETA']
    for col in required:
        if col not in erp_df.columns:
            print(f"  ⚠️ ERP 缺少欄位: {col}")
            return 0, 0, 0

    # 確保「已分配」欄位存在且為 string 型別
    if '已分配' not in erp_df.columns:
        erp_df['已分配'] = ''
    erp_df['已分配'] = erp_df['已分配'].astype(str).replace('nan', '')

    # 建立 lookup: (plant, customer, location, partno) → [(erp_df_idx, erp_row_dict), ...]
    lookup = {}
    for idx, row in erp_df.iterrows():
        plant = str(row['客戶需求地區']).strip() if pd.notna(row['客戶需求地區']) else ''
        customer = str(row['客戶簡稱']).strip() if pd.notna(row['客戶簡稱']) else ''
        location = str(row['送貨地點']).strip() if pd.notna(row['送貨地點']) else ''
        partno = str(row['客戶料號']).strip() if pd.notna(row['客戶料號']) else ''

        if not plant or not partno:
            continue

        # 未在映射階段配對到地區的跳過
        if plant.lower() == 'nan':
            continue

        key = (plant, customer, location, partno)
        lookup.setdefault(key, []).append((idx, {
            '淨需求': row['淨需求'],
            '排程出貨日期': row['排程出貨日期'],
            '排程出貨日期斷點': row['排程出貨日期斷點'],
            'ETA': row['ETA'],
        }))

    print(f"  ERP lookup: {len(lookup)} 筆 unique keys, {sum(len(v) for v in lookup.values())} 筆記錄")

    filled = 0
    skipped = 0
    matched_rows = 0

    # 只掃描 Supply 列 (column I == "Supply")
    for r in range(2, ws.max_row + 1):
        row_type = ws.cell(row=r, column=9).value  # I 欄 = row type
        if row_type is None or str(row_type).strip() != 'Supply':
            continue

        plant_v = ws.cell(row=r, column=2).value
        customer_v = ws.cell(row=r, column=3).value
        location_v = ws.cell(row=r, column=4).value
        partno_v = ws.cell(row=r, column=5).value

        if plant_v is None or partno_v is None:
            continue

        key = (
            str(plant_v).strip(),
            str(customer_v).strip() if customer_v is not None else '',
            str(location_v).strip() if location_v is not None else '',
            str(partno_v).strip(),
        )

        if key not in lookup:
            continue

        matched_rows += 1
        for erp_idx, erp_row in lookup[key]:
            target = calculate_eta_target_date(
                erp_row['排程出貨日期'],
                erp_row['排程出貨日期斷點'],
                erp_row['ETA'],
            )
            if target is None:
                skipped += 1
                continue
            col = find_fill_col(date_col_map, target)
            if col is None:
                skipped += 1
                continue

            net = erp_row['淨需求']
            if net is None or pd.isna(net):
                skipped += 1
                continue

            try:
                value = float(net) * 1000
            except (TypeError, ValueError):
                skipped += 1
                continue

            cell = ws.cell(row=r, column=col)
            old = cell.value if isinstance(cell.value, (int, float)) else 0
            cell.value = (old or 0) + value
            filled += 1

            # 標記 ERP「已分配」
            erp_df.at[erp_idx, '已分配'] = 'Y'

    # 回寫 ERP 檔案 (更新已分配欄位)
    erp_df.to_excel(erp_file, index=False)
    allocated_count = (erp_df['已分配'] == 'Y').sum()
    print(f"  ✅ ERP 填入 Supply: {filled} 筆 (跳過 {skipped}, 匹配 forecast 列數 {matched_rows})")
    print(f"  ✅ ERP 已分配標記: {allocated_count} 筆")
    return filled, skipped, matched_rows


def process_delta_forecast(forecast_file, erp_file, transit_file, output_file):
    """
    Delta 第 4 步驟主入口: 在 forecast 上依序填入 Transit + ERP 資料。

    Args:
        forecast_file: 經過映射的整合 forecast 檔 (integrated_forecast.xlsx)
        erp_file: 整合後的 ERP (integrated_erp.xlsx)
        transit_file: 整合後的 Transit (integrated_transit.xlsx 或 None)
        output_file: 最終輸出 (forecast_result.xlsx)

    Returns:
        dict of stats
    """
    print("=== Delta Forecast 第 4 步驟：Transit + ERP 填入 ===")
    print(f"  Forecast: {forecast_file}")
    print(f"  ERP: {erp_file}")
    print(f"  Transit: {transit_file or '(無)'}")
    print(f"  Output: {output_file}")

    wb = openpyxl.load_workbook(forecast_file)
    ws = wb.active

    # 掃描日期欄位 (從 J=10 開始, 包含 PASSDUE)
    date_col_map = build_date_col_map(ws, start_col=10)
    weekly_count = sum(1 for _, k, _ in date_col_map if k == 'weekly')
    monthly_count = sum(1 for _, k, _ in date_col_map if k == 'monthly')
    print(f"  日期欄位: {len(date_col_map)} (週: {weekly_count}, 月: {monthly_count})")

    # 1. Transit 填入
    print("\n--- Transit 填入 ---")
    t_filled, t_skipped, t_matched = fill_transit_into_forecast(ws, date_col_map, transit_file)

    # 2. ERP 填入
    print("\n--- ERP 填入 ---")
    e_filled, e_skipped, e_matched = fill_erp_into_forecast(ws, date_col_map, erp_file)

    # 儲存
    wb.save(output_file)
    wb.close()

    stats = {
        'transit_filled': t_filled,
        'transit_skipped': t_skipped,
        'transit_matched_rows': t_matched,
        'erp_filled': e_filled,
        'erp_skipped': e_skipped,
        'erp_matched_rows': e_matched,
    }
    print(f"\n=== 處理完成: {stats} ===")
    return stats
