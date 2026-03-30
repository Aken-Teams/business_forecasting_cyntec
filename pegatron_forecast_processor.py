# -*- coding: utf-8 -*-
"""
Pegatron Forecast Processor
專門處理 Pegatron 的 Transit + ERP -> Forecast 邏輯
使用 LibreOffice 跨平台方案保留 .xls/.xlsx 格式和公式
支援 Windows 和 Linux 環境
"""
import pandas as pd
import os
import shutil
import tempfile
from datetime import datetime, timedelta


class PegatronForecastProcessor:
    """
    Pegatron 專用 Forecast 處理器
    - Transit 和 ERP 都填入 ETA QTY 行
    - 使用 LibreOffice 跨平台方案保留格式和公式
    - 支援累加邏輯
    - 追蹤已分配狀態，避免重複分配
    """

    def __init__(self, forecast_file, erp_file, transit_file=None, output_folder=None, output_filename=None, is_merged=False):
        self.forecast_file = forecast_file
        self.erp_file = erp_file
        self.transit_file = transit_file
        self.output_folder = output_folder
        self.is_merged = is_merged  # 合併檔案才需要 A 欄客戶料號
        # 確保輸出為 xlsx 格式（避免轉換失敗）
        if output_filename:
            self.output_filename = os.path.splitext(output_filename)[0] + '.xlsx'
        else:
            self.output_filename = "forecast_result.xlsx"

        # 統計資料
        self.total_filled = 0
        self.total_skipped = 0
        self.total_transit_filled = 0
        self.total_transit_skipped = 0

        # 追蹤已分配的索引
        self.allocated_erp_indices = []
        self.allocated_transit_indices = []

        # 保存 DataFrame 以便更新已分配狀態
        self.erp_df = None
        self.transit_df = None

    def get_week_end_by_breakpoint(self, schedule_date, breakpoint_text):
        """
        根據排程出貨日期和斷點計算該斷點週的結束日
        斷點就是週的結束日
        """
        breakpoint_map = {
            '週一': 0, '週二': 1, '週三': 2, '週四': 3,
            '週五': 4, '週六': 5, '週日': 6,
            '禮拜一': 0, '禮拜二': 1, '禮拜三': 2, '禮拜四': 3,
            '禮拜五': 4, '禮拜六': 5, '禮拜日': 6,
        }

        breakpoint_weekday = breakpoint_map.get(breakpoint_text, 2)  # 預設週三
        current_weekday = schedule_date.weekday()

        if current_weekday <= breakpoint_weekday:
            days_to_breakpoint = breakpoint_weekday - current_weekday
        else:
            days_to_breakpoint = 7 - (current_weekday - breakpoint_weekday)

        week_end = schedule_date + timedelta(days=days_to_breakpoint)
        return week_end

    def calculate_erp_eta_target_date(self, week_end, eta_text):
        """
        根據 ETA 文字和斷點週結束日計算目標日期
        ETA 格式: 本週X, 下週X, 下下週X
        """
        eta_weekday_map = {
            '一': 0, '二': 1, '三': 2, '四': 3,
            '五': 4, '六': 5, '日': 6, '天': 6,
        }

        if not eta_text or pd.isna(eta_text):
            return None

        eta_text = str(eta_text).strip()

        if '下下週' in eta_text or '下下周' in eta_text:
            weeks_offset = 2
            weekday_char = eta_text.replace('下下週', '').replace('下下周', '').strip()
        elif '下週' in eta_text or '下周' in eta_text:
            weeks_offset = 1
            weekday_char = eta_text.replace('下週', '').replace('下周', '').strip()
        elif '本週' in eta_text or '本周' in eta_text:
            weeks_offset = 0
            weekday_char = eta_text.replace('本週', '').replace('本周', '').strip()
        else:
            return None

        target_weekday = eta_weekday_map.get(weekday_char, 1)  # 預設週二
        week_end_weekday = week_end.weekday()

        days_diff = target_weekday - week_end_weekday
        target_date = week_end + timedelta(days=7 * weeks_offset + days_diff)

        return target_date

    def find_week_column(self, target_date, date_columns):
        """
        找到目標日期所在週的欄位
        Forecast 日期已經是週一
        """
        if isinstance(target_date, datetime):
            target_date = target_date.date()

        days_since_monday = target_date.weekday()
        week_monday = target_date - timedelta(days=days_since_monday)

        if week_monday in date_columns:
            return date_columns[week_monday]

        for week_start, col_idx in date_columns.items():
            week_end = week_start + timedelta(days=6)
            if week_start <= target_date <= week_end:
                return col_idx

        return None

    def process_all_blocks(self):
        """主處理函數"""
        try:
            print("=" * 70)
            print("Pegatron Forecast Processor - Transit + ERP -> ETA QTY")
            print("=" * 70)

            # 1. 讀取資料
            # 根據檔案格式選擇引擎
            forecast_ext = os.path.splitext(self.forecast_file)[1].lower()
            if forecast_ext == '.xls':
                forecast_df = pd.read_excel(self.forecast_file, header=None, engine='xlrd')
            else:
                forecast_df = pd.read_excel(self.forecast_file, header=None)
            print(f"Forecast 行數: {len(forecast_df)}, 欄數: {len(forecast_df.columns)}")

            self.erp_df = pd.read_excel(self.erp_file)
            print(f"ERP 行數: {len(self.erp_df)}")

            # 確保 ERP 有「已分配」欄位
            if '已分配' not in self.erp_df.columns:
                self.erp_df['已分配'] = ''

            self.transit_df = None
            if self.transit_file and os.path.exists(self.transit_file):
                self.transit_df = pd.read_excel(self.transit_file)
                print(f"Transit 行數: {len(self.transit_df)}")
                # 確保 Transit 有「已分配」欄位
                if '已分配' not in self.transit_df.columns:
                    self.transit_df['已分配'] = ''

            # 2. 建立 Forecast 區塊結構
            # F+G 欄位 = Line 客戶採購單號, I 欄位 (row+1) = Ordered Item
            forecast_blocks = []
            row_idx = 2
            while row_idx < len(forecast_df):
                m_val = forecast_df.iloc[row_idx, 12] if pd.notna(forecast_df.iloc[row_idx, 12]) else ''
                if m_val == 'WEEK#':
                    f_val = str(forecast_df.iloc[row_idx, 5]).strip() if pd.notna(forecast_df.iloc[row_idx, 5]) else ''
                    g_val = str(forecast_df.iloc[row_idx, 6]).strip() if pd.notna(forecast_df.iloc[row_idx, 6]) else ''
                    line_po = f"{f_val}-{g_val}" if f_val and g_val else ''

                    ordered_item = ''
                    if row_idx + 1 < len(forecast_df):
                        ordered_item = str(forecast_df.iloc[row_idx + 1, 8]).strip() if pd.notna(forecast_df.iloc[row_idx + 1, 8]) else ''

                    eta_qty_row = row_idx + 4  # ETA QTY 行

                    forecast_blocks.append({
                        'start_row': row_idx,
                        'line_po': line_po,
                        'ordered_item': ordered_item,
                        'eta_qty_row': eta_qty_row + 1,  # Excel 1-based
                    })
                    row_idx += 8
                else:
                    row_idx += 1

            print(f"\n找到 {len(forecast_blocks)} 個 Forecast 區塊")

            # 3. 取得日期欄位對應
            date_columns = {}
            for col_idx in range(14, len(forecast_df.columns)):
                date_val = forecast_df.iloc[1, col_idx]
                if pd.notna(date_val):
                    if isinstance(date_val, str):
                        try:
                            date_obj = pd.to_datetime(date_val)
                            date_columns[date_obj.date()] = col_idx + 1
                        except:
                            pass
                    elif isinstance(date_val, (datetime, pd.Timestamp)):
                        date_columns[date_val.date()] = col_idx + 1

            print(f"日期欄位對應: {len(date_columns)} 個日期")

            all_updates = []

            # 4. 處理 Transit 資料
            if self.transit_df is not None:
                print("\n=== 處理 Transit 資料 ===")
                transit_updates = self._process_transit(self.transit_df, forecast_blocks, date_columns)
                all_updates.extend(transit_updates)
                self.total_transit_filled = len(transit_updates)
                print(f"Transit 更新筆數: {self.total_transit_filled}")

            # 5. 處理 ERP 資料
            print("\n=== 處理 ERP 資料 ===")
            erp_updates = self._process_erp(self.erp_df, forecast_blocks, date_columns)
            all_updates.extend(erp_updates)
            self.total_filled = len(erp_updates)
            print(f"ERP 更新筆數: {self.total_filled}")

            if not all_updates:
                print("\n沒有需要更新的資料，複製原始檔案作為輸出")
                if self.output_folder:
                    output_path = os.path.join(self.output_folder, self.output_filename)
                else:
                    output_path = self.output_filename
                shutil.copy2(self.forecast_file, output_path)
                if self.is_merged:
                    self._fill_part_number_to_column_a(output_path)
                print(f"已輸出到: {output_path}")
                return True

            # 6. 使用 COM 寫入 Excel
            print(f"\n=== 使用 COM 更新 {len(all_updates)} 個儲存格 ===")
            success = self._write_to_excel(all_updates)

            if success:
                # 7. 更新 Commit 欄位 (基於 Balance1/Balance2)
                if self.output_folder:
                    output_path = os.path.join(self.output_folder, self.output_filename)
                else:
                    output_path = self.output_filename
                self._update_commit_column(output_path)

                # 8. 寫入客戶料號到 A 欄 (僅合併檔案，方便篩選)
                if self.is_merged:
                    self._fill_part_number_to_column_a(output_path)

                # 9. 更新並保存已分配狀態
                self._save_allocation_status()

                print("\n" + "=" * 50)
                print("處理完成")
                print("=" * 50)
                print(f"Transit 更新: {self.total_transit_filled} 筆")
                print(f"ERP 更新: {self.total_filled} 筆")

            return success

        except Exception as e:
            print(f"處理失敗: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _process_transit(self, transit_df, forecast_blocks, date_columns):
        """處理 Transit 資料"""
        updates = []

        # 檢查必要欄位
        if 'Line 客戶採購單號' not in transit_df.columns:
            print("Transit 缺少 'Line 客戶採購單號' 欄位")
            return updates

        transit_with_line = transit_df[transit_df['Line 客戶採購單號'].notna()]
        print(f"有 Line 客戶採購單號 的在途記錄: {len(transit_with_line)}")

        for idx, transit_row in transit_with_line.iterrows():
            # 跳過已分配的記錄
            if transit_df.at[idx, '已分配'] == '✓':
                continue

            transit_line_po = str(transit_row['Line 客戶採購單號']).strip()
            transit_ordered_item = str(transit_row['Ordered Item']).strip() if 'Ordered Item' in transit_row and pd.notna(transit_row['Ordered Item']) else ''
            transit_qty = transit_row['Qty'] if 'Qty' in transit_row and pd.notna(transit_row['Qty']) else 0
            transit_eta = transit_row['ETA'] if 'ETA' in transit_row else None

            if not transit_line_po or not transit_ordered_item or transit_qty == 0:
                continue

            # 找到匹配的 Forecast 區塊
            matched_block = None
            for block in forecast_blocks:
                if block['line_po'] == transit_line_po and block['ordered_item'] == transit_ordered_item:
                    matched_block = block
                    break

            if not matched_block:
                print(f"  ⚠️ Transit 跳過: {transit_line_po}, {transit_ordered_item} - 找不到匹配的 Forecast 區塊")
                self.total_transit_skipped += 1
                continue

            if not pd.notna(transit_eta):
                print(f"  ⚠️ Transit 跳過: {transit_line_po}, {transit_ordered_item} - ETA 為空")
                self.total_transit_skipped += 1
                continue

            eta_date = pd.to_datetime(transit_eta).date()
            days_since_monday = eta_date.weekday()
            week_start = eta_date - timedelta(days=days_since_monday)

            if week_start not in date_columns:
                print(f"  ⚠️ Transit 跳過: {transit_line_po}, {transit_ordered_item} - 找不到 ETA 日期 {eta_date} 對應的欄位")
                self.total_transit_skipped += 1
                continue

            excel_col = date_columns[week_start]
            eta_qty_value = transit_qty * 1000

            print(f"  Transit: {transit_line_po}, {transit_ordered_item} -> Row {matched_block['eta_qty_row']}, Col {excel_col}, 值={eta_qty_value}")
            updates.append((matched_block['eta_qty_row'], excel_col, eta_qty_value))

            # 記錄已分配的索引
            self.allocated_transit_indices.append(idx)

        return updates

    def _process_erp(self, erp_df, forecast_blocks, date_columns):
        """處理 ERP 資料"""
        updates = []

        # 檢查必要欄位
        required_cols = ['客戶需求地區', 'Line 客戶採購單號', '客戶料號', '淨需求', '排程出貨日期', '排程出貨日期斷點', 'ETA']
        missing_cols = [c for c in required_cols if c not in erp_df.columns]
        if missing_cols:
            print(f"ERP 缺少欄位: {missing_cols}")
            return updates

        erp_with_mapping = erp_df[erp_df['客戶需求地區'].notna() & (erp_df['客戶需求地區'] != '')]
        print(f"有 mapping 的 ERP 記錄: {len(erp_with_mapping)}")

        for idx, erp_row in erp_with_mapping.iterrows():
            # 跳過已分配的記錄
            if erp_df.at[idx, '已分配'] == '✓':
                continue

            erp_line_po = str(erp_row['Line 客戶採購單號']).strip() if pd.notna(erp_row['Line 客戶採購單號']) else ''
            erp_pn = str(erp_row['客戶料號']).strip() if pd.notna(erp_row['客戶料號']) else ''
            erp_qty = erp_row['淨需求'] if pd.notna(erp_row['淨需求']) else 0
            erp_schedule_date = erp_row['排程出貨日期']
            erp_breakpoint = str(erp_row['排程出貨日期斷點']).strip() if pd.notna(erp_row['排程出貨日期斷點']) else ''
            erp_eta = str(erp_row['ETA']).strip() if pd.notna(erp_row['ETA']) else ''

            if not erp_line_po or not erp_pn or erp_qty == 0:
                continue

            # 找到匹配的 Forecast 區塊
            matched_block = None
            for block in forecast_blocks:
                if block['line_po'] == erp_line_po and block['ordered_item'] == erp_pn:
                    matched_block = block
                    break

            if not matched_block:
                print(f"  ⚠️ ERP 跳過: {erp_line_po}, {erp_pn} - 找不到匹配的 Forecast 區塊")
                self.total_skipped += 1
                continue

            if pd.isna(erp_schedule_date) or not erp_breakpoint or not erp_eta:
                print(f"  ⚠️ ERP 跳過: {erp_line_po}, {erp_pn} - 缺少排程日期/斷點/ETA")
                self.total_skipped += 1
                continue

            schedule_date = pd.to_datetime(erp_schedule_date)
            week_end = self.get_week_end_by_breakpoint(schedule_date, erp_breakpoint)
            target_date = self.calculate_erp_eta_target_date(week_end, erp_eta)

            if target_date is None:
                print(f"  ⚠️ ERP 跳過: {erp_line_po}, {erp_pn} - 無法計算目標日期 (ETA={erp_eta})")
                self.total_skipped += 1
                continue

            excel_col = self.find_week_column(target_date, date_columns)

            if excel_col is None:
                print(f"  ⚠️ ERP 跳過: {erp_line_po}, {erp_pn} - 找不到目標日期 {target_date.date()} 對應的欄位")
                self.total_skipped += 1
                continue

            forecast_value = erp_qty * 1000

            print(f"  ERP: {erp_line_po}, {erp_pn} -> Row {matched_block['eta_qty_row']}, Col {excel_col}, 值={forecast_value}")
            updates.append((matched_block['eta_qty_row'], excel_col, forecast_value))

            # 記錄已分配的索引
            self.allocated_erp_indices.append(idx)

        return updates

    def _write_to_excel(self, updates):
        """使用 LibreOffice 跨平台方案寫入 Excel，保留格式和公式"""
        from libreoffice_utils import write_to_excel_libreoffice

        # 決定輸出路徑
        if self.output_folder:
            output_path = os.path.join(self.output_folder, self.output_filename)
        else:
            output_path = self.output_filename

        try:
            success = write_to_excel_libreoffice(self.forecast_file, updates, output_path)
            if success:
                print(f"\n已輸出到: {output_path}")
            return success
        except Exception as e:
            print(f"寫入 Excel 失敗: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _update_commit_column(self, output_path):
        """
        公式重算後讀取 Balance1/Balance2 的 R~W (cols 18-23) 值，
        若 Balance1 或 Balance2 其中一列 R~W 全部 >= 0 → Commit = "Y"，否則 "N"。
        寫入 Column L (12) 合併儲存格。
        """
        from libreoffice_utils import recalculate_xlsx
        from openpyxl import load_workbook

        print("\n=== 更新 Commit 欄位 ===")

        temp_dir = tempfile.mkdtemp()
        try:
            # Step 1: 用 LibreOffice 重算公式
            recalc_path = os.path.join(temp_dir, 'recalculated.xlsx')
            if not recalculate_xlsx(output_path, recalc_path):
                print("  ⚠️ 公式重算失敗，跳過 Commit 更新")
                return False

            # Step 2: 讀取重算後的 Balance 值 (data_only=True 取得快取數值)
            wb_read = load_workbook(recalc_path, data_only=True)
            ws_read = wb_read.active

            # Step 3: 找群組 (M欄="WEEK#") 及其 Balance1/Balance2 行
            def check_all_non_negative(row_num):
                """檢查 R~W (cols 18-23) 是否全部 >= 0"""
                if row_num is None:
                    return False
                for col in range(18, 24):  # R=18, S=19, T=20, U=21, V=22, W=23
                    val = ws_read.cell(row=row_num, column=col).value
                    if val is None or not isinstance(val, (int, float)) or val < 0:
                        return False
                return True

            commit_values = {}  # {start_row: "Y" or "N"}
            row_idx = 1
            while row_idx <= ws_read.max_row:
                m_val = ws_read.cell(row=row_idx, column=13).value
                if m_val and str(m_val).strip() == "WEEK#":
                    # 固定位置: offset+5 = Balance 第一列, offset+6 = Balance 第二列
                    # 不依賴名稱（各廠區命名不同: Balance1/Balance(VMI)/...）
                    balance1_row = row_idx + 5 if row_idx + 5 <= ws_read.max_row else None
                    balance2_row = row_idx + 6 if row_idx + 6 <= ws_read.max_row else None

                    b1_ok = check_all_non_negative(balance1_row)
                    b2_ok = check_all_non_negative(balance2_row)
                    commit = "Y" if (b1_ok or b2_ok) else "N"
                    commit_values[row_idx] = commit

                    row_idx += 8
                else:
                    row_idx += 1

            wb_read.close()

            if not commit_values:
                print("  沒有找到需要更新的群組")
                return True

            # Step 4: 寫入 Commit 到原輸出檔的 Column L (12)
            wb_write = load_workbook(output_path)
            ws_write = wb_write.active

            y_count = 0
            n_count = 0
            for start_row, commit in commit_values.items():
                ws_write.cell(row=start_row, column=12).value = commit
                if commit == "Y":
                    y_count += 1
                else:
                    n_count += 1

            wb_write.save(output_path)
            wb_write.close()

            print(f"  ✅ Commit 更新完成: {len(commit_values)} 個群組 (Y={y_count}, N={n_count})")
            return True

        except Exception as e:
            print(f"  ❌ Commit 更新失敗: {e}")
            import traceback
            traceback.print_exc()
            return False
        finally:
            try:
                shutil.rmtree(temp_dir)
            except:
                pass

    def _fill_part_number_to_column_a(self, output_path):
        """
        用 Excel COM 在最左邊插入新的 A 欄 (PN Model)，
        將每個群組的客戶料號填入每一列（不合併儲存格）。
        Excel COM 會正確處理公式參照和合併儲存格的位移。
        """
        import platform

        print("\n=== 插入客戶料號欄位 (A 欄) ===")

        abs_path = os.path.abspath(output_path)

        if platform.system() == 'Windows':
            try:
                return self._fill_part_number_with_excel_com(abs_path)
            except ImportError:
                print("  win32com 不可用，改用 openpyxl fallback...")
            except Exception as e:
                print(f"  Excel COM 失敗: {e}，改用 openpyxl fallback...")

        return self._fill_part_number_with_openpyxl(abs_path)

    def _fill_part_number_with_excel_com(self, abs_path):
        """用 Excel COM 插入 A 欄並填入客戶料號"""
        import win32com.client
        import pythoncom

        pythoncom.CoInitialize()
        excel = None
        wb = None
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            wb = excel.Workbooks.Open(abs_path)
            ws = wb.ActiveSheet

            # 1. 插入新的 A 欄（Excel 自動處理公式位移和合併儲存格）
            ws.Columns("A:A").Insert()

            # 2. 寫入標題
            ws.Cells(2, 1).Value = "PN Model"

            # 3. 掃描群組，填入客戶料號到每一列
            #    insert 後原 M 欄 (13) → N 欄 (14)，原 I 欄 (9) → J 欄 (10)
            count = 0
            row_idx = 1
            max_row = ws.UsedRange.Rows.Count + ws.UsedRange.Row - 1
            while row_idx <= max_row:
                m_val = ws.Cells(row_idx, 14).Value  # 原 M 欄 → N 欄
                if m_val and str(m_val).strip() == "WEEK#":
                    part_number = ws.Cells(row_idx + 1, 10).Value  # 原 I 欄 → J 欄
                    if part_number:
                        part_number = str(part_number).strip()
                        for r in range(row_idx, min(row_idx + 8, max_row + 1)):
                            ws.Cells(r, 1).Value = part_number
                            # 確保 A 欄該 cell 不是合併儲存格
                            if ws.Cells(r, 1).MergeCells:
                                ws.Cells(r, 1).UnMerge()
                                ws.Cells(r, 1).Value = part_number
                        count += 1
                    row_idx += 8
                else:
                    row_idx += 1

            wb.Save()
            wb.Close(SaveChanges=False)
            wb = None

            print(f"  ✅ 客戶料號欄位插入完成: {count} 個群組 (Excel COM)")
            return True

        except Exception as e:
            if wb:
                try:
                    wb.Close(SaveChanges=False)
                except:
                    pass
            raise
        finally:
            if excel:
                excel.Quit()
            pythoncom.CoUninitialize()

    def _fill_part_number_with_openpyxl(self, abs_path):
        """
        Fallback: 用 openpyxl 插入新 A 欄並寫入客戶料號。
        insert_cols 不會更新公式參照，所以用 Translator 手動修正。
        適用於 Linux / 無 Excel 環境。
        """
        from openpyxl import load_workbook
        from openpyxl.formula.translate import Translator
        from openpyxl.utils import get_column_letter

        try:
            wb = load_workbook(abs_path)
            ws = wb.active

            # 1. 先收集料號資料（insert 前的欄位位置）
            pn_data = {}
            row_idx = 1
            while row_idx <= ws.max_row:
                m_val = ws.cell(row=row_idx, column=13).value
                if m_val and str(m_val).strip() == "WEEK#":
                    pn = ws.cell(row=row_idx + 1, column=9).value
                    if pn:
                        pn_data[row_idx] = str(pn).strip()
                    row_idx += 8
                else:
                    row_idx += 1

            # 2. 插入新 A 欄（物理位移，但公式文字不會更新）
            ws.insert_cols(1)

            # 3. 解除 A 欄的合併儲存格
            #    insert_cols 可能將原本的合併範圍擴展到包含新的 A 欄，
            #    需要把 A 欄從合併中移除，確保每列獨立可篩選。
            merges_to_fix = []
            for merge_range in list(ws.merged_cells.ranges):
                if merge_range.min_col == 1:
                    merges_to_fix.append(str(merge_range))

            for merge_str in merges_to_fix:
                ws.unmerge_cells(merge_str)
                # 解析範圍，重新合併 B 欄起（排除 A 欄）
                from openpyxl.utils.cell import range_boundaries
                min_c, min_r, max_c, max_r = range_boundaries(merge_str)
                if max_c >= 2:
                    new_range = (f"{get_column_letter(2)}{min_r}:"
                                 f"{get_column_letter(max_c)}{max_r}")
                    ws.merge_cells(new_range)

            # 4. 修正所有公式參照（col +1 位移）
            for row in ws.iter_rows(min_col=2, max_col=ws.max_column,
                                     min_row=1, max_row=ws.max_row):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        orig_col = cell.column - 1
                        origin = f"{get_column_letter(orig_col)}{cell.row}"
                        dest = f"{get_column_letter(cell.column)}{cell.row}"
                        try:
                            cell.value = Translator(
                                cell.value, origin=origin
                            ).translate_formula(dest)
                        except Exception:
                            pass  # 無法翻譯的公式保持原樣

            # 5. 寫入標題和客戶料號
            ws.cell(row=2, column=1, value="PN Model")

            count = 0
            for row_idx, pn in pn_data.items():
                for r in range(row_idx, min(row_idx + 8, ws.max_row + 1)):
                    ws.cell(row=r, column=1, value=pn)
                count += 1

            wb.save(abs_path)
            wb.close()

            print(f"  ✅ 客戶料號欄位插入完成: {count} 個群組 (openpyxl + Translator)")
            return True

        except Exception as e:
            print(f"  ❌ 客戶料號欄位插入失敗: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _save_allocation_status(self):
        """保存已分配狀態到 ERP 和 Transit 檔案"""
        try:
            # 更新 ERP 已分配狀態
            if self.erp_df is not None and self.allocated_erp_indices:
                for idx in self.allocated_erp_indices:
                    self.erp_df.at[idx, '已分配'] = '✓'
                # 保存 ERP 檔案
                self.erp_df.to_excel(self.erp_file, index=False)
                print(f"已更新 ERP 已分配狀態: {len(self.allocated_erp_indices)} 筆")

            # 更新 Transit 已分配狀態
            if self.transit_df is not None and self.allocated_transit_indices:
                for idx in self.allocated_transit_indices:
                    self.transit_df.at[idx, '已分配'] = '✓'
                # 保存 Transit 檔案
                self.transit_df.to_excel(self.transit_file, index=False)
                print(f"已更新 Transit 已分配狀態: {len(self.allocated_transit_indices)} 筆")

        except Exception as e:
            print(f"保存已分配狀態失敗: {e}")
            import traceback
            traceback.print_exc()
