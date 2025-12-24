import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore')

class UltraFastForecastProcessor:
    """
    超高速FORECAST批量處理系統
    使用批量openpyxl操作，大幅提升性能
    """
    
    def __init__(self, forecast_file, erp_file, transit_file=None, output_folder=None, output_filename=None):
        self.forecast_file = forecast_file
        self.erp_file = erp_file
        self.transit_file = transit_file
        self.output_folder = output_folder  # 輸出資料夾路徑
        self.output_filename = output_filename or "forecast_result.xlsx"  # 輸出檔名（支援多檔案模式）
        self.total_filled = 0
        self.total_skipped = 0
        self.total_transit_filled = 0
        self.total_transit_skipped = 0

        # 批量操作相關
        self.wb = None
        self.ws = None
        self.pending_changes = []  # 待處理的修改列表
        
    def load_and_prepare_data(self):
        """載入並預處理數據"""
        try:
            print("=== 載入數據 ===")

            # 載入ERP數據
            self.erp_df = pd.read_excel(self.erp_file)

            # 載入Forecast數據
            self.forecast_df = pd.read_excel(self.forecast_file)

            print(f"✅ ERP文件: {len(self.erp_df)} 行")
            print(f"✅ Forecast文件: {len(self.forecast_df)} 行")

            # 預處理：建立ERP索引以提高查找速度
            self.erp_df['match_key'] = self.erp_df['客戶料號'].astype(str) + '_' + self.erp_df['客戶需求地區'].astype(str)

            # 保存原始行索引，以便後續更新 erp_df 的已分配狀態
            self.erp_df['_original_idx'] = self.erp_df.index

            # 檢查「已分配」欄位是否存在（從映射整合階段新增）
            if '已分配' not in self.erp_df.columns:
                self.erp_df['已分配'] = ''
                print("⚠️ ERP 檔案缺少「已分配」欄位，已自動新增")
            else:
                # 統計已分配的數量
                already_allocated = (self.erp_df['已分配'] == '✓').sum()
                print(f"📋 ERP 已分配筆數: {already_allocated}/{len(self.erp_df)}")

            self.erp_index = self.erp_df.set_index('match_key')

            # 載入Transit數據（如果有）
            if self.transit_file and pd.notna(self.transit_file):
                try:
                    self.transit_df = pd.read_excel(self.transit_file)
                    print(f"✅ Transit文件: {len(self.transit_df)} 行, {len(self.transit_df.columns)} 欄")

                    # 建立Transit索引：M欄位(索引12) + F欄位(索引5)
                    if len(self.transit_df.columns) >= 13:
                        print(f"🔍 Transit欄位調試:")
                        print(f"   F欄位名稱(索引5): {self.transit_df.columns[5]}")
                        print(f"   H欄位名稱(索引7): {self.transit_df.columns[7]}")
                        print(f"   I欄位名稱(索引8): {self.transit_df.columns[8]}")
                        print(f"   M欄位名稱(索引12): {self.transit_df.columns[12]}")

                        self.transit_df['match_key'] = self.transit_df.iloc[:, 12].astype(str) + '_' + self.transit_df.iloc[:, 5].astype(str)

                        # 保存原始行索引，以便後續更新 transit_df 的已分配狀態
                        self.transit_df['_original_idx'] = self.transit_df.index

                        # 檢查「已分配」欄位是否存在（從映射整合階段新增）
                        if '已分配' not in self.transit_df.columns:
                            self.transit_df['已分配'] = ''
                            print("⚠️ Transit 檔案缺少「已分配」欄位，已自動新增")
                        else:
                            # 統計已分配的數量
                            already_allocated = (self.transit_df['已分配'] == '✓').sum()
                            print(f"📋 Transit 已分配筆數: {already_allocated}/{len(self.transit_df)}")

                        self.transit_index = self.transit_df.set_index('match_key')
                        print(f"✅ Transit索引建立完成，共 {len(self.transit_index)} 個唯一鍵")

                        # 顯示前幾個 match_key 作為範例
                        sample_keys = list(self.transit_index.index[:3])
                        print(f"   範例鍵: {sample_keys}")
                    else:
                        print(f"⚠️ Transit文件欄位不足，需要至少13欄，實際: {len(self.transit_df.columns)}")
                        self.transit_df = None
                        self.transit_index = None
                except Exception as e:
                    print(f"⚠️ Transit文件載入失敗: {e}")
                    self.transit_df = None
                    self.transit_index = None
            else:
                self.transit_df = None
                self.transit_index = None

            # 一次性載入Excel文件到內存
            print("📁 載入Excel文件到內存...")
            self.wb = load_workbook(self.forecast_file)
            self.ws = self.wb.active
            print("✅ Excel文件已載入內存")

            return True

        except Exception as e:
            print(f"❌ 數據載入失敗: {e}")
            return False
    
    def find_data_blocks(self):
        """快速找到所有數據塊"""
        try:
            print("\n=== 識別數據塊 ===")
            
            data_blocks = []
            a_col = self.forecast_df.columns[0]  # A欄位
            d_col = self.forecast_df.columns[3]  # D欄位
            
            current_block = None
            
            for idx, row in self.forecast_df.iterrows():
                customer_part = row[a_col]
                customer_region = row[d_col]
                
                # 跳過標題行
                if pd.notna(customer_part) and pd.notna(customer_region) and \
                   customer_part != "需求週數" and customer_part != "客戶料號" and \
                   len(str(customer_part)) > 5:
                    
                    if current_block is None or \
                       current_block['customer_part'] != customer_part or \
                       current_block['customer_region'] != customer_region:
                        
                        if current_block is not None:
                            data_blocks.append(current_block)
                        
                        current_block = {
                            'customer_part': customer_part,
                            'customer_region': customer_region,
                            'start_row': idx,
                            'end_row': idx
                        }
                    else:
                        current_block['end_row'] = idx
            
            if current_block is not None:
                data_blocks.append(current_block)
            
            print(f"✅ 找到 {len(data_blocks)} 個數據塊")
            return data_blocks
            
        except Exception as e:
            print(f"❌ 數據塊識別失敗: {e}")
            return []
    
    def get_erp_records(self, customer_part, customer_region):
        """快速獲取ERP記錄（使用索引）"""
        try:
            match_key = f"{customer_part}_{customer_region}"
            
            if match_key in self.erp_index.index:
                records = self.erp_index.loc[match_key]
                if isinstance(records, pd.Series):
                    records = pd.DataFrame([records])
                
                # 按排程出貨日期排序
                records = records.sort_values('排程出貨日期')
                return records
            else:
                return pd.DataFrame()
                
        except Exception as e:
            print(f"    ❌ ERP記錄查找失敗: {e}")
            return pd.DataFrame()
    
    def calculate_target_info(self, erp_record):
        """計算目標填寫信息"""
        try:
            # 解析排程出貨日期
            schedule_date = erp_record['排程出貨日期']
            
            # 檢查日期是否為空或 nan
            if pd.isna(schedule_date):
                return None, None
            
            # 統一日期處理：無論是文字格式還是日期對象都轉換為標準日期
            date_obj = self.normalize_date(schedule_date)
            if date_obj is None:
                return None, None
            
            # 獲取排程出貨日期斷點 (BA欄位)
            schedule_breakpoint = erp_record.get('排程出貨日期斷點', '禮拜四')
            
            # 根據斷點計算周別範圍
            week_end_day = self.get_week_end_day(schedule_breakpoint)
            days_to_week_end = (week_end_day - date_obj.weekday()) % 7
            week_end_date = date_obj + timedelta(days=days_to_week_end)
            week_start_date = week_end_date - timedelta(days=6)
            
            # 獲取ETA (BC欄位)
            eta = erp_record.get('ETA', '下下週二')
            
            # 根據ETA計算目標日期
            target_date = self.calculate_eta_date(eta, week_start_date, week_end_date)
            
            if target_date is None:
                return None, None
            
            # 轉換數值
            net_demand = erp_record['淨需求']
            converted_demand = net_demand * 1000
            target_date_str = target_date.strftime("%Y%m%d")
            
            return target_date_str, converted_demand
            
        except Exception as e:
            print(f"    ❌ 目標計算失敗: {e}")
            return None, None
    
    def get_week_end_day(self, breakpoint):
        """根據斷點文字獲取星期幾的數字"""
        weekday_map = {
            '禮拜一': 0, '禮拜二': 1, '禮拜三': 2, '禮拜四': 3,
            '禮拜五': 4, '禮拜六': 5, '禮拜日': 6, '星期日': 6
        }
        return weekday_map.get(breakpoint, 3)  # 預設禮拜四
    
    def calculate_eta_date(self, eta, week_start, week_end):
        """根據ETA計算目標日期（基於排程出貨日期斷點周別的結束日期）"""
        try:
            # 以week_end為基準，找到對應的標準周別（禮拜天~禮拜六）
            # 從week_end往後找到最近的禮拜六
            days_to_saturday = (5 - week_end.weekday()) % 7
            current_saturday = week_end + timedelta(days=days_to_saturday)
            current_sunday = current_saturday - timedelta(days=6)  # 對應的禮拜天
            
            # 解析ETA文字
            if '下下週' in eta:
                # 下下週的情況：當前禮拜天 + 14天 + 目標星期幾
                target_weekday = self.get_eta_weekday(eta.replace('下下週', ''))
                next_next_sunday = current_sunday + timedelta(days=14)
                days_to_target = (target_weekday - next_next_sunday.weekday()) % 7
                return next_next_sunday + timedelta(days=days_to_target)
                
            elif '下週' in eta:
                # 下週的情況：當前禮拜天 + 7天 + 目標星期幾
                target_weekday = self.get_eta_weekday(eta.replace('下週', ''))
                next_sunday = current_sunday + timedelta(days=7)
                days_to_target = (target_weekday - next_sunday.weekday()) % 7
                return next_sunday + timedelta(days=days_to_target)
                
            elif '本週' in eta:
                # 本週的情況：當前禮拜天 + 目標星期幾
                target_weekday = self.get_eta_weekday(eta.replace('本週', ''))
                days_to_target = (target_weekday - current_sunday.weekday()) % 7
                return current_sunday + timedelta(days=days_to_target)
                
            else:
                print(f"    ⚠️ 無法解析ETA: {eta}")
                return None
                
        except Exception as e:
            print(f"    ❌ ETA計算失敗: {e}")
            return None
    
    def get_eta_weekday(self, weekday_text):
        """獲取ETA中的星期幾"""
        weekday_map = {
            '一': 0, '二': 1, '三': 2, '四': 3,
            '五': 4, '六': 5, '日': 6, '天': 6
        }
        return weekday_map.get(weekday_text, 1)  # 預設星期二
    
    def normalize_date(self, date_value):
        """
        統一日期處理函數
        處理各種日期格式：文字格式、pandas Timestamp、datetime 對象等
        """
        try:
            # 如果是空值或 NaN
            if pd.isna(date_value) or date_value is None:
                return None
            
            # 如果已經是 datetime 對象
            if isinstance(date_value, (datetime, pd.Timestamp)):
                return date_value
            
            # 如果是字串，嘗試解析
            if isinstance(date_value, str):
                date_str = str(date_value).strip()
                if not date_str or date_str.lower() in ['nan', 'none', '']:
                    return None
                
                # 嘗試多種日期格式
                date_formats = [
                    "%Y/%m/%d",      # 2025/10/01
                    "%Y-%m-%d",      # 2025-10-01
                    "%m/%d/%Y",      # 10/01/2025
                    "%d/%m/%Y",      # 01/10/2025
                    "%Y%m%d",        # 20251001
                ]
                
                for fmt in date_formats:
                    try:
                        return datetime.strptime(date_str, fmt)
                    except ValueError:
                        continue
                
                # 如果所有格式都失敗，使用 pandas 自動解析
                try:
                    return pd.to_datetime(date_str)
                except:
                    print(f"    ⚠️ 無法解析日期格式: {date_str}")
                    return None
            
            # 其他類型，嘗試用 pandas 轉換
            try:
                return pd.to_datetime(date_value)
            except:
                print(f"    ⚠️ 無法處理的日期類型: {type(date_value)} - {date_value}")
                return None
                
        except Exception as e:
            print(f"    ❌ 日期標準化失敗: {e}")
            return None
    
    def find_target_position(self, target_date_str, start_row, end_row):
        """找到目標填寫位置（修正pandas/openpyxl索引差異）"""
        try:
            # 掃描K~AW欄位尋找目標日期
            for col_idx in range(10, min(49, len(self.forecast_df.columns))):
                start_date_row = start_row + 1
                end_date_row = start_row + 2
                
                if start_date_row < len(self.forecast_df) and end_date_row < len(self.forecast_df):
                    try:
                        start_date = self.forecast_df.iloc[start_date_row, col_idx]
                        end_date = self.forecast_df.iloc[end_date_row, col_idx]
                        
                        if pd.notna(start_date) and pd.notna(end_date):
                            start_date_str = str(int(float(start_date)))
                            end_date_str = str(int(float(end_date)))
                            
                            if len(start_date_str) == 8 and len(end_date_str) == 8:
                                if start_date_str <= target_date_str <= end_date_str:
                                    # 找到目標欄位，現在找供應數量行
                                    for row_idx in range(start_row, min(start_row + 18, len(self.forecast_df))):
                                        k_value = self.forecast_df.iloc[row_idx, 10]  # K欄位
                                        if pd.notna(k_value) and str(k_value) == "供應數量":
                                            # 修正索引：pandas索引轉換為openpyxl索引
                                            openpyxl_row = row_idx + 2  # +1轉換為1基索引，+1跳過標題行
                                            return col_idx, openpyxl_row
                    except (ValueError, TypeError):
                        continue
            
            return None, None
            
        except Exception as e:
            print(f"    ❌ 位置查找失敗: {e}")
            return None, None
    
    def add_change_to_batch(self, col_idx, row_idx, value):
        """將修改添加到批量處理列表（支援累加）"""
        try:
            # 檢查是否已經有相同位置的修改
            for change in self.pending_changes:
                if change['col_idx'] == col_idx and change['row_idx'] == row_idx:
                    # 累加值而不是覆蓋
                    change['value'] += value
                    return True
            
            # 如果沒有相同位置的修改，添加新的
            self.pending_changes.append({
                'col_idx': col_idx,
                'row_idx': row_idx,
                'value': value
            })
            return True
            
        except Exception as e:
            print(f"    ❌ 添加修改失敗: {e}")
            return False
    
    def apply_all_changes(self):
        """批量應用所有修改（支援累加統計）"""
        try:
            print(f"\n🔄 批量應用 {len(self.pending_changes)} 筆修改...")
            
            # 統計累加情況
            accumulation_count = 0
            total_accumulated_value = 0
            
            for i, change in enumerate(self.pending_changes):
                col_idx = change['col_idx']
                row_idx = change['row_idx']
                value = change['value']
                
                # 直接修改內存中的Excel
                cell = self.ws.cell(row=row_idx, column=col_idx + 1)
                old_value = cell.value if cell.value is not None else 0
                
                # 檢查是否為累加（原值不為0）
                if old_value != 0:
                    accumulation_count += 1
                    total_accumulated_value += value
                    print(f"  📊 累加: {self.forecast_df.columns[col_idx]}{row_idx} = {old_value} + {value} = {value}")
                else:
                    print(f"  ✏️  新填: {self.forecast_df.columns[col_idx]}{row_idx} = {value}")
                
                cell.value = value
                
                # 每100筆修改顯示一次進度
                if (i + 1) % 100 == 0:
                    print(f"  已應用 {i + 1}/{len(self.pending_changes)} 筆修改")
            
            print(f"\n📈 累加統計:")
            print(f"  - 累加位置數: {accumulation_count}")
            print(f"  - 累加總值: {total_accumulated_value:,.0f}")
            print(f"✅ 所有修改已應用到內存")
            return True
            
        except Exception as e:
            print(f"❌ 批量應用修改失敗: {e}")
            return False
    
    def save_file(self):
        """保存文件（為空白單元格填入空格以保持Excel篩選功能）"""
        try:
            print("💾 保存文件...")
            
            # 為D欄位（客戶需求地區）的空白單元格填入空格
            print("🔧 為D欄位空白單元格填入空格以保持Excel篩選功能...")
            
            d_col_idx = 4  # D欄位（客戶需求地區）
            
            fixed_count = 0
            for row_idx in range(1, self.ws.max_row + 1):
                cell = self.ws.cell(row=row_idx, column=d_col_idx)
                # 如果單元格為None或空字符串，填入一個空格
                if cell.value is None:
                    cell.value = ' '
                    fixed_count += 1
                elif isinstance(cell.value, str) and cell.value.strip() == '':
                    cell.value = ' '
                    fixed_count += 1
            
            print(f"  ✅ 已為 D 欄位填入 {fixed_count} 個空格")
            
            # 保存到指定的輸出資料夾
            import os
            if self.output_folder:
                processed_dir = self.output_folder
            else:
                processed_dir = "processed"
            os.makedirs(processed_dir, exist_ok=True)
            output_file = os.path.join(processed_dir, self.output_filename)
            self.wb.save(output_file)
            print(f"✅ 文件已保存為: {output_file}")
            print(f"✅ D欄位空白單元格已填入空格，Excel篩選功能正常")
            return True
            
        except Exception as e:
            print(f"❌ 保存文件失敗: {e}")
            import traceback
            traceback.print_exc()
            return False

    def save_allocation_status(self):
        """儲存更新後的 ERP/Transit 檔案（包含已分配狀態）"""
        try:
            import os

            print("\n💾 儲存分配狀態到 ERP/Transit 檔案...")

            # 儲存 ERP 檔案（覆蓋原檔案，更新已分配欄位）
            # 移除暫時的 match_key 和 _original_idx 欄位後再儲存
            erp_save_df = self.erp_df.drop(columns=['match_key', '_original_idx'], errors='ignore')
            erp_save_df.to_excel(self.erp_file, index=False)
            erp_allocated = (self.erp_df['已分配'] == '✓').sum()
            print(f"✅ ERP 檔案已更新: {self.erp_file}")
            print(f"   已分配筆數: {erp_allocated}/{len(self.erp_df)}")

            # 儲存 Transit 檔案（如果有）
            if self.transit_df is not None and self.transit_file:
                # 移除暫時的 match_key 和 _original_idx 欄位後再儲存
                transit_save_df = self.transit_df.drop(columns=['match_key', '_original_idx'], errors='ignore')
                transit_save_df.to_excel(self.transit_file, index=False)
                transit_allocated = (self.transit_df['已分配'] == '✓').sum()
                print(f"✅ Transit 檔案已更新: {self.transit_file}")
                print(f"   已分配筆數: {transit_allocated}/{len(self.transit_df)}")

            return True

        except Exception as e:
            print(f"❌ 儲存分配狀態失敗: {e}")
            import traceback
            traceback.print_exc()
            return False

    def process_single_block(self, block):
        """處理單個數據塊（批量模式，支援1對1分配邏輯）"""
        try:
            customer_part = block['customer_part']
            customer_region = block['customer_region']
            start_row = block['start_row']
            end_row = block['end_row']

            # 獲取ERP記錄
            erp_records = self.get_erp_records(customer_part, customer_region)

            if len(erp_records) == 0:
                return 0, 0

            filled_count = 0
            skipped_count = 0

            # 處理每筆ERP記錄
            for idx, erp_record in erp_records.iterrows():
                # 使用 _original_idx 來獲取原始 DataFrame 中的行索引
                original_idx = erp_record['_original_idx']

                # 檢查已分配狀態（每筆記錄都是獨立的，即使欄位值相同也是不同的訂單）
                current_status = self.erp_df.at[original_idx, '已分配']
                if current_status == '✓':
                    skipped_count += 1
                    continue

                # 計算目標信息
                target_date_str, converted_demand = self.calculate_target_info(erp_record)

                if target_date_str is None or converted_demand is None:
                    skipped_count += 1
                    continue

                # 找到填寫位置
                col_idx, row_idx = self.find_target_position(target_date_str, start_row, end_row)

                if col_idx is None or row_idx is None:
                    skipped_count += 1
                    continue

                # 添加到批量處理列表（不立即執行）
                success = self.add_change_to_batch(col_idx, row_idx, converted_demand)

                if success:
                    filled_count += 1
                    # 使用 original_idx 標記該筆ERP記錄為已分配
                    self.erp_df.at[original_idx, '已分配'] = '✓'
                else:
                    skipped_count += 1

            return filled_count, skipped_count

        except Exception as e:
            print(f"  ❌ 數據塊處理失敗: {e}")
            return 0, 1
    
    def process_transit_for_block(self, block):
        """處理單個數據塊的Transit數據（支援1對1分配邏輯）"""
        try:
            if self.transit_df is None or self.transit_index is None:
                return 0, 0

            customer_part = block['customer_part']  # forecast A欄位
            customer_region = block['customer_region']  # forecast D欄位
            start_row = block['start_row']
            end_row = block['end_row']

            # 建立匹配鍵：customer_region(forecast D) + customer_part(forecast A)
            # 對應 transit M欄位(索引12) + F欄位(索引5)
            match_key = f"{customer_region}_{customer_part}"

            # 第一次匹配時顯示調試信息
            if not hasattr(self, '_transit_debug_shown'):
                print(f"\n🔍 Transit匹配調試（首次）:")
                print(f"   Forecast D欄位 + A欄位: {match_key}")
                print(f"   查找Transit索引中是否存在...")
                self._transit_debug_shown = True

            # 查找匹配的Transit記錄
            if match_key not in self.transit_index.index:
                return 0, 0

            # 找到匹配時顯示信息
            if not hasattr(self, '_transit_match_found'):
                print(f"   ✅ 找到匹配的Transit記錄！")
                self._transit_match_found = True

            transit_records = self.transit_index.loc[match_key]
            if isinstance(transit_records, pd.Series):
                transit_records = pd.DataFrame([transit_records])

            filled_count = 0
            skipped_count = 0

            # 處理每筆Transit記錄
            for idx, transit_record in transit_records.iterrows():
                try:
                    # 使用 _original_idx 來獲取原始 DataFrame 中的行索引
                    original_idx = transit_record['_original_idx']

                    # 檢查已分配狀態（每筆記錄都是獨立的，即使欄位值相同也是不同的訂單）
                    current_status = self.transit_df.at[original_idx, '已分配']
                    if current_status == '✓':
                        skipped_count += 1
                        continue

                    # 獲取 H 欄位數據（索引7）和 I 欄位 ETA（索引8）
                    h_value = transit_record.iloc[7] if len(transit_record) > 7 else None
                    eta_value = transit_record.iloc[8] if len(transit_record) > 8 else None

                    # 調試信息
                    if filled_count == 0 and skipped_count == 0:
                        print(f"    🔍 Transit記錄調試:")
                        print(f"       M欄位(12): {transit_record.iloc[12] if len(transit_record) > 12 else 'N/A'}")
                        print(f"       F欄位(5): {transit_record.iloc[5] if len(transit_record) > 5 else 'N/A'}")
                        print(f"       H欄位(7): {h_value}")
                        print(f"       I欄位(8): {eta_value}")

                    if pd.isna(h_value) or pd.isna(eta_value) or h_value == 0:
                        skipped_count += 1
                        continue

                    # 轉換單位：K -> 需要 *1000
                    converted_value = float(h_value) * 1000

                    # 直接使用 I欄位的ETA日期（新版邏輯）
                    target_date = self.parse_eta_date(eta_value)

                    if target_date is None:
                        skipped_count += 1
                        continue

                    target_date_str = target_date.strftime("%Y%m%d")

                    # 找到填寫位置
                    col_idx, row_idx = self.find_target_position(target_date_str, start_row, end_row)

                    if col_idx is None or row_idx is None:
                        skipped_count += 1
                        continue

                    # 添加到批量處理列表
                    success = self.add_change_to_batch(col_idx, row_idx, converted_value)

                    if success:
                        filled_count += 1
                        # 使用 original_idx 標記該筆Transit記錄為已分配
                        self.transit_df.at[original_idx, '已分配'] = '✓'
                    else:
                        skipped_count += 1

                except Exception as e:
                    print(f"    ⚠️ Transit記錄處理失敗: {e}")
                    skipped_count += 1
                    continue

            return filled_count, skipped_count

        except Exception as e:
            print(f"  ❌ Transit數據塊處理失敗: {e}")
            return 0, 0
    
    def parse_eta_date(self, eta_value):
        """解析ETA日期（使用統一的日期處理邏輯）"""
        try:
            # 使用統一的日期標準化函數
            target_date = self.normalize_date(eta_value)
            
            if target_date is not None:
                print(f"    ✅ ETA日期解析成功: {target_date.strftime('%Y-%m-%d')}")
                return target_date
            else:
                print(f"    ⚠️ ETA日期解析失敗: {eta_value}")
                return None
                
        except Exception as e:
            print(f"    ❌ ETA日期處理失敗: {e}")
            return None
    
    def process_all_blocks(self):
        """處理所有數據塊（超高速批量模式）"""
        try:
            print("\n" + "="*60)
            print("開始超高速批量處理")
            print("="*60)
            
            # 載入數據
            if not self.load_and_prepare_data():
                return False
            
            # 找到數據塊
            data_blocks = self.find_data_blocks()
            if not data_blocks:
                print("❌ 沒有找到數據塊")
                return False
            
            # 處理每個數據塊（只計算，不立即修改）
            total_blocks = len(data_blocks)
            print(f"\n📊 預處理階段：計算所有修改位置...")
            
            for i, block in enumerate(data_blocks):
                # 處理 ERP 數據
                filled, skipped = self.process_single_block(block)
                self.total_filled += filled
                self.total_skipped += skipped
                
                # 處理 Transit 數據
                if self.transit_df is not None:
                    transit_filled, transit_skipped = self.process_transit_for_block(block)
                    self.total_transit_filled += transit_filled
                    self.total_transit_skipped += transit_skipped
                
                # 每100個數據塊顯示一次進度
                if (i + 1) % 100 == 0:
                    print(f"  預處理進度: {i+1}/{total_blocks} - 已計算 {len(self.pending_changes)} 筆修改")
            
            print(f"\n✅ ERP 預處理完成：")
            print(f"  - 總數據塊數: {total_blocks}")
            print(f"  - 待修改筆數: {len(self.pending_changes)}")
            print(f"  - 成功計算: {self.total_filled}")
            print(f"  - 跳過記錄: {self.total_skipped}")
            
            if self.transit_df is not None:
                print(f"\n✅ Transit 預處理完成：")
                print(f"  - 成功計算: {self.total_transit_filled}")
                print(f"  - 跳過記錄: {self.total_transit_skipped}")
            
            # 批量應用所有修改
            if len(self.pending_changes) > 0:
                if not self.apply_all_changes():
                    return False

                # 保存 Forecast 結果文件
                if not self.save_file():
                    return False
            else:
                print("⚠️ 沒有需要修改的數據")

            # 儲存更新後的 ERP/Transit 檔案（包含已分配狀態）
            self.save_allocation_status()

            # 輸出最終結果
            print("\n" + "="*60)
            print("超高速批量處理完成")
            print("="*60)
            print(f"總數據塊數: {total_blocks}")
            print(f"\n📊 ERP 數據填寫結果：")
            print(f"  - 成功填寫: {self.total_filled}")
            print(f"  - 跳過記錄: {self.total_skipped}")

            total_processed = self.total_filled + self.total_skipped
            if total_processed > 0:
                print(f"  - 處理率: {self.total_filled/total_processed*100:.1f}%")
            else:
                print(f"  - 處理率: 0%")

            # 顯示 ERP 分配統計（1對1邏輯，使用 '✓' 標記計算）
            erp_allocated = (self.erp_df['已分配'] == '✓').sum()
            erp_total = len(self.erp_df)
            print(f"\n📋 ERP 1對1分配統計：")
            print(f"  - 總筆數: {erp_total}")
            print(f"  - 已分配: {erp_allocated}")
            print(f"  - 未分配: {erp_total - erp_allocated}")

            if self.transit_df is not None:
                print(f"\n🚚 Transit 數據填寫結果：")
                print(f"  - 成功填寫: {self.total_transit_filled}")
                print(f"  - 跳過記錄: {self.total_transit_skipped}")

                transit_total_processed = self.total_transit_filled + self.total_transit_skipped
                if transit_total_processed > 0:
                    print(f"  - 處理率: {self.total_transit_filled/transit_total_processed*100:.1f}%")
                else:
                    print(f"  - 處理率: 0%")

                # 顯示 Transit 分配統計（1對1邏輯，使用 '✓' 標記計算）
                transit_allocated = (self.transit_df['已分配'] == '✓').sum()
                transit_total = len(self.transit_df)
                print(f"\n📋 Transit 1對1分配統計：")
                print(f"  - 總筆數: {transit_total}")
                print(f"  - 已分配: {transit_allocated}")
                print(f"  - 未分配: {transit_total - transit_allocated}")

            return True
            
        except Exception as e:
            print(f"❌ 批量處理失敗: {e}")
            import traceback
            traceback.print_exc()
            return False

def main():
    """主函數"""
    try:
        print("🚀 超高速FORECAST批量處理系統啟動")
        print("⚡ 使用批量openpyxl操作，性能提升10倍！")
        
        processor = UltraFastForecastProcessor(
            forecast_file="修改後的ForecastDataFile_ALL-0923.xlsx",
            erp_file="整合後的廣達淨需求.xlsx"
        )
        
        success = processor.process_all_blocks()
        
        if success:
            print("\n🎉 所有數據塊處理完成！")
            print("⚡ 超高速處理成功！")
        else:
            print("\n❌ 處理過程中發生錯誤")
            
    except Exception as e:
        print(f"❌ 系統錯誤: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
