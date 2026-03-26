# -*- coding: utf-8 -*-
"""
LibreOffice Utils - 跨平台 Excel 處理工具
使用 LibreOffice 無頭模式處理 .xls/.xlsx 檔案，保留公式和格式
支援 Windows 和 Linux 環境
"""
import subprocess
import os
import sys
import shutil
import tempfile
import platform
from openpyxl import load_workbook


def get_libreoffice_path():
    """
    取得 LibreOffice 執行檔路徑
    支援 Windows 和 Linux
    """
    if platform.system() == 'Windows':
        # Windows 常見安裝路徑
        possible_paths = [
            r'C:\Program Files\LibreOffice\program\soffice.exe',
            r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',
            os.path.expandvars(r'%PROGRAMFILES%\LibreOffice\program\soffice.exe'),
            os.path.expandvars(r'%PROGRAMFILES(X86)%\LibreOffice\program\soffice.exe'),
        ]
        for path in possible_paths:
            if os.path.exists(path):
                return path
        # 嘗試從 PATH 環境變數找
        return 'soffice'
    else:
        # Linux
        possible_paths = [
            '/usr/bin/libreoffice',
            '/usr/bin/soffice',
            '/usr/local/bin/libreoffice',
            '/usr/local/bin/soffice',
            '/snap/bin/libreoffice',
        ]
        for path in possible_paths:
            if os.path.exists(path):
                return path
        return 'libreoffice'


def check_libreoffice_installed():
    """
    檢查 LibreOffice 是否已安裝
    """
    try:
        libreoffice = get_libreoffice_path()
        result = subprocess.run(
            [libreoffice, '--version'],
            capture_output=True,
            text=True,
            timeout=10
        )
        if result.returncode == 0:
            print(f"LibreOffice 已安裝: {result.stdout.strip()}")
            return True
    except Exception as e:
        print(f"檢查 LibreOffice 失敗: {e}")
    return False


def convert_xls_to_xlsx(input_path, output_dir=None):
    """
    使用 LibreOffice 將 .xls 轉換為 .xlsx（保留公式）

    Args:
        input_path: 輸入的 .xls 檔案路徑
        output_dir: 輸出目錄，預設為輸入檔案所在目錄

    Returns:
        輸出的 .xlsx 檔案路徑
    """
    if output_dir is None:
        output_dir = os.path.dirname(input_path)

    libreoffice = get_libreoffice_path()
    abs_input = os.path.abspath(input_path)
    abs_output_dir = os.path.abspath(output_dir)

    cmd = [
        libreoffice,
        '--headless',
        '--convert-to', 'xlsx',
        '--outdir', abs_output_dir,
        abs_input
    ]

    print(f"  執行轉換: {' '.join(cmd)}")
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)

    if result.returncode != 0:
        raise Exception(f"LibreOffice 轉換失敗: {result.stderr}")

    # 返回輸出檔案路徑
    basename = os.path.splitext(os.path.basename(input_path))[0]
    output_path = os.path.join(abs_output_dir, f"{basename}.xlsx")

    if not os.path.exists(output_path):
        raise Exception(f"轉換後的檔案不存在: {output_path}")

    print(f"  轉換完成: {output_path}")
    return output_path


def convert_xlsx_to_xls(input_path, output_dir=None):
    """
    使用 LibreOffice 將 .xlsx 轉換為 .xls（保留公式）

    Args:
        input_path: 輸入的 .xlsx 檔案路徑
        output_dir: 輸出目錄，預設為輸入檔案所在目錄

    Returns:
        輸出的 .xls 檔案路徑
    """
    if output_dir is None:
        output_dir = os.path.dirname(input_path)

    libreoffice = get_libreoffice_path()
    abs_input = os.path.abspath(input_path)
    abs_output_dir = os.path.abspath(output_dir)

    # 使用 xls:MS Excel 97 明確指定格式
    cmd = [
        libreoffice,
        '--headless',
        '--convert-to', 'xls:MS Excel 97',
        '--outdir', abs_output_dir,
        abs_input
    ]

    print(f"  執行轉換: {' '.join(cmd)}")
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)

    # LibreOffice 有時 returncode 為 0 但仍有輸出問題，檢查 stderr
    if result.stderr:
        print(f"  LibreOffice stderr: {result.stderr}")

    # 返回輸出檔案路徑
    basename = os.path.splitext(os.path.basename(input_path))[0]
    output_path = os.path.join(abs_output_dir, f"{basename}.xls")

    # 等待檔案生成（某些情況下可能需要短暫等待）
    import time
    for _ in range(10):
        if os.path.exists(output_path):
            break
        time.sleep(0.5)

    if not os.path.exists(output_path):
        # 列出目錄內容以便除錯
        print(f"  目錄內容: {os.listdir(abs_output_dir)}")
        raise Exception(f"轉換後的檔案不存在: {output_path}")

    print(f"  轉換完成: {output_path}")
    return output_path


def cleanup_xls_file_libreoffice(file_path, output_path, username):
    """
    清理 .xls 格式的檔案（使用 LibreOffice + openpyxl）
    完整保留格式、公式，只修改指定儲存格的值為 0

    流程：
    1. 使用 LibreOffice 將 .xls 轉換為 .xlsx
    2. 使用 openpyxl 進行清理
    3. 輸出為 .xlsx 格式（避免轉換回 .xls 可能失敗的問題）

    注意：輸出檔案會是 .xlsx 格式，即使 output_path 指定為 .xls

    Args:
        file_path: 輸入的 .xls 檔案路徑
        output_path: 輸出的檔案路徑（實際會輸出為 .xlsx）
        username: 用戶名稱（決定清理邏輯）

    Returns:
        清理的儲存格數量
    """
    print(f"  🔄 開始清理 .xls 檔案（使用 LibreOffice）...")
    print(f"  📂 輸入: {file_path}")
    # 確保輸出為 xlsx 格式
    output_path_xlsx = os.path.splitext(output_path)[0] + '.xlsx'
    print(f"  📂 輸出: {output_path_xlsx}")

    cleaned_count = 0
    temp_dir = tempfile.mkdtemp()

    try:
        # 步驟 1：轉換為 xlsx
        print(f"  📤 步驟 1: 轉換 .xls -> .xlsx...")
        xlsx_path = convert_xls_to_xlsx(file_path, temp_dir)

        # 步驟 2：使用 openpyxl 清理
        print(f"  🧹 步驟 2: 清理資料...")
        wb = load_workbook(xlsx_path)
        ws = wb.active

        max_row = ws.max_row
        max_col = ws.max_column
        print(f"  📊 檔案大小: {max_row} 行 x {max_col} 欄")

        if username == 'liteon':
            # ========== liteon 專屬清理邏輯 ==========
            # 指定讀取 Daily+Weekly+Monthly sheet
            # 清理條件：C欄(column 3) = "Commit" 時，清零 J~BY 欄(column 10~77)
            if 'Daily+Weekly+Monthly' in wb.sheetnames:
                ws = wb['Daily+Weekly+Monthly']
                max_row = ws.max_row
                max_col = ws.max_column
            for row_idx in range(1, max_row + 1):
                c_cell = ws.cell(row=row_idx, column=3)
                if c_cell.value and str(c_cell.value).strip() == "Commit":
                    for col_idx in range(10, min(78, max_col + 1)):  # J=10, BY=77
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value is not None and cell.value != 0 and cell.value != '':
                            cell.value = 0
                            cleaned_count += 1
        else:
            for row_idx in range(1, max_row + 1):
                if row_idx % 100 == 0:
                    print(f"    處理進度: {row_idx}/{max_row} 行...")

                if username == 'pegatron':
                    # 檢查M欄位（第13欄）是否為 "ETA QTY"
                    m_cell = ws.cell(row=row_idx, column=13)
                    if m_cell.value and str(m_cell.value).strip() == "ETA QTY":
                        # 清空N~DN欄位（第14欄到第118欄）設為 0
                        for col_idx in range(14, min(119, max_col + 1)):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            if cell.value is not None and cell.value != 0 and cell.value != '':
                                cell.value = 0
                                cleaned_count += 1
                else:
                    # quanta 清理邏輯
                    k_cell = ws.cell(row=row_idx, column=11)
                    if k_cell.value and str(k_cell.value) == "供應數量":
                        for col_idx in range(12, min(50, max_col + 1)):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            if cell.value != 0 and cell.value != '':
                                cell.value = 0
                                cleaned_count += 1

                    i_cell = ws.cell(row=row_idx, column=9)
                    if i_cell.value and "庫存數量" in str(i_cell.value):
                        if row_idx + 1 <= max_row:
                            next_cell = ws.cell(row=row_idx + 1, column=9)
                            if next_cell.value != 0 and next_cell.value != '':
                                next_cell.value = 0
                                cleaned_count += 1

        # 保存修改後的 xlsx
        cleaned_xlsx = os.path.join(temp_dir, 'cleaned_temp.xlsx')
        wb.save(cleaned_xlsx)
        wb.close()

        # 步驟 3：直接輸出 xlsx（避免轉換回 xls 可能失敗）
        print(f"  📥 步驟 3: 輸出 .xlsx 檔案...")
        shutil.copy2(cleaned_xlsx, output_path_xlsx)

        print(f"  ✅ 清理完成，共清理 {cleaned_count} 個儲存格")
        return cleaned_count

    except Exception as e:
        print(f"  ❌ 清理失敗: {e}")
        import traceback
        traceback.print_exc()
        raise e
    finally:
        # 清理暫存目錄
        try:
            shutil.rmtree(temp_dir)
        except:
            pass


def write_to_excel_libreoffice(forecast_file, updates, output_path):
    """
    使用 LibreOffice + openpyxl 寫入 Excel，保留格式和公式

    流程：
    1. 如果是 .xls，先轉換為 .xlsx
    2. 使用 openpyxl 進行更新
    3. 如果原檔是 .xls，轉換回 .xls

    Args:
        forecast_file: 原始 forecast 檔案路徑
        updates: 更新列表 [(row, col, value), ...]
        output_path: 輸出檔案路徑

    Returns:
        True if success, False otherwise
    """
    print(f"  🔄 開始更新 Excel 檔案（使用 LibreOffice）...")

    input_ext = os.path.splitext(forecast_file)[1].lower()
    output_ext = os.path.splitext(output_path)[1].lower()
    is_xls_input = (input_ext == '.xls')
    is_xls_output = (output_ext == '.xls')

    temp_dir = tempfile.mkdtemp()

    try:
        # 步驟 1：準備 xlsx 檔案進行編輯
        if is_xls_input:
            print(f"  📤 轉換輸入檔案 .xls -> .xlsx...")
            xlsx_path = convert_xls_to_xlsx(forecast_file, temp_dir)
        else:
            # 複製 xlsx 到暫存目錄
            xlsx_path = os.path.join(temp_dir, 'working.xlsx')
            shutil.copy2(forecast_file, xlsx_path)

        # 步驟 2：使用 openpyxl 更新
        print(f"  📝 更新 {len(updates)} 個儲存格...")
        wb = load_workbook(xlsx_path)
        ws = wb.active

        # 合併相同位置的值（累加）
        update_dict = {}
        for row, col, value in updates:
            key = (row, col)
            if key in update_dict:
                update_dict[key] += value
            else:
                update_dict[key] = value

        # 更新儲存格
        for (row, col), value in update_dict.items():
            cell = ws.cell(row=row, column=col)
            current_val = cell.value
            if current_val is None or current_val == '' or current_val == 0:
                cell.value = value
            else:
                try:
                    cell.value = float(current_val) + value
                except (ValueError, TypeError):
                    cell.value = value
            print(f"    更新 Row {row}, Col {col} = {cell.value}")

        # 保存 xlsx
        updated_xlsx = os.path.join(temp_dir, 'updated.xlsx')
        wb.save(updated_xlsx)
        wb.close()

        # 步驟 3：輸出最終檔案（統一輸出 xlsx 避免轉換問題）
        # 如果指定輸出為 .xls，改為輸出 .xlsx
        if is_xls_output:
            output_path_xlsx = os.path.splitext(output_path)[0] + '.xlsx'
            print(f"  📥 輸出為 .xlsx 格式（避免轉換失敗）: {output_path_xlsx}")
            shutil.copy2(updated_xlsx, output_path_xlsx)
            print(f"  ✅ 已輸出到: {output_path_xlsx}")
        else:
            shutil.copy2(updated_xlsx, output_path)
            print(f"  ✅ 已輸出到: {output_path}")

        return True

    except Exception as e:
        print(f"  ❌ 更新失敗: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        # 清理暫存目錄
        try:
            shutil.rmtree(temp_dir)
        except:
            pass


def recalculate_xlsx(file_path, output_path=None):
    """
    開啟 Excel 檔案並強制公式重新計算。
    Windows + Office → 使用 win32com (Excel COM)
    Linux / 無 Office → fallback 到 LibreOffice headless

    Args:
        file_path: 輸入 xlsx 檔案路徑
        output_path: 輸出路徑。若為 None，覆蓋原檔案。

    Returns:
        True if success, False otherwise
    """
    target = output_path or file_path
    abs_input = os.path.abspath(file_path)
    abs_target = os.path.abspath(target)

    # 優先嘗試 win32com (Windows + Office)
    if platform.system() == 'Windows':
        try:
            return _recalculate_with_excel_com(abs_input, abs_target)
        except ImportError:
            print("  win32com 不可用，改用 LibreOffice...")
        except Exception as e:
            print(f"  Excel COM 失敗: {e}，改用 LibreOffice...")

    # Fallback: LibreOffice headless
    return _recalculate_with_libreoffice(abs_input, abs_target)


def _recalculate_with_excel_com(abs_input, abs_target):
    """使用 Excel COM (win32com) 重算公式"""
    import win32com.client
    import pythoncom

    print(f"  🔄 重新計算公式 (Excel COM): {os.path.basename(abs_input)}")

    # 如果 input != target，先複製
    if abs_input != abs_target:
        shutil.copy2(abs_input, abs_target)

    pythoncom.CoInitialize()
    excel = None
    wb = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(abs_target)
        wb.Application.CalculateFull()  # 強制全部重算
        wb.Save()
        wb.Close(SaveChanges=False)
        wb = None

        print(f"  ✅ 公式重算完成 (Excel COM)")
        return True
    except Exception as e:
        print(f"  ❌ Excel COM 重算失敗: {e}")
        if wb:
            try:
                wb.Close(SaveChanges=False)
            except:
                pass
        raise
    finally:
        if excel:
            try:
                excel.Quit()
            except:
                pass
        pythoncom.CoUninitialize()


def _recalculate_with_libreoffice(abs_input, abs_target):
    """使用 LibreOffice headless 重算公式"""
    temp_dir = tempfile.mkdtemp()
    try:
        libreoffice = get_libreoffice_path()

        cmd = [
            libreoffice,
            '--headless',
            '--calc',
            '--convert-to', 'xlsx',
            '--outdir', temp_dir,
            abs_input
        ]

        print(f"  🔄 重新計算公式 (LibreOffice): {os.path.basename(abs_input)}")
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)

        basename = os.path.splitext(os.path.basename(abs_input))[0]
        recalc_path = os.path.join(temp_dir, f"{basename}.xlsx")

        if not os.path.exists(recalc_path):
            print(f"  ❌ LibreOffice 公式重算失敗: {result.stderr}")
            return False

        shutil.copy2(recalc_path, abs_target)
        print(f"  ✅ 公式重算完成 (LibreOffice)")
        return True

    except Exception as e:
        print(f"  ❌ LibreOffice 公式重算失敗: {e}")
        return False
    finally:
        try:
            shutil.rmtree(temp_dir)
        except:
            pass


def merge_excel_files_libreoffice(file_paths, output_path, skip_header=True):
    """
    使用 LibreOffice + openpyxl 合併多個 Excel 檔案，保留格式

    Args:
        file_paths: 要合併的檔案路徑列表
        output_path: 輸出檔案路徑
        skip_header: 是否跳過後續檔案的標題行（第一行）

    Returns:
        True if success, False otherwise
    """
    if not file_paths:
        return False

    print(f"  🔄 開始合併 {len(file_paths)} 個 Excel 檔案（使用 LibreOffice）...")

    output_ext = os.path.splitext(output_path)[1].lower()
    is_xls_output = (output_ext == '.xls')

    temp_dir = tempfile.mkdtemp()

    try:
        # 轉換所有檔案為 xlsx
        xlsx_files = []
        for i, file_path in enumerate(file_paths):
            ext = os.path.splitext(file_path)[1].lower()
            if ext == '.xls':
                print(f"    轉換檔案 {i+1}: {os.path.basename(file_path)}")
                xlsx_path = convert_xls_to_xlsx(file_path, temp_dir)
                # 重命名避免衝突
                new_path = os.path.join(temp_dir, f"file_{i}.xlsx")
                shutil.move(xlsx_path, new_path)
                xlsx_files.append(new_path)
            else:
                # 複製 xlsx
                new_path = os.path.join(temp_dir, f"file_{i}.xlsx")
                shutil.copy2(file_path, new_path)
                xlsx_files.append(new_path)

        # 使用第一個檔案作為基礎
        print(f"  📝 合併資料...")
        base_wb = load_workbook(xlsx_files[0])
        base_ws = base_wb.active

        # 合併其他檔案
        for i, xlsx_path in enumerate(xlsx_files[1:], start=2):
            src_wb = load_workbook(xlsx_path)
            src_ws = src_wb.active

            # 找到目標工作表的下一個空白行
            dest_start_row = base_ws.max_row + 1

            # 複製資料（跳過標題行）
            start_row = 2 if skip_header else 1
            rows_copied = 0
            row_offset = dest_start_row - start_row  # 計算行偏移量

            from openpyxl.formula.translate import Translator
            from openpyxl.utils import get_column_letter as gcl

            dest_row = dest_start_row
            for row_idx in range(start_row, src_ws.max_row + 1):
                for col_idx in range(1, src_ws.max_column + 1):
                    src_cell = src_ws.cell(row=row_idx, column=col_idx)
                    dest_cell = base_ws.cell(row=dest_row, column=col_idx)

                    # 公式平移：調整列號參照（如同 Excel 複製貼上）
                    if isinstance(src_cell.value, str) and src_cell.value.startswith('='):
                        col_letter = gcl(col_idx)
                        origin = f"{col_letter}{row_idx}"
                        dest_ref = f"{col_letter}{dest_row}"
                        try:
                            dest_cell.value = Translator(src_cell.value, origin=origin).translate_formula(dest_ref)
                        except Exception:
                            dest_cell.value = src_cell.value  # fallback: 原樣複製
                    else:
                        dest_cell.value = src_cell.value

                    # 複製樣式（如果有）
                    if src_cell.has_style:
                        dest_cell.font = src_cell.font.copy()
                        dest_cell.border = src_cell.border.copy()
                        dest_cell.fill = src_cell.fill.copy()
                        dest_cell.number_format = src_cell.number_format
                        dest_cell.protection = src_cell.protection.copy()
                        dest_cell.alignment = src_cell.alignment.copy()

                dest_row += 1
                rows_copied += 1

            # 複製合併儲存格
            from openpyxl.utils import get_column_letter
            merged_count = 0
            for merged_range in src_ws.merged_cells.ranges:
                # 只複製在複製範圍內的合併儲存格
                if merged_range.min_row >= start_row:
                    new_min_row = merged_range.min_row + row_offset
                    new_max_row = merged_range.max_row + row_offset
                    new_range = f"{get_column_letter(merged_range.min_col)}{new_min_row}:{get_column_letter(merged_range.max_col)}{new_max_row}"
                    base_ws.merge_cells(new_range)
                    merged_count += 1

            src_wb.close()
            print(f"    合併檔案 {i} 完成（{rows_copied} 行，{merged_count} 個合併儲存格）")

        # 保存合併後的 xlsx
        merged_xlsx = os.path.join(temp_dir, 'merged.xlsx')
        base_wb.save(merged_xlsx)
        base_wb.close()

        # 輸出最終檔案
        if is_xls_output:
            print(f"  📥 轉換輸出檔案 .xlsx -> .xls...")
            temp_xls = convert_xlsx_to_xls(merged_xlsx, temp_dir)
            shutil.copy2(temp_xls, output_path)
        else:
            shutil.copy2(merged_xlsx, output_path)

        print(f"  ✅ 合併完成: {output_path}")
        return True

    except Exception as e:
        print(f"  ❌ 合併失敗: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        # 清理暫存目錄
        try:
            shutil.rmtree(temp_dir)
        except:
            pass


# Windows 環境下的備用方案：嘗試使用 xlwings，失敗則使用 LibreOffice
def is_windows():
    """檢查是否為 Windows 環境"""
    return platform.system() == 'Windows'


def try_xlwings_or_libreoffice(func_name, *args, **kwargs):
    """
    優先嘗試 xlwings（Windows），失敗則使用 LibreOffice
    """
    if is_windows():
        try:
            import xlwings
            # 嘗試使用 xlwings
            print(f"  嘗試使用 xlwings...")
            # 這裡需要根據 func_name 調用對應的 xlwings 函數
            # 但為了簡化，我們直接使用 LibreOffice
        except ImportError:
            pass

    # 使用 LibreOffice
    print(f"  使用 LibreOffice...")
    if func_name == 'cleanup':
        return cleanup_xls_file_libreoffice(*args, **kwargs)
    elif func_name == 'write':
        return write_to_excel_libreoffice(*args, **kwargs)
    elif func_name == 'merge':
        return merge_excel_files_libreoffice(*args, **kwargs)


if __name__ == '__main__':
    # 測試 LibreOffice 是否安裝
    print("檢查 LibreOffice 安裝狀態...")
    if check_libreoffice_installed():
        print("✅ LibreOffice 已正確安裝")
    else:
        print("❌ LibreOffice 未安裝或無法執行")
        print("\n請安裝 LibreOffice:")
        print("  Ubuntu/Debian: sudo apt-get install libreoffice-calc")
        print("  CentOS/RHEL:   sudo yum install libreoffice-calc")
        print("  Windows:       下載 https://www.libreoffice.org/download/")
