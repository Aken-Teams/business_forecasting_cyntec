"""
台達 Buyer Forecast 統一 reader (PoC v2)

設計原則:
  1. 表頭關鍵字定位欄位 (不依賴位置)
  2. Marker 欄位用「資料驅動」偵測 (掃描每欄含 Demand/Supply/Balance 字樣的數量)
  3. 日期欄自動辨識 (YYYYMMDD/YYYY-MM-DD/MMM)
  4. 自動選 sheet (找含 PARTNO 與日期欄者)
  5. 找不到必要欄位 → raise ValueError，由上層決定 skip/abort
  6. 單 PLANT 檔 (file 中無 PLANT col) → 從檔名 fallback (需傳 plant_codes)
"""
import os
import re
import openpyxl
from datetime import datetime

# ============ 表頭關鍵字 (按優先序) ============
# 兩階段比對:
#   Phase A: 嚴格 = exact match (整個 cell 內容 == keyword, 大小寫不敏感)
#   Phase B: 寬鬆 = substring match
# 短/通用關鍵字 (PLANT, PN, ITEM) 只放在 EXACT，避免誤抓 'PLANT STOCK' 之類
HEADER_KEYWORDS_EXACT = {
    'partno':      ['PARTNO', 'PART NO', 'PART_NO', 'PART#', 'PN', 'PART NUMBER',
                    'CUSTOMER PART', 'CUSTOMER PARTNO', '料號',
                    'RAW MATERIAL(P/N)', 'RAW MATERIAL', 'P/N', 'PART P/N'],
    'plant':       ['PLANT', '廠區', '廠別', 'WAREHOUSE'],
    'vendor_part': ['VENDOR PARTNO', 'VENDOR PART', 'VENDOR P/N', 'VENDOR ITEM',
                    'VENDOR PN', '廠商料號', 'SUPPLIER PART'],
    'stock':       ['STOCK', '庫存', 'INV', 'INVENTORY', 'ON HAND', 'ONHAND',
                    'PLANT STOCK', 'TOTAL STOCK', 'STOCK QTY'],
    'on_way':      ['ON-WAY', 'ON WAY', 'ONWAY', 'OTW', 'ON THE WAY',
                    'IN-TRANSIT', 'IN TRANSIT', 'TRANSIT'],
}
HEADER_KEYWORDS_SUBSTR = {
    'partno':      ['PARTNO', 'PART NO', 'PART NUMBER', '料號', '物料號',
                    'RAW MATERIAL'],
    'plant':       ['廠區', '廠別'],
    'vendor_part': ['VENDOR PART', '廠商料號'],
    'stock':       ['STOCK', '庫存', 'INVENTORY'],
    'on_way':      ['ON-WAY', 'ONWAY', 'OTW', 'IN-TRANSIT', 'IN TRANSIT'],
}

# ============ Marker 值分類 (row 屬於 Demand/Supply/Balance/Skip) ============
# 優先序: skip > balance > supply > demand
# 原因: 'Net Demand' 含 'demand' 字樣但實為 balance；先比對更具體的 pattern
MARKER_PATTERNS = {
    'skip': [
        r'remark', r'firm\s*order', r'firmed\s*order', r'^shipment$',
        r'ship\s*in\s*transit', r'shipping\s*mode', r'header',
        r'^etd$', r'd-etd',
    ],
    'balance': [
        r'\bbalance\b', r'c-bal', r'c-net', r'3\s*balance', r'3\.\s*net',
        r'net\s*avail', r'^net$', r'po\s*balance', r'net\s*demand',
    ],
    'supply': [
        r'\bsupply\b', r'b-cfm', r'b-supply', r'2\s*supply', r'2\.\s*supply',
        r'vendor\s*cfm', r'confirm\s*ship', r'forecast\s*conf', r'^cfm$',
    ],
    'demand': [
        r'\bdemand\b', r'a-demand', r'1\s*demand', r'1\.\s*demand',
        r'gross\s*req', r'^request$', r'forecast\s*demand', r'req\s*ship',
    ],
}

# ============ 日期辨識 ============
DATE_PAT_8DIGIT = re.compile(r'^\s*(\d{8})\s*$')
DATE_PAT_HYPHEN = re.compile(r'^\s*(\d{4})[-/](\d{1,2})[-/](\d{1,2})\s*$')
DATE_PAT_MDY = re.compile(r'^\s*(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})\s*$')  # MM/DD/YY 或 MM/DD/YYYY
DATE_PAT_YEAR_MONTH = re.compile(r'^\s*(\d{4})[-/](JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\s*$')
MONTH_NAMES = {'JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC'}


def is_date_header(value):
    if value is None:
        return False
    if isinstance(value, datetime):
        return True
    s = str(value).strip().upper()
    if not s:
        return False
    if DATE_PAT_8DIGIT.match(s):
        return True
    if DATE_PAT_HYPHEN.match(s):
        return True
    if DATE_PAT_MDY.match(s):
        return True
    if DATE_PAT_YEAR_MONTH.match(s):
        return True
    if s in MONTH_NAMES:
        return True
    if 'PASSDUE' in s or 'PAST DUE' in s or 'PAST_DUE' in s or 'PASTDUE' in s:
        return True
    return False


# ============ 表頭定位 ============
def _read_header_row(ws, row_idx=1):
    """讀指定列。預設第 1 列。"""
    for row_values in ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True):
        return list(row_values)
    return []


def find_header_row(ws, max_scan=10):
    """掃描前 N 列, 找最像 header 的列 (含 PARTNO 關鍵字 + 日期欄)。
    回傳 (row_idx, header_values) 或 (None, None)。
    """
    best = None
    best_score = 0
    for r_idx, row_values in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
        h_list = list(row_values)
        if not any(v for v in h_list):
            continue
        found, headers = scan_headers(h_list)
        if 'partno' not in found:
            continue
        date_start = find_first_date_col(headers, h_list)
        if date_start is None:
            continue
        score = len(found) * 100 + len(headers)
        if score > best_score:
            best_score = score
            best = (r_idx, h_list)
    return best if best else (None, None)


def scan_headers(header_values):
    """
    掃描第 1 列，回傳 ({field_name: col_index}, {col: header_str_upper})
    兩階段比對:
      Phase A: exact match (整個 cell == kw)
      Phase B: substring match (kw in header)
    """
    headers = {}
    for idx, v in enumerate(header_values, start=1):
        if v is None:
            continue
        headers[idx] = str(v).strip().upper()

    found = {}

    # Phase A: exact match
    for field, keywords in HEADER_KEYWORDS_EXACT.items():
        for kw in keywords:
            kw_u = kw.upper()
            for c in sorted(headers.keys()):
                if headers[c] == kw_u:
                    if field not in found:
                        found[field] = c
                    break
            if field in found:
                break

    # Phase B: substring match (補上未找到的欄位)
    for field, keywords in HEADER_KEYWORDS_SUBSTR.items():
        if field in found:
            continue
        for kw in keywords:
            kw_u = kw.upper()
            for c in sorted(headers.keys()):
                if kw_u in headers[c]:
                    found[field] = c
                    break
            if field in found:
                break

    return found, headers


def find_first_date_col(headers, header_values):
    for c in sorted(headers.keys()):
        if c - 1 < len(header_values) and is_date_header(header_values[c - 1]):
            return c
    return None


def collect_date_cols(start_col, header_values):
    result = {}
    if start_col is None:
        return result
    for c in range(start_col, len(header_values) + 1):
        v = header_values[c - 1]
        if v is None:
            continue
        if isinstance(v, datetime):
            result[c] = v.strftime('%Y%m%d')
            continue
        s = str(v).strip().upper()
        if DATE_PAT_8DIGIT.match(s):
            result[c] = DATE_PAT_8DIGIT.match(s).group(1)
        elif DATE_PAT_HYPHEN.match(s):
            m = DATE_PAT_HYPHEN.match(s)
            result[c] = f"{m.group(1)}{int(m.group(2)):02d}{int(m.group(3)):02d}"
        elif DATE_PAT_MDY.match(s):
            m = DATE_PAT_MDY.match(s)
            mm, dd, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
            yyyy = yy if yy >= 1000 else (2000 + yy if yy < 70 else 1900 + yy)
            result[c] = f"{yyyy}{mm:02d}{dd:02d}"
        elif DATE_PAT_YEAR_MONTH.match(s):
            result[c] = DATE_PAT_YEAR_MONTH.match(s).group(2)
        elif s in MONTH_NAMES:
            result[c] = s
        elif 'PASSDUE' in s or 'PAST DUE' in s or 'PAST_DUE' in s or 'PASTDUE' in s:
            result[c] = 'PASSDUE'
    return result


# ============ Marker 分類 ============
def classify_marker(value):
    if value is None:
        return None
    s = str(value).strip().lower()
    if not s:
        return None
    for category, patterns in MARKER_PATTERNS.items():
        for pat in patterns:
            if re.search(pat, s, re.IGNORECASE):
                return category
    return None


# ============ Marker 欄位偵測 (資料驅動) ============
def find_marker_col(rows_sample):
    """
    掃描前 N 列所有欄位，找包含最多 demand/supply/balance 字樣的欄。
    rows_sample: list of tuples (iter_rows values_only 結果)
    """
    col_counts = {}
    for row_values in rows_sample:
        for idx, v in enumerate(row_values, start=1):
            cat = classify_marker(v)
            if cat in ('demand', 'supply', 'balance'):
                col_counts[idx] = col_counts.get(idx, 0) + 1
    if not col_counts:
        return None
    # 取出現次數最高且 >= 2 的欄 (避免單一 cell 誤判)
    best = max(col_counts, key=col_counts.get)
    if col_counts[best] < 2:
        return None
    return best


# ============ Sheet 自動選擇 ============
def find_valid_sheets(wb):
    """找出所有含 PARTNO + 日期欄的 sheet, 並過濾為 schema 一致的子集合 (用於多 sheet 合併)。

    策略:
      1. 收集所有合格 sheet (含 PARTNO+日期欄, header 可能在第 1~10 列)
      2. 取「找到欄位數 × max_row」最高為 best
      3. 其他 sheet 若 (partno_col 相同 AND date_cols 數量在 ±30% 範圍) → 一起納入
      4. 否則 → 視為次要資料 sheet, 不納入
    回傳 list of (sheet_name, header_row_idx)
    """
    candidates = []  # (sname, header_row_idx, found, date_count, max_row)
    for sname in wb.sheetnames:
        ws = wb[sname]
        if ws.max_row is None or ws.max_row < 2:
            continue
        h_row, header_values = find_header_row(ws)
        if header_values is None:
            continue
        found, headers = scan_headers(header_values)
        date_start = find_first_date_col(headers, header_values)
        date_col_map = collect_date_cols(date_start, header_values)
        if not date_col_map:
            continue
        candidates.append((sname, h_row, found, len(date_col_map), ws.max_row))

    if not candidates:
        return []
    if len(candidates) == 1:
        return [(candidates[0][0], candidates[0][1])]

    # 取分數最高 (找到欄位數 × max_row)
    candidates.sort(key=lambda x: (-len(x[2]), -x[4]))
    best = candidates[0]
    best_partno = best[2]['partno']
    best_dates = best[3]

    result = [(best[0], best[1])]
    for c in candidates[1:]:
        same_partno = c[2]['partno'] == best_partno
        ratio = c[3] / best_dates if best_dates else 0
        if same_partno and 0.7 <= ratio <= 1.3:
            result.append((c[0], c[1]))
    return result


def find_best_sheet(wb):
    sheets = find_valid_sheets(wb)
    return sheets[0][0] if sheets else None


# ============ 取值 ============
def _get_cell_value(row_values, col_idx):
    if col_idx is None or col_idx - 1 >= len(row_values):
        return None
    return row_values[col_idx - 1]


def _to_partno(v):
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _read_dates_from_row(row_values, date_col_map, valid_keys=None, conversions=None):
    """從一列中讀所有日期欄數值 → {date_str: num}

    Args:
        valid_keys: 若給定，只保留此 set 內的 date keys
        conversions: dict {raw_key: target_key 或 None (丟棄)}
    """
    result = {}
    for col, date_str in date_col_map.items():
        v = _get_cell_value(row_values, col)
        if v is None or v == '':
            continue
        try:
            n = float(v)
        except (TypeError, ValueError):
            continue
        if n == 0:
            continue
        key = date_str
        if conversions and key in conversions:
            key = conversions[key]
            if key is None:
                continue
        if valid_keys is not None and key not in valid_keys:
            continue
        result[key] = result.get(key, 0) + n
    return result


# ============ 檔名 PLANT 比對 ============
def _match_plant_in_filename(filepath, plant_codes):
    if not plant_codes:
        return None
    fn = os.path.splitext(os.path.basename(filepath))[0].upper()
    matched = [p for p in plant_codes if str(p).upper() in fn]
    matched.sort(key=len, reverse=True)
    return matched[0] if matched else None


# ============ 主入口 ============
def read_buyer_file(filepath, plant_codes=None, date_cols=None, conversions=None):
    """
    讀取 Buyer Forecast 檔案，回傳統一格式資料。

    Args:
        filepath: Buyer Forecast .xlsx 路徑
        plant_codes: mapping 表 PLANT 代碼清單 (用於檔內無 PLANT 欄時 fallback 檔名比對)
        date_cols: 主日期清單 (set/list)，若給定，只保留此清單內 date
        conversions: 日期轉換表 {raw: target 或 None}

    Returns:
        list of dict, 每個 dict 含 buyer/plant/part_no/vendor_part/stock/on_way/
        demand/supply/balance_override

    Raises:
        ValueError: 找不到必要欄位 (PARTNO 或日期欄)
    """
    valid_keys = set(date_cols) if date_cols is not None else None
    fname = os.path.basename(filepath)
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    sheet_specs = find_valid_sheets(wb)
    if not sheet_specs:
        wb.close()
        raise ValueError(f'{fname}: 找不到含 PARTNO 與日期欄的 sheet')

    buyer_label = os.path.splitext(fname)[0]
    all_results = []

    for sheet_name, header_row_idx in sheet_specs:
        ws = wb[sheet_name]
        header_values = _read_header_row(ws, header_row_idx)
        found, headers = scan_headers(header_values)
        date_start = find_first_date_col(headers, header_values)
        date_col_map = collect_date_cols(date_start, header_values)

        if not date_col_map:
            continue

        data_start_row = header_row_idx + 1

        # 抽前 50 列做 marker 偵測 (資料驅動)
        rows_sample = []
        cnt = 0
        for row_values in ws.iter_rows(min_row=data_start_row, values_only=True):
            rows_sample.append(row_values)
            cnt += 1
            if cnt >= 50:
                break
        marker_col = find_marker_col(rows_sample)

        # 檔名 PLANT fallback (適用於檔內無 PLANT col 的單 PLANT 檔)
        filename_plant = None
        if 'plant' not in found:
            filename_plant = _match_plant_in_filename(filepath, plant_codes)

        all_rows = list(ws.iter_rows(min_row=data_start_row, values_only=True))

        if marker_col:
            res = _read_multirow(all_rows, found, date_col_map, marker_col,
                                 buyer_label, filename_plant, valid_keys, conversions)
        else:
            res = _read_flat(all_rows, found, date_col_map,
                             buyer_label, filename_plant, valid_keys, conversions)
        all_results.extend(res)

    wb.close()
    return all_results


def _read_flat(all_rows, found, date_col_map, buyer_label, filename_plant,
               valid_keys=None, conversions=None):
    """單列格式: 每列一個料號 (全部視為 Demand)"""
    results = []
    partno_col = found['partno']
    plant_col = found.get('plant')
    vendor_col = found.get('vendor_part')
    stock_col = found.get('stock')
    onway_col = found.get('on_way')

    for row_values in all_rows:
        partno = _get_cell_value(row_values, partno_col)
        partno_str = _to_partno(partno)
        if not partno_str:
            continue

        # 過濾明顯的非料號 (純英文、太短、像表頭/註解)
        if not _looks_like_partno(partno_str):
            continue

        demand = _read_dates_from_row(row_values, date_col_map, valid_keys, conversions)

        plant_raw = _get_cell_value(row_values, plant_col)
        plant = str(plant_raw).strip() if plant_raw else (filename_plant or '')

        vendor = _get_cell_value(row_values, vendor_col)
        stock = _get_cell_value(row_values, stock_col)
        onway = _get_cell_value(row_values, onway_col)

        results.append({
            'buyer': buyer_label,
            'plant': plant,
            'part_no': partno_str,
            'vendor_part': str(vendor) if vendor else '',
            'stock': stock or 0,
            'on_way': onway,
            'demand': demand, 'supply': {},
            'balance_override': {},
        })
    return results


def _read_multirow(all_rows, found, date_col_map, marker_col,
                   buyer_label, filename_plant, valid_keys=None, conversions=None):
    """多列格式: 每料號有多列，按 marker 區分"""
    results = []
    partno_col = found['partno']
    plant_col = found.get('plant')
    vendor_col = found.get('vendor_part')
    stock_col = found.get('stock')
    onway_col = found.get('on_way')

    last_partno = None
    last_plant = ''
    last_vendor = ''
    last_stock = 0
    last_onway = None
    groups = {}  # (partno, plant) -> {demand, supply, balance, meta}
    order = []   # 保留首次出現順序

    for row_values in all_rows:
        partno_raw = _get_cell_value(row_values, partno_col)
        partno = _to_partno(partno_raw)
        marker = classify_marker(_get_cell_value(row_values, marker_col))

        # 新 partno → 更新 last_*
        if partno and _looks_like_partno(partno):
            last_partno = partno
            plant_raw = _get_cell_value(row_values, plant_col)
            if plant_raw:
                last_plant = str(plant_raw).strip()
            elif filename_plant:
                last_plant = filename_plant
            vendor = _get_cell_value(row_values, vendor_col)
            if vendor:
                last_vendor = str(vendor)
            stock = _get_cell_value(row_values, stock_col)
            if stock is not None:
                last_stock = stock
            onway = _get_cell_value(row_values, onway_col)
            if onway is not None:
                last_onway = onway

        if not last_partno or not marker or marker == 'skip':
            continue

        key = (last_partno, last_plant)
        if key not in groups:
            groups[key] = {
                'partno': last_partno, 'plant': last_plant,
                'vendor': last_vendor, 'stock': last_stock, 'onway': last_onway,
                'demand': {}, 'supply': {}, 'balance': {},
            }
            order.append(key)

        target = groups[key][marker]
        date_data = _read_dates_from_row(row_values, date_col_map, valid_keys, conversions)
        for d, v in date_data.items():
            target[d] = target.get(d, 0) + v

    for key in order:
        g = groups[key]
        results.append({
            'buyer': buyer_label,
            'plant': g['plant'] or '',
            'part_no': g['partno'],
            'vendor_part': g['vendor'] or '',
            'stock': g['stock'] or 0,
            'on_way': g['onway'],
            'demand': g['demand'],
            'supply': g['supply'],
            'balance_override': g['balance'],
        })
    return results


def _looks_like_partno(s):
    """判斷字串是否像料號 (排除註解/聯絡人/標題)"""
    if not s or len(s) < 4:
        return False
    s = s.strip()
    # 純中文/姓名 → 排除 (但允許含中文+數字)
    if all('\u4e00' <= c <= '\u9fff' or c.isspace() for c in s):
        return False
    # 純英文字母且 < 8 字 → 排除 (像 'Bobbi', 'PLANT' 等)
    if s.isalpha() and len(s) < 8:
        return False
    return True
